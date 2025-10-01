from django.core.management.base import BaseCommand
from django.utils.timezone import now
from django.db.models import Q
from vehicle_history.models import VINHistory
from supplies.models import TraceData
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime
from django.utils.timezone import is_naive

class Command(BaseCommand):
    help = "Экспорт отчёта по УУД в Excel (за выбранную дату)"

    def add_arguments(self, parser):
        parser.add_argument(
            "--date",
            type=str,
            help="Дата в формате YYYY-MM-DD (по умолчанию сегодня).",
        )

    @staticmethod
    def _parse_dt(dt_str):
        """Безопасный парсер ISO8601 в datetime (возвращает None при ошибке)."""
        if not dt_str:
            return None
        try:
            # Поддержка вариантов ...Z
            return datetime.fromisoformat(dt_str.replace("Z", "+00:00"))
        except Exception:
            return None

    @staticmethod
    def _date_or_none(dt):
        """Дата из datetime, если есть."""
        try:
            return dt.date()
        except Exception:
            return None

    @staticmethod
    def _status_as_of(extra, target_date):
        """
        Возвращает шаг, на котором машина находилась на конец target_date.
        Логика: шаг считается выполненным, если его *_at <= target_date.
        """
        d1 = Command._date_or_none(Command._parse_dt(extra.get("step1_at")))
        d2 = Command._date_or_none(Command._parse_dt(extra.get("step2_at")))
        d3 = Command._date_or_none(Command._parse_dt(extra.get("step3_at")))
        d4 = Command._date_or_none(Command._parse_dt(extra.get("step4_at")))

        # Проверяем от большего шага к меньшему
        if d4 and d4 <= target_date:
            return "done"
        if d3 and d3 <= target_date:
            return "step3"
        if d2 and d2 <= target_date:
            return "step2"
        if d1 and d1 <= target_date:
            return "step1"
        # Если даже step1 ещё не был на эту дату — возвращаем None (не учитываем)
        return None

    def handle(self, *args, **options):
        date_str = options.get("date")
        if date_str:
            try:
                target_date = datetime.strptime(date_str, "%Y-%m-%d").date()
            except ValueError:
                self.stderr.write(self.style.ERROR("❌ Неверный формат даты. Используй YYYY-MM-DD"))
                return
        else:
            target_date = now().date()

        # Результаты по шагам на конец выбранной даты
        grouped = {
            "step1": [],
            "step2": [],
            "step3": [],
            "done": [],
        }

        for vh in VINHistory.objects.all():
            history = vh.history or {}
            uud_data = history.get("УУД", {}).get("УУД", []) or []

            for entry in uud_data:
                # Фильтруем только записи, созданные в выбранную дату
                created_at = self._parse_dt(entry.get("created_at"))
                created_day = self._date_or_none(created_at)
                if created_day != target_date:
                    continue

                extra = entry.get("extra_data", {}) or {}
                status = self._status_as_of(extra, target_date)
                if not status:
                    # На выбранную дату машина ещё не была даже на step1 — пропускаем
                    continue

                vin = vh.vin
                # Получаем бренд/модель по последней TraceData
                trace = (
                    TraceData.objects
                    .filter(Q(vin_c__iexact=vin) | Q(vin_rk__iexact=vin))
                    .order_by("-date_added")
                    .first()
                )
                brand = getattr(trace, "brand", "") if trace else ""
                model = getattr(trace, "model", "") if trace else ""

                row = [vin, brand, model, target_date.isoformat()]

                grouped[status].append(row)

        # Создаём Excel
        wb = openpyxl.Workbook()
        ws_map = {
            "step1": wb.active,
            "step2": wb.create_sheet("Step2"),
            "step3": wb.create_sheet("Step3"),
            "done": wb.create_sheet("Done"),
        }
        ws_map["step1"].title = "Step1"

        headers = ["VIN", "Бренд", "Модель", "Дата"]

        for step, ws in ws_map.items():
            ws.append(headers)
            for row in grouped[step]:
                ws.append(row)

            # автоширина
            for col in ws.columns:
                max_len = 0
                col_idx = col[0].column
                for cell in col:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

        filename = f"uud_report_{target_date.isoformat()}.xlsx"
        wb.save(filename)
        self.stdout.write(self.style.SUCCESS(f"✅ Отчёт сохранён: {filename}"))
