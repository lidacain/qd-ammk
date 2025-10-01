from django.core.management.base import BaseCommand
from django.utils.timezone import now
from pathlib import Path
from collections import defaultdict
import json

DEBUG = False
def _dbg(*args):
    if DEBUG:
        print("[DBG]", *args)

# ===== Модели =====
from vehicle_history.models import VINHistory

try:
    from vehicle_history.models import AssemblyPassLog  # где лежит ваш класс
except Exception:
    AssemblyPassLog = None

try:
    from supplies.models import TraceData
except Exception:
    TraceData = None

# ===== Excel =====
import openpyxl
from openpyxl.utils import get_column_letter

# Варианты названий зоны/поста
ASSEMBLY_ZONE_CANDIDATES = {s.lower() for s in ["Цех сборки", "Сборка", "Assembly", "assembly"]}
DOC_POST_CANDIDATES     = {s.lower() for s in ["Документация", "Documentation", "Docs", "Документы"]}



def _iter_json(obj):
    if isinstance(obj, dict):
        yield obj
        for v in obj.values():
            yield from _iter_json(v)
    elif isinstance(obj, list):
        for v in obj:
            yield from _iter_json(v)


def _normalize_vin(v):
    return (v or "").strip().upper()


def _norm(s):
    return str(s or "").strip().lower()

def has_assembly(history_json) -> bool:
    if not history_json:
        return False
    if isinstance(history_json, dict):
        for key in history_json.keys():
            if _norm(key) in ASSEMBLY_ZONE_CANDIDATES and history_json.get(key):
                return True
    for d in _iter_json(history_json):
        if isinstance(d, dict):
            zone_val = _norm(d.get("zone") or d.get("area") or d.get("section") or d.get("zone_name") or d.get("zoneTitle"))
            if zone_val in ASSEMBLY_ZONE_CANDIDATES:
                return True
    return False

def has_documentation(history_json) -> bool:
    if not history_json:
        return False
    def _walk(o) -> bool:
        if isinstance(o, dict):
            # (A) ключи-кандидаты
            for k, v in o.items():
                if _norm(k) in DOC_POST_CANDIDATES and v:
                    return True
            # (B) поля post/name/title
            post_name = _norm(o.get("post") or o.get("name") or o.get("title") or o.get("post_name"))
            if post_name in DOC_POST_CANDIDATES:
                return True
            for v in o.values():
                if _walk(v):
                    return True
        elif isinstance(o, list):
            for it in o:
                if _walk(it):
                    return True
        return False
    return _walk(history_json)


def get_vin_field(obj):
    for fname in ("vin_number", "vin", "VIN", "vinNumber"):
        if hasattr(obj, fname):
            return getattr(obj, fname)
    try:
        h = getattr(obj, "history", None)
        if isinstance(h, dict):
            for d in _iter_json(h):
                if isinstance(d, dict) and "vin" in d:
                    return str(d["vin"]).strip()
    except Exception:
        pass
    return ""


def trace_lookup(vin: str):
    if not vin or TraceData is None:
        return "", ""
    try:
        td = (TraceData.objects.filter(vin_rk__iexact=vin).first()
              or TraceData.objects.filter(vin_c__iexact=vin).first())
        if td:
            brand = getattr(td, "brand", "") or getattr(td, "Brand", "") or ""
            model = getattr(td, "model", "") or getattr(td, "Model", "") or ""
            return str(brand), str(model)
    except Exception:
        pass
    return "", ""


def _trace_fields():
    """
    Возвращает словарь с реальными именами полей в TraceData:
    {'vin_number': <field or None>, 'vin': <field or None>, 'brand': <field>, 'model': <field>}
    Корректно читает список полей через TraceData._meta.get_fields().
    """
    opts = getattr(TraceData, "_meta", None)
    if not opts:
        return {"vin_number": None, "vin": None, "brand": None, "model": None}

    # Собираем имена полей модели (без попытки итерировать по _meta)
    field_names = {getattr(f, "name", None) for f in opts.get_fields()}
    field_names = {n for n in field_names if n}  # убрать None

    def pick(*cands):
        for c in cands:
            if c in field_names:
                return c
        return None

    _dbg("TraceData fields detected:", sorted(field_names))

    return {
        "vin_number": pick("vin_rk", "vin_number", "VIN_NUMBER", "vinNumber"),
        "vin":        pick("vin_c", "vin", "VIN"),
        "brand":      pick("brand", "Brand", "brand_name", "BrandName"),
        "model":      pick("model", "Model", "model_name", "ModelName"),
    }



def bulk_trace_lookup(vins):
    """
    Возвращает dict {VIN_UPPER -> (brand, model)} пакетно,
    учитывая разные варианты имён полей в TraceData.
    """
    out = {v: ("", "") for v in vins}
    if not vins or TraceData is None:
        return out

    f = _trace_fields()
    brand_f, model_f = f["brand"], f["model"]
    vin_num_f, vin_f = f["vin_number"], f["vin"]

    _dbg("Using fields ->",
         "vin_number:", vin_num_f, "vin:", vin_f, "brand:", brand_f, "model:", model_f)

    vins_list = list(vins)

    _dbg("Lookup VINs count:", len(vins_list))

    # ---- helpers for batched case-insensitive fetch ----
    from django.db.models import Q

    def _iter_chunks(seq, size=500):
        buf = []
        for x in seq:
            buf.append(x)
            if len(buf) >= size:
                yield buf
                buf = []
        if buf:
            yield buf

    def _fetch_case_insensitive(field_name, vin_list):
        """
        Возвращает генератор кортежей (vin_val, brand, model), найденных по field_name__iexact.
        Выполняет запросы батчами, чтобы избежать слишком длинных WHERE.
        """
        for chunk in _iter_chunks(vin_list, size=500):
            q = Q()
            for v in chunk:
                q |= Q(**{f"{field_name}__iexact": v})
            if not q:
                continue
            for rec in TraceData.objects.filter(q).values_list(field_name, brand_f, model_f):
                yield rec

    found1 = 0
    if vin_num_f:
        for vin_val, brand, model in _fetch_case_insensitive(vin_num_f, vins_list):
            out[_normalize_vin(vin_val)] = (str(brand or ""), str(model or ""))
            found1 += 1
        _dbg("Found by", vin_num_f, ":", found1)

    if vin_f:
        missing = [v for v, (b, m) in out.items() if not (b or m)]
        _dbg("Still missing after first pass:", len(missing))
        found2 = 0
        if missing:
            for vin_val, brand, model in _fetch_case_insensitive(vin_f, missing):
                out[_normalize_vin(vin_val)] = (str(brand or ""), str(model or ""))
                found2 += 1
        _dbg("Found by", vin_f, ":", found2)

    return out





class Command(BaseCommand):
    help = (
        "Экспорт VIN-ов, у которых есть следы в Цехе сборки (VINHistory) "
        "или есть скан TRIM IN (AssemblyPassLog), но НЕТ прохождения поста 'Документация' по VINHistory."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--output",
            type=str,
            default=f"no_docs_vins_{now().strftime('%Y%m%d_%H%M')}.xlsx",
            help="Имя Excel-файла для сохранения.",
        )
        parser.add_argument(
            "--debug",
            action="store_true",
            help="Печать диагностики поиска бренда/модели (TraceData поля, совпадения и т.п.).",
        )

    def handle(self, *args, **opts):
        out_path = Path(opts["output"]).resolve()

        global DEBUG
        DEBUG = bool(opts.get("debug"))

        # === 1) Один проход по VINHistory: определяем vin -> has_docs, has_assembly ===
        self.stdout.write("🔎 Сканирую VINHistory...")
        vin_has_docs = {}      # vin -> bool
        vin_has_assembly = {}  # vin -> bool

        qs = VINHistory.objects.all().only("id", "history")
        for vh in qs.iterator():
            vin = _normalize_vin(get_vin_field(vh))
            if not vin:
                continue

            hist = getattr(vh, "history", None)
            if isinstance(hist, str):
                try:
                    hist = json.loads(hist)
                except Exception:
                    hist = None

            vin_has_docs[vin] = vin_has_docs.get(vin, False) or has_documentation(hist)
            vin_has_assembly[vin] = vin_has_assembly.get(vin, False) or has_assembly(hist)

        _dbg("VINHistory parsed:",
             "docs:", sum(1 for v in vin_has_docs.values() if v),
             "assembly:", sum(1 for v in vin_has_assembly.values() if v))

        # === 2) VIN-ы из AssemblyPassLog (TRIM IN) ===
        vins_from_passlog = set()
        if AssemblyPassLog is not None:
            self.stdout.write("📥 Беру VIN-ы из AssemblyPassLog...")
            for v in AssemblyPassLog.objects.values_list("vin", flat=True).iterator():
                vins_from_passlog.add(_normalize_vin(v))
            _dbg("AssemblyPassLog VINs:", len(vins_from_passlog))
        else:
            self.stdout.write("⚠️ Модель AssemblyPassLog недоступна — пропускаю этот источник.")

        # === 3) Объединение источников ===
        vins_from_history_assembly = {vin for vin, has_a in vin_has_assembly.items() if has_a}
        candidate_vins = vins_from_history_assembly.union(vins_from_passlog)

        _dbg("Candidate VINs:", len(candidate_vins))

        # === 4) Фильтр: исключить тех, у кого по VINHistory есть Документация ===
        result_vins = []
        for vin in candidate_vins:
            if vin_has_docs.get(vin, False):
                continue  # прошёл Документацию — исключаем
            # если про vin ничего не знаем в vin_has_docs (нет VINHistory), значит Документации нет — включаем
            result_vins.append(vin)

        _dbg("Result (no docs) VINs:", len(result_vins))

        # === 5) Формируем строки с брендом/моделью ===
        rows = []
        trace_map = bulk_trace_lookup(result_vins)
        for vin in sorted(result_vins):
            brand, model = trace_map.get(vin, ("", ""))
            rows.append((vin, brand, model))

        _dbg("Sample row:", rows[0] if rows else ("<empty>", "", ""))

        # === 6) Пишем Excel ===
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "VIN без Документации"

        headers = ["№", "VIN", "Бренд", "Модель"]
        ws.append(headers)

        for i, (vin, brand, model) in enumerate(rows, start=1):
            ws.append([i, vin, brand, model])

        # Автоширина
        for col_idx in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for cell in ws[col_letter]:
                v = str(cell.value) if cell.value is not None else ""
                if len(v) > max_len:
                    max_len = len(v)
            ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

        wb.save(out_path)
        self.stdout.write(self.style.SUCCESS(
            f"✅ Готово! VIN-ов: {len(rows)}. Файл: {out_path}"
        ))
