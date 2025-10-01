from django.contrib import admin, messages
from .models import VINHistory, ContainerHistory, VINHistoryBackup, AssemblyPassLog, VESPassLog, TrimOutPassLog, VehicleIdentifiers
from django.utils.html import format_html
from django.utils.safestring import mark_safe
from django.http import HttpResponse
from django.urls import reverse
import csv
import json
from supplies.models import TraceData
from django.utils.timezone import now
# vehicle_history/admin.py
from django import forms
from django.contrib import admin
from django.contrib.admin.widgets import AdminSplitDateTime
from django.http import HttpResponse
import csv

from .models import TrimOutPassLog
from supplies.models import TraceData


class BrandFilter(admin.SimpleListFilter):
    title = "Бренд"
    parameter_name = "brand"

    def lookups(self, request, model_admin):
        brands = (
            TraceData.objects.values_list("brand", flat=True)
            .distinct().order_by("brand")
        )
        return [(b, b) for b in brands if b]

    def queryset(self, request, queryset):
        value = self.value()
        if value:
            vins = TraceData.objects.filter(brand=value).values_list("vin_rk", flat=True)
            return queryset.filter(vin__in=list(vins))
        return queryset


class ModelFilter(admin.SimpleListFilter):
    title = "Модель"
    parameter_name = "model"

    def lookups(self, request, model_admin):
        models = (
            TraceData.objects.values_list("model", flat=True)
            .distinct().order_by("model")
        )
        return [(m, m) for m in models if m]

    def queryset(self, request, queryset):
        value = self.value()
        if value:
            vins = TraceData.objects.filter(model=value).values_list("vin_rk", flat=True)
            return queryset.filter(vin__in=list(vins))
        return queryset

class VESOpenStatusFilter(admin.SimpleListFilter):
    title = "Статус VES"
    parameter_name = "ves_status"

    def lookups(self, request, model_admin):
        return [
            ("open", "Открытые (ожидают приёма)"),
            ("closed", "Закрытые (приняты)"),
        ]

    def queryset(self, request, queryset):
        value = self.value()
        if value == "open":
            return queryset.filter(received_at__isnull=True)
        if value == "closed":
            return queryset.filter(received_at__isnull=False)
        return queryset


@admin.register(AssemblyPassLog)
class AssemblyPassLogAdmin(admin.ModelAdmin):
    """
    Админка для счётчика проходов (VIN-сохранений):
    - показывает VIN, бренд/модель (из TraceData), кто сохранил и когда;
    - быстрый поиск по VIN;
    - фильтры по дате, бренду и модели;
    - экспорт CSV.
    """
    list_display = ("vin", 'line', "brand", "model_name", "saved_by", "scanned_at")
    search_fields = ("vin",)
    list_filter = (("scanned_at", admin.DateFieldListFilter), BrandFilter, ModelFilter)
    date_hierarchy = "scanned_at"
    ordering = ("-scanned_at",)
    list_per_page = 50
    save_on_top = True
    actions = ("export_csv",)

    @admin.display(description="Бренд")
    def brand(self, obj):
        td = TraceData.objects.filter(vin_rk=obj.vin).only("brand").first()
        return td.brand if td and td.brand else "–"

    @admin.display(description="Модель")
    def model_name(self, obj):
        td = TraceData.objects.filter(vin_rk=obj.vin).only("model").first()
        return td.model if td and td.model else "–"

    @admin.action(description="Экспортировать выборку в CSV")
    def export_csv(self, request, queryset):
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="assembly_pass_log.csv"'
        writer = csv.writer(response, delimiter=';')
        writer.writerow(["vin", "brand", "model", "saved_by", "scanned_at"])
        for o in queryset.select_related("saved_by"):
            td = TraceData.objects.filter(vin_rk=o.vin).only("brand", "model").first()
            brand = getattr(td, 'brand', '') or ''
            model = getattr(td, 'model', '') or ''
            writer.writerow([
                o.vin,
                brand,
                model,
                (o.saved_by.get_full_name() or o.saved_by.username) if o.saved_by else "",
                o.scanned_at.strftime("%Y-%m-%d %H:%M:%S") if o.scanned_at else "",
            ])
        return response


class TrimOutPassLogAdminForm(forms.ModelForm):
    class Meta:
        model = TrimOutPassLog
        fields = "__all__"
        widgets = {
            "scanned_at": AdminSplitDateTime(),  # удобный сплит-виджет
        }


@admin.register(TrimOutPassLog)
class TrimOutPassLogAdmin(admin.ModelAdmin):
    """
    Админка для TRIM OUT (VIN-сохранений):
    - показывает VIN, линию, бренд/модель (из TraceData), кто сохранил и когда;
    - быстрый поиск по VIN;
    - фильтры по дате, бренду и модели;
    - экспорт CSV;
    - можно изменять дату/время сохранения (scanned_at).
    """
    form = TrimOutPassLogAdminForm

    list_display = ("vin", "line", "brand", "model_name", "saved_by", "scanned_at")
    list_display_links = ("vin",)                       # клик по VIN открывает форму
    list_editable = ("scanned_at",)                     # редактирование прямо со списка (при желании)
    search_fields = ("vin",)
    list_filter = (("scanned_at", admin.DateFieldListFilter), BrandFilter, ModelFilter)
    date_hierarchy = "scanned_at"
    ordering = ("-scanned_at",)
    list_per_page = 50
    save_on_top = True
    actions = ("export_csv",)

    # (Опционально) разрешить правку scanned_at только суперпользователям:
    def get_readonly_fields(self, request, obj=None):
        ro = super().get_readonly_fields(request, obj)
        if request.user.is_superuser:
            return ro
        # для несуперпользователей делаем scanned_at только для чтения
        return (*ro, "scanned_at",)

    @admin.display(description="Бренд")
    def brand(self, obj):
        td = TraceData.objects.filter(vin_rk=obj.vin).only("brand").first()
        return td.brand if td and td.brand else "–"

    @admin.display(description="Модель")
    def model_name(self, obj):
        td = TraceData.objects.filter(vin_rk=obj.vin).only("model").first()
        return td.model if td and td.model else "–"

    @admin.action(description="Экспортировать выборку в CSV")
    def export_csv(self, request, queryset):
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="trimout_pass_log.csv"'
        writer = csv.writer(response, delimiter=';')
        writer.writerow(["vin", "brand", "model", "saved_by", "scanned_at"])
        for o in queryset.select_related("saved_by"):
            td = TraceData.objects.filter(vin_rk=o.vin).only("brand", "model").first()
            brand = getattr(td, 'brand', '') or ''
            model = getattr(td, 'model', '') or ''
            writer.writerow([
                o.vin,
                brand,
                model,
                (o.saved_by.get_full_name() or o.saved_by.username) if o.saved_by else "",
                o.scanned_at.strftime("%Y-%m-%d %H:%M:%S") if o.scanned_at else "",
            ])
        return response


@admin.register(VINHistory)
class VINHistoryAdmin(admin.ModelAdmin):
    list_display = ("vin", "brand", "model_name", "posts_summary", "inspections_total", "created_at", "updated_at")
    search_fields = ("vin",)
    readonly_fields = ("formatted_history",)
    list_filter = (("created_at", admin.DateFieldListFilter), ("updated_at", admin.DateFieldListFilter), BrandFilter, ModelFilter)
    date_hierarchy = "created_at"
    ordering = ("-created_at",)
    list_per_page = 50
    save_on_top = True
    actions = ("export_csv", "export_json",)

    class Media:
        js = [
            "https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.7.0/highlight.min.js",
        ]
        css = {
            "all": [
                "https://cdnjs.cloudflare.com/ajax/libs/highlight.js/11.7.0/styles/github-dark.min.css"
            ]
        }

    # === Computed columns ===
    @admin.display(description="Бренд")
    def brand(self, obj):
        td = TraceData.objects.filter(vin_rk=obj.vin).only("brand").first()
        return td.brand if td and td.brand else "–"

    @admin.display(description="Модель")
    def model_name(self, obj):
        td = TraceData.objects.filter(vin_rk=obj.vin).only("model").first()
        return td.model if td and td.model else "–"

    @admin.display(description="Посты")
    def posts_summary(self, obj):
        try:
            zones = obj.history or {}
            post_names = set()
            for zone_posts in zones.values():
                for post_name in zone_posts.keys():
                    post_names.add(post_name)
            if not post_names:
                return "–"
            return ", ".join(sorted(list(post_names))[:4]) + (" …" if len(post_names) > 4 else "")
        except Exception:
            return "–"

    @admin.display(description="Осмотров всего")
    def inspections_total(self, obj):
        try:
            total = 0
            for zone_posts in (obj.history or {}).values():
                for entries in zone_posts.values():
                    if isinstance(entries, list):
                        total += len(entries)
            return total
        except Exception:
            return 0

    def formatted_history(self, obj):
        try:
            pretty_json = json.dumps(obj.history, indent=4, ensure_ascii=False)
        except Exception as e:
            pretty_json = f"Ошибка форматирования JSON: {e}"

        return format_html(
            """
            <pre><code class="json">{}</code></pre>
            <script>
                document.addEventListener("DOMContentLoaded", function() {{
                    hljs.highlightAll();
                }});
            </script>
        """,
            mark_safe(pretty_json),
        )
    formatted_history.short_description = "История контроля"

    def changelist_view(self, request, extra_context=None):
        def _post_exists(history_dict, post_name: str) -> bool:
            try:
                zones = history_dict or {}
                for zone_posts in zones.values():
                    if isinstance(zone_posts, dict):
                        entries = zone_posts.get(post_name)
                        if isinstance(entries, list) and len(entries) > 0:
                            return True
                return False
            except Exception:
                return False

        def _uud_wip(history_dict) -> bool:
            """Возвращает True, если по истории видно, что машина находится на УУД (не завершено)."""
            try:
                zones = history_dict or {}
                uud_zone = zones.get("УУД", {})
                if not isinstance(uud_zone, dict):
                    return False
                entries = uud_zone.get("УУД") or []
                if not isinstance(entries, list) or not entries:
                    return False
                # считаем WIP, если нет явного завершения шага 4
                for e in entries:
                    extra = (e or {}).get("extra_data", {}) or {}
                    # если step4_at отсутствует или пустой — трактуем как незавершённый УУД
                    if not extra.get("step4_at"):
                        return True
                return False
            except Exception:
                return False

        vins_in_history = set(VINHistory.objects.values_list("vin", flat=True))
        vins_in_trace = set(TraceData.objects.values_list("vin_rk", flat=True))
        missing_vins = vins_in_history - vins_in_trace

        if missing_vins:
            sample = list(missing_vins)[:3]
            sample_text = ', '.join(sample)
            messages.warning(
                request,
                f"⚠️ В базе VINHistory найдено {len(missing_vins)} VIN-номеров, которых нет в TraceData. Примеры: {sample_text}"
            )

        # Проверка на неконсистентный статус:
        # 1) прошли «Документация», но числятся на УУД (незавершённый УУД)
        # 2) прошли «Документация», но числятся на VES с открытой передачей
        inconsistent_uud_vins = []
        inconsistent_ves_vins = []
        qs = VINHistory.objects.all().only("vin", "history")
        for o in qs:
            try:
                has_docs = _post_exists(o.history, "Документация")
                if not has_docs:
                    continue
                # УУД WIP
                if _uud_wip(o.history):
                    inconsistent_uud_vins.append(o.vin)
                # Открытая VES-передача
                if VESPassLog.objects.filter(vin=o.vin, received_at__isnull=True).exists():
                    inconsistent_ves_vins.append(o.vin)
            except Exception:
                continue

        if inconsistent_uud_vins:
            sample_uud = ", ".join(inconsistent_uud_vins[:3])
            messages.warning(
                request,
                (
                    f"⚠️ VIN-номера ({len(inconsistent_uud_vins)} шт.) прошли пост «Документация», "
                    f"но по статусу числятся на УУД (незавершено). Примеры: {sample_uud}"
                ),
            )

        if inconsistent_ves_vins:
            sample_ves = ", ".join(inconsistent_ves_vins[:3])
            messages.warning(
                request,
                (
                    f"⚠️ VIN-номера ({len(inconsistent_ves_vins)} шт.) прошли пост «Документация», "
                    f"но имеют открытую передачу на VES. Примеры: {sample_ves}"
                ),
            )

        return super().changelist_view(request, extra_context=extra_context)

    # === Actions ===
    @admin.action(description="Экспортировать выборку в CSV")
    def export_csv(self, request, queryset):
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="vin_history.csv"'
        writer = csv.writer(response, delimiter=';')
        writer.writerow(["vin", "brand", "model", "inspections_total", "created_at", "updated_at"])
        for o in queryset:
            td = TraceData.objects.filter(vin_rk=o.vin).only("brand", "model").first()
            brand = getattr(td, 'brand', '') or ''
            model = getattr(td, 'model', '') or ''
            writer.writerow([
                o.vin,
                brand,
                model,
                self.inspections_total(o),
                o.created_at.strftime("%Y-%m-%d %H:%M:%S") if o.created_at else "",
                o.updated_at.strftime("%Y-%m-%d %H:%M:%S") if o.updated_at else "",
            ])
        return response

    @admin.action(description="Экспортировать выборку в JSON")
    def export_json(self, request, queryset):
        payload = []
        for o in queryset:
            item = {
                "vin": o.vin,
                "created_at": o.created_at.isoformat() if o.created_at else None,
                "updated_at": o.updated_at.isoformat() if o.updated_at else None,
                "history": o.history,
            }
            td = TraceData.objects.filter(vin_rk=o.vin).only("brand", "model").first()
            if td:
                item["brand"] = td.brand
                item["model"] = td.model
            payload.append(item)
        data = json.dumps(payload, ensure_ascii=False, indent=2)
        response = HttpResponse(data, content_type="application/json; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="vin_history.json"'
        return response


@admin.register(ContainerHistory)
class ContainerHistoryAdmin(admin.ModelAdmin):
    list_display = ("container_number", "records_count", "created_at", "updated_at")
    search_fields = ("container_number",)
    list_filter = (("created_at", admin.DateFieldListFilter), ("updated_at", admin.DateFieldListFilter))
    ordering = ("-created_at",)
    list_per_page = 50
    save_on_top = True
    actions = ("export_csv", 'export_excel')

    @admin.display(description="Записей")
    def records_count(self, obj):
        try:
            total = 0
            for zone_posts in (obj.history or {}).values():
                for entries in zone_posts.values():
                    if isinstance(entries, list):
                        total += len(entries)
            return total
        except Exception:
            return 0

    @admin.action(description="Экспортировать выборку в CSV")
    def export_csv(self, request, queryset):
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="container_history.csv"'
        writer = csv.writer(response, delimiter=';')
        writer.writerow(["container_number", "records_count", "created_at", "updated_at"])
        for o in queryset:
            writer.writerow([
                o.container_number,
                self.records_count(o),
                o.created_at.strftime("%Y-%m-%d %H:%M:%S") if o.created_at else "",
                o.updated_at.strftime("%Y-%m-%d %H:%M:%S") if o.updated_at else "",
            ])
        return response

    @admin.action(description="Экспортировать выборку в Excel (с фото в строку)")
    def export_excel(self, request, queryset):
        from collections import defaultdict
        import os
        from django.http import HttpResponse
        from django.conf import settings
        from django.utils.dateparse import parse_datetime
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
        from openpyxl.utils import get_column_letter
        from openpyxl.drawing.image import Image as XLImage

        # --------- Утилиты ---------
        def resolve_media_path(p: str) -> str:
            """Преобразует относительный путь из JSON в абсолютный путь в MEDIA_ROOT."""
            if not p:
                return ""
            s = str(p).replace("\\", "/")
            if s.startswith("/media/"):
                s = s[len("/media/"):]
            elif s.startswith("media/"):
                s = s[len("media/"):]
            if os.path.isabs(s):
                return s
            return os.path.join(settings.MEDIA_ROOT, s.lstrip("/"))

        def insert_image_fit(ws, row_idx: int, col_idx: int, path: str,
                             max_w_px=300, max_h_px=170):
            """Вставляет картинку в ячейку и масштабирует под заданные размеры (по большей стороне)."""
            try:
                if not (path and os.path.exists(path)):
                    return
                img = XLImage(path)
                if img.width and img.height:
                    k = min(max_w_px / img.width, max_h_px / img.height, 1.0)
                    img.width = img.width * k
                    img.height = img.height * k
                ws.add_image(img, f"{get_column_letter(col_idx)}{row_idx}")
            except Exception:
                # не валим экспорт из-за одного изображения
                pass

        def norm_dt(raw):
            """Безопасный парсер даты из JSON (date_added|date). Возвращает строку 'ДД.ММ.ГГГГ ЧЧ:ММ'."""
            if not raw:
                return ""
            dt = parse_datetime(raw)
            if not dt:
                return ""
            # если timezone-aware, можно снять tzinfo или форматировать как есть — оставим просто локальный формат
            return dt.strftime("%d.%m.%Y %H:%M")

        # --------- Сбор данных (как в export_assembly_excel: строки по дефекту/осмотру) ---------
        # Структурируем: container_number -> список записей
        grouped = defaultdict(list)
        max_photos = 0

        for obj in queryset:
            container_number = (obj.container_number or "").strip()
            history = obj.history or {}

            # ожидаем структуру как: { "Зона": { "Пост": [ entries... ] }, ... }
            for zone, posts in history.items():
                if not isinstance(posts, dict):
                    continue
                for post_name, entries in posts.items():
                    if not isinstance(entries, list):
                        continue
                    for e in entries:
                        # поля, которые могут быть в JSON (подстраиваемся под ваши формы)
                        raw_date = e.get("date_added") or e.get("date")
                        date_str = norm_dt(raw_date)
                        controller = (
                                e.get("controller")
                                or (e.get("extra_data") or {}).get("controller")
                                or ""
                        )
                        # Статус контейнера: "поврежден/не поврежден" или что-то подобное в e["status"]
                        status = e.get("status", "")
                        description = (e.get("description") or e.get("comment") or "").strip()

                        # Фото: могут быть в e["photos"] / e["defect_photos"]
                        photos_raw = e.get("photos") or e.get("defect_photos") or []
                        photos_abs = []
                        for ph in photos_raw:
                            p = resolve_media_path(ph)
                            if p and os.path.exists(p):
                                photos_abs.append(p)

                        max_photos = max(max_photos, len(photos_abs))
                        grouped[container_number].append({
                            "date": date_str,
                            "zone": zone or "",
                            "post": post_name or "",
                            "controller": controller,
                            "status": status,
                            "description": description,
                            "photos": photos_abs,
                        })

        # --------- Готовим Excel ---------
        wb = Workbook()
        ws = wb.active
        ws.title = "Контейнеры"

        # Заголовки как в «сборке», но адаптированы под контейнеры
        headers = [
                      "Контейнер",  # аналог VIN-колонки (будем объединять для блока)
                      "Дата",
                      "Зона",
                      "Пост",
                      "Контроллер",
                      "Статус",
                      "Комментарий",
                  ] + [f"Фото {i + 1}" for i in range(max_photos)]

        for i, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Ширины столбцов (настроены под фото в строку)
        widths = {
            1: 24,  # Контейнер
            2: 18,  # Дата
            3: 24,  # Зона
            4: 28,  # Пост
            5: 20,  # Контроллер
            6: 20,  # Статус
            7: 45,  # Комментарий
        }
        for idx, w in widths.items():
            ws.column_dimensions[get_column_letter(idx)].width = w
        # фото-колонки
        for i in range(max_photos):
            ws.column_dimensions[get_column_letter(8 + i)].width = 32

        # Высота строки и целевые размеры изображений
        ROW_H_PT = 95  # ~высота строки (по опыту: 90–110 pt)
        IMG_W_PX = 300
        IMG_H_PX = 170

        # Заполняем построчно
        for container_number, entries in grouped.items():
            if not entries:
                continue

            start_row = ws.max_row + 1

            for item in entries:
                ws.append([
                              container_number,
                              item["date"],
                              item["zone"],
                              item["post"],
                              item["controller"],
                              item["status"],
                              item["description"],
                          ] + [""] * max_photos)

                r = ws.max_row
                ws.row_dimensions[r].height = ROW_H_PT

                # Фото по строке вправо
                for j, path in enumerate(item["photos"]):
                    insert_image_fit(ws, r, 8 + j, path, IMG_W_PX, IMG_H_PX)

            # Объединяем ячейки "Контейнер" для блока записей по одному контейнеру
            if len(entries) > 1:
                end_row = ws.max_row
                ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
                top_cell = ws.cell(row=start_row, column=1)
                top_cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Выравнивание всего листа (кроме заголовка уже сделано)
        for col in ws.columns:
            for cell in col:
                if cell.row == 1:
                    continue
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Ответ
        resp = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        resp["Content-Disposition"] = 'attachment; filename="container_history.xlsx"'
        wb.save(resp)
        return resp


@admin.register(VINHistoryBackup)
class VINHistoryBackupAdmin(admin.ModelAdmin):
    list_display = ("vin", "action_badge", "deleted_at")
    search_fields = ("vin",)
    list_filter = ("action", ("deleted_at", admin.DateFieldListFilter))
    ordering = ("-deleted_at",)
    date_hierarchy = "deleted_at"
    list_per_page = 50
    save_on_top = True
    actions = ("export_csv",)

    @admin.display(description="Действие")
    def action_badge(self, obj):
        color = "#ef4444" if str(obj.action).lower() == "delete" else "#6b7280"
        return format_html('<span style="padding:2px 8px;border-radius:999px;border:1px solid {0};color:{0}">{1}</span>', color, obj.action)

    def has_add_permission(self, request):
        return False

    def has_delete_permission(self, request, obj=None):
        return False

    @admin.action(description="Экспортировать выборку в CSV")
    def export_csv(self, request, queryset):
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="vin_history_backup.csv"'
        writer = csv.writer(response, delimiter=';')
        writer.writerow(["vin", "action", "deleted_at"])
        for o in queryset:
            writer.writerow([
                o.vin,
                o.action,
                o.deleted_at.strftime("%Y-%m-%d %H:%M:%S") if o.deleted_at else "",
            ])
        return response


@admin.register(VESPassLog)
class VESPassLogAdmin(admin.ModelAdmin):
    """
    Админка для VES-передач/приёмов:
    - показывает VIN, бренд/модель (из TraceData), кто отдал/когда, кто принял/когда;
    - фильтры по дате, бренду/модели, статусу (открыт/закрыт);
    - быстрый поиск по VIN и пользователям;
    - действия: экспорт CSV, «Отметить как принят сейчас», «Снять приём».
    """
    list_display = (
        "vin", "brand", "model_name",
        "given_by", "given_at",
        "received_by", "received_at",
        "status_badge", "duration_fmt",
    )
    search_fields = ("vin", "given_by__username", "received_by__username", "given_by__first_name", "given_by__last_name", "received_by__first_name", "received_by__last_name")
    list_filter = (
        ("given_at", admin.DateFieldListFilter),
        ("received_at", admin.DateFieldListFilter),
        BrandFilter,
        ModelFilter,
        VESOpenStatusFilter,
    )
    date_hierarchy = "given_at"
    ordering = ("-given_at",)
    list_per_page = 50
    save_on_top = True
    actions = ("export_csv", "mark_received_now", "clear_received")

    # === Computed columns ===
    @admin.display(description="Бренд")
    def brand(self, obj):
        td = TraceData.objects.filter(vin_rk=obj.vin).only("brand").first()
        return td.brand if td and td.brand else "–"

    @admin.display(description="Модель")
    def model_name(self, obj):
        td = TraceData.objects.filter(vin_rk=obj.vin).only("model").first()
        return td.model if td and td.model else "–"

    @admin.display(description="Статус")
    def status_badge(self, obj):
        if obj.received_at:
            return format_html('<span class="badge bg-success">принят</span>')
        return format_html('<span class="badge bg-warning text-dark">в ожидании</span>')

    @admin.display(description="Время на VES")
    def duration_fmt(self, obj):
        if not obj.given_at:
            return "–"
        end_dt = obj.received_at or now()
        total = int((end_dt - obj.given_at).total_seconds())
        if total < 60:
            return f"{total}s"
        mins, secs = divmod(total, 60)
        if mins < 60:
            return f"{mins}m {secs}s"
        hours, mins = divmod(mins, 60)
        return f"{hours}h {mins}m"

    # === Actions ===
    @admin.action(description="Экспортировать выборку в CSV")
    def export_csv(self, request, queryset):
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = 'attachment; filename="ves_pass_log.csv"'
        writer = csv.writer(response, delimiter=';')
        writer.writerow(["vin", "brand", "model", "given_by", "given_at", "received_by", "received_at", "status", "duration_seconds"])
        for o in queryset.select_related("given_by", "received_by"):
            td = TraceData.objects.filter(vin_rk=o.vin).only("brand", "model").first()
            brand = getattr(td, "brand", "") or ""
            model = getattr(td, "model", "") or ""
            status = "closed" if o.received_at else "open"
            duration = ""
            try:
                end_dt = o.received_at or now()
                duration = int((end_dt - o.given_at).total_seconds()) if o.given_at else ""
            except Exception:
                duration = ""
            writer.writerow([
                o.vin,
                brand,
                model,
                (o.given_by.get_full_name() or o.given_by.username) if o.given_by else "",
                o.given_at.strftime("%Y-%m-%d %H:%M:%S") if o.given_at else "",
                (o.received_by.get_full_name() or o.received_by.username) if o.received_by else "",
                o.received_at.strftime("%Y-%m-%d %H:%M:%S") if o.received_at else "",
                status,
                duration,
            ])
        return response

    @admin.action(description="Отметить как принят сейчас")
    def mark_received_now(self, request, queryset):
        updated = 0
        for o in queryset.filter(received_at__isnull=True):
            o.received_at = now()
            o.received_by = request.user if not o.received_by else o.received_by
            o.save(update_fields=["received_at", "received_by", "updated_at"])
            updated += 1
        if updated:
            self.message_user(request, f"✅ Обновлено записей: {updated}", messages.SUCCESS)
        else:
            self.message_user(request, "Нет открытых записей для обновления.", messages.WARNING)

    @admin.action(description="Снять приём (сделать открытой)")
    def clear_received(self, request, queryset):
        cleared = 0
        for o in queryset.filter(received_at__isnull=False):
            o.received_at = None
            o.received_by = None
            o.save(update_fields=["received_at", "received_by", "updated_at"])
            cleared += 1
        if cleared:
            self.message_user(request, f"♻️ Снято признаков приёма: {cleared}", messages.SUCCESS)
        else:
            self.message_user(request, "Нет закрытых записей для очистки.", messages.WARNING)


# uud1234!

@admin.register(VehicleIdentifiers)
class VehicleIdentifiersAdmin(admin.ModelAdmin):
    list_display = (
        "vin",
        "engine_number",
        "transmission_number",
        "saved_by",
        "created_at",
        "updated_at",
    )
    list_display_links = ("vin",)
    search_fields = ("vin", "engine_number", "transmission_number")
    list_filter = ("saved_by", "created_at", "updated_at")
    readonly_fields = ("created_at", "updated_at", "saved_by")
    ordering = ("-updated_at",)
    date_hierarchy = "created_at"
    list_per_page = 50

    def save_model(self, request, obj, form, change):
        # Автозаполнение "кто сохранил" текущим пользователем (если не задано)
        if not obj.saved_by:
            obj.saved_by = request.user
        super().save_model(request, obj, form, change)

