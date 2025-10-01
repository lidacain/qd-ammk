# views.py
import os
from django.http import HttpResponse
from django.conf import settings
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from .models import VINHistory, ContainerHistory
from collections import defaultdict
from PIL import Image as PILImage
from io import BytesIO
import json
from django.utils.dateparse import parse_date
from django.utils.timezone import make_aware
from datetime import datetime
import io
import zipfile
from django.contrib.auth.decorators import login_required
from users.decorators import role_required
from django.utils.dateparse import parse_datetime
from supplies.models import TraceData


def export_all_vin_histories_to_excel(request):
    wb = Workbook()
    wb.remove(wb.active)

    all_histories = VINHistory.objects.all()
    post_data = defaultdict(list)

    image_fields = ["photos", "defect_photos", "body_photos", "vin_photo", "engine_photo", "component_photos"]

    for obj in all_histories:
        vin = obj.vin
        history = obj.history

        for department, posts in history.items():
            for post_name, entries in posts.items():
                for entry in entries:
                    base_row = {
                        "vin": vin,
                        "department": department,
                        "post": post_name,
                        "controller": entry.get("controller", ""),
                        "date_added": entry.get("date_added", ""),
                    }

                    for field in image_fields:
                        val = entry.get(field)
                        if isinstance(val, list):
                            base_row[field] = val
                        elif isinstance(val, str):
                            base_row[field] = [val]
                        else:
                            base_row[field] = []

                    for k, v in entry.items():
                        if k in base_row or k in image_fields or k == "defects":
                            continue
                        base_row[k] = ", ".join(str(i) for i in v) if isinstance(v, list) else v

                    defects = entry.get("defects")
                    if defects:
                        for defect in defects:
                            row = base_row.copy()
                            for dk, dv in defect.items():
                                row[f"defect_{dk}"] = ", ".join(dv) if isinstance(dv, list) else dv
                            post_data[post_name].append(row)
                    else:
                        post_data[post_name].append(base_row)

    for post_name, rows in post_data.items():
        ws = wb.create_sheet(title=post_name[:31])
        if not rows:
            continue

        headers = list(sorted(set().union(*(r.keys() for r in rows))))
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)
            ws.cell(row=1, column=col_idx).alignment = Alignment(wrap_text=True)

        for row_idx, row in enumerate(rows, start=2):
            for col_idx, header in enumerate(headers, 1):
                val = row.get(header, "")

                if header in image_fields and isinstance(val, list):
                    for image_path in val:
                        full_path = os.path.join(settings.MEDIA_ROOT, image_path.strip("/"))
                        if os.path.exists(full_path):
                            try:
                                with PILImage.open(full_path) as img:
                                    img.thumbnail((150, 150))
                                    img_byte_arr = BytesIO()
                                    img.save(img_byte_arr, format='PNG')  # ✅ формат PNG
                                    img_byte_arr.seek(0)
                                    xl_img = XLImage(img_byte_arr)
                                    col_letter = get_column_letter(col_idx)
                                    img_anchor = f"{col_letter}{row_idx}"
                                    ws.add_image(xl_img, img_anchor)
                                    break  # Только первое изображение
                            except Exception as e:
                                ws.cell(row=row_idx, column=col_idx).value = f"Ошибка: {str(e)}"
                        else:
                            ws.cell(row=row_idx, column=col_idx).value = "Файл не найден"
                else:
                    if isinstance(val, dict):
                        val = json.dumps(val, ensure_ascii=False)
                    ws.cell(row=row_idx, column=col_idx).value = val
                    ws.cell(row=row_idx, column=col_idx).alignment = Alignment(wrap_text=True, vertical="top")

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=ALL_VIN_history_with_images.xlsx'
    wb.save(response)
    return response


@login_required
@role_required(["master", "head_area"])
def export_incoming_photos_zip(request):
    start_date = parse_date(request.GET.get("start_date"))
    end_date = parse_date(request.GET.get("end_date"))

    if not start_date or not end_date:
        return HttpResponse("❌ Ошибка: укажите диапазон дат.", status=400)

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:

        for history in VINHistory.objects.all():
            vin = history.vin
            zones = history.history

            # Работать только с "Цех поставки"
            zone_data = zones.get("Цех поставки")
            if not zone_data:
                continue

            for post_name, entries in zone_data.items():
                for entry in entries:
                    raw_date = entry.get("date_added")
                    if not raw_date:
                        continue

                    entry_date = parse_date(raw_date[:10])
                    if start_date and (entry_date < start_date or entry_date > end_date):
                        continue

                    base_path = f"Входной контроль/{post_name}/{vin}"

                    # Фото кузова
                    body_photos = entry.get("body_photos", [])
                    if body_photos:
                        for photo in body_photos:
                            add_photo_to_zip(zip_file, photo, f"{base_path}/body_photos")

                    # Фото компонентов
                    component_photos = entry.get("component_photos", [])
                    if component_photos:
                        for photo in component_photos:
                            add_photo_to_zip(zip_file, photo, f"{base_path}/component_photos")

                    # Фото двигателя
                    engine_photo = entry.get("engine_photo")
                    if engine_photo:
                        add_photo_to_zip(zip_file, engine_photo, f"{base_path}/engine_photo")

                    # Фото дефектов
                    defects = entry.get("defects", [])
                    for defect in defects:
                        defect_photos = defect.get("defect_photos", [])
                        if defect_photos:
                            for photo in defect_photos:
                                add_photo_to_zip(zip_file, photo, f"{base_path}/defect_photos")

    zip_buffer.seek(0)
    response = HttpResponse(zip_buffer, content_type="application/zip")
    filename = f"vhodnoy_kontrol_photos_{start_date}_{end_date}.zip"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def add_photo_to_zip(zip_file, photo_url, zip_path_prefix):
    if not photo_url:
        return
    rel_path = photo_url.replace(settings.MEDIA_URL, "")
    abs_path = os.path.join(settings.MEDIA_ROOT, rel_path)

    if os.path.exists(abs_path):
        filename = os.path.basename(abs_path)
        arcname = f"{zip_path_prefix}/{filename}"
        zip_file.write(abs_path, arcname)


@login_required
@role_required(["master", "head_area"])
def export_body_inspection_excel(request):
    start_date = parse_date(request.GET.get("start_date"))
    end_date = parse_date(request.GET.get("end_date"))

    if not (start_date and end_date):
        return HttpResponse("❌ Ошибка: укажите диапазон дат.", status=400)

    zone = "Цех поставки"
    post = "Зона первичного осмотра кузовов, DKD"

    grouped_by_vin = defaultdict(list)
    max_defect_photos = 0
    max_body_photos = 0

    # Сбор данных, как в export_assembly_excel: по строке на каждый дефект, с учётом формата полей
    for history in VINHistory.objects.all():
        vin = history.vin
        post_entries = (history.history or {}).get(zone, {}).get(post, []) or []

        for entry in post_entries:
            raw_date = entry.get("date_added") or entry.get("date")
            if not raw_date:
                continue

            dt = parse_datetime(raw_date)
            if not dt:
                # попытка ISO (на случай старого формата)
                try:
                    dt = datetime.fromisoformat(raw_date)
                except Exception:
                    continue

            if not (start_date <= dt.date() <= end_date):
                continue

            controller = entry.get("controller", "") or entry.get("extra_data", {}).get("controller", "")
            body_photos = entry.get("body_photos", []) or entry.get("extra_data", {}).get("body_photos", [])
            max_body_photos = max(max_body_photos, len(body_photos))

            defects = entry.get("defects")
            if isinstance(defects, list) and defects:
                for defect in defects:
                    defect_photos = defect.get("defect_photos", []) or []
                    max_defect_photos = max(max_defect_photos, len(defect_photos))
                    grouped_by_vin[vin].append({
                        "date": dt,
                        "controller": controller,
                        "defect": defect.get("defect"),
                        "detail": defect.get("detail"),
                        "quantity": defect.get("quantity"),
                        "grade": defect.get("grade"),
                        "repair_type": defect.get("repair_type"),
                        "responsible": defect.get("responsible"),
                        "comment": defect.get("custom_detail_explanation"),
                        "extra_comment": defect.get("custom_defect_explanation"),
                        "defect_photos": defect_photos,
                        "body_photos": body_photos,
                    })
            else:
                grouped_by_vin[vin].append({
                    "date": dt,
                    "controller": controller,
                    "defect": "Без дефектов",
                    "detail": "None",
                    "quantity": "None",
                    "grade": "None",
                    "repair_type": "None",
                    "responsible": "None",
                    "comment": "None",
                    "extra_comment": "None",
                    "defect_photos": [],
                    "body_photos": body_photos,
                })

    # TraceData для подписи VIN (бренд/модель/код) — как в export_assembly_excel
    trace_data = TraceData.objects.filter(vin_rk__in=grouped_by_vin.keys())
    trace_map = {
        td.vin_rk: {"brand": td.brand, "model": td.model, "config_code": td.config_code}
        for td in trace_data
    }

    # --- Excel ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Пост кузовов"

    headers = [
        "VIN", "Дата", "Контроллер", "Дефект", "Деталь", "Кол-во", "Степень",
        "Тип ремонта", "Ответственный", "Комментарий"
    ] + [f"Фото дефекта {i+1}" for i in range(max_defect_photos)] \
      + [f"Фото кузова {i+1}" for i in range(max_body_photos)]
    ws.append(headers)

    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    # Заполнение строк
    for vin, entries in grouped_by_vin.items():
        if not entries:
            continue
        start_row = ws.max_row + 1

        for item in entries:
            row = [
                vin,
                item["date"].strftime("%d.%m.%Y %H:%M"),
                item["controller"],
                item["defect"],
                item["detail"],
                item["quantity"],
                item["grade"],
                item["repair_type"],
                item["responsible"],
                item["comment"],
            ]
            # пустые ячейки под фото
            row += [""] * max_defect_photos
            row += [""] * max_body_photos
            ws.append(row)

            row_idx = ws.max_row

            # Вставка фото (дефекты)
            for i, img_path in enumerate(item["defect_photos"]):
                insert_single_image(ws, row_idx, 10 + i + 1, img_path)  # 11-й столбец — первый фото дефекта

            # Вставка фото (кузов)
            for j, img_path in enumerate(item["body_photos"]):
                insert_single_image(ws, row_idx, 10 + max_defect_photos + j + 1, img_path)

            ws.row_dimensions[row_idx].height = 75

        # объединяем VIN-ячейки и пишем бренд/модель/код
        if len(entries) > 1:
            ws.merge_cells(start_row=start_row, end_row=ws.max_row, start_column=1, end_column=1)

        trace = trace_map.get(vin, {})
        brand_info = f'{trace.get("brand", "")}\n{trace.get("model", "")}\n{trace.get("config_code", "")}'
        cell = ws.cell(row=start_row, column=1)
        cell.value = f"{vin}\n{brand_info}"
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Выравнивание
    for col in ws.columns:
        for cell in col:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # --- Сводка (как в export_assembly_excel) ---
    summary_ws = wb.create_sheet("Сводка")
    summary_ws.append(["Показатель", "Значение"])
    vins = list(grouped_by_vin.keys())
    total_rows = sum(len(records) for records in grouped_by_vin.values())
    total_defects = sum(
        1 for records in grouped_by_vin.values() for r in records
        if (r.get("defect") or "").strip() and (r.get("defect") != "Без дефектов")
    )
    dpu = round(total_defects / len(vins), 2) if vins else 0

    summary_ws.append(["Всего машин (VIN)", len(vins)])
    summary_ws.append(["Всего строк (осмотров)", total_rows])
    summary_ws.append(["Всего дефектов", total_defects])
    summary_ws.append(["DPU (дефектов на VIN)", dpu])

    # Ответ
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    filename = f"post_kuzovov_{start_date}_{end_date}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


def write_row_with_horizontal_images(ws, static_values, defect_photos, body_photos, max_defect_photos, max_body_photos):
    """
    static_values — список значений до фото
    defect_photos — список путей к фото дефектов
    body_photos — список путей к фото кузова
    """
    row = static_values + [""] * max_defect_photos + [""] * max_body_photos
    ws.append(row)
    row_idx = ws.max_row
    ws.row_dimensions[row_idx].height = 90  # ✅ увеличим высоту для фото

    # ✅ Устанавливаем ширину всех колонок с фото (дефекты)
    for i in range(max_defect_photos):
        col_idx = len(static_values) + i + 1
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 18  # под 85–90px

    # ✅ Фото дефектов
    for i, path in enumerate(defect_photos):
        if i >= max_defect_photos:
            break
        insert_single_image(ws, row_idx, len(static_values) + i + 1, path)

    # ✅ Устанавливаем ширину всех колонок с фото (кузов)
    for j in range(max_body_photos):
        col_idx = len(static_values) + max_defect_photos + j + 1
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 18

    # ✅ Фото кузова
    for j, path in enumerate(body_photos):
        if j >= max_body_photos:
            break
        insert_single_image(ws, row_idx, len(static_values) + max_defect_photos + j + 1, path)


def insert_single_image(ws, row_idx, col_idx, path):
    try:
        rel_path = path.replace(settings.MEDIA_URL, "")
        abs_path = os.path.join(settings.MEDIA_ROOT, rel_path)

        if os.path.exists(abs_path):
            img = XLImage(abs_path)
            img.width = 85
            img.height = 85

            col_letter = get_column_letter(col_idx)
            ws.add_image(img, f"{col_letter}{row_idx}")

            # 💡 Установим ширину колонки (один раз)
            if col_letter not in ws.column_dimensions:
                ws.column_dimensions[col_letter].width = 17  # под 90px картинки

    except Exception as e:
        print(f"❌ Ошибка изображения {path}: {e}")


@login_required
@role_required(["master", "head_area"])
def export_parts_acceptance_excel(request):
    start_date = parse_date(request.GET.get("start_date"))
    end_date = parse_date(request.GET.get("end_date"))

    if not (start_date and end_date):
        return HttpResponse("❌ Ошибка: укажите диапазон дат.", status=400)

    zone = "Цех поставки"
    post = "Зона выгрузки комплектующих, DKD"

    grouped_by_vin = defaultdict(list)
    max_defect_photos = 0
    max_component_photos = 0

    # Сбор данных в "единый" стиль: по строке на каждый дефект, иначе — строка "Без дефектов"
    for history in VINHistory.objects.all():
        vin = history.vin
        post_entries = (history.history or {}).get(zone, {}).get(post, []) or []

        for entry in post_entries:
            raw_date = entry.get("date_added") or entry.get("date")
            if not raw_date:
                continue

            dt = parse_datetime(raw_date)
            if not dt:
                try:
                    dt = datetime.fromisoformat(raw_date)
                except Exception:
                    continue

            if not (start_date <= dt.date() <= end_date):
                continue

            # Базовые поля + данные из extra_data при наличии
            controller = entry.get("controller", "") or entry.get("extra_data", {}).get("controller", "")
            engine_number = entry.get("engine_number", "") or entry.get("extra_data", {}).get("engine_number", "")
            engine_photo = entry.get("engine_photo") or entry.get("extra_data", {}).get("engine_photo")
            component_photos = entry.get("component_photos", []) or entry.get("extra_data", {}).get("component_photos", []) or []
            max_component_photos = max(max_component_photos, len(component_photos))

            defects = entry.get("defects")
            if isinstance(defects, list) and defects:
                for defect in defects:
                    defect_photos = defect.get("defect_photos", []) or []
                    max_defect_photos = max(max_defect_photos, len(defect_photos))
                    grouped_by_vin[vin].append({
                        "date": dt,
                        "controller": controller,
                        "engine_number": engine_number,
                        "engine_photo": engine_photo,
                        "defect": defect.get("defect"),
                        "detail": defect.get("detail"),
                        "quantity": defect.get("quantity"),
                        "grade": defect.get("grade"),
                        "repair_type": defect.get("repair_type"),
                        "responsible": defect.get("responsible"),
                        "comment": defect.get("custom_detail_explanation"),
                        "defect_photos": defect_photos,
                        "component_photos": component_photos,
                    })
            else:
                grouped_by_vin[vin].append({
                    "date": dt,
                    "controller": controller,
                    "engine_number": engine_number,
                    "engine_photo": engine_photo,
                    "defect": "Без дефектов",
                    "detail": "None",
                    "quantity": "None",
                    "grade": "None",
                    "repair_type": "None",
                    "responsible": "None",
                    "comment": "None",
                    "defect_photos": [],
                    "component_photos": component_photos,
                })

    # TraceData для подписи VIN (бренд/модель/код) и объединения ячейки VIN
    trace_data = TraceData.objects.filter(vin_rk__in=grouped_by_vin.keys())
    trace_map = {
        td.vin_rk: {"brand": td.brand, "model": td.model, "config_code": td.config_code}
        for td in trace_data
    }

    # --- Excel ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Пост комплектующих"

    headers = [
        "VIN", "Дата", "Контроллер", "Номер двигателя", "Фото двигателя",
        "Дефект", "Деталь", "Кол-во", "Степень", "Тип ремонта", "Ответственный", "Комментарий"
    ] + [f"Фото дефекта {i+1}" for i in range(max_defect_photos)] \
      + [f"Фото комплектующих {i+1}" for i in range(max_component_photos)]
    ws.append(headers)

    # ширины — как в других экспортерах
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    # Заполнение строк по VIN
    for vin, entries in grouped_by_vin.items():
        if not entries:
            continue
        start_row = ws.max_row + 1

        for item in entries:
            row = [
                vin,
                item["date"].strftime("%d.%m.%Y %H:%M"),
                item["controller"],
                item["engine_number"],
                "",  # фото двигателя вставим далее
                item["defect"],
                item["detail"],
                item["quantity"],
                item["grade"],
                item["repair_type"],
                item["responsible"],
                item["comment"],
            ]
            row += [""] * max_defect_photos
            row += [""] * max_component_photos
            ws.append(row)
            row_idx = ws.max_row

            # Фото двигателя в 5-й колонке
            if item["engine_photo"]:
                insert_single_image(ws, row_idx, 5, item["engine_photo"])

            # Фото дефектов (начиная с 13-й колонки)
            for i, img_path in enumerate(item["defect_photos"]):
                insert_single_image(ws, row_idx, 12 + i + 1, img_path)  # 13-й столбец — первый фото дефекта

            # Фото комплектующих после блока фото дефектов
            for j, img_path in enumerate(item["component_photos"]):
                insert_single_image(ws, row_idx, 12 + max_defect_photos + j + 1, img_path)

            ws.row_dimensions[row_idx].height = 75

        # Объединяем VIN-ячейки и вписываем бренд/модель/код
        if len(entries) > 1:
            ws.merge_cells(start_row=start_row, end_row=ws.max_row, start_column=1, end_column=1)

        trace = trace_map.get(vin, {})
        brand_info = f'{trace.get("brand", "")}\n{trace.get("model", "")}\n{trace.get("config_code", "")}'
        cell = ws.cell(row=start_row, column=1)
        cell.value = f"{vin}\n{brand_info}"
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Выравнивание
    for col in ws.columns:
        for cell in col:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Сводка — по аналогии с другими экспортами
    summary_ws = wb.create_sheet("Сводка")
    summary_ws.append(["Показатель", "Значение"])
    vins = list(grouped_by_vin.keys())
    total_rows = sum(len(records) for records in grouped_by_vin.values())
    total_defects = sum(
        1 for records in grouped_by_vin.values() for r in records
        if (r.get("defect") or "").strip() and (r.get("defect") != "Без дефектов")
    )
    dpu = round(total_defects / len(vins), 2) if vins else 0

    summary_ws.append(["Всего машин (VIN)", len(vins)])
    summary_ws.append(["Всего строк (осмотров)", total_rows])
    summary_ws.append(["Всего дефектов", total_defects])
    summary_ws.append(["DPU (дефектов на VIN)", dpu])

    # Ответ
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"post_parts_acceptance_{start_date}_{end_date}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
@role_required(["master", "head_area"])
def export_final_acceptance_excel(request):
    start_date = parse_date(request.GET.get("start_date"))
    end_date = parse_date(request.GET.get("end_date"))

    if not (start_date and end_date):
        return HttpResponse("❌ Ошибка: укажите диапазон дат.", status=400)

    zone = "Цех поставки"
    post = "Зона основной приемки DKD"

    grouped_by_vin = defaultdict(list)
    max_defect_photos = 0
    max_body_photos = 0

    # Сбор данных — в стиле export_body_inspection_excel
    for history in VINHistory.objects.all():
        vin = history.vin
        post_entries = (history.history or {}).get(zone, {}).get(post, []) or []

        for entry in post_entries:
            raw_date = entry.get("date_added") or entry.get("date")
            if not raw_date:
                continue

            dt = parse_datetime(raw_date)
            if not dt:
                try:
                    dt = datetime.fromisoformat(raw_date)
                except Exception:
                    continue

            if not (start_date <= dt.date() <= end_date):
                continue

            controller = entry.get("controller", "") or entry.get("extra_data", {}).get("controller", "")
            inspection_time = entry.get("inspection_duration_seconds", "") or entry.get("extra_data", {}).get("inspection_duration_seconds", "")
            body_photos = entry.get("body_photos", []) or entry.get("extra_data", {}).get("body_photos", [])
            max_body_photos = max(max_body_photos, len(body_photos))

            defects = entry.get("defects")
            if isinstance(defects, list) and defects:
                for defect in defects:
                    defect_photos = defect.get("defect_photos", []) or []
                    max_defect_photos = max(max_defect_photos, len(defect_photos))
                    grouped_by_vin[vin].append({
                        "date": dt,
                        "controller": controller,
                        "inspection_time": inspection_time,
                        "detail": defect.get("detail"),
                        "comment_detail": defect.get("custom_detail_explanation"),
                        "defect": defect.get("defect"),
                        "comment_defect": defect.get("custom_defect_explanation"),
                        "quantity": defect.get("quantity"),
                        "grade": defect.get("grade"),
                        "repair_type": defect.get("repair_type"),
                        "responsible": defect.get("responsible"),
                        "defect_photos": defect_photos,
                        "body_photos": body_photos,
                    })
            else:
                grouped_by_vin[vin].append({
                    "date": dt,
                    "controller": controller,
                    "inspection_time": inspection_time,
                    "detail": "None",
                    "comment_detail": "None",
                    "defect": "Без дефектов",
                    "comment_defect": "None",
                    "quantity": "None",
                    "grade": "None",
                    "repair_type": "None",
                    "responsible": "None",
                    "defect_photos": [],
                    "body_photos": body_photos,
                })

    # Данные TraceData для подписи VIN (бренд/модель/код)
    trace_data = TraceData.objects.filter(vin_rk__in=grouped_by_vin.keys())
    trace_map = {
        td.vin_rk: {"brand": td.brand, "model": td.model, "config_code": td.config_code}
        for td in trace_data
    }

    # --- Excel ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Пост основной приемки"

    headers = [
        "VIN", "Дата", "Контроллер", "Время осмотра (сек)",
        "Деталь", "Комментарий (деталь)", "Дефект", "Комментарий (дефект)",
        "Кол-во", "Степень", "Тип ремонта", "Ответственный"
    ] + [f"Фото дефекта {i+1}" for i in range(max_defect_photos)] \
      + [f"Фото кузова {i+1}" for i in range(max_body_photos)]
    ws.append(headers)

    # ширины колонок
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    # Заполнение строк
    for vin, entries in grouped_by_vin.items():
        if not entries:
            continue
        start_row = ws.max_row + 1

        for item in entries:
            row = [
                vin,
                item["date"].strftime("%d.%m.%Y %H:%M"),
                item["controller"],
                item["inspection_time"],
                item["detail"],
                item["comment_detail"],
                item["defect"],
                item["comment_defect"],
                item["quantity"],
                item["grade"],
                item["repair_type"],
                item["responsible"],
            ]
            # пустые ячейки под фото
            row += [""] * max_defect_photos
            row += [""] * max_body_photos
            ws.append(row)
            row_idx = ws.max_row

            # Фото дефектов (первый фото-столбец — 13)
            for i, img_path in enumerate(item["defect_photos"]):
                insert_single_image(ws, row_idx, 12 + i + 1, img_path)

            # Фото кузова — сразу после блока дефектов
            for j, img_path in enumerate(item["body_photos"]):
                insert_single_image(ws, row_idx, 12 + max_defect_photos + j + 1, img_path)

            ws.row_dimensions[row_idx].height = 75

        # Объединяем VIN-ячейки и вписываем бренд/модель/код
        if len(entries) > 1:
            ws.merge_cells(start_row=start_row, end_row=ws.max_row, start_column=1, end_column=1)

        trace = trace_map.get(vin, {})
        brand_info = f'{trace.get("brand", "")}\n{trace.get("model", "")}\n{trace.get("config_code", "")}'
        cell = ws.cell(row=start_row, column=1)
        cell.value = f"{vin}\n{brand_info}"
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Выравнивание
    for col in ws.columns:
        for cell in col:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Сводка
    summary_ws = wb.create_sheet("Сводка")
    summary_ws.append(["Показатель", "Значение"])
    vins = list(grouped_by_vin.keys())
    total_rows = sum(len(records) for records in grouped_by_vin.values())
    total_defects = sum(
        1 for records in grouped_by_vin.values() for r in records
        if (r.get("defect") or "").strip() and (r.get("defect") != "Без дефектов")
    )
    dpu = round(total_defects / len(vins), 2) if vins else 0

    summary_ws.append(["Всего машин (VIN)", len(vins)])
    summary_ws.append(["Всего строк (осмотров)", total_rows])
    summary_ws.append(["Всего дефектов", total_defects])
    summary_ws.append(["DPU (дефектов на VIN)", dpu])

    # Ответ
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"post_final_acceptance_{start_date}_{end_date}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
@role_required(["master", "head_area"])
def export_assembly_excel(request):
    start_date = parse_date(request.GET.get("start_date"))
    end_date = parse_date(request.GET.get("end_date"))
    post = request.GET.get("post")

    if not (start_date and end_date and post):
        return HttpResponse("❌ Укажите диапазон дат и пост.", status=400)

    zone = "Цех сборки"
    grouped_by_vin = defaultdict(list)
    max_photos = 0

    def get_entry_base(entry, dt):
        controller = entry.get("controller", "") or entry.get("extra_data", {}).get("controller", "")
        line = entry.get("line", "")
        duration = entry.get("inspection_duration_seconds", "")
        has_defect = entry.get("has_defect")
        if has_defect is True:
            has_defect_str = "Да"
        elif has_defect is False:
            has_defect_str = "Нет"
        else:
            has_defect_str = ""  # поле может отсутствовать в старом формате
        return controller, line, duration, has_defect_str

    for history in VINHistory.objects.all():
        vin = history.vin
        post_entries = (history.history or {}).get(zone, {}).get(post, []) or []

        for entry in post_entries:
            raw_date = entry.get("date_added") or entry.get("date")
            if not raw_date:
                continue

            dt = parse_datetime(raw_date)
            if not dt:
                continue
            if not (start_date <= dt.date() <= end_date):
                continue

            controller, line, duration, has_defect_str = get_entry_base(entry, dt)

            # НОВЫЙ ФОРМАТ: список defects (может быть пустым)
            if isinstance(entry.get("defects"), list):
                defects = entry["defects"]

                if defects:  # по одному ряду на каждый дефект
                    for defect in defects:
                        photos = defect.get("photos", []) or []
                        max_photos = max(max_photos, len(photos))
                        grouped_by_vin[vin].append({
                            "date": dt,
                            "controller": controller,
                            "has_defect": has_defect_str,
                            "defect": defect.get("name", ""),
                            "unit": defect.get("unit", ""),
                            "grade": defect.get("grade", ""),
                            "comment": defect.get("comment", ""),
                            "responsible": defect.get("responsible", ""),
                            "line": line,
                            "duration": duration,
                            "photos": photos,
                        })
                else:
                    # дефектов нет — всё равно пишем строку осмотра
                    photos = entry.get("photos", []) or entry.get("defect_photos", []) or []
                    max_photos = max(max_photos, len(photos))
                    grouped_by_vin[vin].append({
                        "date": dt,
                        "controller": controller,
                        "has_defect": has_defect_str,
                        "defect": "",
                        "unit": "",
                        "grade": "",
                        "comment": "",
                        "responsible": "",
                        "line": line,
                        "duration": duration,
                        "photos": photos,
                    })

            # СТАРЫЙ ФОРМАТ: одиночные поля дефекта/фото (или вообще без них)
            else:
                photos = entry.get("defect_photos", []) or entry.get("photos", []) or []
                max_photos = max(max_photos, len(photos))
                grouped_by_vin[vin].append({
                    "date": dt,
                    "controller": controller,
                    "has_defect": has_defect_str,
                    "defect": entry.get("defect_description", ""),
                    "unit": entry.get("unit", ""),
                    "grade": entry.get("grade", ""),
                    "comment": entry.get("comment", ""),
                    "responsible": entry.get("responsible", ""),
                    "line": line,
                    "duration": duration,
                    "photos": photos,
                })

    # --- Бренд/модель/код берём из TraceData ---
    trace_data = TraceData.objects.filter(vin_rk__in=grouped_by_vin.keys())
    trace_map = {
        td.vin_rk: {
            "brand": td.brand or "",
            "model": td.model or "",
            "config_code": td.config_code or "",
        }
        for td in trace_data
    }

    # --- Excel ---
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = post

    # Добавили отдельные столбцы "Бренд", "Модель", "Код комплектации"
    headers = [
        "VIN", "Бренд", "Модель", "Код комплектации",   # <-- новые отдельные колонки
        "Дата", "Линия", "Контроллер", "Есть дефект",
        "Дефект", "Комментарий", "Грейд", "Единица", "Кто виноват",
        "Время осмотра (сек)",
    ] + [f"Фото {i+1}" for i in range(max_photos)]
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    # Фото теперь начинаются с 15-го столбца (было 12-й, добавили 3 колонки)
    first_photo_col_idx = 15

    for vin, entries in grouped_by_vin.items():
        if not entries:
            continue

        # Достанем данные из trace_map (или пустые строки)
        tr = trace_map.get(vin, {})
        brand = tr.get("brand", "")
        model = tr.get("model", "")
        config_code = tr.get("config_code", "")

        start_row = ws.max_row + 1

        for entry in entries:
            row = [
                vin,                # col 1
                brand,              # col 2
                model,              # col 3
                config_code,        # col 4
                entry["date"].strftime("%d.%m.%Y %H:%M"),  # col 5
                entry["line"],      # col 6
                entry["controller"],# col 7
                entry["has_defect"],# col 8
                entry["defect"],    # col 9
                entry["comment"],   # col 10
                entry["grade"],     # col 11
                entry["unit"],      # col 12
                entry["responsible"],# col 13
                entry["duration"],  # col 14
            ]
            row += [""] * max_photos
            ws.append(row)

            # Вставка фото (insert_single_image — как у тебя)
            row_idx = ws.max_row
            for i, img_path in enumerate(entry["photos"]):
                insert_single_image(ws, row_idx, first_photo_col_idx + i, img_path)

            ws.row_dimensions[row_idx].height = 75

        # Объединяем по VIN также столбцы Бренд/Модель/Код (1..4)
        if len(entries) > 1:
            for col_idx in (1, 2, 3, 4):
                ws.merge_cells(start_row=start_row, end_row=ws.max_row, start_column=col_idx, end_column=col_idx)

        # Выравнивание для объединённых шапок VIN/бренд/модель/код
        for col_idx in (1, 2, 3, 4):
            cell = ws.cell(row=start_row, column=col_idx)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Общее выравнивание
    for col in ws.columns:
        for cell in col:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # --- Сводка ---
    summary_ws = wb.create_sheet("Сводка")
    summary_ws.append(["Показатель", "Значение"])
    vins = list(grouped_by_vin.keys())
    total_rows = sum(len(records) for records in grouped_by_vin.values())
    total_defects = sum(1 for records in grouped_by_vin.values() for r in records if (r.get("defect") or "").strip())
    dpu = round(total_defects / len(vins), 2) if vins else 0

    summary_ws.append(["Всего машин (VIN)", len(vins)])
    summary_ws.append(["Всего строк (осмотров)", total_rows])
    summary_ws.append(["Всего дефектов", total_defects])
    summary_ws.append(["DPU (дефектов на VIN)", dpu])

    # --- Ответ ---
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    filename = f"export_{post.replace(' ', '_')}_{start_date}_{end_date}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response



@login_required
@role_required(["master", "head_area"])
def export_containers_acceptance_excel(request):
    start_date = parse_date(request.GET.get("start_date"))
    end_date = parse_date(request.GET.get("end_date"))

    if not (start_date and end_date):
        return HttpResponse("❌ Ошибка: укажите диапазон дат.", status=400)

    zone = "Цех поставки"
    post = "Приемка контейнеров"

    rows = []
    max_photos = 0

    # Сбор данных в плоский список по датам
    for history in ContainerHistory.objects.all():
        post_entries = (history.history or {}).get(zone, {}).get(post, []) or []

        for entry in post_entries:
            raw_date = entry.get("date_added") or entry.get("date")
            if not raw_date:
                continue

            dt = parse_datetime(raw_date)
            if not dt:
                try:
                    dt = datetime.fromisoformat(raw_date)
                except Exception:
                    continue

            if not (start_date <= dt.date() <= end_date):
                continue

            controller = entry.get("controller", "") or entry.get("extra_data", {}).get("controller", "")
            container_number = entry.get("container_number", "") or entry.get("extra_data", {}).get("container_number", "")
            photos = entry.get("photos", []) or []
            description = entry.get("description", "") or ""

            has_defect_val = str(entry.get("has_defect", "")).lower()
            has_defect = "Да" if has_defect_val in ("yes", "true", "1") else "Нет"

            max_photos = max(max_photos, len(photos))

            rows.append({
                "date": dt,
                "controller": controller,
                "container_number": container_number,
                "has_defect": has_defect,
                "description": description,
                "photos": photos,
            })

    # --- Excel ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Приёмка контейнеров"

    headers = [
        "Дата", "Контролёр", "Номер контейнера", "Есть дефект", "Описание"
    ] + [f"Фото {i+1}" for i in range(max_photos)]
    ws.append(headers)

    # ширины колонок
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22

    # Заполнение строк
    for item in sorted(rows, key=lambda r: r["date"] or datetime.min, reverse=False):
        row = [
            item["date"].strftime("%d.%m.%Y %H:%M") if item["date"] else "",
            item["controller"],
            item["container_number"],
            item["has_defect"],
            item["description"],
        ]
        row += [""] * max_photos
        ws.append(row)

        row_idx = ws.max_row
        # Вставка фото начиная с 6-й колонки
        for i, img_path in enumerate(item["photos"]):
            insert_single_image(ws, row_idx, 5 + i + 1, img_path)

        ws.row_dimensions[row_idx].height = 75

    # Выравнивание
    for col in ws.columns:
        for cell in col:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Сводка
    summary_ws = wb.create_sheet("Сводка")
    summary_ws.append(["Показатель", "Значение"])
    summary_ws.append(["Всего записей", len(rows)])
    total_with_defect = sum(1 for r in rows if r["has_defect"] == "Да")
    summary_ws.append(["С дефектом", total_with_defect])
    summary_ws.append(["Без дефекта", len(rows) - total_with_defect])

    # Ответ
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    filename = f"containers_acceptance_{start_date}_{end_date}.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
