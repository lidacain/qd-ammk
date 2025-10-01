# <app>/management/commands/export_assembly_refbooks.py
from django.core.management.base import BaseCommand
from django.utils.timezone import now
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from assembly.models import AssemblyDefect, AssemblyZone, AssemblyUnit


class Command(BaseCommand):
    help = "Экспорт справочников сборки (Defects, Zones, Units) в один Excel с отдельными листами."

    def add_arguments(self, parser):
        parser.add_argument(
            "--out",
            default=f"assembly_refbooks_{now().date().isoformat()}.xlsx",
            help="Путь/имя выходного файла .xlsx",
        )

    def handle(self, *args, **opts):
        out_path = opts["out"]
        wb = Workbook()

        # --- Defects ---
        ws_def = wb.active
        ws_def.title = "Defects"
        def_headers = ["ID", "Название дефекта", "Defect name (ENG)", "Посты"]
        ws_def.append(def_headers)

        defects = AssemblyDefect.objects.prefetch_related("posts").order_by("name")
        for d in defects:
            post_names = ", ".join(d.posts.values_list("name", flat=True))
            ws_def.append([d.id, d.name or "", d.nameENG or "", post_names])

        # --- Zones ---
        ws_zone = wb.create_sheet("Zones")
        zone_headers = ["ID", "Название зоны", "Посты"]
        ws_zone.append(zone_headers)

        zones = AssemblyZone.objects.prefetch_related("posts").order_by("name")
        for z in zones:
            post_names = ", ".join(z.posts.values_list("name", flat=True))
            ws_zone.append([z.id, z.name or "", post_names])

        # --- Units ---
        ws_unit = wb.create_sheet("Units")
        unit_headers = ["ID", "Название узла", "Зона", "Посты"]
        ws_unit.append(unit_headers)

        units = (
            AssemblyUnit.objects.select_related("zone")
            .prefetch_related("posts")
            .order_by("zone__name", "name", "id")
        )
        for u in units:
            zone_name = u.zone.name if u.zone else ""
            post_names = ", ".join(u.posts.values_list("name", flat=True))
            ws_unit.append([u.id, u.name or "", zone_name, post_names])

        # автоширина и фиксация шапки
        for ws in [ws_def, ws_zone, ws_unit]:
            for col_idx in range(1, ws.max_column + 1):
                col_letter = get_column_letter(col_idx)
                max_len = 0
                for cell in ws[col_letter]:
                    v = "" if cell.value is None else str(cell.value)
                    if len(v) > max_len:
                        max_len = len(v)
                ws.column_dimensions[col_letter].width = min(max_len + 2, 80)
            ws.freeze_panes = "A2"

        wb.save(out_path)
        self.stdout.write(self.style.SUCCESS(f"✅ Готово: {out_path}"))
