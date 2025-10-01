from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponseNotFound
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.conf import settings
from django.utils.timezone import now
import os
from collections import Counter
from supplies.models import TraceData
from .models import AssemblyPart, DefectAssembly, PostAssembly, AssemblyDefect, AssemblyUnit, AssemblyDefectGrade, AssemblyDefectResponsible, AssemblyZone
from .forms import (TorqueControlForm, TorqueGraphForm, AssemblyTemplateForm)
from users.decorators import role_required
from vehicle_history.models import VINHistory, AssemblyPassLog, VESPassLog, TrimOutPassLog, VehicleIdentifiers  # ✅ Модель для истории VIN
import json
from django.utils.dateparse import parse_datetime
from datetime import datetime
from uuid import uuid4
from .forms import TorqueControlForm, UUDDKDForm, UUDCheckForm
from django.views.decorators.csrf import csrf_exempt
from vehicle_history.utils import now_almaty_iso
from django.views.decorators.http import require_GET
from PIL import Image, ImageOps
from io import BytesIO
import sys
import time
from django.core.files.uploadedfile import InMemoryUploadedFile
from django.core.serializers.json import DjangoJSONEncoder
from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponseNotFound
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.conf import settings
from django.utils.timezone import now
import os
from collections import Counter
from supplies.models import TraceData
from .models import AssemblyPart, DefectAssembly, PostAssembly, AssemblyDefect, AssemblyUnit, AssemblyDefectGrade, AssemblyDefectResponsible, AssemblyZone
from .forms import (TorqueControlForm, TorqueGraphForm, AssemblyTemplateForm)
from users.decorators import role_required
from vehicle_history.models import VINHistory  # ✅ Модель для истории VIN
import json
from django.utils.dateparse import parse_datetime
from datetime import datetime
from uuid import uuid4
from .forms import TorqueControlForm, UUDDKDForm, UUDCheckForm
from django.views.decorators.csrf import csrf_exempt
from vehicle_history.utils import now_almaty_iso
from django.views.decorators.http import require_GET
from PIL import Image, ImageOps
from io import BytesIO
import sys
import time
from django.core.files.uploadedfile import InMemoryUploadedFile
from django.core.serializers.json import DjangoJSONEncoder
from django.views.decorators.http import require_POST

from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.shortcuts import get_object_or_404
from django.utils.timezone import now as dj_now
import uuid
from django.core.files.storage import default_storage
from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponseNotFound
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.conf import settings
from django.utils.timezone import now
import os
from collections import Counter
from supplies.models import TraceData
from .models import AssemblyPart, DefectAssembly, PostAssembly, AssemblyDefect, AssemblyUnit, AssemblyDefectGrade, AssemblyDefectResponsible, AssemblyZone
from .forms import (TorqueControlForm, TorqueGraphForm, AssemblyTemplateForm)
from users.decorators import role_required
from vehicle_history.models import VINHistory, AssemblyPassLog, VESPassLog  # ✅ История VIN, сканы VIN и VES-логи
import json
from django.utils.dateparse import parse_datetime
from datetime import datetime
from uuid import uuid4
from .forms import TorqueControlForm, UUDDKDForm, UUDCheckForm
from django.views.decorators.csrf import csrf_exempt
from vehicle_history.utils import now_almaty_iso
from django.views.decorators.http import require_GET
from PIL import Image, ImageOps
from io import BytesIO
import sys
import time
from django.core.files.uploadedfile import InMemoryUploadedFile
from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponseNotFound
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.conf import settings
from django.utils.timezone import now
import os
from collections import Counter
from supplies.models import TraceData
from .models import AssemblyPart, DefectAssembly, PostAssembly, AssemblyDefect, AssemblyUnit, AssemblyDefectGrade, AssemblyDefectResponsible, AssemblyZone
from .forms import (TorqueControlForm, TorqueGraphForm, AssemblyTemplateForm)
from users.decorators import role_required
from vehicle_history.models import VINHistory  # ✅ Модель для истории VIN
import json
from django.utils.dateparse import parse_datetime
from datetime import datetime
from uuid import uuid4
from .forms import TorqueControlForm, UUDDKDForm, UUDCheckForm
from django.views.decorators.csrf import csrf_exempt
from vehicle_history.utils import now_almaty_iso
from django.views.decorators.http import require_GET
from PIL import Image, ImageOps
from io import BytesIO
import sys
import time
from django.core.files.uploadedfile import InMemoryUploadedFile
from django.core.serializers.json import DjangoJSONEncoder
from django.views.decorators.http import require_POST

from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.shortcuts import get_object_or_404
from django.utils.timezone import now as dj_now
import uuid
from django.core.files.storage import default_storage
from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse, HttpResponseNotFound
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.conf import settings
from django.utils.timezone import now
import os
from collections import Counter
from supplies.models import TraceData
from .models import AssemblyPart, DefectAssembly, PostAssembly, AssemblyDefect, AssemblyUnit, AssemblyDefectGrade, AssemblyDefectResponsible, AssemblyZone
from .forms import (TorqueControlForm, TorqueGraphForm, AssemblyTemplateForm)
from users.decorators import role_required
from vehicle_history.models import VINHistory  # ✅ Модель для истории VIN
import json
from django.utils.dateparse import parse_datetime
from datetime import datetime
from uuid import uuid4
from .forms import TorqueControlForm, UUDDKDForm, UUDCheckForm
from django.views.decorators.csrf import csrf_exempt
from vehicle_history.utils import now_almaty_iso
from django.views.decorators.http import require_GET
from PIL import Image, ImageOps
from io import BytesIO
import sys
import time
from django.core.files.uploadedfile import InMemoryUploadedFile
from django.core.serializers.json import DjangoJSONEncoder
from django.views.decorators.http import require_POST

from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.shortcuts import get_object_or_404
from django.utils.timezone import now as dj_now
import uuid
from django.core.files.storage import default_storage
import io, json, os
from datetime import datetime
from typing import Dict, Any, List, Tuple

from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, HttpResponseBadRequest
from django.views.decorators.http import require_POST
from collections import OrderedDict
from PIL import Image
import xlsxwriter
from django.utils import timezone
from django.contrib.auth.decorators import permission_required


def compress_uploaded_image(uploaded_file, quality=60, max_width=1600):
    try:
        # 🛠️ Важно: перемотать в начало
        uploaded_file.seek(0)

        # ✅ Если это уже InMemoryUploadedFile, обернутый ранее, перемотка обязательна
        image = Image.open(uploaded_file)
        image = ImageOps.exif_transpose(image)

        if image.mode in ("RGBA", "P"):
            image = image.convert("RGB")

        if image.width > max_width:
            ratio = max_width / float(image.width)
            height = int(float(image.height) * float(ratio))
            image = image.resize((max_width, height), Image.Resampling.LANCZOS)

        output_io = BytesIO()
        image.save(output_io, format="JPEG", quality=quality, optimize=True)
        output_io.seek(0)

        new_filename = f"photo_{now_almaty_iso().replace(':', '-').replace('T', '_')}.jpg"
        return InMemoryUploadedFile(
            output_io,
            'ImageField',
            new_filename,
            'image/jpeg',
            output_io.getbuffer().nbytes,
            None
        )
    except Exception as e:
        print("Ошибка при сжатии:", e)
        return uploaded_file  # Возвращаем оригинал, если не удалось сжать


def save_check_photos(photos):
    saved_photos = []

    for photo in photos:
        filename = f"{now().strftime('%Y_%m_%d')}_{photo.name}"
        path = f"images/checks/{filename}"
        full_path = os.path.join(settings.MEDIA_ROOT, path)

        os.makedirs(os.path.dirname(full_path), exist_ok=True)

        with open(full_path, "wb+") as destination:
            for chunk in photo.chunks():
                destination.write(chunk)

        saved_photos.append(f"{settings.MEDIA_URL}{path}")

    return saved_photos


@login_required
@role_required(["master", "head_area"])
def torque_graph_view(request):
    form = TorqueGraphForm(request.GET or None)
    labels, data, chart_info = [], [], []
    min_torque = max_torque = 0
    seen_combinations = set()

    if form.is_valid():
        assembly_part = form.cleaned_data["assembly_part"]
        start_date = form.cleaned_data["start_date"]
        end_date = form.cleaned_data["end_date"]

        min_torque = assembly_part.min_torque
        max_torque = assembly_part.max_torque

        for history in VINHistory.objects.all():
            zone_data = history.history.get("Цех сборки", {})
            post_data = zone_data.get("Пост момента затяжки, DKD", [])

            for entry in post_data:
                part_name = entry.get("assembly_part", "")
                vin = entry.get("vin_number", "")

                if part_name != str(assembly_part):
                    continue

                combination_key = (vin, part_name)
                if combination_key in seen_combinations:
                    continue
                seen_combinations.add(combination_key)

                timestamp = parse_datetime(entry.get("date_added"))
                if timestamp is None or not (start_date <= timestamp.date() <= end_date):
                    continue

                for torque in entry.get("torque_values", []):
                    labels.append(timestamp.strftime("%Y-%m-%d %H:%M:%S"))
                    data.append(torque)
                    chart_info.append({
                        "vin": vin,
                        "torque": torque,
                        "timestamp": timestamp.strftime("%Y-%m-%d %H:%M:%S")
                    })

    has_data = bool(data)

    return render(request, "assembly/torque_graph_dkd.html", {
        "form": form,
        "labels": labels,
        "data": data,
        "chart_info": chart_info,
        "min_torque": min_torque,
        "max_torque": max_torque,
        "has_data": has_data,
    })


def vin_lookup(request):
    """
    Универсальный API для поиска VIN.
    - Если передан параметр `q`, выполняется поиск VIN'ов по подстроке (автодополнение).
    - Если передан параметр `vin`, возвращаются полные данные по VIN.
    """
    query = request.GET.get("q", "").strip().upper()
    vin_number = request.GET.get("vin", "").strip().upper()

    # 🔍 Поиск по подстроке (для автодополнения)
    if query:
        vins = TraceData.objects.filter(vin_rk__icontains=query)[:5]
        results = [
            {
                "vin": vin.vin_rk.strip(),
                "engine_number": vin.engine_number,
                "model": vin.model,
                "body_color": vin.body_color,
                "drive_type": vin.modification,
            }
            for vin in vins
        ]
        return JsonResponse({"results": results})

    # 📌 Полный поиск по VIN (для получения всех данных)
    if vin_number:
        trace = TraceData.objects.filter(vin_rk=vin_number).first()
        if not trace:
            return JsonResponse({"error": "VIN не найден"}, status=404)

        data = {
            "vin": trace.vin_rk,
            "engine_number": trace.engine_number,
            "model": trace.model,
            "body_color": trace.body_color,
            "drive_type": trace.modification,
        }
        return JsonResponse(data)

    return JsonResponse({"error": "Не указан параметр `q` или `vin`"}, status=400)


@login_required
@role_required(["controller"])
def torque_control_dkd(request):
    post = get_object_or_404(PostAssembly, name="Пост момента затяжки, DKD")

    if request.method == "POST":
        form = TorqueControlForm(request.POST, request.FILES)

        if form.is_valid():
            vin = form.cleaned_data.get("vin_number")
            modification = form.cleaned_data.get("modification")
            assembly_part = str(form.cleaned_data.get("assembly_part"))
            torque_values = form.cleaned_data.get("torque_values")

            history_entry, _ = VINHistory.objects.get_or_create(vin=vin)

            zone = "Цех сборки"
            post_name = "Пост момента затяжки, DKD"

            if zone not in history_entry.history:
                history_entry.history[zone] = {}

            if post_name not in history_entry.history[zone]:
                history_entry.history[zone][post_name] = []

            has_defect = request.POST.get("has_defect")

            if has_defect == "no":
                inspection_data = {
                    "vin_number": vin,
                    "modification": modification,
                    "assembly_part": assembly_part,
                    "torque_values": torque_values,
                    "controller": request.user.username,
                    "defects": [],
                    "date_added": now_almaty_iso(),
                }
                history_entry.history[zone][post_name].append(inspection_data)

            else:
                index = 1
                while f"defect_type_{index}" in request.POST:
                    defect_type = request.POST.get(f"defect_type_{index}", "").strip()
                    defect_quantity = request.POST.get(f"defect_quantity_{index}", "1").strip()

                    if not defect_type:
                        index += 1
                        continue

                    defect_data = {
                        "type": defect_type,
                        "quantity": defect_quantity,
                        "photos": [],
                    }

                    defect_photos = request.FILES.getlist(f"defect_photo_{index}")
                    for photo in defect_photos:
                        path = f"images/defects/{photo.name}"
                        full_path = os.path.join(settings.MEDIA_ROOT, path)
                        os.makedirs(os.path.dirname(full_path), exist_ok=True)
                        with open(full_path, "wb+") as dest:
                            for chunk in photo.chunks():
                                dest.write(chunk)
                        defect_data["photos"].append(f"{settings.MEDIA_URL}{path}")

                    inspection_data = {
                        "vin_number": vin,
                        "modification": modification,
                        "assembly_part": assembly_part,
                        "torque_values": torque_values,
                        "controller": request.user.username,
                        "defects": [defect_data],
                        "date_added": now_almaty_iso(),
                    }

                    history_entry.history[zone][post_name].append(inspection_data)
                    index += 1

            history_entry.save()

            messages.success(request, "✅ Инспекция успешно сохранена!")
            return redirect("/assembly/torque_control_dkd/")

        else:
            messages.error(request, "❌ Ошибка: Проверьте корректность заполнения формы.")
            print("❌ Ошибки формы:", form.errors)

    else:
        form = TorqueControlForm()

    return render(request, "assembly/torque_control_dkd.html", {
        "form": form,
        "post": post,
    })


# ✅ Функция сохранения загружаемых файлов
def save_uploaded_file(file, folder):
    file_path = os.path.join(settings.MEDIA_ROOT, folder, file.name)
    os.makedirs(os.path.dirname(file_path), exist_ok=True)

    with open(file_path, "wb+") as destination:
        for chunk in file.chunks():
            destination.write(chunk)

    return f"{settings.MEDIA_URL}{folder}{file.name}"


def get_assembly_parts(request):
    """
    API для получения узлов/деталей.
    - Если передан `drive_type`, фильтруем по приводу (2WD/4WD).
    - Если нет параметров, возвращаем все узлы/детали.
    """
    drive_type = request.GET.get("drive_type", "").strip().upper()

    if drive_type in ["2WD", "4WD"]:
        parts = AssemblyPart.objects.filter(modification=drive_type)
    else:
        parts = AssemblyPart.objects.all()

    # 📌 Теперь API возвращает `modification` для каждого узла
    results = [
        {
            "id": part.id,
            "name": part.name,
            "size": part.size,
            "modification": part.modification  # ✅ Добавлено!
        }
        for part in parts
    ]
    return JsonResponse(results, safe=False)


def get_part_details(request):
    """
    API для получения информации о детали.
    """
    part_id = request.GET.get("part_id")

    if not part_id:
        return JsonResponse({"error": "Деталь не выбрана"}, status=400)

    part = AssemblyPart.objects.filter(id=part_id).first()

    if not part:
        return JsonResponse({"error": "Деталь не найдена"}, status=404)

    result = {
        "size": part.size,
        "min_quantity": part.min_quantity,
        "max_quantity": part.max_quantity,
        "min_torque": part.min_torque,
        "max_torque": part.max_torque
    }

    return JsonResponse(result)


def vin_defects_api(request):
    vin = request.GET.get("vin", "").strip().upper()

    if not vin:
        return JsonResponse({"error": "VIN не указан"}, status=400)

    defects = DefectAssembly.objects.filter(posts__name="Контроль затяжки (DKD)")

    defect_counts = Counter(defect.name for defect in defects)

    results = [
        {
            "post": "Контроль затяжки (DKD)",
            "defect_type": defect_name,
            "quantity": count,
            "defect_photos": []
        }
        for defect_name, count in defect_counts.items()
    ]

    return JsonResponse({"defects": results})


def view_image(request, image_path):
    """
    Отображает изображение дефекта в увеличенном формате.
    """
    full_image_path = os.path.join(settings.MEDIA_ROOT, image_path)
    if not os.path.exists(full_image_path):
        messages.error(request, "Ошибка: Файл не найден.")
        return redirect("post_list")  # ✅ Перенаправляем на список постов, если файла нет

    return render(request, "assembly/view_image.html", {"image_url": f"{settings.MEDIA_URL}{image_path}"})


@login_required
@role_required(["controller"])
def uud_dkd(request):
    post_id = request.GET.get("post_id")

    if not post_id:
        return HttpResponseNotFound("❌ Не указан post_id в URL")

    post = get_object_or_404(PostAssembly, id=post_id)

    if request.method == "POST":
        form = UUDDKDForm(request.POST, request.FILES)

        if form.is_valid():
            vin_number = form.cleaned_data.get("vin_number")
            repair_description = form.cleaned_data.get("repair_description")
            repair_photos = form.cleaned_data.get("repair_photos")

            # Получаем или создаем историю VIN
            history_entry, _ = VINHistory.objects.get_or_create(vin=vin_number)

            # ✅ Сохраняем фотографии
            saved_photos = []
            for photo in repair_photos:
                path = f"images/repairs/{now().strftime('%Y_%m_%d')}_{photo.name}"
                full_path = os.path.join(settings.MEDIA_ROOT, path)
                os.makedirs(os.path.dirname(full_path), exist_ok=True)
                with open(full_path, "wb+") as dest:
                    for chunk in photo.chunks():
                        dest.write(chunk)
                saved_photos.append(f"{settings.MEDIA_URL}{path}")

            # ✅ Сохраняем запись в историю VIN
            history_entry.add_entry(
                post=post,
                defects=[],
                extra_data={
                    "repair_description": repair_description,
                    "repair_photos": saved_photos,
                    "controller": request.user.username
                }
            )

            # ✅ Создаем уведомление для мастеров
            from users.models import CustomUser, Notification

            message = f"УУД завершил ремонт по VIN {vin_number}. Ознакомьтесь с результатом."

            for user in CustomUser.objects.filter(role="master"):
                Notification.objects.create(
                    recipient=user,
                    message=message,
                    vin_number=vin_number
                )

            messages.success(request, "✅ Данные успешно сохранены в историю VIN")
            return redirect(request.path + f"?post_id={post_id}")

        else:
            messages.error(request, "❌ Ошибка в форме, проверьте данные.")
            print(form.errors)

    else:
        form = UUDDKDForm()

    return render(request, "assembly/uud_dkd.html", {
        "form": form,
        "post": post,
    })


@login_required
@csrf_exempt
def offline_defects_api(request):
    vin = request.GET.get("vin")

    if not vin:
        return JsonResponse({"error": "VIN не передан"}, status=400)

    history_entry = VINHistory.objects.filter(vin=vin).first()

    if not history_entry:
        return JsonResponse({"error": "История VIN не найдена"}, status=404)

    offline_defects = []

    # проходим по истории
    for zone, posts in history_entry.history.items():
        for post, inspections in posts.items():
            for inspection in inspections:
                for defect in inspection.get("defects", []):
                    if defect.get("repair_type") == "offline":
                        offline_defects.append({
                            "zone": zone,
                            "post": post,
                            "date": inspection.get("date_added"),
                            "detail": defect.get("detail"),
                            "defect": defect.get("defect"),
                            "grade": defect.get("grade"),
                            "quantity": defect.get("quantity"),
                            "responsible": defect.get("responsible"),
                            "repair_type": defect.get("repair_type"),
                            "repair_type_description": defect.get("repair_type_description"),
                            "photos": defect.get("defect_photos", [])
                        })

    return JsonResponse({"offline_defects": offline_defects}, status=200)


def save_check_photos(photos):
    saved_photos = []

    for photo in photos:
        filename = f"{now().strftime('%Y_%m_%d')}_{photo.name}"
        path = f"images/checks/{filename}"
        full_path = os.path.join(settings.MEDIA_ROOT, path)

        os.makedirs(os.path.dirname(full_path), exist_ok=True)

        with open(full_path, "wb+") as destination:
            for chunk in photo.chunks():
                destination.write(chunk)

        saved_photos.append(f"{settings.MEDIA_URL}{path}")

    return saved_photos


def get_last_uud_entry(history_entry):
    """
    Возвращает последнюю запись из поста 'Участок устранения дефектов, DKD'
    в зоне 'Цех УУД' для объекта VINHistory
    """
    zone_data = history_entry.history.get("Цех УУД", {})
    post_data = zone_data.get("Участок устранения дефектов, DKD", [])

    if post_data:
        return post_data[-1]  # последняя запись
    return None


@login_required
@role_required(["controller"])
def uud_check_dkd(request):
    if request.method == "POST":
        form = UUDCheckForm(request.POST, request.FILES)

        if form.is_valid():
            vin_number = form.cleaned_data["vin_number"]
            check_status = form.cleaned_data["check_status"]
            comment = form.cleaned_data.get("comment", "")
            check_photos = form.cleaned_data.get("check_photos")

            # Получаем историю VIN
            history_entry = VINHistory.objects.filter(vin=vin_number).first()
            if not history_entry:
                messages.error(request, "❌ История VIN не найдена.")
                return redirect("uud_check_dkd")

            # Ищем последнюю запись поста УУД
            last_uud_entry = get_last_uud_entry(history_entry)
            if not last_uud_entry:
                messages.error(request, "❌ Не найдена запись поста УУД для данного VIN.")
                return redirect("uud_check_dkd")

            # Сохраняем фото, если проверка не пройдена
            saved_photos = save_check_photos(check_photos) if check_status == "not_passed" else []

            # Обновляем поле check_result
            last_uud_entry.setdefault("extra_data", {})["check_result"] = {
                "status": check_status,
                "checked_by": request.user.username,
                "comment": comment if check_status == "not_passed" else "",
                "checked_at": now().strftime("%Y-%m-%d %H:%M:%S"),
                "check_photos": saved_photos,
            }

            history_entry.save()
            messages.success(request, "✅ Проверка успешно сохранена.")
            return redirect("assembly:uud_check_dkd")
        else:
            messages.error(request, "❌ Ошибка в форме, проверьте данные.")
    else:
        form = UUDCheckForm()

    return render(request, "assembly/uud_check_dkd.html", {"form": form})


@login_required
@require_GET
def uud_zone_data_api(request):
    vin = request.GET.get("vin")
    if not vin:
        return JsonResponse({"error": "VIN не указан."}, status=400)

    try:
        history_entry = VINHistory.objects.get(vin=vin)
    except VINHistory.DoesNotExist:
        return JsonResponse({"error": "История VIN не найдена."}, status=404)

    history = history_entry.history
    zone_data = history.get("Цех УУД", {})
    post_data = zone_data.get("Участок устранения дефектов, DKD", [])

    if not post_data:
        return JsonResponse({
            "uud_data": None
        })

    last_entry = post_data[-1]
    extra_data = last_entry.get("extra_data", {})

    check_result = extra_data.get("check_result", {})

    result = {
        "repair_description": extra_data.get("repair_description", ""),
        "repair_photos": extra_data.get("repair_photos", []),
        "date_added": last_entry.get("date_added", ""),
        "controller": extra_data.get("controller", ""),
        "check_status": check_result.get("status", ""),
        "checked_by": check_result.get("checked_by", ""),
        "checked_at": check_result.get("checked_at", ""),
        "check_comment": check_result.get("comment", ""),
        "check_photos": check_result.get("check_photos", []),
    }

    return JsonResponse({
        "uud_data": result
    })


def search_vin(request):
    """
    Поиск информации о кузове по VIN.
    """
    query = request.GET.get("q", "").strip()
    if not query:
        return JsonResponse({"results": []})

    # Поиск по VIN в TraceData (ограничиваем 5 результатами)
    vins = TraceData.objects.filter(vin_rk__icontains=query)[:5]

    if not vins.exists():
        return JsonResponse({"results": [], "error": "Нет данных по этому VIN"})

    results = [
        {
            "vin": vin.vin_rk,
            "engine_number": vin.engine_number,
            "config_code": vin.config_code,
            "model": vin.model,
            "body_color": vin.body_color,
            "drive_type": vin.modification,
        }
        for vin in vins
    ]

    return JsonResponse({"results": results})



@login_required
@role_required(["controller", "master", 'head_area'])
def assembly_post_view(request, line, post_suffix="", template_name="assembly/assembly_post_template.html"):
    post_id = request.GET.get("post_id")
    post_obj = PostAssembly.objects.filter(id=post_id).first()
    post_name = post_obj.name if post_obj else "Неизвестный пост"

    # 🔽 Справочники
    defects = AssemblyDefect.objects.all()
    zones = AssemblyZone.objects.all()
    units = AssemblyUnit.objects.all()
    grades = AssemblyDefectGrade.objects.all()
    responsibles = AssemblyDefectResponsible.objects.all()

    if request.method == "POST":
        form = AssemblyTemplateForm(request.POST, request.FILES)

        if form.is_valid():
            vin_number = form.cleaned_data.get("vin_number")
            has_defect = form.cleaned_data.get("has_defect", "")
            duration_seconds = request.POST.get("inspection_duration_seconds")

            if not vin_number:
                messages.error(request, "❌ VIN-номер обязателен.")
                return redirect(request.get_full_path())

            history_entry, _ = VINHistory.objects.get_or_create(vin=vin_number)
            zone = "Цех сборки"

            inspection_data = {
                "vin_number": vin_number,
                "controller": request.user.username,
                "has_defect": has_defect,
                "defects": [],
                "date_added": now_almaty_iso(),
                "line": line.lower(),
            }

            try:
                inspection_data["inspection_duration_seconds"] = int(duration_seconds)
            except (ValueError, TypeError):
                inspection_data["inspection_duration_seconds"] = None

            if has_defect == "no":
                # Просто сохраняем осмотр без дефектов
                pass
            else:
                # Обработка дефектов
                defects_list = []
                defect_index = 1
                while f"defect_name_{defect_index}" in request.POST:
                    def_id = lambda field: request.POST.get(f"{field}_{defect_index}")
                    comment = request.POST.get(f"defect_comment_{defect_index}", "").strip()
                    repair_type = request.POST.get(f"defect_repair_type_{defect_index}", "").strip().lower()
                    if repair_type not in ["online", "offline"]:
                        repair_type = "online"  # значение по умолчанию
                    try:
                        quantity = int(request.POST.get(f"defect_quantity_{defect_index}", "1"))
                    except ValueError:
                        quantity = 1
                    images = request.FILES.getlist(f"defect_photo_{defect_index}")
                    zone_name = request.POST.get(f"defect_zone_{defect_index}", "").strip()
                    photo_urls = []

                    for file in images:
                        try:
                            compressed = compress_uploaded_image(file)
                            filename = compressed.name
                            path = f"images/defects/{filename}"
                            full_path = os.path.join(settings.MEDIA_ROOT, path)
                            os.makedirs(os.path.dirname(full_path), exist_ok=True)
                            with open(full_path, "wb+") as destination:
                                for chunk in compressed.chunks():
                                    destination.write(chunk)
                            photo_urls.append(f"{settings.MEDIA_URL}{path}")
                        except Exception as e:
                            print(f"❌ Ошибка при сохранении фото дефекта {defect_index}: {e}")


                    defects_list.append({
                        "zone": zone_name,
                        "unit": AssemblyUnit.objects.filter(pk=def_id("defect_unit")).first().name if def_id(
                            "defect_unit") else "",
                        "name": AssemblyDefect.objects.filter(pk=def_id("defect_name")).first().name if def_id(
                            "defect_name") else "",
                        "nameENG": AssemblyDefect.objects.filter(pk=def_id("defect_name")).first().nameENG if def_id(
                            "defect_name") else "",
                        "grade": AssemblyDefectGrade.objects.filter(pk=def_id("defect_grade")).first().name if def_id(
                            "defect_grade") else "",
                        "responsible": AssemblyDefectResponsible.objects.filter(
                            pk=def_id("defect_responsible")).first().name if def_id("defect_responsible") else "",
                        "quantity": quantity,
                        "repair_type": repair_type,
                        "comment": comment,
                        "photos": photo_urls,
                    })
                    defect_index += 1

                inspection_data["defects"] = defects_list

            history_entry.history.setdefault(zone, {}).setdefault(post_name, []).append(inspection_data)
            history_entry.save()

            messages.success(request, "✅ Данные успешно сохранены.")
            return redirect(request.get_full_path())
        else:
            messages.error(request, "❌ Ошибка в форме. Проверьте данные.")
            print("Форма не прошла валидацию:", form.errors)

    else:
        form = AssemblyTemplateForm()

    return render(request, template_name, {
        "form": form,
        "post_id": post_id,
        "post_name": post_name,
        "user_role": request.user.role if hasattr(request.user, 'role') else '',
        "defects": defects,
        "zones": zones,
        "units": units,
        "grades": grades,
        "responsibles": responsibles,
        "line": line.lower(),
    })


@require_GET
@login_required
def vin_status_api(request):
    """
    GET /assembly/api/vin-status/?vin=...
    Возвращает статусы для УУД и VES (пока VES пустой).
    Формат:
    {
      "vin": "...",
      "uud": {"step": "step1|step2|step3|done|None", "label": "...", "color": "red|yellow|green"},
      "ves": null
    }
    """
    vin = (request.GET.get("vin") or "").strip().upper()
    res = {"vin": vin, "uud": None, "ves": None}

    if not vin:
        return JsonResponse(res)

    vh = VINHistory.objects.filter(vin__iexact=vin).first()
    history = vh.history if vh else {}

    # --- UUD ---
    # ожидаем структуру: history["УУД"]["УУД"] = [ { "steps": "step1|step2|step3|done", ... }, ... ]
    def _latest_uud_step(hist: dict):
        try:
            uud_block = (hist.get("УУД") or {}).get("УУД") or []
            if not isinstance(uud_block, list) or not uud_block:
                return None
            # берём последний по entry_index или updated_at/created_at
            def _key(e):
                # приоритет: entry_index (числовой), затем updated_at, затем created_at
                idx = e.get("entry_index") or 0
                upd = e.get("updated_at") or ""
                crt = e.get("created_at") or ""
                return (int(idx), str(upd), str(crt))
            latest = sorted(uud_block, key=_key)[-1]
            return str(latest.get("steps") or "").lower() or None
        except Exception:
            return None

    step = _latest_uud_step(history)

    # маппинг step -> label/color
    if step in ("step1", "step2"):
        res["uud"] = {"step": step, "label": "Находится на УУД", "color": "red"}
    elif step == "step3":
        res["uud"] = {"step": step, "label": "Ждёт приёма на линию", "color": "yellow"}
    else:
        # done или записей нет → зелёный
        res["uud"] = {"step": step or None, "label": "не числится на УУД", "color": "green"}

    # --- VES (пока пусто) ---
    res["ves"] = None

    return JsonResponse(res)



@login_required
@role_required(["controller", "master", 'head_area'])
def documentation_views(request, line="", post_suffix="", template_name="assembly/documentation.html"):
    ZONE_NAME = "Цех сборки"
    POST_NAME = "Документация"
    DOC_DIR   = os.path.join("images", "assembly_documentation_photos")

    # ───────────────────────────────────────────────────────────────
    # Локальные помощники (чтобы не зависеть от наличия глобальных)
    # ───────────────────────────────────────────────────────────────
    def _find_defect_in_history(history: dict, defect_id: str):
        """
        Возвращает (zone_key, post_key, entry_dict, defect_dict) для дефекта с указанным id.
        Если не найдено — (None, None, None, None).
        """
        if not isinstance(history, dict):
            return None, None, None, None

        for zone_key, posts in history.items():
            if not isinstance(posts, dict):
                continue
            for post_key, entries in posts.items():
                if not isinstance(entries, list):
                    continue
                for entry in entries:
                    defects = (entry or {}).get("defects") or []
                    for defect in defects:
                        # В истории у дефекта ключ обычно "id"
                        if (defect or {}).get("id") == defect_id:
                            return zone_key, post_key, entry, defect
        return None, None, None, None

    def _ensure_uud_container(defect: dict):
        """
        Гарантирует наличие extra.UUD со структурой {"status":..., "history":[...]} и возвращает ссылку на UUD.
        """
        extra = defect.setdefault("extra", {})
        if not isinstance(extra, dict):
            extra = {}
            defect["extra"] = extra
        uud = extra.setdefault("UUD", {})
        if not isinstance(uud, dict):
            uud = {}
            extra["UUD"] = uud
        uud.setdefault("status", None)
        uud.setdefault("history", [])
        if not isinstance(uud["history"], list):
            uud["history"] = []
        return uud

    # ───────────────────────────────────────────────────────────────

    if request.method == "POST":
        vin_number = (
            request.POST.get("vin_number")
            or request.POST.get("vin")
            or request.GET.get("vin")
            or ""
        ).strip().upper()

        if not vin_number:
            messages.error(request, "❌ VIN-номер обязателен.")
            return redirect(request.path)

        # собрать отмеченные как устранённые дефекты (из скрытого поля)
        raw_fixed = (request.POST.get("fixed_defects") or "").strip()
        fixed_ids = [i for i in (raw_fixed.split(",") if raw_fixed else []) if i]

        # сохранить фото
        def iter_image_files(files_dict):
            for key in files_dict:
                for f in files_dict.getlist(key):
                    if getattr(f, "content_type", "").startswith("image/"):
                        yield f

        photo_urls = []
        for file in iter_image_files(request.FILES):
            try:
                img = compress_uploaded_image(file)
                ts_iso = now_almaty_iso()
                ts_safe = ts_iso.replace(":", "-").replace("T", "_").replace("+", "")
                _, ext = os.path.splitext(img.name or "")
                ext = (ext or ".jpg").lower()
                safe_name = f"{vin_number}_{ts_safe}_{uuid4().hex}{ext}"

                rel_path = os.path.join(DOC_DIR, safe_name)
                abs_path = os.path.join(settings.MEDIA_ROOT, rel_path)
                os.makedirs(os.path.dirname(abs_path), exist_ok=True)

                with open(abs_path, "wb+") as dst:
                    for chunk in img.chunks():
                        dst.write(chunk)

                photo_urls.append(f"{settings.MEDIA_URL}{rel_path}".replace("\\", "/"))
            except Exception as e:
                print(f"[documentation_wiews] Ошибка сохранения фото: {e}")

        # запись в историю VIN
        vh, _ = VINHistory.objects.get_or_create(vin=vin_number)

        # ── 1) Применяем решения по отмеченным дефектам ──────────────────────
        applied_ids = []
        if fixed_ids:
            history = vh.history if isinstance(vh.history, dict) else {}
            changed = False

            for d_id in fixed_ids:
                zone_key, post_key, entry, defect = _find_defect_in_history(history, d_id)
                if not defect:
                    continue

                uud = _ensure_uud_container(defect)
                status = uud.get("status")
                # уже закрытые/невозможные не трогаем
                if status in ("resolved", "impossible"):
                    continue

                ts = now_almaty_iso()

                # если статус "checking": дописываем решение в последний тул
                if status == "checking":
                    if uud["history"]:
                        last = uud["history"][-1]
                    else:
                        last = {"id": 1}
                        uud["history"].append(last)

                    if not isinstance(last.get("decision"), dict):
                        last["decision"] = {
                            "by": request.user.username,
                            "at": ts,
                            "status": "resolved",
                            "comment": "Принято на посте Документация",
                        }
                        uud["status"] = "resolved"
                        applied_ids.append(d_id)
                        changed = True

                # если истории нет ИЛИ статус пустой/не устранено — новая запись-решение
                elif (not uud["history"]) or status in (None, "", "not_resolved"):
                    next_id = (uud["history"][-1]["id"] + 1) if uud["history"] else 1
                    uud["history"].append({
                        "id": next_id,
                        "decision": {
                            "by": request.user.username,
                            "at": ts,
                            "status": "resolved",
                            "comment": "Принято на посте Документация",
                        }
                    })
                    uud["status"] = "resolved"
                    applied_ids.append(d_id)
                    changed = True

                # иные статусы не трогаем

            if changed:
                vh.history = history
                vh.save(update_fields=["history", "updated_at"])

        # ── 2) Создаём запись поста «Документация» и кладём applied_ids ──────
        class _PostStub:
            location = ZONE_NAME
            name = POST_NAME

        entry_id = vh.add_entry(post=_PostStub(), defects=[], extra_data={})
        vh.update_entry(
            entry_id,
            added_at=now_almaty_iso(),
            added_by=request.user.username,
            photos=photo_urls,
            VIN=vin_number,
            fixed_defects=applied_ids,   # <-- сохраняем только реально применённые id
        )

        # подчистим вспомогательные поля в самой записи
        zone, post, entry, entries_list = vh.get_entry_by_id(entry_id)
        if entry is not None:
            entry.pop("defects", None)
            entry.pop("extra_data", None)
            vh.save(update_fields=["history", "updated_at"])

        messages.success(request, "✅ Документация сохранена.")
        # Переходим с ?vin=..., чтобы сервер отрисовал дефекты;
        # ваш JS затем уберёт ?vin из адресной строки.
        return redirect(request.path)

    # ===== GET: серверный рендер VIN-инфо и дефектов =====
    vin = (
        request.GET.get("vin")
        or request.POST.get("vin_number")
        or ""
    ).strip().upper()

    context = {
        "post_name": POST_NAME,
        "line": (line or "").lower(),
        "vin": vin,
        "vehicle_info": None,
        "defect_tree": {},
        "defect_list": [],
        "defects_count": 0,
        "history_updated_at": None,
    }

    if vin:
        vh = VINHistory.objects.filter(vin__iexact=vin).first()
        if vh:
            context["defect_tree"] = vh.history or {}
            # Берём дефекты из UUD-истории (как в UUD)
            context["defect_list"] = _build_defects_index(vh.history or {})
            context["defects_count"] = len(context["defect_list"])
            context["history_updated_at"] = vh.updated_at
        try:
            context["vehicle_info"] = _resolve_vehicle_info(vin)
        except Exception:
            pass

    return render(request, template_name, context)


# --- Новый view для простого сканирования VIN и записи только VIN, дату, автора и линию
@login_required
@role_required(["controller", "master", "head_area"])
def assembly_vin_scan_view(request, template_name="assembly/assembly_vin_scan.html", line=None):
    """
    Простой пост: сканируют VIN и жмут «Сохранить».
    Сохраняем VIN, автора, дату и линию.
    """
    if request.method == "POST":
        vin = (request.POST.get("vin") or request.POST.get("vin_number") or "").strip().upper()
        if not vin:
            messages.error(request, "❌ VIN-номер обязателен.")
            return redirect(request.get_full_path())

        obj, created = AssemblyPassLog.record_scan(vin=vin, user=request.user, line=line)
        if created:
            messages.success(request, f"✅ VIN {obj.vin} зафиксирован.")
        else:
            messages.info(request, f"ℹ️ VIN {obj.vin} уже был зафиксирован ранее.")
        return redirect(request.get_full_path())

    # GET — просто рендерим страницу с одним полем VIN
    return render(request, template_name, {"line": line})


@login_required
@role_required(["controller", "master", "head_area"])
def assembly_vin_trimout_view(request, template_name="assembly/assembly_vin_trimout.html", line=None):
    """
    Простой пост: сканируют VIN для TRIM OUT.
    Сохраняем VIN, автора, дату и линию.
    """
    if request.method == "POST":
        vin = (request.POST.get("vin") or request.POST.get("vin_number") or "").strip().upper()
        if not vin:
            messages.error(request, "❌ VIN-номер обязателен.")
            return redirect(request.get_full_path())

        obj, _ = TrimOutPassLog.record_scan(vin=vin, user=request.user, line=line)
        messages.success(request, f"✅ TRIM OUT: VIN {obj.vin} зафиксирован.")
        return redirect(request.get_full_path())

    return render(request, template_name, {"line": line})


@login_required
@role_required(["controller", "master", 'head_area'])
def ves_views(request, line, post_suffix="", template_name="assembly/ves.html"):
    post_id = request.GET.get("post_id")
    post_obj = PostAssembly.objects.filter(id=post_id).first()
    post_name = post_obj.name if post_obj else "Неизвестный пост"

    # 🔽 Справочники
    defects = AssemblyDefect.objects.all()
    zones = AssemblyZone.objects.all()
    units = AssemblyUnit.objects.all()
    grades = AssemblyDefectGrade.objects.all()
    responsibles = AssemblyDefectResponsible.objects.all()

    if request.method == "POST":
        form = AssemblyTemplateForm(request.POST, request.FILES)

        if form.is_valid():
            vin_number = form.cleaned_data.get("vin_number")
            has_defect = form.cleaned_data.get("has_defect", "")
            duration_seconds = request.POST.get("inspection_duration_seconds")

            if not vin_number:
                messages.error(request, "❌ VIN-номер обязателен.")
                return redirect(request.get_full_path())

            history_entry, _ = VINHistory.objects.get_or_create(vin=vin_number)
            zone = "VES"

            inspection_data = {
                "vin_number": vin_number,
                "controller": request.user.username,
                "has_defect": has_defect,
                "defects": [],
                "date_added": now_almaty_iso(),
                "line": line.lower(),
            }

            try:
                inspection_data["inspection_duration_seconds"] = int(duration_seconds)
            except (ValueError, TypeError):
                inspection_data["inspection_duration_seconds"] = None

            if has_defect == "no":
                pass
            else:
                defects_list = []
                defect_index = 1
                while f"defect_name_{defect_index}" in request.POST:
                    def_id = lambda field: request.POST.get(f"{field}_{defect_index}")
                    comment = request.POST.get(f"defect_comment_{defect_index}", "").strip()
                    repair_type = request.POST.get(f"defect_repair_type_{defect_index}", "").strip().lower()
                    if repair_type not in ["online", "offline"]:
                        repair_type = "online"
                    try:
                        quantity = int(request.POST.get(f"defect_quantity_{defect_index}", "1"))
                    except ValueError:
                        quantity = 1
                    images = request.FILES.getlist(f"defect_photo_{defect_index}")
                    zone_name = request.POST.get(f"defect_zone_{defect_index}", "").strip()
                    photo_urls = []

                    for file in images:
                        try:
                            compressed = compress_uploaded_image(file)
                            filename = compressed.name
                            path = f"images/defects/{filename}"
                            full_path = os.path.join(settings.MEDIA_ROOT, path)
                            os.makedirs(os.path.dirname(full_path), exist_ok=True)
                            with open(full_path, "wb+") as destination:
                                for chunk in compressed.chunks():
                                    destination.write(chunk)
                            photo_urls.append(f"{settings.MEDIA_URL}{path}")
                        except Exception as e:
                            print(f"❌ Ошибка при сохранении фото дефекта {defect_index}: {e}")

                    defects_list.append({
                        "zone": zone_name,
                        "unit": AssemblyUnit.objects.filter(pk=def_id("defect_unit")).first().name if def_id("defect_unit") else "",
                        "name": AssemblyDefect.objects.filter(pk=def_id("defect_name")).first().name if def_id("defect_name") else "",
                        "nameENG": AssemblyDefect.objects.filter(pk=def_id("defect_name")).first().nameENG if def_id("defect_name") else "",
                        "grade": AssemblyDefectGrade.objects.filter(pk=def_id("defect_grade")).first().name if def_id("defect_grade") else "",
                        "responsible": AssemblyDefectResponsible.objects.filter(pk=def_id("defect_responsible")).first().name if def_id("defect_responsible") else "",
                        "quantity": quantity,
                        "repair_type": repair_type,
                        "comment": comment,
                        "photos": photo_urls,
                    })
                    defect_index += 1

                inspection_data["defects"] = defects_list

            history_entry.history.setdefault(zone, {}).setdefault(post_name, []).append(inspection_data)
            history_entry.save()

            messages.success(request, "✅ Данные успешно сохранены.")
            return redirect(request.get_full_path())
        else:
            messages.error(request, "❌ Ошибка в форме. Проверьте данные.")
            print("Форма не прошла валидацию:", form.errors)

    else:
        form = AssemblyTemplateForm()

    # --- Универсальный флэттер истории: поддерживает старый и новый форматы
    def _flatten_history(history_root):
        """
        Возвращает список дефектов вида:
        {
            "id": ..., "unit": ..., "name": ..., "grade": ..., "photos": [...],
            "controller": ..., "date_added": ..., "zone": ..., "post": ...
        }
        Поддерживает:
          1) Новый формат: { "ZONE": { "POST": [entry, ...] }, ... }
          2) Промежуточный формат: { "POST": [entry, ...] }
          3) Старый формат: [entry, entry, ...]
        где entry = {"controller","date_added","defects":[...], ...}
        """
        res = []
        if not history_root:
            return res

        def _yield_entries(root, zone_name=None, post_name=None):
            # Список записей сразу
            if isinstance(root, list):
                for entry in root:
                    if isinstance(entry, dict):
                        yield (zone_name or ""), (post_name or ""), entry
                return

            # Словарь
            if isinstance(root, dict):
                for k, v in root.items():
                    # Ветка: пост -> [entries]
                    if isinstance(v, list):
                        for entry in v:
                            if isinstance(entry, dict):
                                yield (zone_name or ""), k, entry
                    # Ветка: зона -> {пост: [entries]}
                    elif isinstance(v, dict):
                        for post_k, entries in v.items():
                            if isinstance(entries, list):
                                for entry in entries:
                                    if isinstance(entry, dict):
                                        yield k, post_k, entry
                            # Иногда бывает пост -> dict (не список) — игнорируем мусор
                    # Иное — пропускаем

        for zone, post, entry in _yield_entries(history_root):
            controller = entry.get("controller") or ""
            when = entry.get("date_added") or ""
            defects = entry.get("defects") or []

            # Нормализуем случай, когда дефект один и пришёл dict
            if isinstance(defects, dict):
                defects = [defects]

            if not isinstance(defects, list):
                continue

            for d in defects:
                if not isinstance(d, dict):
                    continue
                res.append({
                    "id": d.get("id") or "",
                    "unit": d.get("unit") or "",
                    "name": d.get("name") or "",
                    "grade": d.get("grade") or "",
                    "photos": d.get("photos") or [],
                    "controller": controller,
                    "date_added": when,
                    "zone": (d.get("zone") or zone or ""),
                    "post": (post or ""),
                })
        return res

    # --- Собираем карту VIN → список дефектов (без дублирования кода)
    vin_defects_map = {}
    for h in VINHistory.objects.all().only("vin", "history"):
        key = (h.vin or "").upper()
        vin_defects_map[key] = _flatten_history(getattr(h, "history", {}) or {})

    vin_defects_json = json.dumps(vin_defects_map, cls=DjangoJSONEncoder, ensure_ascii=False)

    return render(request, template_name, {
        "form": form,
        "post_id": post_id,
        "post_name": post_name,
        "user_role": request.user.role if hasattr(request.user, 'role') else '',
        "defects": defects,
        "zones": zones,
        "units": units,
        "grades": grades,
        "responsibles": responsibles,
        "line": line.lower(),
        "vin_defects_json": vin_defects_json,
        "vin_defects_data": vin_defects_map,
    })


# --- Новый view: VES-передача/приём ---
@login_required
@role_required(["controller", "master", 'head_area'])
def ves_pass_view(request, template_name="assembly/ves_pass.html"):
    """
    Пост "VES-передача/приём":
    - Сканируют VIN.
    - Если по VIN нет открытой передачи (received_at is NULL) — показываем кнопку "Отдать на VES".
    - Если есть открытая передача — показываем кнопку "Принять с VES".
    - POST с action=give/receive записывает событие.
    """
    # --- Обработка действий ---
    if request.method == "POST":
        vin = (request.POST.get("vin") or request.POST.get("vin_number") or "").strip().upper()
        action = (request.POST.get("action") or "").strip().lower()

        if not vin:
            messages.error(request, "❌ VIN-номер обязателен.")
            return redirect(request.get_full_path())

        try:
            if action == "give":
                VESPassLog.record_give(vin=vin, user=request.user)
                messages.success(request, f"✅ VIN {vin} передан на VES.")
            elif action == "receive":
                obj, updated = VESPassLog.record_receive(vin=vin, user=request.user)
                if updated:
                    messages.success(request, f"✅ VIN {vin} принят с VES.")
                else:
                    messages.info(request, f"ℹ️ Открытой передачи не было. Создана запись с мгновенной отдачей/приёмом для {vin}.")
            else:
                messages.error(request, "❌ Неизвестное действие.")
        except Exception as e:
            messages.error(request, f"❌ Ошибка: {e}")

        # Возврат на ту же страницу, чтобы можно было сканировать следующий VIN
        return redirect(request.get_full_path())

    # --- Рендер страницы / определение статуса по VIN ---
    vin = (request.GET.get("vin") or "").strip().upper()
    open_log = None
    last_log = None
    trace = None

    if vin:
        open_log = VESPassLog.objects.filter(vin=vin, received_at__isnull=True).order_by("-given_at").first()
        last_log = VESPassLog.objects.filter(vin=vin).order_by("-given_at").first()
        trace = TraceData.objects.filter(vin_rk=vin).first()

    # Подготовим краткую информацию для отображения (без жёсткой завязки на шаблон)
    trace_info = None
    if trace:
        trace_info = {
            "vin": trace.vin_rk,
            "model": getattr(trace, "model", ""),
            "body_color": getattr(trace, "body_color", ""),
            "modification": getattr(trace, "modification", ""),
            "engine_number": getattr(trace, "engine_number", ""),
            "config_code": getattr(trace, "config_code", ""),
        }

    context = {
        "vin": vin,
        "open_log": open_log,     # если есть открытая передача (ожидает приёма)
        "last_log": last_log,     # последняя запись по VIN
        "trace": trace_info,      # данные трейсинга (модель/цвет/привод и т.д.)
        "show_give_button": bool(vin and not open_log),
        "show_receive_button": bool(vin and open_log),
    }
    return render(request, template_name, context)


# counter_gwm
# Счетчик TRIM IN
# На линии GWM
# Counter123!


@login_required
@role_required(["controller", "master", "uud_controller", 'head_area'])
def uud_uniq(request, template_name="assembly/uud_uniq.html"):
    """
    Вариант без сохранения: при вводе/скане VIN показываем инфо и дефекты.
    """
    vin = (
        request.POST.get("vin")
        or request.POST.get("vin_number")
        or request.GET.get("vin")
        or ""
    ).strip().upper()

    # Базовый контекст
    context = {
        "vin": vin,
        "vehicle_info": None,   # сюда позже подставим данные машины
        "defect_tree": {},      # полный JSON из VINHistory.history
        "defect_list": [],      # плоский список дефектов (удобно для таблиц/карт)
        "defects_count": 0,
        "history_updated_at": None,
    }

    if request.method == "POST":
        # Пока НИЧЕГО не сохраняем — просто покажем данные
        if not vin:
            messages.error(request, "❌ VIN-номер обязателен.")
            return render(request, template_name, context)

    # Падаем дальше к сбору данных по vin
    if vin:
        # 1) Дефекты из VINHistory
        vh = VINHistory.objects.filter(vin__iexact=vin).first()
        if vh:
            context["defect_tree"] = vh.history or {}
            context["defect_list"] = _build_defects_index(vh.history or {})
            context["defects_count"] = len(context["defect_list"])
            context["history_updated_at"] = vh.updated_at

        # 2) Данные машины (опционально; если в проекте есть соответствующая модель)
        # Если у тебя уже есть точная модель (например, Vehicle/Tracing/Stock и т.п.),
        # просто поправь реализацию _resolve_vehicle_info().
        context["vehicle_info"] = _resolve_vehicle_info(vin)

    # GET без vin или POST с ошибкой — просто рендер
    return render(request, template_name, context)


def _build_defects_index(history: dict) -> list[dict]:
    items: list[dict] = []

    for zone, posts in (history or {}).items():
        for post_name, entries in (posts or {}).items():
            for entry in (entries or []):
                if not isinstance(entry, dict):   # ← защита
                    continue

                defects = entry.get("defects") or []
                if not isinstance(defects, list):  # ← на всякий
                    continue

                for d in defects:
                    if not isinstance(d, dict):    # ← защита
                        continue

                    name = (d.get("name") or d.get("defect") or d.get("nameENG") or "")
                    unit = (d.get("unit") or d.get("detail") or "")
                    photos = (d.get("photos") or d.get("defect_photos") or []) or []
                    grade = d.get("grade") or (d.get("extra") or {}).get("qrr_grade")
                    comment = (d.get("comment") or d.get("custom_defect_explanation") or "")

                    extra = d.get("extra") or {}
                    custom_detail_note = d.get("custom_detail_explanation")
                    if custom_detail_note:
                        extra = {**extra, "detail_note": custom_detail_note}

                    items.append({
                        "zone": zone,
                        "post": post_name,
                        "entry_id": entry.get("id"),
                        "entry_index": entry.get("entry_index"),
                        "date_added": entry.get("date_added"),
                        "controller": entry.get("controller"),

                        "defect_id": d.get("id"),
                        "name": name,
                        "unit": unit,
                        "grade": grade,
                        "photos": photos,
                        "comment": comment,
                        "responsible": d.get("responsible"),
                        "zone_inner": d.get("zone"),
                        "quantity": d.get("quantity"),
                        "repair_type": d.get("repair_type"),
                        "nameENG": d.get("nameENG"),
                        "extra": extra,
                    })

    items.sort(key=lambda x: ((x.get("zone") or ""), (x.get("post") or ""), (x.get("date_added") or ""), (x.get("defect_id") or "")))
    return items




def _resolve_vehicle_info(vin: str) -> dict | None:
    # 1) сначала TraceData — это твой основной источник
    t = (TraceData.objects
         .filter(vin_rk__iexact=vin)
         .order_by('-id')
         .first())
    if t:
        return {
            "vin": t.vin_rk,
            "engine_number": t.engine_number,
            "model": t.model,
            "body_color": t.body_color,
        }

    # 2) дальше — мягкие фоллбэки, если они вообще есть в проекте
    try:
        from supplies.models import VehicleInfo
        v = (VehicleInfo.objects
             .filter(Q(vin__iexact=vin) | Q(vin_number__iexact=vin))
             .order_by("-id").first())
        if v:
            return {
                "vin": vin,
                "engine_number": getattr(v, "engine_number", None),
                "model": getattr(v, "model", None),
                "body_color": getattr(v, "body_color", None),
            }
    except Exception:
        pass

    return None



# Если утилита в другом месте — поправь импорт
try:
    from vehicle_history.utils import now_almaty_iso  # пример
except Exception:
    # запасной вариант — серверное время; позже можно заменить на локальное Алма-Аты
    def now_almaty_iso() -> str:
        return dj_now().isoformat()

ZONE = "УУД"
POST = "УУД"

def _payload(request) -> dict:
    """Читает JSON или form-data и возвращает dict."""
    ctype = (request.content_type or "").lower()
    if "application/json" in ctype:
        try:
            return json.loads((request.body or b"").decode("utf-8") or "{}")
        except Exception:
            return {}
    return request.POST

def _ok(state: dict, message: str | None = None, **extra):
    data = {"status": "ok", "uud_state": state}
    if message:
        data["message"] = message
    data.update(extra)
    return JsonResponse(data)

def _err(message: str, code: int = 400, **extra):
    return JsonResponse({"status": "error", "message": message, **extra}, status=code)

def _get_uud_sessions(history: dict, create: bool = False) -> list:
    """
    Достаёт массив сессий УУД из history[ZONE][POST].
    При create=True — мягко инициализирует структуру.
    """
    if not isinstance(history, dict):
        if not create:
            return []
        history = {}
    zone = history.get(ZONE)
    if zone is None:
        if not create:
            return []
        history[ZONE] = zone = {}
    post = zone.get(POST)
    if post is None:
        if not create:
            return []
        zone[POST] = post = []

    # post здесь — список карточек-сессий
    if not isinstance(post, list):
        if create:
            zone[POST] = []
            return zone[POST]
        return []
    return post

def _active_session(sessions: list) -> tuple[int | None, dict | None]:
    """
    Возвращает (index, session) для последней незавершённой (steps != 'done') сессии.
    """
    for idx in range(len(sessions) - 1, -1, -1):
        s = sessions[idx]
        if isinstance(s, dict) and s.get("steps") != "done":
            return idx, s
    return None, None

def _state_from_sessions(sessions: list) -> dict:
    """
    Собирает компактное состояние для фронта.
    """
    idx, sess = _active_session(sessions)
    if sess:
        return {
            "has_active": True,
            "session_id": (sess.get("id") or (idx + 1)),
            "status": sess.get("status"),
            "step": sess.get("steps"),
            "updated_at": sess.get("updated_at"),
        }
    # нет активной — значит либо нет вообще, либо последний цикл done
    last_done = sessions[-1] if sessions else None
    return {
        "has_active": False,
        "last_session_id": (last_done.get("id") if isinstance(last_done, dict) else None),
        "last_status": (last_done.get("status") if isinstance(last_done, dict) else None),
        "step": None,
        "status": None,
    }

def _require_vin(request):
    vin = (_payload(request).get("vin") or request.GET.get("vin") or "").strip().upper()
    if not vin:
        raise ValueError("VIN обязателен.")
    return vin

# ──────────────────────────────────────────────────────────────────────────────
# 1) Мы отдали машину на УУД (создаёт новую сессию, если нет активной)
# Роли: controller, master
# ──────────────────────────────────────────────────────────────────────────────
@login_required
@role_required(["controller", "master", 'head_area'])
@require_http_methods(["POST"])
def us_to_uud(request):
    try:
        vin = _require_vin(request)
    except ValueError as e:
        return _err(str(e), 400)

    # Достаём историю по VIN с блокировкой строки
    with transaction.atomic():
        try:
            vh = VINHistory.objects.select_for_update().get(vin__iexact=vin)
        except VINHistory.DoesNotExist:
            return _err("VIN не найден в истории.", 404)

        history = vh.history if isinstance(vh.history, dict) else {}
        sessions = _get_uud_sessions(history, create=True)
        idx, active = _active_session(sessions)

        # Если активная сессия уже есть — просто вернём текущее состояние (идемпотентно)
        if active:
            state = _state_from_sessions(sessions)
            return _ok(state, message="Уже есть активная сессия УУД; новый цикл не создан.")

        # Создаём новую сессию
        ts = now_almaty_iso()
        new_id = (sessions[-1].get("id") + 1) if (sessions and isinstance(sessions[-1], dict) and isinstance(sessions[-1].get("id"), int)) else (len(sessions) + 1)
        sess = {
            "id": new_id,
            "status": "hold",
            "steps": "step1",
            "extra_data": {
                "step1_by": request.user.username,
                "step1_at": ts,
            },
            "created_at": ts,
            "updated_at": ts,
        }
        sessions.append(sess)
        # сохранить обратно
        vh.history = history
        vh.save(update_fields=["history", "updated_at"])

        state = _state_from_sessions(sessions)
        return _ok(state, message="Машина отдана на УУД (step1).")

# ──────────────────────────────────────────────────────────────────────────────
# 2) УУД забрал её в зону УУД
# Роль: uud_controller
# ──────────────────────────────────────────────────────────────────────────────
@login_required
@role_required(["uud_controller"])
@require_http_methods(["POST"])
def uud_to_uudzone(request):
    try:
        vin = _require_vin(request)
    except ValueError as e:
        return _err(str(e), 400)

    with transaction.atomic():
        try:
            vh = VINHistory.objects.select_for_update().get(vin__iexact=vin)
        except VINHistory.DoesNotExist:
            return _err("VIN не найден в истории.", 404)

        history = vh.history if isinstance(vh.history, dict) else {}
        sessions = _get_uud_sessions(history, create=False)

        if not sessions:
            return _err("Машина ещё не передана на УУД (нет циклов).", 400)

        idx, active = _active_session(sessions)
        if not active:
            return _err("Активной сессии УУД нет (последний цикл завершён).", 400)

        step = active.get("steps")
        if step == "step1":
            # OK → step2
            ts = now_almaty_iso()
            active["steps"] = "step2"
            active.setdefault("extra_data", {})
            active["extra_data"]["step2_by"] = request.user.username
            active["extra_data"]["step2_at"] = ts
            active["updated_at"] = ts
        elif step in ("step2", "step3"):
            # идемпотентность: если уже step2/step3 — просто вернуть состояние
            pass
        else:
            # неправильный порядок
            return _err("Неверный порядок: сначала нужно отдать на УУД (step1).", 400)

        vh.history = history
        vh.save(update_fields=["history", "updated_at"])
        return _ok(_state_from_sessions(sessions), message="Зафиксировано: забрали на зону УУД (step2).")

# ──────────────────────────────────────────────────────────────────────────────
# 3) С зоны УУД принесли на УУД
# Роль: uud_controller
# ──────────────────────────────────────────────────────────────────────────────
@login_required
@role_required(["uud_controller"])
@require_http_methods(["POST"])
def uudzone_to_uud(request):
    try:
        vin = _require_vin(request)
    except ValueError as e:
        return _err(str(e), 400)

    with transaction.atomic():
        try:
            vh = VINHistory.objects.select_for_update().get(vin__iexact=vin)
        except VINHistory.DoesNotExist:
            return _err("VIN не найден в истории.", 404)

        history = vh.history if isinstance(vh.history, dict) else {}
        sessions = _get_uud_sessions(history, create=False)

        if not sessions:
            return _err("Машина ещё не передана на УУД (нет циклов).", 400)

        idx, active = _active_session(sessions)
        if not active:
            return _err("Активной сессии УУД нет (последний цикл завершён).", 400)

        step = active.get("steps")
        if step == "step2":
            # OK → step3
            ts = now_almaty_iso()
            active["steps"] = "step3"
            active.setdefault("extra_data", {})
            active["extra_data"]["step3_by"] = request.user.username
            active["extra_data"]["step3_at"] = ts
            active["updated_at"] = ts
        elif step == "step3":
            # идемпотентно
            pass
        else:
            return _err("Неверный порядок: требуется шаг 'забрали на зону УУД' (step2).", 400)

        vh.history = history
        vh.save(update_fields=["history", "updated_at"])
        return _ok(_state_from_sessions(sessions), message="Зафиксировано: доставили с зоны на УУД (step3).")

# ──────────────────────────────────────────────────────────────────────────────
# 4) Мы приняли с УУД
# Роли: controller, master
# ──────────────────────────────────────────────────────────────────────────────
@login_required
@role_required(["controller", "master", 'head_area'])
@require_http_methods(["POST"])
def uud_to_us(request):
    try:
        vin = _require_vin(request)
    except ValueError as e:
        return _err(str(e), 400)

    with transaction.atomic():
        try:
            vh = VINHistory.objects.select_for_update().get(vin__iexact=vin)
        except VINHistory.DoesNotExist:
            return _err("VIN не найден в истории.", 404)

        history = vh.history if isinstance(vh.history, dict) else {}
        sessions = _get_uud_sessions(history, create=False)

        if not sessions:
            return _err("Машина ещё не передана на УУД (нет циклов).", 400)

        idx, active = _active_session(sessions)
        if not active:
            return _err("Активной сессии УУД нет (последний цикл завершён).", 400)

        step = active.get("steps")
        if step != "step3":
            return _err("Неверный порядок: принять можно только после возврата с зоны (step3).", 400)

        ts = now_almaty_iso()
        # фиксируем step4
        active["steps"] = "step4"
        active.setdefault("extra_data", {})
        active["extra_data"]["step4_by"] = request.user.username
        active["extra_data"]["step4_at"] = ts
        active["updated_at"] = ts

        # завершить цикл
        active["steps"] = "done"
        active["status"] = "on_line"   # только в done — on_line, иначе hold

        vh.history = history
        vh.save(update_fields=["history", "updated_at"])

        return _ok(_state_from_sessions(sessions), message="Зафиксировано: машина принята с УУД (done).")

# ──────────────────────────────────────────────────────────────────────────────
# (Опционально) GET-эндпоинт, чтобы фронт мог спросить текущее состояние по VIN
# ──────────────────────────────────────────────────────────────────────────────
@login_required
@require_http_methods(["GET"])
def uud_current_state(request):
    vin = (request.GET.get("vin") or "").strip().upper()
    if not vin:
        return _err("VIN обязателен.", 400)

    vh = VINHistory.objects.filter(vin__iexact=vin).first()
    if not vh or not isinstance(vh.history, dict):
        return _ok({"has_active": False, "step": None, "status": None}, message="История не найдена.")

    sessions = _get_uud_sessions(vh.history, create=False)
    return _ok(_state_from_sessions(sessions))





# ──────────────────────────────────────────────────────────────────────────────
# Хелперы для работы с дефектом внутри JSON-истории
# ──────────────────────────────────────────────────────────────────────────────
def _find_defect_by_id(history: dict, defect_id: str):
    """
    Возвращает (zone, post_name, entry_dict, defect_dict) для дефекта по ID.
    Если не найдено — (None, None, None, None)
    """
    if not isinstance(history, dict):
        return None, None, None, None

    for zone, posts in history.items():
        if not isinstance(posts, dict):
            continue
        for post_name, entries in posts.items():
            if not isinstance(entries, list):
                continue
            for entry in entries:
                if not isinstance(entry, dict):
                    continue
                defects = entry.get("defects") or []
                if not isinstance(defects, list):
                    continue
                for d in defects:
                    if isinstance(d, dict) and str(d.get("id")) == str(defect_id):
                        return zone, post_name, entry, d
    return None, None, None, None


def _ensure_uud_container(defect: dict) -> dict:
    """
    Гарантирует наличие блока defect['extra']['UUD'].
    Возвращает dict UUD-блока.
    Структура:
      extra: {
        ...,
        "UUD": {
          "status": "checking" | "resolved" | "not_resolved" | None,
          "history": [ { "id": 1, "fix": {...}, "decision": {...}? }, ... ]
        }
      }
    """
    extra = defect.setdefault("extra", {})
    uud = extra.get("UUD")
    if not isinstance(uud, dict):
        uud = {"status": None, "history": []}
        extra["UUD"] = uud
    if "history" not in uud or not isinstance(uud["history"], list):
        uud["history"] = []
    return uud


def _save_uploaded_photos(request, vin: str) -> list[str]:
    """
    Сохраняет все загруженные фото (input name="photos") и возвращает список путей.
    Папка: images/uud_fixes/
    """
    saved = []
    files = request.FILES.getlist("photos") or []
    for f in files:
        # нормализуем расширение
        base, ext = os.path.splitext(getattr(f, "name", "") or "")
        ext = (ext or ".jpg").lower()
        fname = f"{vin}_{uuid.uuid4().hex}{ext}"
        path = os.path.join("images", "uud_fixes", fname).replace("\\", "/")
        saved_path = default_storage.save(path, f)  # type: ignore
        # для единообразия — URL-путь через /media/... если у тебя MEDIA_URL=/media/
        if not saved_path.startswith("/"):
            saved_path = "/" + saved_path
        saved.append(saved_path if saved_path.startswith("/media/") else "/media/" + saved_path.lstrip("/"))
    return saved


# ──────────────────────────────────────────────────────────────────────────────
# 1) UUD-контроллер отправляет результат устранения (создаёт новую запись в history)
#    Видна и доступна ТОЛЬКО роли "uud_controller".
#    Бизнес-правила:
#      - если UUD.status отсутствует или == "not_resolved" → создаём новый history-элемент, ставим status="checking"
#      - если UUD.status == "checking" → 409 (уже на проверке)
#      - если UUD.status == "resolved" → 409 (уже принято)
# ──────────────────────────────────────────────────────────────────────────────
@login_required
@role_required(["uud_controller"])
@require_http_methods(["POST"])
def uud_defect_submit_fix(request):
    """
    POST form-data или JSON:
      vin: str
      defect_id: str
      comment: str (optional)
      photos: <files>[] (optional)
    """
    data = _payload(request)
    vin = (data.get("vin") or "").strip().upper()
    defect_id = (data.get("defect_id") or "").strip()

    if not vin or not defect_id:
        return _err("VIN и defect_id обязательны.", 400)

    with transaction.atomic():
        try:
            vh = VINHistory.objects.select_for_update().get(vin__iexact=vin)
        except VINHistory.DoesNotExist:
            return _err("VIN не найден в истории.", 404)

        history = vh.history if isinstance(vh.history, dict) else {}
        zone, post_name, entry, defect = _find_defect_by_id(history, defect_id)
        if not defect:
            return _err("Дефект не найден.", 404)

        # статус и история UUD
        uud = _ensure_uud_container(defect)
        status = uud.get("status")

        if status == "checking":
            return _err("По данному дефекту уже идёт проверка.", 409)
        if status == "resolved":
            return _err("Дефект уже отмечен как устранённый.", 409)
        # ➕ ДОБАВИТЬ:
        if status == "impossible":
            return _err("Дефект отмечен как невозможный к исправлению.", 409)

        # сохраняем фото
        photos = _save_uploaded_photos(request, vin)

        ts = now_almaty_iso()
        comment = (data.get("comment") or "").strip()

        # новый цикл устранения
        next_id = (uud["history"][-1]["id"] + 1) if uud["history"] else 1
        uud["history"].append({
            "id": next_id,
            "fix": {
                "by": request.user.username,
                "at": ts,
                "comment": comment,
                "photos": photos,
            }
        })
        uud["status"] = "checking"

        vh.history = history
        vh.save(update_fields=["history", "updated_at"])

        # вернём актуальное состояние UUD по дефекту
        return JsonResponse({
            "status": "ok",
            "message": "Отправлено на проверку (checking).",
            "vin": vin,
            "defect_id": defect_id,
            "UUD": uud,
        })


# ──────────────────────────────────────────────────────────────────────────────
# 2) Контроллер/Мастер принимает решение: "устранено" или "не устранено"
#    Видно/доступно ролям "controller", "master".
#    Правила:
#      - работает ТОЛЬКО, если текущий статус == "checking" и есть открытая последняя history-запись без decision
#      - decision = "resolved" → status="resolved"
#      - decision = "not_resolved" (+ reason?) → status="not_resolved"
# ──────────────────────────────────────────────────────────────────────────────
@login_required
@role_required(["controller", "master", 'head_area'])
@require_http_methods(["POST"])
def uud_defect_decide(request):
    """
    POST JSON или form-data:
      vin: str
      defect_id: str
      decision: "resolved" | "not_resolved"
      comment: str (optional, обязательно для 'not_resolved' по UX)
    """
    data = _payload(request)
    vin = (data.get("vin") or "").strip().upper()
    defect_id = (data.get("defect_id") or "").strip()
    decision = (data.get("decision") or "").strip().lower()
    comment = (data.get("comment") or "").strip()

    if not vin or not defect_id or decision not in ("resolved", "not_resolved"):
        return _err("VIN, defect_id и корректное decision обязательны.", 400)

    with transaction.atomic():
        try:
            vh = VINHistory.objects.select_for_update().get(vin__iexact=vin)
        except VINHistory.DoesNotExist:
            return _err("VIN не найден в истории.", 404)

        history = vh.history if isinstance(vh.history, dict) else {}
        zone, post_name, entry, defect = _find_defect_by_id(history, defect_id)
        if not defect:
            return _err("Дефект не найден.", 404)

        uud = _ensure_uud_container(defect)
        if uud.get("status") != "checking":
            return _err("Неверное состояние: решение возможно только из статуса 'checking'.", 409)

        if not uud["history"]:
            return _err("История устранения пуста, нечего принимать.", 409)

        last = uud["history"][-1]
        if "decision" in last and isinstance(last["decision"], dict):
            return _err("Последний цикл уже закрыт решением.", 409)

        ts = now_almaty_iso()
        # фиксируем решение
        last["decision"] = {
            "by": request.user.username,
            "at": ts,
            "status": decision,
            **({"comment": comment} if comment else {}),
        }
        uud["status"] = decision  # 'resolved' или 'not_resolved'

        vh.history = history
        vh.save(update_fields=["history", "updated_at"])

        return JsonResponse({
            "status": "ok",
            "message": ("Принято: устранено." if decision == "resolved" else "Принято: не устранено."),
            "vin": vin,
            "defect_id": defect_id,
            "UUD": uud,
        })


# ──────────────────────────────────────────────────────────────────────────────
# 3) (Опционально) Получить UUD-состояние и историю по конкретному дефекту
#    Удобно для открытия нижней панели/окна без полной перезагрузки.
# ──────────────────────────────────────────────────────────────────────────────
@login_required
@require_http_methods(["GET"])
def uud_defect_info(request):
    """
    GET:
      vin: str
      defect_id: str
    Ответ:
      { status: "ok", vin, defect_id, UUD: {...}, defect: {...минимум...} }
    """
    vin = (request.GET.get("vin") or "").strip().upper()
    defect_id = (request.GET.get("defect_id") or "").strip()

    if not vin or not defect_id:
        return _err("VIN и defect_id обязательны.", 400)

    vh = VINHistory.objects.filter(vin__iexact=vin).first()
    if not vh or not isinstance(vh.history, dict):
        return _err("История по VIN не найдена.", 404)

    zone, post_name, entry, defect = _find_defect_by_id(vh.history, defect_id)
    if not defect:
        return _err("Дефект не найден.", 404)

    uud = (_ensure_uud_container(defect) or {})
    # отдаём компактную карточку дефекта для UI
    card = {
        "id": defect.get("id"),
        "name": defect.get("name"),
        "unit": defect.get("unit"),
        "grade": defect.get("grade"),
        "photos": defect.get("photos") or [],
        "zone": zone,
        "post": post_name,
        "entry_id": entry.get("id") if isinstance(entry, dict) else None,
    }

    return JsonResponse({
        "status": "ok",
        "vin": vin,
        "defect_id": defect_id,
        "UUD": uud,
        "defect": card,
    })


@login_required
@role_required(["uud_controller"])
@require_http_methods(["POST"])
def uud_defect_mark_impossible(request):
    """
    Пометить дефект как 'невозможно устранить' со стороны УУД.

    POST form-data или JSON:
      vin: str
      defect_id: str
      comment: str (optional)
      photos: <files>[] (optional)
    """
    data = _payload(request)
    vin = (data.get("vin") or "").strip().upper()
    defect_id = (data.get("defect_id") or "").strip()
    comment = (data.get("comment") or "").strip()

    if not vin or not defect_id:
        return _err("VIN и defect_id обязательны.", 400)

    with transaction.atomic():
        try:
            vh = VINHistory.objects.select_for_update().get(vin__iexact=vin)
        except VINHistory.DoesNotExist:
            return _err("VIN не найден в истории.", 404)

        history = vh.history if isinstance(vh.history, dict) else {}
        zone, post_name, entry, defect = _find_defect_by_id(history, defect_id)
        if not defect:
            return _err("Дефект не найден.", 404)

        uud = _ensure_uud_container(defect)
        status = uud.get("status")

        if status == "checking":
            return _err("По данному дефекту уже идёт проверка.", 409)
        if status == "resolved":
            return _err("Дефект уже отмечен как устранённый.", 409)
        if status == "impossible":
            # Идемпотентность: уже помечено — просто вернём текущее состояние
            return JsonResponse({
                "status": "ok",
                "message": "Уже отмечено как невозможное к исправлению.",
                "vin": vin,
                "defect_id": defect_id,
                "UUD": uud,
            })

        # Можно (необязательно) приложить фото-доказательства
        photos = _save_uploaded_photos(request, vin)

        ts = now_almaty_iso()
        next_id = (uud["history"][-1]["id"] + 1) if uud["history"] else 1
        uud["history"].append({
            "id": next_id,
            "fix": {
                "kind": "impossible",           # ← важно
                "by": request.user.username,
                "at": ts,
                "comment": comment,
                "photos": photos,
            }
        })
        uud["status"] = "impossible"

        vh.history = history
        vh.save(update_fields=["history", "updated_at"])

        return JsonResponse({
            "status": "ok",
            "message": "Отмечено как невозможное к исправлению.",
            "vin": vin,
            "defect_id": defect_id,
            "UUD": uud,
        })

def uud_defect_decide(request):
    """
    POST JSON или form-data:
      vin: str
      defect_id: str
      decision: "resolved" | "not_resolved" | "impossible"
      comment: str (optional)
    """
    data = _payload(request)
    vin = (data.get("vin") or "").strip().upper()
    defect_id = (data.get("defect_id") or "").strip()
    decision = (data.get("decision") or "").strip().lower()
    comment = (data.get("comment") or "").strip()

    if not vin or not defect_id or decision not in ("resolved", "not_resolved", "impossible"):
        return _err("VIN, defect_id и корректный decision обязательны.", 400)

    with transaction.atomic():
        try:
            vh = VINHistory.objects.select_for_update().get(vin__iexact=vin)
        except VINHistory.DoesNotExist:
            return _err("VIN не найден в истории.", 404)

        history = vh.history if isinstance(vh.history, dict) else {}
        zone, post_name, entry, defect = _find_defect_by_id(history, defect_id)
        if not defect:
            return _err("Дефект не найден.", 404)

        uud = _ensure_uud_container(defect)

        if not uud["history"]:
            return _err("История устранения пуста, нечего принимать.", 409)

        last = uud["history"][-1]
        if "decision" in last and isinstance(last["decision"], dict):
            return _err("Последний цикл уже закрыт решением.", 409)

        ts = now_almaty_iso()

        if decision in ("resolved", "not_resolved"):
            if uud.get("status") != "checking":
                return _err("Решение возможно только из статуса 'checking'.", 409)

            last["decision"] = {
                "by": request.user.username,
                "at": ts,
                "status": decision,
                **({"comment": comment} if comment else {}),
            }
            # здесь статус принимает значение решения
            uud["status"] = decision

        else:  # decision == "impossible"
            if uud.get("status") != "impossible":
                return _err("Подтверждение 'невозможно' доступно только при статусе 'impossible'.", 409)
            fix = last.get("fix") or {}
            if fix.get("kind") != "impossible":
                return _err("Последний цикл не является 'невозможно к исправлению'.", 409)

            last["decision"] = {
                "by": request.user.username,
                "at": ts,
                "status": "impossible",
                **({"comment": comment} if comment else {}),
            }
            # статус остаётся 'impossible'
            # uud["status"] = "impossible"

        vh.history = history
        vh.save(update_fields=["history", "updated_at"])

        return JsonResponse({
            "status": "ok",
            "message": (
                "Принято: устранено."
                if decision == "resolved" else
                ("Принято: не устранено." if decision == "not_resolved" else "Подтверждено: невозможно устранить.")
            ),
            "vin": vin,
            "defect_id": defect_id,
            "UUD": uud,
        })






#таблица документации настройка -----------------------------------------------------------------------------------------


ZONE_DOC   = "Цех сборки"
POST_DOC   = "Документация"


def _doc_parse_iso_no_convert(value):
    """Парсим ISO-дату без конвертации TZ и без сдвигов."""
    if not value:
        return None
    from datetime import datetime
    try:
        s = value.replace("Z", "+00:00") if isinstance(value, str) else value
        return datetime.fromisoformat(s)
    except Exception:
        return None


def _doc_count_impossible_in_zone(history_dict, zone_name):
    """
    Считает дефекты со статусом РОВНО 'impossible' в указанной зоне (по всем постам зоны).
    """
    if not isinstance(history_dict, dict):
        return 0
    zone = history_dict.get(zone_name, {})
    total = 0
    if isinstance(zone, dict):
        for _post, entries in zone.items():
            if not isinstance(entries, list):
                continue
            for entry in entries:
                if not isinstance(entry, dict):
                    continue
                defects = entry.get("defects") or []
                if not isinstance(defects, list):
                    continue
                for d in defects:
                    if not isinstance(d, dict):
                        continue
                    if (d.get("status") or "").strip().lower() == "impossible":
                        total += 1
    return total


def _doc_build_trace_map(vins: set[str]) -> dict[str, tuple[str, str, str, str]]:
    """
    Возвращает {VIN: (brand, model, body_color, config_code)} из TraceData.
    Не падает, если таблицы/данных нет.
    """
    result: dict[str, tuple[str, str, str, str]] = {}
    if not vins:
        return result
    try:
        from supplies.models import TraceData
        qs = TraceData.objects.filter(vin_rk__in=vins).only(
            "vin_rk", "brand", "model", "body_color", "config_code"
        )
        for t in qs:
            brand = (t.brand or "").strip()
            model = (t.model or "").strip()
            body_color = (getattr(t, "body_color", "") or "").strip()
            config_code = (getattr(t, "config_code", "") or "").strip()
            result[t.vin_rk] = (brand, model, body_color, config_code)
    except Exception:
        # Тихо игнорируем любые ошибки и возвращаем то, что успели собрать.
        pass
    return result


@login_required
@role_required(["controller", "master", "head_area"])
def documentation_table_view(request, template_name="assembly/documentation_table.html"):
    """
    Серверная таблица по посту «Документация».
    Колонки: VIN, бренд, модель, дата, время (как в базе), added_by,
             Да/Нет (есть ли impossible), количество impossible (последняя колонка),
             added_at (сырое ISO из entry).
    """
    from vehicle_history.models import VINHistory

    # Локальный хелпер: посчитать количество impossible по всей истории VIN
    def _count_impossible_in_history(history: dict) -> int:
        if not isinstance(history, dict):
            return 0
        total = 0
        for zone_posts in history.values():
            if not isinstance(zone_posts, dict):
                continue
            for post_entries in zone_posts.values():
                if not isinstance(post_entries, list):
                    continue
                for entry in post_entries:
                    defects = entry.get("defects") or []
                    if not isinstance(defects, list):
                        continue
                    for d in defects:
                        status = (
                            ((d or {}).get("extra") or {})
                            .get("UUD", {})
                            .get("status", "")
                        )
                        s = str(status).lower()
                        if s in ("impossible", "imposible"):
                            total += 1
        return total

    rows = []
    vins_for_trace: set[str] = set()
    impossible_by_vin: dict[str, int] = {}

    # Берём записи, сортируем по updated_at у VINHistory
    for vh in VINHistory.objects.only("vin", "history", "updated_at").order_by("-updated_at"):
        vin = (vh.vin or "").strip()
        history = vh.history or {}

        # Посчитать impossible один раз на VIN
        if vin and vin not in impossible_by_vin:
            impossible_by_vin[vin] = _count_impossible_in_history(history)

        # Документация хранится в зоне/посте
        zone_block = history.get(ZONE_DOC, {})
        posts = zone_block.get(POST_DOC, [])
        if not isinstance(posts, list):
            continue

        if vin:
            vins_for_trace.add(vin)

        for entry in posts:
            if not isinstance(entry, dict):
                continue

            added_by = entry.get("added_by") or ""
            # ✅ Берём сырое added_at (если нет — date_added), БЕЗ сорт-ключей
            added_at_raw = entry.get("added_at") or entry.get("date_added") or ""

            # Для удобства отображения отдельно выведем date/time (как раньше), но считаем из added_at_raw
            dt = _doc_parse_iso_no_convert(added_at_raw)
            date_str = dt.strftime("%d.%m.%Y") if dt else ""
            time_str = dt.strftime("%H:%M") if dt else ""

            imp_count = impossible_by_vin.get(vin, 0)

            rows.append({
                "vin": entry.get("VIN") or vin,
                "brand": "",
                "model": "",
                "body_color": "",  # ← новое поле
                "config_code": "",  # ← новое поле
                "date": date_str,
                "time": time_str,
                "added_by": added_by,
                "impossible_yesno": "Да" if imp_count > 0 else "Нет",
                "impossible_defects": imp_count,
                "added_at": added_at_raw,
            })

    # Подтягиваем бренд/модель пачкой
    trace_map = _doc_build_trace_map(vins_for_trace)  # обновим сам хелпер на шаге 3
    for r in rows:
        data = trace_map.get(r["vin"])
        if isinstance(data, (list, tuple)):
            # обратная совместимость: 2 или 4 значения
            if len(data) == 2:
                brand, model = data
                body_color = ""
                config_code = ""
            else:
                brand, model, body_color, config_code = (list(data) + ["", "", "", ""])[:4]
        elif isinstance(data, dict):
            # если вдруг хелпер вернёт dict
            brand = data.get("brand", "")
            model = data.get("model", "")
            body_color = data.get("body_color", "")
            config_code = data.get("config_code", "")
        else:
            brand = model = body_color = config_code = ""

        r["brand"] = brand
        r["model"] = model
        r["body_color"] = body_color
        r["config_code"] = config_code
    # Сортировка по сырому added_at (ISO-строка сортируется лексикографически корректно), затем по VIN
    rows.sort(key=lambda r: (r.get("added_at", ""), r.get("vin", "")), reverse=True)

    return render(request, template_name, {"rows": rows})



# ====== ЗАМЕНИ ЭТИ 2 ФУНКЦИИ ПОД СВОЮ БД ======
def fetch_bundle_for_vin(vin: str) -> Dict[str, Any]:
    from vehicle_history.models import VINHistory
    vh = (
        VINHistory.objects
        .filter(vin__iexact=str(vin).strip())
        .only("history", "updated_at")
        .order_by("-updated_at")  # ← всегда самая свежая история
        .first()
    )
    return (vh.history or {}) if vh else {}

def get_base_attrs_for_vin(vin: str) -> Dict[str, str]:
    """Бренд/модель/цвет/код берём из supplies.TraceData по vin_rk."""
    from supplies.models import TraceData
    tr = (TraceData.objects
          .filter(vin_rk__iexact=str(vin).strip())
          .only("brand", "model", "body_color", "config_code")
          .order_by("-date_added")
          .first())
    return {
        "brand": (getattr(tr, "brand", "") or ""),
        "model": (getattr(tr, "model", "") or ""),
        "body_color": (getattr(tr, "body_color", "") or ""),
        "config_code": (getattr(tr, "config_code", "") or ""),
    }


# ====== ХЕЛПЕРЫ ПАРСИНГА/ФОРМАТОВ ======
def split_dt(dt_str: str) -> Tuple[str, str]:
    """'2025-09-11T11:48:55.153472+00:00' -> ('11.09.2025', '11:48')"""
    if not dt_str:
        return "", ""
    try:
        # поддержка ISO с таймзоной
        dt = datetime.fromisoformat(dt_str.replace("Z", "+00:00"))
    except Exception:
        # на всякий случай простой разбор
        try:
            date_part, time_part = dt_str.split("T", 1)
            dt = datetime.fromisoformat(date_part + "T" + time_part[:8])
        except Exception:
            return "", ""
    return dt.strftime("%d.%m.%Y"), dt.strftime("%H:%M")


def image_stream_from_media(
    rel_or_abs_path: str,
    max_width_px: int = 140,   # ─ эти аргументы теперь не влияют на ресэмплинг,
    max_height_px: int = 110,  #   оставлены для совместимости сигнатуры
    jpeg_quality: int = 92,
) -> io.BytesIO | None:
    """
    Открывает изображение из MEDIA_ROOT (или абсолютного пути) и кодирует в JPEG.
    Больше НЕ уменьшаем картинку до 140×110 — сохраняем высокое разрешение,
    лишь мягко ограничиваем длинную сторону (pixel_cap_long_side) и даём умеренную
    компрессию для разумного веса. Это позволяет растягивать картинку в Excel
    без сильной пикселизации.

    Возвращает поток JPEG (BytesIO) или None.
    """
    pixel_cap_long_side = 2048  # мягкий «потолок» по длинной стороне

    if not rel_or_abs_path:
        return None

    p = rel_or_abs_path
    if p.startswith("/media/"):
        p = p[len("/media/"):]
    if not os.path.isabs(p):
        p = os.path.join(settings.MEDIA_ROOT, p)
    if not os.path.exists(p):
        return None

    try:
        im = Image.open(p).convert("RGB")  # Excel не любит PNG с альфой
        w, h = im.size
        if not w or not h:
            return None

        # Мягкий кап по длинной стороне
        long_side = max(w, h)
        if pixel_cap_long_side and long_side > pixel_cap_long_side:
            scale = pixel_cap_long_side / float(long_side)
            new_w, new_h = max(1, int(w * scale)), max(1, int(h * scale))
            im = im.resize((new_w, new_h), Image.LANCZOS)

        bio = io.BytesIO()
        im.save(
            bio,
            format="JPEG",
            quality=jpeg_quality,   # ~92: хорошее качество и умеренный вес
            optimize=True,
            subsampling=1           # 4:2:2 — компромисс чёткости/веса
        )
        bio.seek(0)
        return bio
    except Exception:
        return None


def _split_dt_iso_to_date_time(dt_str: str) -> Tuple[str, str]:
    # безопасный парсер: '2025-09-13T11:09:09.279269+00:00' → ('13.09.2025','11:09')
    try:
        dt = datetime.fromisoformat((dt_str or "").replace("Z", "+00:00"))
        return dt.strftime("%d.%m.%Y"), dt.strftime("%H:%M")
    except Exception:
        return "", ""


def iter_normalized_defects(bundle: dict):
    """
    Генератор по дефектам из ОБОИХ цехов с унификацией полей.
    Для каждой найденной записи-осмотра (inspection/post entry) с дефектами
    yield словарь:
      {
        "detail": str, "defect": str, "grade": str,
        "def_photos": List[str], "controller": str,
        "found_date": "DD.MM.YYYY", "found_time": "HH:MM",
        "uud": dict | {},
      }
    """
    if not isinstance(bundle, dict):
        return

    # 1) Цех поставки — ТЕКУЩЕЕ поведение (оставляем полностью)
    supply = bundle.get("Цех поставки") or {}
    if isinstance(supply, dict):
        for _zone_name, inspections in supply.items():
            for insp in (inspections or []):
                controller = (insp.get("controller") or "").strip()
                found_date, found_time = split_dt(insp.get("date_added") or "")
                for d in (insp.get("defects") or []):
                    uud = ((d.get("extra") or {}).get("UUD") or {})
                    defect = (d.get("defect") or d.get("name") or d.get("nameENG") or "").strip()
                    detail = (d.get("detail") or d.get("unit") or "").strip()
                    grade  = (d.get("grade") or "").strip()
                    def_photos_all = list(d.get("defect_photos") or d.get("photos") or [])
                    def_photos = list(dict.fromkeys([str(p) for p in def_photos_all]))
                    yield {
                        "detail": detail, "defect": defect, "grade": grade,
                        "def_photos": def_photos, "controller": controller,
                        "found_date": found_date, "found_time": found_time,
                        "uud": uud,
                    }

    # 2) Цех сборки — НОВОЕ (с учётом другой схемы полей)
    assembly = bundle.get("Цех сборки") or {}
    if isinstance(assembly, dict):
        for _post_name, entries in assembly.items():
            for insp in (entries or []):
                controller = (insp.get("controller") or "").strip()
                # unify дата/время нахождения дефекта по inspection
                found_date, found_time = _split_dt_iso_to_date_time(insp.get("date_added") or "")
                for d in (insp.get("defects") or []):
                    # UUD структура такая же (внутри d["extra"]["UUD"])
                    uud = ((d.get("extra") or {}).get("UUD") or {})
                    # Маппинг имён полей
                    defect = (d.get("defect") or d.get("name") or d.get("nameENG") or "").strip()
                    detail = (d.get("detail") or d.get("unit") or "").strip()
                    grade  = (d.get("grade") or "").strip()
                    def_photos_all = list(d.get("defect_photos") or d.get("photos") or [])
                    def_photos = list(dict.fromkeys([str(p) for p in def_photos_all]))
                    yield {
                        "detail": detail, "defect": defect, "grade": grade,
                        "def_photos": def_photos, "controller": controller,
                        "found_date": found_date, "found_time": found_time,
                        "uud": uud,
                    }


# ====== СБОР СТРОК ДЛЯ XLSX ======
def collect_rows(
    vins: List[str],
    grade_filter: set[str] | None = None,
) -> Tuple[List[Dict[str, Any]], int, int, int]:
    """
    В отчёт попадают:
      - все дефекты со статусом UUD = impossible/imposible (каждый дефект — отдельной строкой),
        дополнительно можно отфильтровать по грейдам через grade_filter (набор {'V1+','V1','V2','V3'});
      - если у VIN нет таких дефектов:
          * без фильтра — ОДНА строка только с блоком 'Документация';
          * с фильтром — VIN пропускается (не попадает в отчёт).
    """

    # нормализуем фильтр грейдов
    norm_grade_filter: set[str] | None = None
    if grade_filter:
        norm_grade_filter = {str(g).strip().upper().replace(" ", "") for g in grade_filter} or None

    rows: List[Dict[str, Any]] = []
    max_doc_ph = max_def_ph = max_uud_ph = 0

    MIN_DOC_PHOTO_COLS = 2
    MIN_DEF_PHOTO_COLS = 2
    MIN_UUD_PHOTO_COLS = 2

    seen: set[tuple] = set()

    # ИСПОЛЬЗУЕМ ВНЕШНИЙ iter_normalized_defects(bundle)
    for vin in vins:
        bundle = fetch_bundle_for_vin(vin) or {}
        base = get_base_attrs_for_vin(vin)

        # --- Документация: берём последний блок
        doc_list = (bundle.get("Цех сборки") or {}).get("Документация") or []
        doc_block = {}
        if doc_list:
            doc_block = sorted(
                doc_list,
                key=lambda x: x.get("added_at") or x.get("date_added") or "",
                reverse=True,
            )[0]

        doc_src_ts = (doc_block.get("added_at") or doc_block.get("date_added") or "")
        doc_date, doc_time = split_dt(doc_src_ts)
        doc_by = (doc_block.get("added_by") or "").strip()

        # фото документации
        doc_photos_all = list(doc_block.get("photos") or [])
        doc_photos = list(dict.fromkeys([str(p) for p in doc_photos_all]))
        max_doc_ph = max(max_doc_ph, len(doc_photos))

        # --- Дефекты: только impossible + фильтр по грейдам (если задан)
        has_impossible = False
        for item in iter_normalized_defects(bundle):
            uud = item.get("uud") or {}
            status = (uud.get("status") or "").lower()
            if status not in {"impossible", "imposible"}:
                continue

            detail = item.get("detail", "")
            defect = item.get("defect", "")
            grade  = item.get("grade", "")
            controller = item.get("controller", "")
            found_date = item.get("found_date", "")
            found_time = item.get("found_time", "")

            # фильтр по грейдам
            if norm_grade_filter is not None:
                grade_norm = (grade or "").strip().upper().replace(" ", "")
                if grade_norm not in norm_grade_filter:
                    continue

            has_impossible = True

            # фото дефекта
            def_photos = item.get("def_photos") or []
            max_def_ph = max(max_def_ph, len(def_photos))

            # --- UUD (fix/decision + фото)
            uud_by = uud_date = uud_time = uud_comment = ""
            uud_photos: List[str] = []
            accept_by = accept_date = accept_time = ""

            for h in (uud.get("history") or []):
                if h.get("fix"):
                    fix = h["fix"]
                    if fix.get("kind") in (None, "impossible", "not_resolved"):
                        uud_by = (fix.get("by") or uud_by or "").strip()
                        uud_comment = (fix.get("comment") or uud_comment or "").strip()
                        if not (uud_date and uud_time):
                            uud_date, uud_time = split_dt(fix.get("at"))
                        _ph_all = list(fix.get("photos") or [])
                        if _ph_all:
                            uud_photos = list(dict.fromkeys([str(p) for p in _ph_all]))
                if h.get("decision"):
                    dec = h["decision"]
                    accept_by = (dec.get("by") or accept_by or "").strip()
                    accept_date, accept_time = split_dt(dec.get("at"))

            max_uud_ph = max(max_uud_ph, len(uud_photos))

            # ключ дедупликации строки
            row_key = (
                vin.strip().upper(),
                (detail or "").lower(), (defect or "").lower(), (grade or "").lower(),
                found_date, found_time, (controller or "").lower(),
                (uud_by or "").lower(), uud_date, uud_time, (uud_comment or "").lower(),
                (accept_by or "").lower(), accept_date, accept_time,
            )
            if row_key in seen:
                continue
            seen.add(row_key)

            rows.append({
                # Документация / базовые
                "vin": vin,
                "brand": base.get("brand", ""),
                "model": base.get("model", ""),
                "body_color": base.get("body_color", ""),
                "config_code": base.get("config_code", ""),
                "doc_date": doc_date, "doc_time": doc_time, "doc_by": doc_by,
                "doc_photos": doc_photos,

                # Дефект (фильтрованный)
                "detail": detail, "defect": defect, "grade": grade,
                "found_date": found_date, "found_time": found_time, "found_by": controller,
                "def_photos": def_photos,

                # УУД
                "uud_by": uud_by, "uud_date": uud_date, "uud_time": uud_time, "uud_comment": uud_comment,
                "uud_photos": uud_photos,
                "uud_accept_by": accept_by, "uud_accept_date": accept_date, "uud_accept_time": accept_time,
            })

        # ---- ВАЖНО: если (с учётом фильтра) impossible не нашли
        if not has_impossible:
            # без фильтра — добавляем базовую строку документации (как раньше)
            if norm_grade_filter is None:
                base_key = (vin.strip().upper(), "__base__", doc_date, doc_time, (doc_by or "").lower())
                if base_key not in seen:
                    seen.add(base_key)
                    rows.append({
                        "vin": vin,
                        "brand": base.get("brand", ""),
                        "model": base.get("model", ""),
                        "body_color": base.get("body_color", ""),
                        "config_code": base.get("config_code", ""),
                        "doc_date": doc_date, "doc_time": doc_time, "doc_by": doc_by,
                        "doc_photos": doc_photos,

                        # блоки Дефект/УУД — пустые
                        "detail": "", "defect": "", "grade": "",
                        "found_date": "", "found_time": "", "found_by": "",
                        "def_photos": [],
                        "uud_by": "", "uud_date": "", "uud_time": "", "uud_comment": "",
                        "uud_photos": [],
                        "uud_accept_by": "", "uud_accept_date": "", "uud_accept_time": "",
                    })
            # с фильтром — просто пропускаем VIN (ничего не добавляем)

    # Минимальные колонки для фото (чтобы колонки всегда были)
    max_doc_ph = max(max_doc_ph, MIN_DOC_PHOTO_COLS)
    max_def_ph = max(max_def_ph, MIN_DEF_PHOTO_COLS)
    max_uud_ph = max(max_uud_ph, MIN_UUD_PHOTO_COLS)

    return rows, max_doc_ph, max_def_ph, max_uud_ph

def image_scale_for_draw(
    image_stream: io.BytesIO,
    draw_max_width_px: int,
    draw_max_height_px: int,
) -> tuple[float, float]:
    """
    Определяет коэффициенты x_scale/y_scale для вставки в Excel так,
    чтобы картинка умещалась в заданный прямоугольник отображения
    (например, 140×110 px) без потери пропорций.
    """
    try:
        pos = image_stream.tell()
        image_stream.seek(0)
        im = Image.open(image_stream)
        w, h = im.size
        image_stream.seek(pos)  # вернуть указатель
        if not w or not h:
            return 1.0, 1.0
        sx = min(draw_max_width_px / float(w), 1.0)
        sy = min(draw_max_height_px / float(h), 1.0)
        return sx, sy
    except Exception:
        return 1.0, 1.0

# ====== ВЬЮШКА ЭКСПОРТА ======
@require_POST
@login_required
def export_documentation_extended(request):
    # ---- 0) принять список VIN'ов (JSON или form-data)
    vins = None
    grades_raw = None

    ctype = (request.content_type or "").lower()
    if "application/json" in ctype:
        try:
            data = json.loads((request.body or b"").decode("utf-8") or "{}")
        except Exception:
            data = {}
        vins = data.get("vins_json")
        grades_raw = data.get("grades")  # ← принимаем грейды из JSON
    else:
        vins = request.POST.get("vins_json")
        # поддержим и grades[]=... и grades=... (через запятую)
        grades_raw = (request.POST.getlist("grades[]")
                      or request.POST.getlist("grades")
                      or request.POST.get("grades"))

    # vins_json может прийти как строка JSON
    if isinstance(vins, str):
        try:
            vins = json.loads(vins)
        except json.JSONDecodeError:
            return HttpResponseBadRequest("Bad vins_json")

    if not isinstance(vins, list) or not vins:
        return HttpResponseBadRequest("vins_json (list) is required")

    vins = [str(v).strip() for v in vins if str(v).strip()]
    if not vins:
        return HttpResponseBadRequest("Empty VINs")

    # ---- 1) распарсить выбранные грейды (опционально)
    def _parse_grades(raw):
        """
        Принимаем: ["V1+","V1"] ИЛИ "V1+,V1" ИЛИ JSON-строку.
        Возвращаем множество {'V1+','V1','V2','V3'} либо None (значит: все).
        """
        allowed = {"V1+", "V1", "V2", "V3"}
        if not raw:
            return None

        vals = None
        if isinstance(raw, (list, tuple)):
            vals = list(raw)
        elif isinstance(raw, str):
            # пробуем как JSON
            try:
                j = json.loads(raw)
                if isinstance(j, (list, tuple)):
                    vals = list(j)
                else:
                    vals = [s.strip() for s in raw.split(",") if s.strip()]
            except Exception:
                vals = [s.strip() for s in raw.split(",") if s.strip()]
        else:
            return None

        norm = {str(x).strip().upper().replace(" ", "") for x in vals}
        norm &= allowed
        return norm or None

    grade_filter = _parse_grades(grades_raw)

    # ---- 2) собрать плоские строки (по одной на дефект) и максимумы по фото
    rows, max_doc_ph, max_def_ph, max_uud_ph = collect_rows(vins, grade_filter=grade_filter)

    # ---- сгруппировать по VIN в том же порядке, в каком запрошены
    vin_groups = OrderedDict()
    for v in vins:
        vin_groups[v] = []
    for r in rows:
        v = r.get("vin") or ""
        if v not in vin_groups:
            vin_groups[v] = []
        vin_groups[v].append(r)

    # ---- 3) создать книгу
    output = io.BytesIO()
    wb = xlsxwriter.Workbook(
        output,
        {"in_memory": True, "strings_to_urls": False, "constant_memory": True},
    )
    ws = wb.add_worksheet("Документация")

    # ---- форматы
    fmt_head_group = wb.add_format({
        "bold": True, "align": "center", "valign": "vcenter",
        "bg_color": "#E6F2FF", "border": 1
    })
    fmt_head = wb.add_format({
        "bold": True, "align": "center", "valign": "vcenter",
        "text_wrap": True, "bg_color": "#F2F2F2", "border": 1
    })
    fmt_text = wb.add_format({"text_wrap": True, "valign": "top", "border": 1})
    fmt_center = wb.add_format({"align": "center", "valign": "top", "border": 1})
    fmt_merge = wb.add_format({"text_wrap": True, "align": "center", "valign": "vcenter", "border": 1})

    # ---- 4) шапка (группы и колонки)
    # Группа 1: Документация
    doc_fixed_headers = [
        "VIN", "Бренд", "Модель", "Цвет", "Код комплектации",
        "Дата прохождения", "Время прохождения", "Кто проводил",
    ]
    doc_photo_headers = [f"Фото  {i}" for i in range(1, max_doc_ph + 1)]

    # Группа 2: Цех (дефект)
    plant_fixed_headers = [
        "Деталь", "Дефект", "Грейд",
        "Дата нахождения", "Время нахождения", "Кто нашёл",
    ]
    plant_photo_headers = [f"Фото дефекта {i}" for i in range(1, max_def_ph + 1)]

    # Группа 3: УУД
    uud_fixed_headers  = ["Кто с УУД отметил невозможным", "Дата с УУД", "Время с УУД", "Комментарий с УУД"]
    uud_photo_headers  = [f"Фото с УУД {i}" for i in range(1, max_uud_ph + 1)]
    uud_tail_headers   = ["Кто принял с УУД", "Дата принятия", "Время принятия"]

    headers = (
        doc_fixed_headers + doc_photo_headers +
        plant_fixed_headers + plant_photo_headers +
        uud_fixed_headers + uud_photo_headers + uud_tail_headers
    )

    # ширины (примерно)
    col_widths = []
    col_widths += [20, 14, 18, 14, 18, 12, 10, 14]                      # док фикс
    col_widths += [18] * len(doc_photo_headers)                          # док фото
    col_widths += [28, 22, 10, 14, 10, 16]                               # цех фикс
    col_widths += [18] * len(plant_photo_headers)                        # цех фото
    col_widths += [24, 12, 10, 28]                                       # uud фикс
    col_widths += [18] * len(uud_photo_headers)                          # uud фото
    col_widths += [18, 12, 10]                                           # uud хвост

    # строка 0 — заголовки групп
    col = 0
    g1_start = col; col += len(doc_fixed_headers) + len(doc_photo_headers); g1_end = col - 1
    if g1_end >= g1_start:
        ws.merge_range(0, g1_start, 0, g1_end, "Данные с поста документации", fmt_head_group)

    g2_start = col; col += len(plant_fixed_headers) + len(plant_photo_headers); g2_end = col - 1
    if g2_end >= g2_start:
        ws.merge_range(0, g2_start, 0, g2_end, "Данные дефекта", fmt_head_group)

    g3_start = col; col += len(uud_fixed_headers) + len(uud_photo_headers) + len(uud_tail_headers); g3_end = col - 1
    if g3_end >= g3_start:
        ws.merge_range(0, g3_start, 0, g3_end, "Данные с УУД", fmt_head_group)

    # строка 1 — названия колонок
    for c, h in enumerate(headers):
        ws.write(1, c, h, fmt_head)
        if c < len(col_widths):
            ws.set_column(c, c, col_widths[c])

    # индексы колонок (поможет писать)
    def calc_slices():
        c = 0
        doc_fixed = list(range(c, c + len(doc_fixed_headers)));   c += len(doc_fixed_headers)
        doc_ph    = list(range(c, c + len(doc_photo_headers)));   c += len(doc_photo_headers)
        p_fixed   = list(range(c, c + len(plant_fixed_headers))); c += len(plant_fixed_headers)
        p_ph      = list(range(c, c + len(plant_photo_headers))); c += len(plant_photo_headers)
        uud_fixed = list(range(c, c + len(uud_fixed_headers)));   c += len(uud_fixed_headers)
        uud_ph    = list(range(c, c + len(uud_photo_headers)));   c += len(uud_photo_headers)
        uud_tail  = list(range(c, c + len(uud_tail_headers)));    c += len(uud_tail_headers)
        return doc_fixed, doc_ph, p_fixed, p_ph, uud_fixed, uud_ph, uud_tail

    doc_fixed_idx, doc_ph_idx, p_fixed_idx, p_ph_idx, uud_fixed_idx, uud_ph_idx, uud_tail_idx = calc_slices()

    # колонки, которые нужно объединять по VIN-группе
    merge_cols = doc_fixed_idx[:5]  # VIN, Бренд, Модель, Цвет, Код комплектации

    # высота строки под фото и размеры отображения (в пикселях)
    ws.set_default_row(120)  # ~высота под 110px изображения
    DOC_PHOTO_MAX_W = 120
    DOC_PHOTO_MAX_H = 120
    DEF_PHOTO_MAX_W = 120
    DEF_PHOTO_MAX_H = 120
    UUD_PHOTO_MAX_W = 120
    UUD_PHOTO_MAX_H = 110

    # ---- 5) данные с объединением по VIN
    row_xlsx = 2

    for vin, items in vin_groups.items():
        if not items:
            continue

        group_start = row_xlsx

        # значения для объединяемых колонок берём из первой строки группы
        first = items[0]
        merge_values = [
            first.get("vin") or "",
            first.get("brand") or "",
            first.get("model") or "",
            first.get("body_color") or "",
            first.get("config_code") or "",
        ]

        for i, r in enumerate(items):
            # --- Документация фикс
            if i == 0:
                base5 = merge_values
            else:
                base5 = ["", "", "", "", ""]

            rest_doc_fixed = [
                r.get("doc_date", ""), r.get("doc_time", ""), r.get("doc_by", "")
            ]
            values_doc_fixed = base5 + rest_doc_fixed

            for c, val in zip(doc_fixed_idx, values_doc_fixed):
                is_center = (c in (doc_fixed_idx[5], doc_fixed_idx[6]))
                ws.write(row_xlsx, c, val, fmt_center if is_center else fmt_text)

            # --- Фото «Документация»: вставляем со scale
            for j, c in enumerate(doc_ph_idx):
                p = r["doc_photos"][j] if j < len(r["doc_photos"]) else None
                if p:
                    bio = image_stream_from_media(p, jpeg_quality=92)
                    if bio:
                        sx, sy = image_scale_for_draw(bio, DOC_PHOTO_MAX_W, DOC_PHOTO_MAX_H)
                        ws.insert_image(
                            row_xlsx, c, "doc.jpg",
                            {"image_data": bio, "x_offset": 4, "y_offset": 4, "x_scale": sx, "y_scale": sy}
                        )

            # --- Цех фикс
            pfx = [
                r.get("detail",""), r.get("defect",""), r.get("grade",""),
                r.get("found_date",""), r.get("found_time",""), r.get("found_by",""),
            ]
            for c, val in zip(p_fixed_idx, pfx):
                is_center = (c in (p_fixed_idx[3], p_fixed_idx[4]))
                ws.write(row_xlsx, c, val, fmt_center if is_center else fmt_text)

            # --- Фото дефекта (со scale)
            for j, c in enumerate(p_ph_idx):
                p = r["def_photos"][j] if j < len(r["def_photos"]) else None
                if p:
                    bio = image_stream_from_media(p, jpeg_quality=92)
                    if bio:
                        sx, sy = image_scale_for_draw(bio, DEF_PHOTO_MAX_W, DEF_PHOTO_MAX_H)
                        ws.insert_image(
                            row_xlsx, c, "defect.jpg",
                            {"image_data": bio, "x_offset": 4, "y_offset": 4, "x_scale": sx, "y_scale": sy}
                        )

            # --- UUD фикс
            ufx = [r.get("uud_by",""), r.get("uud_date",""), r.get("uud_time",""), r.get("uud_comment","")]
            for c, val in zip(uud_fixed_idx, ufx):
                is_center = (c in (uud_fixed_idx[1], uud_fixed_idx[2]))
                ws.write(row_xlsx, c, val, fmt_center if is_center else fmt_text)

            # --- Фото UUD (со scale)
            for j, c in enumerate(uud_ph_idx):
                p = r["uud_photos"][j] if j < len(r["uud_photos"]) else None
                if p:
                    bio = image_stream_from_media(p, jpeg_quality=92)
                    if bio:
                        sx, sy = image_scale_for_draw(bio, UUD_PHOTO_MAX_W, UUD_PHOTO_MAX_H)
                        ws.insert_image(
                            row_xlsx, c, "uud.jpg",
                            {"image_data": bio, "x_offset": 4, "y_offset": 4, "x_scale": sx, "y_scale": sy}
                        )

            # --- UUD хвост
            utail = [r.get("uud_accept_by",""), r.get("uud_accept_date",""), r.get("uud_accept_time","")]
            for c, val in zip(uud_tail_idx, utail):
                is_center = (c in (uud_tail_idx[1], uud_tail_idx[2]))
                ws.write(row_xlsx, c, val, fmt_center if is_center else fmt_text)

            row_xlsx += 1

        group_end = row_xlsx - 1
        # объединяем VIN/Бренд/Модель/Цвет/Код комплектации, если строк > 1
        if group_end > group_start:
            for c, val in zip(merge_cols, merge_values):
                ws.merge_range(group_start, c, group_end, c, val, fmt_merge)

    # ---- 6) отдать файл
    wb.close()
    output.seek(0)

    now = datetime.now()
    fname = f"documentation_extended_{now:%Y-%m-%d_%H-%M}.xlsx"
    resp = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{fname}"'
    return resp







# marriage post, table and export
from django.db.models import OuterRef, Subquery
from django.db.models.functions import Coalesce
from django.db.models import OuterRef, Subquery, Q
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


@login_required
@role_required(["controller", "master", "head_area"])
def marriage_view(request):
    """
    GET  -> страница с формой (VIN + ДВС + КПП)
    POST -> сохранить/обновить VehicleIdentifiers
    """
    if request.method == "POST":
        vin = (request.POST.get("vin_number") or "").strip().upper()
        eng = (request.POST.get("engine_number") or "").strip()
        trans = (request.POST.get("transmission_number") or "").strip()

        if len(vin) != 17:
            messages.error(request, "Введите корректный VIN (17 символов).")
            return redirect(request.path)

        obj, created = VehicleIdentifiers.objects.get_or_create(
            vin=vin,
            defaults={
                "engine_number": eng or None,
                "transmission_number": trans or None,
                "saved_by": request.user,
            },
        )
        if not created:
            # обновляем только то, что пришло
            if eng:
                obj.engine_number = eng
            if trans:
                obj.transmission_number = trans
            if not obj.saved_by:
                obj.saved_by = request.user
            obj.save(update_fields=["engine_number", "transmission_number", "saved_by", "updated_at"])

        messages.success(request, "Данные сохранены.")
        return redirect(f"{request.path}?vin={vin}")

    # GET
    ctx = {"prefill_vin": (request.GET.get("vin") or "").strip().upper()}
    return render(request, "assembly/marriage.html", ctx)


@login_required
@permission_required('users.access_to_the_marriage_table', raise_exception=True)
def marriage_table_view(request):
    """
    Таблица записей VehicleIdentifiers c подтяжкой brand/model из TraceData по VIN.
    Фильтры сохраняются/читаются на стороне клиента (query-string), см. шаблон.
    """
    # Подзапросы: берём ПОСЛЕДНЮЮ запись из TraceData по vin_rk или vin_c
    td_by_rk = (TraceData.objects
                .filter(vin_rk=OuterRef("vin"))
                .order_by("-date_added"))
    td_by_c = (TraceData.objects
               .filter(vin_c=OuterRef("vin"))
               .order_by("-date_added"))

    qs = (VehicleIdentifiers.objects
          .select_related("saved_by")
          .annotate(
              brand_rk=Subquery(td_by_rk.values("brand")[:1]),
              brand_c=Subquery(td_by_c.values("brand")[:1]),
              brand=Coalesce("brand_rk", "brand_c"),

              model_rk=Subquery(td_by_rk.values("model")[:1]),
              model_c=Subquery(td_by_c.values("model")[:1]),
              model=Coalesce("model_rk", "model_c"),
          )
          .order_by("-updated_at"))

    context = {
        "records": qs,  # шаблон читает .vin .brand .model .engine_number .transmission_number .created_at .saved_by
    }
    return render(request, "assembly/marriage_table.html", context)






@login_required
@permission_required('users.access_to_the_marriage_table', raise_exception=True)
def marriage_table_export(request):
    """
    XLSX экспорт «Marriage — таблица».
    Колонки: VIN RK, VIN China, Бренд, Модель, Код комплектации, Двигатель (код), Двигатель, Трансмиссия (код), Трансмиссия, Привод (код), Привод, № ДВС, № Трансмиссии, Мат-код, День, Время, Кто сохранил.
    День/Время берём ИЗ updated_at (локально, по settings.TIME_ZONE).
    Диапазоны day_from/day_to и time_from/time_to применяем к updated_at.
    """
    # ----- подзапросы по трейсингу -----
    td_by_rk = TraceData.objects.filter(vin_rk=OuterRef("vin")).order_by("-date_added")
    td_by_c  = TraceData.objects.filter(vin_c =OuterRef("vin")).order_by("-date_added")

    qs = (VehicleIdentifiers.objects
          .select_related("saved_by")
          .annotate(
              brand_rk = Subquery(td_by_rk.values("brand")[:1]),
              brand_c  = Subquery(td_by_c.values("brand")[:1]),
              model_rk = Subquery(td_by_rk.values("model")[:1]),
              model_c  = Subquery(td_by_c.values("model")[:1]),
              config_code_rk = Subquery(td_by_rk.values("config_code")[:1]),
              config_code_c  = Subquery(td_by_c.values("config_code")[:1]),
              # вместо vin_c теперь тянем body_number для VIN China
              body_number_rk = Subquery(td_by_rk.values("body_number")[:1]),
              body_number_c  = Subquery(td_by_c.values("body_number")[:1]),
              engine_volume_rk = Subquery(td_by_rk.values("engine_volume")[:1]),
              engine_volume_c  = Subquery(td_by_c.values("engine_volume")[:1]),
          ))

    # ----- декодирование двигателя/трансмиссии/привода из config_code -----
    ENGINE_MAP = {
        "E4G": "SQRE4G15C",
        "E4T": "SQRE4T15C",
        "J16": "SQRF4J16",
        "J20": "SQRF4J20",
        "F4J": "SQRF4J16C",
        "J15": "SQRH4J15",
    }
    TRANS_MAP = {
        "C8": "CVT18",
        "D6": "6DCT",
        "D7": "7DCT",
        "A8": "8AT",
    }
    DRIVE_MAP = {
        "2": "2WD",
        "A": "AWD",
        "4": "4WD",
    }
    ROOF_MAP = {
        "P": "Panoramic",
        "N": "Normal",
        "S": "Sunroof",
    }
    SPEC_MAP = {
        "S": "Style",
        "E": "Elite",
        "R": "Prime",
        "U": "Ultimate",
        "F": "Flagship",
        "A": "Ultra",
        "C": "Comfort",
        "P": "Premium",
        "Y": "Luxury",
    }

    def decode_from_config(cfg: str) -> tuple[str, str, str, str, str, str, str, str, str, str]:
        """Возвращает (eng_code, eng_full, trans_code, trans_full, drive_code, drive_full, roof_code, roof_full, spec_code, spec_full).
        Позиции (индексация с 1): 8–10 двигатель, 11–12 трансмиссия, 13 привод, 14 roof, 15 spec.
        """
        if not cfg:
            return "—", "—", "—", "—", "—", "—", "—", "—", "—", "—"
        s = str(cfg).strip()
        if len(s) < 15:
            return "—", "—", "—", "—", "—", "—", "—", "—", "—", "—"
        eng_code = s[7:10]
        trans_code = s[10:12]
        drive_code = s[12]
        roof_code = s[13]
        spec_code = s[14]
        eng_full = ENGINE_MAP.get(eng_code, "—")
        trans_full = TRANS_MAP.get(trans_code, "—")
        drive_full = DRIVE_MAP.get(drive_code, "—")
        roof_full = ROOF_MAP.get(roof_code, "—")
        spec_full = SPEC_MAP.get(spec_code, "—")
        return (
            eng_code or "—", eng_full,
            trans_code or "—", trans_full,
            drive_code or "—", drive_full,
            roof_code or "—", roof_full,
            spec_code or "—", spec_full,
        )

    # ----- чтение фильтров -----
    p = request.GET
    q_vin = (p.get("q") or "").strip()
    brands = [s for s in (p.get("brand") or "").split(",") if s]
    models = [s for s in (p.get("model") or "").split(",") if s]
    saved_by_vals = [s for s in (p.get("saved_by") or "").split(",") if s]
    day_from = (p.get("day_from") or "").strip()
    day_to   = (p.get("day_to") or "").strip()
    time_from = (p.get("time_from") or "").strip()
    time_to   = (p.get("time_to") or "").strip()

    # ----- применение фильтров -----
    if q_vin:
        qs = qs.filter(vin__icontains=q_vin)

    if brands:
        cond = Q(brand_rk__in=brands) | Q(brand_c__in=brands)
        if "—" in brands:
            cond |= (Q(brand_rk__isnull=True) & Q(brand_c__isnull=True))
        qs = qs.filter(cond)

    if models:
        cond = Q(model_rk__in=models) | Q(model_c__in=models)
        if "—" in models:
            cond |= (Q(model_rk__isnull=True) & Q(model_c__isnull=True))
        qs = qs.filter(cond)

    if saved_by_vals:
        cond = Q(saved_by__username__in=saved_by_vals)
        if "—" in saved_by_vals:
            cond |= Q(saved_by__isnull=True)
        qs = qs.filter(cond)

    # диапазон по ДНЮ (updated_at)
    if day_from:
        qs = qs.filter(updated_at__date__gte=day_from)
    if day_to:
        qs = qs.filter(updated_at__date__lte=day_to)

    # диапазон по ВРЕМЕНИ (updated_at)
    def _parse_time(s):
        if not s: return None
        for fmt in ("%H:%M:%S", "%H:%M"):
            try: return datetime.strptime(s, fmt).time()
            except ValueError: pass
        return None

    tf = _parse_time(time_from)
    tt = _parse_time(time_to)
    if tf:
        qs = qs.filter(updated_at__time__gte=tf)
    if tt:
        qs = qs.filter(updated_at__time__lte=tt)

    qs = qs.order_by("-updated_at")

    # ----- формирование XLSX -----
    wb = Workbook()
    ws = wb.active
    ws.title = "Marriage"

    headers = [
        "VIN RK", "VIN China", "Бренд", "Модель", "Код комплектации",
        "Двигатель (код)", "Двигатель", "Трансмиссия (код)", "Трансмиссия", "Привод (код)", "Привод",
        "Roof (код)", "Roof", "Spec (код)", "Spec",
        "Номенклатура",
        "№ ДВС", "№ Трансмиссии", "Мат-код", "День", "Время", "Кто сохранил"
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"

    def _calc_mat_code(model_val: str | None, vin_val: str) -> str:
        m = (model_val or "").strip().lower()
        try:
            if "tiggo 4" in m:
                code = make_mat_code_tiggo4(vin_val)
                return code or "—"
            if "tiggo 7" in m:
                code = make_mat_code_tiggo7(vin_val)
                return code or "—"
            if "tiggo 45" in m:
                code = make_mat_code_tiggo45(vin_val)  # noqa: F821
                return code or "—"
        except NameError:
            return "—"
        return "—"

    for obj in qs.iterator():
        brand = getattr(obj, "brand_rk", None) or getattr(obj, "brand_c", None) or "—"
        model = getattr(obj, "model_rk", None) or getattr(obj, "model_c", None) or "—"

        # VIN China = body_number из той же записи TraceData
        vin_china = getattr(obj, "body_number_rk", None) or getattr(obj, "body_number_c", None) or "—"

        config_code = getattr(obj, "config_code_rk", None) or getattr(obj, "config_code_c", None) or "—"

        # decode engine/trans/drive/roof/spec from config_code
        eng_code, eng_full, trans_code, trans_full, drive_code, drive_full, roof_code, roof_full, spec_code, spec_full = decode_from_config(config_code if isinstance(config_code, str) else str(config_code or ""))

        # --- ENGINE VOLUME ---
        engine_volume_val = getattr(obj, "engine_volume_rk", None)
        if engine_volume_val in (None, ""):
            engine_volume_val = getattr(obj, "engine_volume_c", None)
        engine_volume_str = str(engine_volume_val) if engine_volume_val not in (None, "") else "—"

        # nomenclature string
        nomenclature = f"Автомобиль {brand} {model} {config_code}, {spec_full}, {trans_full}, {engine_volume_str}, {drive_full}"

        mat_code = _calc_mat_code(model, obj.vin)

        dt_local = timezone.localtime(obj.updated_at)
        day = dt_local.date().isoformat()
        time_str = dt_local.strftime("%H:%M")

        saved = getattr(obj.saved_by, "username", None) or "—"

        ws.append([
            obj.vin,                 # VIN RK
            vin_china,               # VIN China (body_number)
            brand,
            model,
            config_code,             # Код комплектации
            eng_code,                # Двигатель (код)
            eng_full,                # Двигатель (расшифровка)
            trans_code,              # Трансмиссия (код)
            trans_full,              # Трансмиссия (расшифровка)
            drive_code,              # Привод (код)
            drive_full,              # Привод (расшифровка)
            roof_code,               # Roof (код)
            roof_full,               # Roof (расшифровка)
            spec_code,               # Spec (код)
            spec_full,               # Spec (расшифровка)
            nomenclature,            # Номенклатура
            obj.engine_number or "—",
            obj.transmission_number or "—",
            mat_code,
            day,
            time_str,
            saved,
        ])

    # ширины столбцов (учтена колонка VIN China и новая "Номенклатура")
    widths = [23, 23, 14, 20, 18, 12, 20, 14, 16, 10, 12, 10, 14, 10, 14, 40, 18, 18, 16, 12, 10, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64+i)].width = w

    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = 'attachment; filename="marriage_table.xlsx"'
    wb.save(resp)
    return resp






def make_mat_code_tiggo4(vin: str) -> str | None:
    """
    Заготовка мат-кода для модели Tiggo 45 по VIN (та же механика, что и Tiggo 7):
      - Берём самую свежую TraceData по VIN (vin_rk -> vin_c).
      - По config_code выбираем ШАБЛОН (внутри '**' — место для 2-симв. кода цвета).
      - Цвет: берём body_color, оставляем только буквы/цифры, upper.
        Если РОВНО 2 символа — подставляем вместо '**', иначе оставляем '**'.

    ВАЖНО: ниже нужно заполнить правила выбора шаблона по config_code (template).
    Дай условия (позиции/символы) и строки шаблонов — вставлю.
    """
    from supplies.models import TraceData

    vin = (vin or "").strip().upper()
    if not vin:
        return None

    # 1) свежая запись трейсинга
    td = (TraceData.objects
          .filter(vin_rk=vin)
          .order_by("-date_added")
          .first())
    if not td:
        td = (TraceData.objects
              .filter(vin_c=vin)
              .order_by("-date_added")
              .first())
    if not td:
        return None

    cfg = (td.config_code or "").strip()
    if len(cfg) < 15:
        return None

    # 2) ВЫБОР ШАБЛОНА ПО КОМПЛЕКТАЦИИ (ЗАПОЛНИТЬ!)
    # Примеры, как было у других моделей:
    # - if cfg[14].upper() == 'X': template = "T71XXK?**AB0001"
    # - elif cfg[13:15].upper() == 'YZ': template = "T71YYK?**CD0002"
    # - else: return None
    template = None  # <<< TODO: задать шаблон(ы) по правилам config_code

    if not template:
        return None

    # 3) Цвет: подставляем ТОЛЬКО если ровно 2 символа
    filtered_color = "".join(ch for ch in (td.body_color or "") if ch.isalnum()).upper()
    if len(filtered_color) == 2:
        return template.replace("**", filtered_color)
    return template



def make_mat_code_tiggo7(vin: str) -> str | None:
    """
    Мат-код для Tiggo 7 по VIN.

    Комплектация из config_code:
      - PRIME:   14–15-й символы == 'PR' -> шаблон 'T7161K2**FF0008'
      - ULTIMATE: 15-й символ == 'U'     -> шаблон 'T7161K3**FG0008'

    Цвет:
      - Берём body_color из TraceData, оставляем только буквы/цифры, upper.
      - Если РОВНО 2 символа — подставляем вместо '**'.
      - Если символов != 2 (в т.ч. > 2) — оставляем '**' как есть.

    Возвращает строку или None, если не смогли определить комплектацию/трейсинг.
    """
    from supplies.models import TraceData

    vin = (vin or "").strip().upper()
    if not vin:
        return None

    # свежая запись трейсинга (сначала vin_rk, потом vin_c)
    td = (TraceData.objects
          .filter(vin_rk=vin)
          .order_by("-date_added")
          .first())
    if not td:
        td = (TraceData.objects
              .filter(vin_c=vin)
              .order_by("-date_added")
              .first())
    if not td:
        return None

    cfg = (td.config_code or "").strip()
    if len(cfg) < 15:
        return None

    # определяем комплектацию (важно сперва проверить 'PR', затем 'U')
    is_prime = cfg[13:15].upper() == "PR"   # 14–15-й символы
    is_ultimate = cfg[14].upper() == "U"    # 15-й символ

    if not (is_prime or is_ultimate):
        return None

    template = "T7161K2**FF0008" if is_prime else "T7161K3**FG0008"

    # извлекаем цвет (только если ровно 2 символа — подставляем)
    filtered = "".join(ch for ch in (td.body_color or "") if ch.isalnum()).upper()
    if len(filtered) == 2:
        return template.replace("**", filtered)
    return template











