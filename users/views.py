from django.contrib.auth import authenticate, login, logout
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.decorators import login_required
from .decorators import role_required
from django.http import JsonResponse
from django.db.models import Q, Count, F
from django.utils.timezone import now, timedelta
from supplies.models import ContainerUnloadingZone2Inspection as Inspection, Post, Detail, Defect
from .models import Notification, HelpdeskContact, ExportHistory, Selection, Employee, CustomUser, OvertimeRecord, KTVDefect
from django.http import HttpResponse
from supplies.models import ContainerUnloadingZoneSBInspection
from supplies.forms import MainUnloadingZoneDKDForm, BodyUnloadingZoneDKDForm
from assembly.forms import AssemblyTemplateForm
from vehicle_history.models import VINHistory, VINHistoryBackup, ContainerHistory, AssemblyPassLog, VESPassLog
from django.db.models.functions import Lower
from django.db.models import Q
from django.urls import reverse
from django.utils.dateparse import parse_datetime
from django.utils import timezone
from django.utils.dateparse import parse_date
from collections import Counter
import json
from .forms import (ControllerCreationForm, ControllerPasswordChangeForm, ProfileUpdateForm, EmployeeSearchForm,
                    EmployeeSelectionForm, CustomPasswordChangeForm, ControllerEditForm,
                    AssemblyZoneForm, AssemblyUnitForm, AssemblyDefectForm)
from assembly.models import AssemblyZone, AssemblyUnit, AssemblyDefect, PostAssembly
from django.contrib.auth import get_user_model
from django.contrib import messages
from collections import defaultdict
from datetime import datetime, date, time
from supplies.models import Post, BodyDetail, Defect, DefectGrade, DefectResponsible
from django.conf import settings
import os
from vehicle_history.utils import now_almaty_iso
from supplies.models import TraceData
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash
from django.utils.timezone import localtime
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from datetime import datetime, time
from django.utils.timezone import make_aware
from PIL import Image, ImageOps
from io import BytesIO
from django.core.files.uploadedfile import InMemoryUploadedFile
import sys
from django.utils.dateparse import parse_datetime
from django.utils.timezone import get_current_timezone
from datetime import timedelta
import pandas as pd
import urllib.parse  # для urlencode
import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font
from docx import Document
from urllib.parse import quote
from django.http import HttpResponseForbidden
import io
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from users.reports.report_generator import generate_summary_report
from itertools import chain
from django.db import IntegrityError, transaction
from django.views.decorators.http import require_POST
from django.core.serializers.json import DjangoJSONEncoder
from datetime import datetime
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.utils.timezone import localdate
from django.http import JsonResponse, HttpResponseBadRequest, HttpResponseNotAllowed
from django.views.decorators.http import require_POST, require_http_methods
from django.contrib.auth import authenticate, login, logout
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.decorators import login_required
from .decorators import role_required
from django.http import JsonResponse
from django.db.models import Q, Count, F
from django.utils.timezone import now, timedelta
from supplies.models import ContainerUnloadingZone2Inspection as Inspection, Post, Detail, Defect
from .models import Notification, HelpdeskContact, ExportHistory, Selection, Employee, CustomUser, OvertimeRecord
from django.http import HttpResponse
from supplies.models import ContainerUnloadingZoneSBInspection
from supplies.forms import MainUnloadingZoneDKDForm, BodyUnloadingZoneDKDForm
from assembly.forms import AssemblyTemplateForm
from vehicle_history.models import VINHistory, VINHistoryBackup
from django.db.models.functions import Lower
from django.db.models import Q
from django.urls import reverse
from django.utils.dateparse import parse_datetime
from django.utils import timezone
from django.utils.dateparse import parse_date
from collections import Counter
import json
from .forms import (ControllerCreationForm, ControllerPasswordChangeForm, ProfileUpdateForm, EmployeeSearchForm,
                    EmployeeSelectionForm, CustomPasswordChangeForm, ControllerEditForm,
                    AssemblyZoneForm, AssemblyUnitForm, AssemblyDefectForm)
from assembly.models import AssemblyZone, AssemblyUnit, AssemblyDefect, PostAssembly
from django.contrib.auth import get_user_model
from django.contrib import messages
from collections import defaultdict
from datetime import datetime, date, time
from supplies.models import Post, BodyDetail, Defect, DefectGrade, DefectResponsible
from django.conf import settings
import os
from vehicle_history.utils import now_almaty_iso
from supplies.models import TraceData
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash
from django.utils.timezone import localtime
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from datetime import datetime, time
from django.utils.timezone import make_aware
from PIL import Image, ImageOps
from io import BytesIO
from django.core.files.uploadedfile import InMemoryUploadedFile
import sys
from django.utils.dateparse import parse_datetime
from django.utils.timezone import get_current_timezone
from datetime import timedelta
import pandas as pd
import urllib.parse  # для urlencode
import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font
from docx import Document
from urllib.parse import quote
from django.http import HttpResponseForbidden
import io
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from users.reports.report_generator import generate_summary_report
from itertools import chain
from django.db import IntegrityError, transaction
from django.views.decorators.http import require_POST
from django.core.serializers.json import DjangoJSONEncoder
import re
from django.db import DatabaseError
from django.db.models import Q, F, Value, CharField
from django.db.models.functions import Lower
from django.db.models import Func
from qrr.views import QRR_WHITELIST_USERS, ensure_qrr_or_whitelist
from django.urls import resolve, Resolver404
from urllib.parse import urlparse, parse_qs
from django.core.exceptions import PermissionDenied
from django.contrib.auth.decorators import permission_required



MASTER_DASHBOARD_PERMISSION = f"{CustomUser._meta.app_label}.access_to_the_master_dashboard"
APP_LABEL = CustomUser._meta.app_label


# Удаляет все нецифры
def normalize_digits(s: str) -> str:
    return re.sub(r"\D+", "", s or "")



# Аналог PostgreSQL REGEXP_REPLACE(text, pattern, replacement, flags)
class RegexReplace(Func):
    function = "REGEXP_REPLACE"
    arity = 4  # text, pattern, replacement, flags
    output_field = CharField()


few_roles = ['m.adylbayev', 't.tokshekenov', 'ai.kaldybek', 'lidacain', 'p.kalinin', 'e.fedorova']


# === Недавние посты сборки (храним в сессии пользователя) ===
RECENT_POST_LABELS = {
    "assembly:gaps_and_drops": "Зазоры и перепады",
    "assembly:exterior": "Экстерьер",
    "assembly:interior": "Интерьер",
    "assembly:trunk": "Багажник",
    "assembly:the_motor": "Моторный отсек",
    "assembly:functional": "Функционал",
    "assembly:geometry_of_wheels": "Геометрия колес",
    "assembly:adjusting_the_headlights_and_calibrating_the_steering_wheel": "Регулировка света фар и калибровка руля",
    "assembly:breaking_system": "Тормозная система",
    "assembly:underbody": "Underbody",
    "assembly:adas_chery": "ADAS (Chery)",
    "assembly:adas_gwm": "ADAS (GWM)",
    "assembly:adas_changan": "ADAS (Changan)",
    "assembly:avm_chery": "AVM (Chery)",
    "assembly:avm_gwm": "AVM (GWM)",
    "assembly:avm_changan": "AVM (Changan)",
    "assembly:tightness_of_the_body": "Герметичность кузова",
    "assembly:diagnostics": "Диагностика",
    "assembly:test_track": "Тест трек",
    "assembly:documentation": "Документация",
    "assembly:final_current_control_chery": "Финал текущий контроль (Chery)",
    "assembly:final_current_control_gwm": "Финал текущий контроль (GWM)",
    "assembly:final_current_control_changan": "Финал текущий контроль (Changan)",
    "assembly:chassis_chery": "Chassis (Chery)",
    "assembly:chassis_gwm": "Chassis (GWM)",
    "assembly:chassis_changan": "Chassis (Changan)",
    "assembly:torque_control_chery": "Момент затяжек (Chery)",
    "assembly:torque_control_gwm": "Момент затяжек (GWM)",
    "assembly:torque_control_changan": "Момент затяжек (Changan)",
    "assembly:ves_pass_view": "VES передача/прием",
}


def _push_recent_post(request, view_name: str, post_id: str, max_items: int = 6) -> None:
    """Сохраняем «недавние посты» в сессии — элементы вида {"label": ..., "href": ...}."""
    try:
        label = RECENT_POST_LABELS.get(view_name, view_name)
        href = f"{reverse(view_name)}?post_id={post_id}" if post_id else reverse(view_name)
    except Exception:
        return  # если reverse не удался, молча выходим

    item = {"label": label, "href": href}
    current = request.session.get("recent_posts", [])
    current = [it for it in current if it.get("href") != item["href"]]  # убрать дубликат
    current.insert(0, item)
    request.session["recent_posts"] = current[:max_items]  # максимум N штук


@login_required
@require_POST
def remember_recent_post(request):
    """
    AJAX-эндпоинт: принимает {"url": "/assembly/...?...post_id=18"}, резолвит view_name + post_id
    и записывает пост в сессию как «недавний».
    """
    try:
        payload = json.loads(request.body.decode("utf-8"))
        raw_url = (payload.get("url") or "").strip()
        if not raw_url:
            return JsonResponse({"ok": False, "error": "empty url"}, status=400)

        parsed = urlparse(raw_url)
        path = parsed.path
        query = parse_qs(parsed.query)

        try:
            match = resolve(path)
            view_name = match.view_name
        except Resolver404:
            return JsonResponse({"ok": False, "error": "unresolvable url"}, status=400)

        post_id = ""
        if "post_id" in query and query["post_id"]:
            post_id = str(query["post_id"][0])
        elif "post_id" in match.kwargs:
            post_id = str(match.kwargs.get("post_id") or "")

        _push_recent_post(request, view_name, post_id)
        return JsonResponse({"ok": True})
    except Exception as e:
        return JsonResponse({"ok": False, "error": str(e)}, status=500)



def compress_uploaded_image(uploaded_file, quality=60, max_width=1600):
    """Сжимает изображение JPEG/PNG и возвращает InMemoryUploadedFile"""
    try:
        image = Image.open(uploaded_file)

        # Преобразуем в RGB (если PNG с прозрачностью)
        if image.mode in ("RGBA", "P"):
            image = image.convert("RGB")

        # Уменьшаем размер (если слишком большое изображение)
        if image.width > max_width:
            ratio = max_width / float(image.width)
            height = int(float(image.height) * float(ratio))
            image = image.resize((max_width, height), Image.Resampling.LANCZOS)

        output_io = BytesIO()
        image.save(output_io, format="JPEG", quality=quality, optimize=True)

        # Генерируем уникальное имя
        new_filename = f"photo_{time.strftime('%Y-%m-%dT%H-%M-%S')}.jpg"

        return InMemoryUploadedFile(
            output_io,
            'ImageField',
            new_filename,
            'image/jpeg',
            sys.getsizeof(output_io),
            None
        )
    except Exception as e:
        print("Ошибка при сжатии:", e)
        return uploaded_file  # Вернуть оригинал, если не удалось

@login_required
def profile(request):
    form = ProfileUpdateForm(instance=request.user)
    password_form = CustomPasswordChangeForm(user=request.user)

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'update_profile':
            form = ProfileUpdateForm(request.POST, request.FILES, instance=request.user)
            if form.is_valid():
                form.save()
                messages.success(request, "✅ Профиль обновлён.")
                return redirect('profile')
            else:
                messages.error(request, "❌ Проверьте правильность введённых данных.")
        elif action == 'change_password':
            password_form = CustomPasswordChangeForm(user=request.user, data=request.POST)
            if password_form.is_valid():
                user = password_form.save()
                update_session_auth_hash(request, user)
                messages.success(request, "✅ Пароль успешно изменён.")
                return redirect('profile')
            else:
                messages.error(request, "❌ Пароль не изменён. Проверьте ошибки ниже.")

    return render(request, "users/profile.html", {
        "form": form,
        "password_form": password_form,
        "enable_particles": True,
        "enable_background": True,
        "enable_white_square": False
    })


def in_development(request):
    return render(request, "users/in_development.html")


# ✅ Контроллер имеет доступ только к своей панели
@login_required
def controller_dashboard(request):
    # Проверка: только controller + право
    if not (
        request.user.role == "controller"
        or request.user.has_perm("users.access_to_post_of_the_controller")
    ):
        raise PermissionDenied

    allowed_usernames = ["QC-006", "QC-007", "QC-003", "QC-043", "QC-002", "QC-034", "QC-035"]
    master_usernames = ['lidacain_m', "lidacain_c", 'tima_c', 'a.boranchayev', 'an.zhakupov', 'g.mukhanbetov', 'M-02', 'M-06', 'm.akhmetzhanov', 'miko_m', 'n.myrzagul', 'd.zabridze']
    head_area_usernames = ['n.tikkeldiyev', 'nu.mukhambetov', 'p.kalinin', 'p.shen', 't.bostanov', 't.tokshekenov', 'a.yerikov', 'lidacain', 'm.adylbayev', 'miko', 'a.bykov', 'a.alpysbay']


    counter_changan = ['counter_changan']
    show_button_for_counter_changan = request.user.username in counter_changan

    counter_chery = ['counter_chery']
    show_button_for_counter_chery = request.user.username in counter_chery

    counter_gwm = ['counter_gwm']
    show_button_for_counter_gwm = request.user.username in counter_gwm

    counter_frame = ['counter_frame']
    show_button_for_counter_frame = request.user.username in counter_frame


    counter_trim_out_changan = ['counter_trim_out_changan']
    show_button_for_counter_trim_out_changan = request.user.username in counter_trim_out_changan

    counter_trim_out_chery = ['counter_trim_out_chery']
    show_button_for_counter_trim_out_chery = request.user.username in counter_trim_out_chery

    counter_trim_out_gwm = ['counter_trim_out_gwm']
    show_button_for_counter_trim_out_gwm = request.user.username in counter_trim_out_gwm

    counter_trim_out_frame = ['counter_trim_out_frame']
    show_button_for_counter_trim_out_frame = request.user.username in counter_trim_out_frame


    show_manual_buttons = request.user.username in allowed_usernames
    show_head_area_posts_buttons = request.user.username in head_area_usernames
    show_master_posts_buttons = request.user.username in master_usernames

    return render(request, "users/controller_dashboard.html", {
        "show_manual_buttons": show_manual_buttons,
        "show_head_area_posts_buttons": show_head_area_posts_buttons,
        "show_master_posts_buttons": show_master_posts_buttons,

        'show_button_for_counter_changan': show_button_for_counter_changan,
        'show_button_for_counter_chery': show_button_for_counter_chery,
        'show_button_for_counter_gwm': show_button_for_counter_gwm,
        'show_button_for_counter_frame': show_button_for_counter_frame,

        'show_button_for_counter_trim_out_changan': show_button_for_counter_trim_out_changan,
        'show_button_for_counter_trim_out_chery': show_button_for_counter_trim_out_chery,
        'show_button_for_counter_trim_out_gwm': show_button_for_counter_trim_out_gwm,
        'show_button_for_counter_trim_out_frame': show_button_for_counter_trim_out_frame,

        "recent_posts": request.session.get("recent_posts", []),
        "is_controller": getattr(request.user, "role", "") == "controller",
    })

    # Counter123!

@login_required
@role_required(["uud_controller"])
def uud_uniq(request):
    from assembly.views import uud_uniq as assembly_uud_uniq
    return assembly_uud_uniq(request)



# ✅ Мастер имеет доступ только к своей панели
@login_required
@permission_required('users.access_to_the_master_dashboard', raise_exception=True)
def master_dashboard(request):
    can_view_qrr = request.user.has_perm(f"{APP_LABEL}.access_to_the_qrr_responsible")

    show_head_area_posts_buttons = request.user.has_perm(f"{APP_LABEL}.access_to_post_of_the_controller")
    show_uud_tables_username      = request.user.has_perm(f"{APP_LABEL}.access_to_the_uud_table")
    show_marriage_tables_username = request.user.has_perm(f"{APP_LABEL}.access_to_the_marriage_table")

    special_usernames = request.user.has_perm(f"{APP_LABEL}.access_to_dp_rvd")

    return render(request, "users/master_dashboard.html", {
        "special_usernames": special_usernames,
        "can_view_qrr": can_view_qrr,
        'show_head_area_posts_buttons': show_head_area_posts_buttons,
        'show_uud_tables_username': show_uud_tables_username,
        'show_marriage_tables_username': show_marriage_tables_username,
        "enable_particles": True,
        "enable_background": True,
        "enable_white_square": False
    })


# ✅ Админ имеет доступ только к своей панели
@login_required
@role_required(["admin"])
def admin_dashboard(request):
    return render(request, "users/admin_dashboard.html")


def csrf_failure(request, reason=""):
    return render(request, "users/403_csrf.html", {"reason": reason}, status=403)


# ✅ Только мастера могут заходить в таблицу дефектов
@login_required
@role_required(["master", "head_area"])
def defect_table(request):
    """Страница с таблицей всех дефектов (только для мастера)"""
    search_query = request.GET.get("search", "")
    sort_by = request.GET.get("sort_by", "-created_at")
    start_date = request.GET.get("start_date", "")
    end_date = request.GET.get("end_date", "")

    defects = Inspection.objects.all()  # ✅ Переключили на новую модель

    if search_query:
        defects = defects.filter(
            Q(container_number__icontains=search_query) |
            Q(pallet_number__icontains=search_query) |
            Q(detail__name__icontains=search_query) |
            Q(defect__name__icontains=search_query)
        )

    if start_date:
        defects = defects.filter(created_at__gte=start_date)
    if end_date:
        defects = defects.filter(created_at__lte=end_date)

    defects = defects.order_by(sort_by)

    return render(request, "users/master_defects_table.html", {
        "defects": defects,
        "search_query": search_query,
        "sort_by": sort_by,
        "start_date": start_date,
        "end_date": end_date
    })


# ✅ API для получения статистики дефектов в реальном времени
@login_required
@role_required(["master", "head_area"])
def get_defect_stats(request):
    """API для получения статистики дефектов (обновляется автоматически)"""
    today = now().date()
    start_of_week = today - timedelta(days=today.weekday())
    start_of_month = today.replace(day=1)

    # 🟢 Топ-5 дефектов
    top_all_time = list(
        Inspection.objects.values("detail__name", "defect__name")
        .annotate(total=Count("id"))
        .order_by("-total")[:5]
    )

    top_today = list(
        Inspection.objects.filter(created_at__gte=today)
        .values("detail__name", "defect__name")
        .annotate(total=Count("id"))
        .order_by("-total")[:5]
    )

    top_week = list(
        Inspection.objects.filter(created_at__gte=start_of_week)
        .values("detail__name", "defect__name")
        .annotate(total=Count("id"))
        .order_by("-total")[:5]
    )

    top_month = list(
        Inspection.objects.filter(created_at__gte=start_of_month)
        .values("detail__name", "defect__name")
        .annotate(total=Count("id"))
        .order_by("-total")[:5]
    )

    return JsonResponse({
        "top_all_time": top_all_time,
        "top_today": top_today,
        "top_week": top_week,
        "top_month": top_month,
    }, safe=False)


# ✅ API для обновления таблицы дефектов
@login_required
@role_required(["master", "head_area"])
def get_defects(request):
    """API для получения данных без перезагрузки страницы"""
    defects = Inspection.objects.all().order_by("-created_at")

    data = []
    for defect in defects:
        data.append({
            "container_number": defect.container_number,
            "pallet_number": defect.pallet_number,
            "detail_name": defect.detail.name if defect.detail else "-",
            "defect_name": defect.defect.name if defect.defect else "-",
            "controller_username": defect.controller.username if defect.controller else "-",
            "created_at": defect.created_at.strftime("%Y-%m-%d %H:%M:%S"),
            "container_image": request.build_absolute_uri(defect.container_image.url) if defect.container_image else None,
            "pallet_image": request.build_absolute_uri(defect.pallet_image.url) if defect.pallet_image else None,
            "defect_images": [request.build_absolute_uri(img) for img in defect.get_defect_images()]
        })

    return JsonResponse(data, safe=False)


# ✅ Страница входа с редиректом по ролям
def login_view(request):
    form = AuthenticationForm(request, data=request.POST or None)

    # Если уже залогинен как guest — сразу в MES
    if request.user.is_authenticated and request.user.get_username() == "guest":
        return redirect("mes_dashboard")

    if request.method == "POST":
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            # Спец-правило: пользователь "guest" всегда уходит на MES дашборд
            if user.get_username() == "guest":
                return redirect("mes_dashboard")
            role = user.role

            # 🔁 Перенаправление в зависимости от роли
            if role in ["qrr_specialist", "qrr_supervisor", "qrr_shift_lead"]:
                return redirect(reverse("qrr_dashboard"))
            elif role == "controller":
                return redirect("controller_dashboard")
            elif role in ["master", "head_area", "dp"]:
                return redirect("master_dashboard")
            elif role == "admin":
                return redirect("admin_dashboard")
            elif user.role == 'uud_controller':
                return redirect('uud_uniq')
            return redirect("/")  # fallback
        else:
            messages.error(request, "Неверное имя пользователя или пароль.")

    return render(request, "users/login.html", {"form": form})


# ✅ Выход из аккаунта
@login_required
def logout_view(request):
    logout(request)
    return redirect("login")


# ✅ Редирект на страницу мастера
def master_redirect(request):
    return redirect("master_dashboard")


# ✅ Детальная страница поста
def post_detail(request, post_id):
    return render(request, "users/post_detail.html", {"post_id": post_id})


# ✅ Страница для создания контроллера (если требуется)
@login_required
@role_required(["admin"])
def create_controller(request):
    return render(request, "users/create_controller.html")


def defect_reports(request):
    return render(request, "users/master_defect_reports.html")


@login_required
@role_required(["master", "head_area"])
def notifications_view(request):
    notifications = Notification.objects.filter(recipient=request.user).order_by('-created_at')
    return render(request, "users/notifications.html", {"notifications": notifications})


@login_required
def get_notifications(request):
    notifications = Notification.objects.filter(recipient=request.user, is_read=False).order_by('-created_at')
    unread_count = notifications.count()
    latest_message = notifications.first().message if notifications.exists() else ""

    return JsonResponse({
        "unread_count": unread_count,
        "latest_message": latest_message
    })


@login_required
def sb_defect_details_view(request, defect_id):
    defect = get_object_or_404(ContainerUnloadingZoneSBInspection, id=defect_id)

    # Отмечаем уведомление как прочитанное, если оно есть
    Notification.objects.filter(recipient=request.user, message__icontains=defect.container_number).update(is_read=True)

    return render(request, "users/sb_notification_defect.html", {"defect": defect})


@login_required
@role_required(["master", "head_area"])
def uud_defect_details_view(request, vin_number):
    history_entry = get_object_or_404(VINHistory, vin=vin_number)

    # Отмечаем уведомления как прочитанные
    Notification.objects.filter(recipient=request.user, vin_number=vin_number).update(is_read=True)

    history = history_entry.history
    uud_data = None

    # Безопасно получаем данные из структуры
    zone_data = history.get("Цех УУД", {})
    post_data = zone_data.get("Участок устранения дефектов, DKD", [])

    if isinstance(post_data, list) and post_data:
        # Берем последнюю запись (можно также перебрать, если надо)
        uud_data = post_data[-1]

    return render(request, "users/uud_notification.html", {
        "vin": vin_number,
        "uud_data": uud_data,
    })

@login_required
@role_required(["master", "head_area"])
def master_panel(request):
    vins = VINHistory.objects.values_list('vin', flat=True).distinct()
    return render(request, 'users/master_dashboard.html', {'vin_list': vins})


@login_required
@permission_required('users.access_to_the_guide', raise_exception=True)
def helpdesk_directory(request):
    query = request.GET.get("q", "").strip()
    contacts = HelpdeskContact.objects.all().order_by(Lower("department"))

    if query:
        q_objects = Q()
        for variant in (query, query.lower(), query.title(), query.upper()):
            q_objects |= Q(employee_name__icontains=variant)
            q_objects |= Q(department__icontains=variant)
            q_objects |= Q(position__icontains=variant)
            q_objects |= Q(email__icontains=variant)

        q_digits = normalize_digits(query)
        if q_digits:
            try:
                contacts = contacts.annotate(
                    phone_digits=RegexReplace(F("phone_number"), Value(r"\D+"), Value(""), Value("g"))
                )
                q_objects |= Q(phone_digits__icontains=q_digits)
                contacts = contacts.filter(q_objects)
            except DatabaseError:
                ids = [c.id for c in contacts if q_digits in normalize_digits(c.phone_number)]
                contacts = contacts.filter(q_objects | Q(id__in=ids))
        else:
            contacts = contacts.filter(q_objects)

    return render(request, "users/helpdesk_directory.html", {
        "contacts": contacts,
        "query": query,
        "enable_particles": True,
        "enable_background": True,
        "enable_white_square": False
    })








@login_required
@permission_required('users.access_to_tracking_by_vin', raise_exception=True)
def vin_tracking_overview(request):
    today = localtime(now()).date()
    start_of_day = make_aware(datetime.combine(today, time.min))
    end_of_day = make_aware(datetime.combine(today, time.max))

    posts_by_zone = {
        "Цех поставки": [
            "Зона первичного осмотра кузовов, DKD",
            "Зона выгрузки комплектующих, DKD",
            "Зона основной приемки DKD"
        ],
        "Цех сборки": [
            "Пост момента затяжек",
            "Chassis Chery",
            "Финал текущий контроль",
            "Первая инспекция",
            "Геометрия колес",
            "Регулировка света фар",
            "Тормозная система",
            "ADAS",
            "AVM",
            "Герметичность кузова",
            "Финал",
            "ТЕСТ трек",
            "Underbody",
            "Shower Test",
        ]
    }

    counters = {(zone, post): 0 for zone, posts in posts_by_zone.items() for post in posts}

    for history in VINHistory.objects.all():
        for zone, posts in posts_by_zone.items():
            zone_data = history.history.get(zone, {})
            for post in posts:
                entries = zone_data.get(post, [])
                for entry in entries:
                    date_str = entry.get("date_added")
                    if not date_str:
                        continue
                    dt = parse_datetime(date_str)
                    if dt and start_of_day <= dt <= end_of_day:
                        counters[(zone, post)] += 1
                        break

    return render(request, "users/tracking/vin_tracking_overview.html", {
        "posts_by_zone": posts_by_zone,
        "counters": counters,
    })


@login_required
@permission_required('users.access_to_tracking_by_vin', raise_exception=True)
def vin_tracking_post_vins(request, post):
    today = now().date()
    vins_today = set()

    for history in VINHistory.objects.all():
        for zone_data in history.history.values():
            entries = zone_data.get(post, [])
            for entry in entries:
                date_str = entry.get("date_added")
                if date_str and parse_datetime(date_str).date() == today:
                    vins_today.add(history.vin)
                    break

    return render(request, "users/tracking/vin_tracking_post_vins.html", {
        "post": post,
        "vins": sorted(vins_today),
    })


@login_required
@permission_required('users.access_to_tracking_by_vin', raise_exception=True)
def vin_tracking_select(request):
    vins = VINHistory.objects.values_list("vin", flat=True).distinct()
    return render(request, "users/tracking/vin_tracking_select.html", {
        "vins": vins,
        "enable_particles": True,
        "enable_background": True,
        "enable_white_square": False,
    })


def get_worst_grade(defects):
    grades_order = ["V1+", "V1", "V2", "V3"]
    for grade in grades_order:
        for defect in defects:
            if defect.get("grade") == grade:
                return grade
    return None


from collections import defaultdict
from django.utils.dateparse import parse_datetime




@login_required
@permission_required('users.access_to_tracking_by_vin', raise_exception=True)
def vin_tracking_view(request, vin):
    # --- TRIM IN пост ---
    TRIM_IN_POST = "TRIM IN"

    # Карта пост → участок
    base_mapping = {
        # Текущий контроль
        "Пост момента затяжек": "Текущий контроль",
        "Chassis": "Текущий контроль",
        "Финал текущий контроль": "Текущий контроль",

        # Первая инспекция
        "Зазоры и перепады": "Первая инспекция",
        "Экстерьер": "Первая инспекция",
        "Интерьер": "Первая инспекция",
        "Багажник": "Первая инспекция",
        "Мотор": "Первая инспекция",
        "Функцонал": "Первая инспекция",

        # Тестовая линия
        "Геометрия колес": "Тестовая линия",
        "Регулировка света фар и калибровка руля": "Тестовая линия",
        "Тормозная система": "Тестовая линия",
        "Underbody": "Тестовая линия",
        "ADAS": "Тестовая линия",
        "AVM": "Тестовая линия",
        "Герметичность кузова": "Тестовая линия",

        # Финальная инспекция
        "Диагностика": "Финальная инспекция",
        "Тест трек": "Финальная инспекция",
        "Документация": "Финальная инспекция",
    }
    # Вставляем TRIM IN первым в «Текущий контроль»
    POST_AREA_MAPPING = {TRIM_IN_POST: "Текущий контроль", **base_mapping}

    posts_by_zone = {
        "Цех поставки": [
            "Зона первичного осмотра кузовов, DKD",
            "Зона выгрузки комплектующих, DKD",
            "Зона основной приемки DKD",
        ],
        "Цех сборки": list(POST_AREA_MAPPING.keys()),  # TRIM IN тоже в этом списке
    }

    # Сайдбар VIN-ов
    vins = list(VINHistory.objects.values_list("vin", flat=True).order_by("vin"))
    try:
        current_index = vins.index(vin)
    except ValueError:
        current_index = -1

    trace = TraceData.objects.filter(vin_rk=vin).first()
    # Метаданные модели/цвета для ВСЕХ VIN в списке (для сайдбара)
    traces_qs = TraceData.objects.filter(vin_rk__in=vins)
    trace_map = {t.vin_rk: t for t in traces_qs}
    vin_items = [
        {
            "vin": v,
            "model": (trace_map.get(v).model if trace_map.get(v) else ""),
            "color_1c": (trace_map.get(v).color_1c if trace_map.get(v) else ""),
        }
        for v in vins
    ]
    vin_history = VINHistory.objects.filter(vin=vin).first()
    post_statuses = []                    # карточки: Цех поставки
    grouped_statuses = defaultdict(list)  # карточки: Цех сборки (по участкам)
    timeline = []                         # общая лента событий со step’ами

    # Default initializations for certain status variables
    last_step_post = None
    is_on_uud = False
    is_on_ves = False
    uud_latest_name = None
    uud_latest_time = None
    ves_latest_name = None
    ves_latest_time = None
    is_on_sgp = False
    sgp_latest_time = None

    if vin_history:
        history_data = vin_history.history or {}

        # === Flag: was there ANY defect with UUD status "impossible" at any time? ===
        has_uud_impossible = False
        try:
            for zone_dict in (history_data or {}).values():
                if has_uud_impossible:
                    break
                for entries in (zone_dict or {}).values():
                    if has_uud_impossible:
                        break
                    for entry in (entries or []):
                        defects = entry.get("defects", []) or []
                        if not isinstance(defects, list):
                            continue
                        for d in defects:
                            if not isinstance(d, dict):
                                continue
                            extra = d.get("extra", {}) or {}
                            uud = extra.get("UUD", {}) or {}
                            status = (uud.get("status") or "").lower()
                            if status == "impossible":
                                has_uud_impossible = True
                                break
                        if has_uud_impossible:
                            break
        except Exception:
            has_uud_impossible = False

        # --- Собираем ВСЕ события в raw_events ---

        raw_events = []  # (dt, zone_name, post_name)
        last_at = {}  # будет заполнен позже, оставляем имя для читабельности
        post_steps = defaultdict(list)
        timeline = []

        # 1) TRIM IN из AssemblyPassLog
        trim_in_datetimes = []
        if AssemblyPassLog is not None:
            for log in AssemblyPassLog.objects.filter(vin=vin).order_by("scanned_at"):
                if log.scanned_at:
                    raw_events.append((log.scanned_at, "Цех сборки", TRIM_IN_POST))
                    trim_in_datetimes.append(log.scanned_at)

        # 2) Посты из VINHistory (Цех поставки + Цех сборки)
        for zone_name, posts in posts_by_zone.items():
            zone_dict = history_data.get(zone_name, {}) or {}
            for post in posts:
                if post == TRIM_IN_POST:
                    continue
                for entry in zone_dict.get(post, []) or []:
                    dt = parse_datetime(entry.get("date_added") or entry.get("date") or "")
                    if dt:
                        raw_events.append((dt, zone_name, post))

        # 3) УУД: 4 шага
        uud_zone = history_data.get("УУД", {}) or {}
        uud_entries = uud_zone.get("УУД", []) or []
        uud_posts = {
            "step1_at": "УУД — Шаг 1 (отдана на УУД)",
            "step2_at": "УУД — Шаг 2 (принята на УУД)",
            "step3_at": "УУД — Шаг 3 (отдана на линию)",
            "step4_at": "УУД — Шаг 4 (принята на линию)",
        }
        for entry in uud_entries:
            extra = entry.get("extra_data", {}) or {}
            for step_key, post_name in uud_posts.items():
                dt_str = extra.get(step_key)
                if not dt_str:
                    continue
                dt = parse_datetime(dt_str)
                if dt:
                    raw_events.append((dt, "УУД", post_name))

        # 4) VES: отдан/принят
        ves_logs = VESPassLog.objects.filter(vin=vin).order_by("given_at")
        ves_posts = {"given": "VES — Отдан", "received": "VES — Принят"}
        for log in ves_logs:
            if log.given_at:
                raw_events.append((log.given_at, "VES", ves_posts["given"]))
            if log.received_at:
                raw_events.append((log.received_at, "VES", ves_posts["received"]))

        from datetime import timedelta

        def _minute(dt):
            return dt.replace(second=0, microsecond=0)

        # Индекс статусов конкретных записей: (zone, post, dt_minute) -> 'ok' | 'defect'
        entry_status_by_minute = {}

        for zname in ("Цех сборки", "Цех поставки"):
            zdict = history_data.get(zname, {}) or {}
            for pname, entries in zdict.items():
                for e in entries or []:
                    dt_e = parse_datetime(e.get("date_added") or e.get("date") or "")
                    if not dt_e:
                        continue
                    status_e = "defect" if (e.get("has_defect") == "yes" or e.get("defects")) else "ok"
                    entry_status_by_minute[(zname, pname, _minute(dt_e))] = status_e



        # --- ЕДИНСТВЕННЫЙ пересчёт глобальной шкалы шагов ---
        raw_events.sort(key=lambda t: t[0])
        post_steps.clear()
        last_at = {}
        timeline.clear()

        # Допуск, если минуты не совпадают (можно оставить 0 минут, если совпадают точно)
        TOL = timedelta(minutes=3)

        for step, (dt, zone_name, post_name) in enumerate(raw_events, start=1):
            status = "ok"
            if zone_name in {"Цех сборки", "Цех поставки"}:
                key = (zone_name, post_name, _minute(dt))
                if key in entry_status_by_minute:
                    status = entry_status_by_minute[key]
                else:
                    # необязательный «поиск ближайшей записи» в пределах TOL
                    nearest = None
                    nearest_diff = None
                    # перебор вокруг нужного времени (несколько минут назад/вперёд)
                    for shift in range(-int(TOL.total_seconds() // 60), int(TOL.total_seconds() // 60) + 1):
                        k = (zone_name, post_name, _minute(dt + timedelta(minutes=shift)))
                        if k in entry_status_by_minute:
                            diff = abs((dt - (dt + timedelta(minutes=shift))).total_seconds())
                            if nearest_diff is None or diff < nearest_diff:
                                nearest = entry_status_by_minute[k]
                                nearest_diff = diff
                    if nearest is not None:
                        status = nearest
                    # иначе останется 'ok'

            timeline.append({"step": step, "dt": dt, "zone": zone_name, "post": post_name, "status": status})
            post_steps[post_name].append({"step": step, "dt": dt, "status": status})
            last_at[post_name] = dt

        # --- Карточки для VES (отдан / принят) ---
        ves_given = "VES — Отдан"
        ves_received = "VES — Принят"

        grouped_statuses["VES"] = [
            {
                "zone": "VES",
                "post": ves_given,
                "status": "ok" if last_at.get(ves_given) else "missing",
                "grade": None,
                "last_at": last_at.get(ves_given),
                "steps": post_steps.get(ves_given, []),
            },
            {
                "zone": "VES",
                "post": ves_received,
                "status": "ok" if last_at.get(ves_received) else "missing",
                "grade": None,
                "last_at": last_at.get(ves_received),
                "steps": post_steps.get(ves_received, []),
            },
        ]

        uud_posts_map = {
            "УУД — Шаг 1 (отдана на УУД)": "step1_at",
            "УУД — Шаг 2 (принята на УУД)": "step2_at",
            "УУД — Шаг 3 (отдана на линию)": "step3_at",
            "УУД — Шаг 4 (принята на линию)": "step4_at",
        }
        grouped_statuses["УУД"] = [
            {
                "zone": "УУД",
                "post": post_name,
                "status": "ok" if last_at.get(post_name) else "missing",
                "grade": None,
                "last_at": last_at.get(post_name),
                "steps": post_steps.get(post_name, []),
            }
            for post_name in uud_posts_map.keys()
        ]


        # --- Карточки «Цех поставки» ---
        post_statuses = []
        for post in posts_by_zone["Цех поставки"]:
            entries = (history_data.get("Цех поставки", {}) or {}).get(post, []) or []
            if entries:
                defects = []
                has_defect_flag = any(e.get("has_defect") == "yes" for e in entries)
                for e in entries:
                    d = e.get("defects", [])
                    if isinstance(d, list):
                        defects.extend(d)
                worst_grade = get_worst_grade(defects)
                overall_status = "defect" if defects or has_defect_flag else "ok"
            else:
                overall_status = "missing"
                worst_grade = None

            post_statuses.append({
                "zone": "Цех поставки",
                "post": post,
                "status": overall_status,
                "grade": worst_grade,
                "last_at": last_at.get(post),
                "steps": post_steps.get(post, []),  # ГЛОБАЛЬНЫЕ шаги
            })

        # --- Карточки «Цех сборки», включая TRIM IN ---
        for post, area in POST_AREA_MAPPING.items():
            if post == TRIM_IN_POST:
                present = bool(trim_in_datetimes)
                status = "ok" if present else "missing"
                grouped_statuses[area].append({
                    "zone": "Цех сборки",
                    "post": post,
                    "status": status,
                    "grade": None,
                    "last_at": last_at.get(post),
                    "steps": post_steps.get(post, []),  # ГЛОБАЛЬНЫЕ шаги
                })
                continue

            entries = (history_data.get("Цех сборки", {}) or {}).get(post, []) or []
            if entries:
                defects = []
                has_defect_flag = any(e.get("has_defect") == "yes" for e in entries)
                for e in entries:
                    d = e.get("defects", [])
                    if isinstance(d, list):
                        defects.extend(d)
                worst_grade = get_worst_grade(defects)
                overall_status = "defect" if defects or has_defect_flag else "ok"
            else:
                overall_status = "missing"
                worst_grade = None

            grouped_statuses[area].append({
                "zone": "Цех сборки",
                "post": post,
                "status": overall_status,
                "grade": worst_grade,
                "last_at": last_at.get(post),
                "steps": post_steps.get(post, []),  # ГЛОБАЛЬНЫЕ шаги
            })

        # --- Последний пост ---
        last_step_post = timeline[-1] if timeline else None

        # --- Статусы УУД/ VES / СГП (те же, что были) ---
        uud_step1 = "УУД — Шаг 1 (отдана на УУД)"
        uud_step2 = "УУД — Шаг 2 (принята на УУД)"
        uud_step3 = "УУД — Шаг 3 (отдана на линию)"
        uud_step4 = "УУД — Шаг 4 (принята на линию)"

        def _latest_named(names):
            latest_t = None
            latest_name = None
            for nm in names:
                t = last_at.get(nm)
                if t and (latest_t is None or t > latest_t):
                    latest_t = t
                    latest_name = nm
            return latest_t, latest_name

        uud_latest_time, uud_latest_name = _latest_named([uud_step1, uud_step2, uud_step3, uud_step4])
        is_on_uud = bool(uud_latest_name in {uud_step1, uud_step2})

        ves_given = "VES — Отдан"
        ves_received = "VES — Принят"
        t_given = last_at.get(ves_given)
        t_received = last_at.get(ves_received)
        is_on_ves = bool(t_given and (not t_received or t_given > t_received))
        ves_latest_time = t_given if (t_given and (not t_received or t_given >= t_received)) else t_received
        ves_latest_name = ves_given if (t_given and (not t_received or t_given > t_received)) else (
            ves_received if t_received else None)

        sgp_post = "Документация"
        sgp_latest_time = last_at.get(sgp_post)
        is_on_sgp = bool(sgp_latest_time)

    return render(request, "users/tracking/vin_tracking_view.html", {
        "vin": vin,
        "vins": vins,
        "current_index": current_index,
        "post_statuses": post_statuses,
        "grouped_statuses": dict(grouped_statuses),
        "last_step_post": last_step_post,
        "timeline": timeline,
        "trace": trace,
        "is_on_uud": is_on_uud,
        "is_on_ves": is_on_ves,
        "uud_latest_name": uud_latest_name,
        "uud_latest_time": uud_latest_time,
        "ves_latest_name": ves_latest_name,
        "ves_latest_time": ves_latest_time,
        "is_on_sgp": is_on_sgp,
        "sgp_latest_time": sgp_latest_time,
        "vin_items": vin_items,
        "trace_map": trace_map,
        "has_uud_impossible": has_uud_impossible,
        "enable_particles": True,
        "enable_background": True,
        "enable_white_square": False,
    })






from collections import defaultdict
from django.utils.dateparse import parse_datetime

from collections import defaultdict
from django.contrib.auth.decorators import login_required
from users.utils.post_extractors import extract_entries

@login_required
@permission_required('users.access_to_tracking_by_vin', raise_exception=True)
def vin_post_detail(request, vin, post):
    vin_history = VINHistory.objects.filter(vin=vin).first()
    if not vin_history:
        return render(request, "users/tracking/not_found.html", {"message": "VIN не найден"})

    trace_data   = TraceData.objects.filter(vin_rk=vin).first()
    history_data = vin_history.history or {}

    entries = extract_entries(history_data, post, vin=vin)  # <-- передай vin

    # агрегированные флаги на уровне поста (например, «был невозможный ремонт»)
    has_uud_impossible = any(e.get("flags", {}).get("uud_impossible") for e in entries)

    return render(request, "users/tracking/vin_post_detail.html", {
        "vin": vin,
        "post": post,
        "entries": entries,
        "trace_data": trace_data,
        "has_uud_impossible": has_uud_impossible,
    })









@login_required
@permission_required('users.access_to_tracking_by_vin', raise_exception=True)
def export_vin_excel(request, vin):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment
    from collections import defaultdict
    import os

    vin_history = VINHistory.objects.filter(vin=vin).first()
    if not vin_history:
        return HttpResponse("❌ VIN не найден", status=404)

    trace = TraceData.objects.filter(vin_rk=vin).first()
    brand = trace.brand if trace else ""
    model = trace.model if trace else ""
    config_code = trace.config_code if trace else ""

    data_rows = []
    max_photos = 0

    for zone_name, posts in vin_history.history.items():
        for post_name, entries in posts.items():
            for entry in entries:
                date_raw = entry.get("date_added")
                if not date_raw:
                    continue
                date = parse_datetime(date_raw)
                controller = entry.get("controller", "") or entry.get("extra_data", {}).get("controller", "")
                line = entry.get("line", "")
                duration = entry.get("inspection_duration_seconds", "")

                if "defects" in entry and isinstance(entry["defects"], list) and entry["defects"]:
                    for defect in entry["defects"]:
                        photos = defect.get("photos", [])
                        max_photos = max(max_photos, len(photos))
                        data_rows.append({
                            "zone": zone_name,
                            "post": post_name,
                            "vin": vin,
                            "brand": brand,
                            "model": model,
                            "config_code": config_code,
                            "date": date,
                            "line": line,
                            "controller": controller,
                            "defect": defect.get("name", ""),
                            "comment": defect.get("comment", ""),
                            "grade": defect.get("grade", ""),
                            "unit": defect.get("unit", ""),
                            "responsible": defect.get("responsible", ""),
                            "duration": duration,
                            "photos": photos,
                        })
                elif entry.get("defect_description") or entry.get("defect_photos"):
                    photos = entry.get("defect_photos", [])
                    max_photos = max(max_photos, len(photos))
                    data_rows.append({
                        "zone": zone_name,
                        "post": post_name,
                        "vin": vin,
                        "brand": brand,
                        "model": model,
                        "config_code": config_code,
                        "date": date,
                        "line": line,
                        "controller": controller,
                        "defect": entry.get("defect_description", ""),
                        "comment": "",
                        "grade": "",
                        "unit": "",
                        "responsible": "",
                        "duration": duration,
                        "photos": photos,
                    })

    # 📗 Excel
    wb = Workbook()
    ws = wb.active
    ws.title = f"VIN {vin}"

    headers = [
        "Участок", "Пост", "VIN", "Бренд", "Модель", "Код комплектации", "Дата", "Линия", "Контроллер",
        "Дефект", "Комментарий", "Грейд", "Единица", "Кто виноват", "Время осмотра (сек)"
    ] + [f"Фото {i+1}" for i in range(max_photos)]

    ws.append(headers)

    for row in data_rows:
        excel_row = [
            row["zone"],
            row["post"],
            row["vin"],
            row["brand"],
            row["model"],
            row["config_code"],
            row["date"].strftime("%d.%m.%Y %H:%M"),
            row["line"],
            row["controller"],
            row["defect"],
            row["comment"],
            row["grade"],
            row["unit"],
            row["responsible"],
            row["duration"],
        ]
        excel_row += [""] * max_photos
        ws.append(excel_row)

        row_idx = ws.max_row
        for i, img_path in enumerate(row["photos"]):
            insert_single_image(ws, row_idx, 16 + i, img_path)
        ws.row_dimensions[row_idx].height = 75

    for col in ws.columns:
        for cell in col:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    # Ответ
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="vin_report_{vin}.xlsx"'
    wb.save(response)
    return response















import os
from django.http import HttpResponse
from django.conf import settings
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from vehicle_history.models import VINHistory
from collections import defaultdict
from PIL import Image as PILImage
from io import BytesIO
import json
from django.utils.dateparse import parse_date
from django.utils.timezone import make_aware
from datetime import datetime
import io
import zipfile
from django.contrib.auth.decorators import login_required
from users.decorators import role_required
from django.utils.dateparse import parse_datetime
from supplies.models import TraceData


def insert_single_image(ws, row_idx, col_idx, path):
    try:
        rel_path = path.replace(settings.MEDIA_URL, "")
        abs_path = os.path.join(settings.MEDIA_ROOT, rel_path)

        if os.path.exists(abs_path):
            img = XLImage(abs_path)
            img.width = 85
            img.height = 85

            col_letter = get_column_letter(col_idx)
            ws.add_image(img, f"{col_letter}{row_idx}")

            # 💡 Установим ширину колонки (один раз)
            if col_letter not in ws.column_dimensions:
                ws.column_dimensions[col_letter].width = 17  # под 90px картинки

    except Exception as e:
        print(f"❌ Ошибка изображения {path}: {e}")


@login_required
@permission_required('users.access_to_the_rvd', raise_exception=True)
def employee_search(request):
    search_form = EmployeeSearchForm(request.GET)

    selected_employee_ids = request.GET.getlist('employees') if request.method == 'GET' else []

    # ВАЖНО: поднимаем сюда!
    employees = Employee.objects.all()

    employees_queryset = Employee.objects.filter(id__in=selected_employee_ids)

    if request.method == 'GET':
        form_data = request.GET.copy()
        selection_form = EmployeeSelectionForm(data=form_data, initial={'employees': selected_employee_ids})
        selection_form.fields['employees'].queryset = Employee.objects.filter(
            Q(id__in=selected_employee_ids) | Q(id__in=employees.values_list('id', flat=True))
        )
    else:
        selection_form = EmployeeSelectionForm()

    if search_form.is_valid():
        query = search_form.cleaned_data['search_query']
        position = search_form.cleaned_data['position']
        department = search_form.cleaned_data['department']

        if query:
            employees = employees.filter(name__iregex=query)
        if position:
            employees = employees.filter(position__iregex=position)
        if department:
            employees = employees.filter(department__iregex=department)

    # Обработка отправки выбора
    if request.method == 'POST':
        selection_form = EmployeeSelectionForm(request.POST)
        if selection_form.is_valid():
            try:
                selected_employees = selection_form.cleaned_data['employees']
                selected_dates = selection_form.cleaned_data['selected_dates']
                is_xtk = selection_form.cleaned_data['is_xtk']
                justification = selection_form.cleaned_data['justification']
                start_time_str = selection_form.cleaned_data['start_time']
                end_time_str = selection_form.cleaned_data['end_time']
                start_time = datetime.strptime(start_time_str, "%H:%M").time()
                end_time = datetime.strptime(end_time_str, "%H:%M").time()

                created_count = 0
                for employee in selected_employees:
                    for date_str in selected_dates:
                        selected_date = datetime.strptime(date_str, "%Y-%m-%d").date()
                        if not Selection.objects.filter(
                            manager=request.user,
                            employee=employee,
                            selected_date=selected_date
                        ).exists():
                            Selection.objects.create(
                                manager=request.user,
                                employee=employee,
                                selected_date=selected_date,
                                is_xtk=is_xtk,
                                justification=justification,
                                start_time=start_time,
                                end_time=end_time
                            )
                            created_count += 1

                messages.success(
                    request,
                    f'✅ Добавлено {created_count} записей для {len(selected_employees)} сотрудника(ов)'
                )
                return redirect('employee_search')

            except Exception as e:
                messages.error(request, f'Ошибка при сохранении: {str(e)}')

    selections = Selection.objects.filter(manager=request.user)

    context = {
        'search_form': search_form,
        'selection_form': selection_form,
        'employees': employees,
        'selections': selections,
        'employee_ids': selected_employee_ids,  # для скрытых input в шаблоне
        'enable_particles': True,
        'enable_background': True,
        'enable_white_square': False
    }
    return render(request, 'users/management/search.html', context)


@login_required
@permission_required('users.access_to_the_rvd', raise_exception=True)
def edit_selection(request, selection_id):
    selection = get_object_or_404(Selection, id=selection_id, manager=request.user)

    if request.method == 'POST':
        form = EmployeeSelectionForm(request.POST)
        if form.is_valid():
            employee = form.cleaned_data['employee']
            selected_dates = form.cleaned_data['selected_dates']
            is_xtk = form.cleaned_data['is_xtk']

            if len(selected_dates) != 1:
                messages.error(request, 'Редактировать можно только одну дату')
            else:
                selected_date = datetime.strptime(selected_dates[0], "%Y-%m-%d").date()

                # Проверка на дубликат
                if Selection.objects.filter(
                        manager=request.user,
                        employee=employee,
                        selected_date=selected_date
                ).exclude(id=selection.id).exists():
                    messages.error(request, 'Такой выбор уже существует')
                else:
                    selection.employee = employee
                    selection.selected_date = selected_date
                    selection.is_xtk = is_xtk
                    selection.save()
                    messages.success(request, f'Выбор успешно обновлён')
                    return redirect('employee_search')
    else:
        form = EmployeeSelectionForm(initial={
            'employee': selection.employee,
            'selected_dates': [selection.selected_date.strftime("%Y-%m-%d")],
            'is_xtk': selection.is_xtk
        })

    return render(request, 'users/management/edit_selection.html', {
        'form': form,
        'selection': selection
    })


@login_required
@permission_required('users.access_to_the_rvd', raise_exception=True)
def delete_all_selections(request):
    if request.method == 'POST':
        deleted_count, _ = Selection.objects.filter(manager=request.user).delete()
        messages.success(request, f'🗑 Удалено {deleted_count} выборов.')
    return redirect('employee_search')


@login_required
@permission_required('users.access_to_the_rvd', raise_exception=True)
def delete_selection(request, selection_id):
    selection = get_object_or_404(Selection, id=selection_id, manager=request.user)

    # ✅ Удаляем сразу, без формы
    employee_name = selection.employee.name
    selection.delete()
    messages.success(request, f'Выбор для {employee_name} успешно удален')

    return redirect('employee_search')


@login_required
@permission_required('users.access_to_the_rvd', raise_exception=True)
def export_excel(request):
    try:
        selections = Selection.objects.filter(manager=request.user)

        data = [{
            'Сотрудник': sel.employee.name,
            'Должность': sel.employee.position,
            'Отдел': sel.employee.department or '—',
            'Дата': sel.selected_date.strftime("%d.%m.%Y"),
            'ХТК': 'Да' if sel.is_xtk else 'Нет',
            'Часы': sel.hours
        } for sel in selections]

        if not data:
            messages.warning(request, 'Нет данных для экспорта')
            return redirect('employee_search')

        df = pd.DataFrame(data)
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        filename = f'employee_selections_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        response['Content-Disposition'] = f'attachment; filename={filename}'

        df.to_excel(response, index=False)

        ExportHistory.objects.create(
            manager=request.user,
            file_name=filename,
            record_count=len(data),
            export_data=json.dumps(data, ensure_ascii=False)
        )

        return response

    except Exception as e:
        messages.error(request, f'Ошибка при экспорте: {str(e)}')
        return redirect('employee_search')


@login_required
@permission_required('users.access_to_the_rvd', raise_exception=True)
def export_history(request):
    try:
        history = ExportHistory.objects.filter(manager=request.user)

        for record in history:
            try:
                data = record.get_data_as_list()
            except Exception as e:
                messages.error(request, f'Ошибка при чтении данных записи {record.id}: {str(e)}')

        return render(request, 'users/management/export_history.html', {
            'history': history
        })
    except Exception as e:
        messages.error(request, f'Ошибка при загрузке истории: {str(e)}')
        return redirect('employee_search')


@login_required
@permission_required('users.access_to_the_rvd', raise_exception=True)
def load_export_to_selection(request, history_id):
    try:
        history_record = get_object_or_404(ExportHistory, id=history_id, manager=request.user)
        data = history_record.get_data_as_list()

        # Clear existing selections for this user
        Selection.objects.filter(manager=request.user).delete()

        # Create new selections from the export data
        for item in data:
            # Find or create employee
            employee, created = Employee.objects.get_or_create(
                name=item['Name'],
                defaults={
                    'position': item['Position'],
                    'department': item['Department']
                }
            )

            # Create selection
            Selection.objects.create(
                manager=request.user,
                employee=employee,
                selected_date=datetime.strptime(item['Дата'], "%d.%m.%Y").date(),
                is_xtk=item.get('ХТК') == 'Да',
                hours=int(item.get('Часы', 8))
            )

        messages.success(request, 'Список успешно загружен в текущие выборы')
        return redirect('users:employee_search')
    except Exception as e:
        messages.error(request, f'Ошибка при загрузке списка: {str(e)}')
        return redirect('employee_search')


@login_required
@permission_required('users.access_to_the_rvd', raise_exception=True)
def re_export_excel(request, history_id):
    try:
        history_record = get_object_or_404(ExportHistory, id=history_id, manager=request.user)

        # Get the stored data
        try:
            data = history_record.get_data_as_list()
        except Exception as e:
            messages.error(request, f'Ошибка при чтении данных: {str(e)}')
            return redirect('employee_search')

        if not data:
            messages.warning(request, 'Нет данных для экспорта')
            return redirect('employee_search')

        # Create new filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f're_export_{history_record.file_name}'

        # Create Excel file
        df = pd.DataFrame(data)
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={filename}'

        df.to_excel(response, index=False)

        # Record the re-export in history
        try:
            ExportHistory.objects.create(
                manager=request.user,
                file_name=filename,
                record_count=len(data),
                export_data=json.dumps(data, ensure_ascii=False)
            )
        except Exception as e:
            messages.error(request, f'Ошибка при сохранении истории: {str(e)}')

        return response
    except Exception as e:
        messages.error(request, f'Ошибка при повторном экспорте: {str(e)}')
        return redirect('employee_search')


@login_required
@permission_required('users.access_to_the_rvd', raise_exception=True)
def edit_export_history(request, history_id):
    history_record = get_object_or_404(ExportHistory, id=history_id, manager=request.user)

    if request.method == 'POST':
        try:
            new_filename = request.POST.get('file_name')
            if new_filename:
                history_record.file_name = new_filename
                history_record.save()
                messages.success(request, 'Имя файла успешно обновлено')
                return redirect('users:export_history')
        except Exception as e:
            messages.error(request, f'Ошибка при обновлении: {str(e)}')

    return render(request, 'users/management/edit_export_history.html', {
        'record': history_record
    })


@login_required
@permission_required('users.access_to_the_rvd', raise_exception=True)
def delete_export_history(request, history_id):
    history_record = get_object_or_404(ExportHistory, id=history_id, manager=request.user)

    if request.method == 'POST':
        try:
            filename = history_record.file_name
            history_record.delete()
            messages.success(request, f'Запись {filename} успешно удалена')
            return redirect('users:export_history')
        except Exception as e:
            messages.error(request, f'Ошибка при удалении: {str(e)}')

    return render(request, 'users/management/delete_export_history.html', {
        'record': history_record
    })








@login_required
@permission_required('users.access_to_dp_rvd', raise_exception=True)
def office_overview(request):


    manager_filter = request.GET.get('manager')

    selections = Selection.objects.select_related('employee', 'manager') \
        .order_by('selected_date', 'employee__name')

    if manager_filter:
        selections = selections.filter(manager__id=manager_filter)

    # Группировка по сотрудникам (как раньше)
    grouped = {}
    for sel in selections:
        key = sel.employee.id
        if key not in grouped:
            grouped[key] = {
                'name': sel.employee.name,
                'position': sel.employee.position,
                'department': sel.employee.department,
                'dates': [],
                'is_xtk': sel.is_xtk,
                'justification': sel.justification,
                'manager': sel.manager.get_full_name_with_patronymic(),
                'start_time': sel.start_time.strftime('%H:%M') if sel.start_time else '',
                'end_time': sel.end_time.strftime('%H:%M') if sel.end_time else '',
                'hours': sel.hours,
            }
        grouped[key]['dates'].append(sel.selected_date.strftime("%d.%m.%Y"))

    # Статистика
    manager_ids = (
        Selection.objects.values('manager')
        .annotate(count=Count('employee', distinct=True))
    )

    user_map = {
        user.id: user for user in CustomUser.objects.filter(
            id__in=[item['manager'] for item in manager_ids]
        )
    }

    manager_stats = []
    for item in manager_ids:
        manager = user_map.get(item['manager'])
        manager_stats.append({
            'id': manager.id,
            'manager': manager,
            'count': item['count'],
        })

    return render(request, 'users/management/overview.html', {
        'grouped': grouped,
        'manager_stats': manager_stats,
        'current_manager': int(manager_filter) if manager_filter else None
    })


@login_required
@permission_required('users.access_to_dp_rvd', raise_exception=True)
def export_rvd_excel(request):


    selections = Selection.objects.select_related('employee', 'manager').order_by('employee__name', 'selected_date')

    # Группировка по сотруднику
    grouped = {}
    for sel in selections:
        emp = sel.employee
        key = emp.id
        if key not in grouped:
            grouped[key] = {
                "name": emp.name,
                "position": emp.position,
                "subdivision": "Департамент обеспечения качества",
                "dates": [],
                "xtk": "✅" if sel.is_xtk else "—",
                "justification": sel.justification,
                "total_hours": 0
            }
        time_range = f"({sel.start_time.strftime('%H:%M')}–{sel.end_time.strftime('%H:%M')})" if sel.start_time and sel.end_time else ""
        date_str = f"{sel.selected_date.strftime('%d.%m.%Y')}\n{time_range}"
        grouped[key]["dates"].append(date_str)
        grouped[key]["total_hours"] += sel.hours

    # Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "РВД"

    headers = ["№", "Сотрудник", "Должность", "Подразделение", "Даты", "Часы", "ХТК", "Обоснование"]
    ws.append(headers)

    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    font = Font(name='Arial', size=10)

    for i, (emp_id, data) in enumerate(grouped.items(), 1):
        row = [
            i,
            data["name"],
            data["position"],
            data["subdivision"],
            '\n'.join(data["dates"]),
            data["total_hours"],
            data["xtk"],
            data["justification"]
        ]
        ws.append(row)

    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 7
    ws.column_dimensions['G'].width = 7
    ws.column_dimensions['H'].width = 25

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=8):
        for cell in row:
            cell.alignment = alignment
            cell.border = thin_border
            cell.font = font

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = "РВД от департамента обеспечения качества.xlsx"
    response['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(filename)}"
    wb.save(response)
    return response


@login_required
@permission_required('users.access_to_dp_rvd', raise_exception=True)
def delete_all_selections_dp(request):

    if request.method == 'POST':
        Selection.objects.all().delete()
        messages.success(request, "✅ Все записи успешно удалены.")
    return redirect('office_overview')


@login_required
@permission_required('users.access_to_dp_rvd', raise_exception=True)
def export_word_rvd(request):

    selections = Selection.objects.select_related('employee').all()

    # 📌 Группируем по сотрудникам
    from collections import defaultdict
    grouped = defaultdict(lambda: {"dates": [], "position": "", "name": ""})
    for sel in selections:
        grouped[sel.employee_id]["name"] = sel.employee.name
        grouped[sel.employee_id]["position"] = sel.employee.position
        grouped[sel.employee_id]["dates"].append({
            "date": sel.selected_date,
            "start_time": sel.start_time.strftime("%H:%M") if sel.start_time else "",
            "end_time": sel.end_time.strftime("%H:%M") if sel.end_time else "",
            "hours": sel.hours
        })

    # 📄 Загружаем шаблон
    doc = Document("media/templates/messages_template_v2.docx")

    # ⬇️ Таблица — первая
    table = doc.tables[1]

    # Удалим старые строки кроме заголовка
    while len(table.rows) > 1:
        table._tbl.remove(table.rows[1]._tr)

    # 🧾 Заполняем
    for idx, (emp_id, data) in enumerate(grouped.items(), start=1):
        row = table.add_row().cells
        if len(row) < 9:
            raise ValueError("❌ В шаблоне недостаточно колонок — должно быть 9")

        row[0].text = str(idx)                          # №
        row[1].text = data["name"]                      # ФИО
        row[2].text = data["position"]                  # Должность
        row[3].text = ""                                # Согласие / несогласие
        row[4].text = ""                                # С оплатой
        row[5].text = ""                               # День отдыха
        dates_sorted = sorted(data["dates"], key=lambda x: x["date"])
        row[6].text = "\n".join(
            f"{d['date'].strftime('%d.%m.%Y')}\n({d['start_time']}–{d['end_time']})" for d in dates_sorted
        )
        row[7].text = str(sum(d["hours"] for d in dates_sorted))

    # 📥 Отправляем файл
    buffer = BytesIO()
    doc.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document"
    )

    filename = f"Уведомление_о_работе_в_выходной_{datetime.now().date()}.docx"
    quoted_filename = quote(filename)

    response["Content-Disposition"] = f"attachment; filename*=UTF-8''{quoted_filename}"
    return response



@login_required
@permission_required('users.access_to_dp_rvd', raise_exception=True)
def overtime_overview(request):


    manager_filter = request.GET.get('manager')
    records = OvertimeRecord.objects.select_related('employee', 'added_by').order_by('-date')
    if manager_filter:
        records = records.filter(added_by__id=manager_filter)

    grouped_with_day_off = {}
    grouped_without_day_off = {}

    for r in records:
        key = r.employee.id

        group = {
            'employee': r.employee,
            'manager': r.added_by.get_full_name_with_patronymic()
            if hasattr(r.added_by, 'get_full_name_with_patronymic')
            else r.added_by.get_full_name(),
        }

        if r.compensated_day:
            if key not in grouped_with_day_off:
                grouped_with_day_off[key] = {**group, 'records': [], 'day_off_count': 0}
            grouped_with_day_off[key]['records'].append(r)
            grouped_with_day_off[key]['day_off_count'] += 1
        else:
            if key not in grouped_without_day_off:
                grouped_without_day_off[key] = {**group, 'records': [], 'pending_day_off_count': 0}
            grouped_without_day_off[key]['records'].append(r)
            grouped_without_day_off[key]['pending_day_off_count'] += 1

    # Статистика по менеджерам
    manager_ids = (
        OvertimeRecord.objects.values('added_by')
        .annotate(count=Count('employee', distinct=True))
    )
    user_map = {
        user.id: user for user in CustomUser.objects.filter(
            id__in=[item['added_by'] for item in manager_ids]
        )
    }

    manager_stats = []
    for item in manager_ids:
        manager = user_map.get(item['added_by'])
        if manager:
            manager_stats.append({
                'id': manager.id,
                'manager': manager,
                'count': item['count'],
            })

    return render(request, 'users/management/overtime_overview.html', {
        'with_day_off': grouped_with_day_off,
        'without_day_off': grouped_without_day_off,
        'manager_stats': manager_stats,
        'current_manager': int(manager_filter) if manager_filter else None,
    })


@login_required
@permission_required('users.access_to_dp_rvd', raise_exception=True)
def assign_day_off(request):



    record_id = request.POST.get("record_id")
    day_off = request.POST.get("day_off")

    try:
        record = OvertimeRecord.objects.get(id=record_id)
        record.compensated_day = day_off
        record.save()
        messages.success(request, "Выходной успешно назначен.")
    except OvertimeRecord.DoesNotExist:
        messages.error(request, "Запись не найдена.")
    except Exception as e:
        messages.error(request, f"Ошибка: {str(e)}")

    return redirect(request.META.get('HTTP_REFERER', '/'))


@login_required
@permission_required('users.access_to_dp_rvd', raise_exception=True)
def update_day_off(request):


    record_id = request.POST.get('record_id')
    record = get_object_or_404(OvertimeRecord, id=record_id)

    if 'delete_day_off' in request.POST:
        record.compensated_day = None
    else:
        new_day_off = request.POST.get('new_day_off')
        record.compensated_day = new_day_off

    record.save()
    return redirect(request.META.get('HTTP_REFERER', '/'))


@login_required
@permission_required('users.access_to_dp_rvd', raise_exception=True)
def export_pending_day_off(request):


    records = OvertimeRecord.objects.filter(compensated_day__isnull=True).select_related('employee')

    grouped = {}
    for r in records:
        key = r.employee.id
        if key not in grouped:
            grouped[key] = {
                'name': r.employee.name,
                'dates': [],
                'debt': 0,
            }
        grouped[key]['dates'].append(r.date.strftime('%d.%m.%Y'))
        grouped[key]['debt'] += 1

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Долги по выходным"
    ws.append(["ФИО", "Даты переработок", "Долг (дней)"])

    # Стили заголовка
    header_font = Font(bold=True, size=13)
    for col in range(1, 4):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        ws.column_dimensions[get_column_letter(col)].width = 30

    for i, data in enumerate(grouped.values(), start=2):
        ws.cell(row=i, column=1, value=data['name'])

        dates_str = "\n".join(data['dates'])
        cell_dates = ws.cell(row=i, column=2, value=dates_str)
        cell_dates.alignment = Alignment(wrap_text=True, vertical='top')

        ws.cell(row=i, column=3, value=data['debt'])

        # Увеличение высоты строки
        ws.row_dimensions[i].height = 15 * len(data['dates'])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"overtime_debt_{date.today()}.xlsx"
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response






@login_required
@permission_required('users.access_to_pqa_data', raise_exception=True)
def incoming_workshop_dashboard(request):
    today_str = now().date().isoformat()
    zone = "Цех поставки"

    # Отображаемые названия → ключи постов в истории
    POST_NAME_MAP = {
        "Пост первичного осмотра кузовов": "Зона первичного осмотра кузовов, DKD",
        "Пост приемки комплектующих": "Зона выгрузки комплектующих, DKD",
        "Пост основной приемки": "Зона основной приемки DKD",
        "Пост приемки контейнеров": "Приемка контейнеров",
    }

    # Счётчики за сегодня
    post_vin_set = defaultdict(set)
    post_defect_counts = defaultdict(int)

    for history in VINHistory.objects.all():
        vin = history.vin
        posts = (history.history or {}).get(zone, {})

        # Проходим только по нужным постам
        for display_title, post_key in POST_NAME_MAP.items():
            entries = posts.get(post_key, []) or []
            for entry in entries:
                raw_date = entry.get("date_added") or entry.get("date") or ""
                date_str = raw_date[:10]
                if date_str != today_str:
                    continue

                # Уникальные единицы за сегодня: для контейнеров считаем по номеру контейнера, иначе по VIN
                if post_key == "Приемка контейнеров":
                    key_item = entry.get("container_number") or f"no-container-{vin}"
                else:
                    key_item = vin
                post_vin_set[post_key].add(key_item)

                # Подсчёт дефектов: новый формат (список dict'ов) или флаг has_defect
                defects = entry.get("defects")
                if isinstance(defects, list):
                    post_defect_counts[post_key] += len(defects)
                elif entry.get("has_defect") == "yes":
                    post_defect_counts[post_key] += 1

    # Формируем карточки для дашборда
    posts = [
        {
            "title": "Пост первичного осмотра кузовов",
            "icon": "",
            "report_url": reverse("body_inspection_report"),
            "table_url": reverse("body_inspection_table"),
            "export_url": reverse("body_inspection_export"),
            "form_url": reverse("body_inspection_form"),
            "today_vins": len(post_vin_set.get(POST_NAME_MAP["Пост первичного осмотра кузовов"], set())),
            "today_entries": post_defect_counts.get(POST_NAME_MAP["Пост первичного осмотра кузовов"], 0),
        },
        {
            "title": "Пост приемки комплектующих",
            "icon": "",
            "report_url": reverse("parts_acceptance_report"),
            "table_url": reverse("parts_acceptance_table"),
            "export_url": reverse("parts_acceptance_export"),
            "form_url": reverse("parts_acceptance_form"),
            "today_vins": len(post_vin_set.get(POST_NAME_MAP["Пост приемки комплектующих"], set())),
            "today_entries": post_defect_counts.get(POST_NAME_MAP["Пост приемки комплектующих"], 0),
        },
        {
            "title": "Пост основной приемки",
            "icon": "",
            "report_url": reverse("final_acceptance_report"),
            "table_url": reverse("final_acceptance_table"),
            "export_url": reverse("final_acceptance_export"),
            "form_url": reverse("final_acceptance_form"),
            "today_vins": len(post_vin_set.get(POST_NAME_MAP["Пост основной приемки"], set())),
            "today_entries": post_defect_counts.get(POST_NAME_MAP["Пост основной приемки"], 0),
        },
        {
            "title": "Пост приемки контейнеров",
            "icon": "",
            "report_url": "#",  # заглушка
            "table_url": reverse("containers_acceptance_table"),
            "export_url": reverse("containers_acceptance_export"),
            "form_url": "#",  # заглушка
            "today_vins": len(post_vin_set.get(POST_NAME_MAP["Пост приемки контейнеров"], set())),
            "today_entries": post_defect_counts.get(POST_NAME_MAP["Пост приемки контейнеров"], 0),
        },
    ]

    return render(request, "users/incoming/incoming_workshop_dashboard.html", {
        "posts": posts,
        "enable_particles": True,
        "enable_background": True,
        "enable_white_square": False,
    })


@login_required
@permission_required('users.access_to_pqa_data', raise_exception=True)
def incoming_general_report(request):
    posts_info = {
        "body_inspection": {
            "zone": "Цех поставки",
            "post": "Зона первичного осмотра кузовов, DKD",
            "title": "Пост первичного осмотра кузовов",
        },
        "parts_acceptance": {
            "zone": "Цех поставки",
            "post": "Зона выгрузки комплектующих, DKD",
            "title": "Пост приёмки комплектующих",
        },
        "final_acceptance": {
            "zone": "Цех поставки",
            "post": "Зона основной приемки DKD",
            "title": "Пост основной приёмки",
        }
    }

    # Период
    period = request.GET.get("period", "week")
    today = now().date()

    start_date_str = request.GET.get("start_date")
    end_date_str = request.GET.get("end_date")

    try:
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date() if start_date_str else None
        end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date() if end_date_str else None
    except ValueError:
        start_date = end_date = None

    if not (start_date and end_date):
        if period == "day":
            start_date = end_date = today
        elif period == "month":
            start_date = today - timedelta(days=30)
            end_date = today
        elif period == "all":
            start_date = end_date = None
        else:
            start_date = today - timedelta(days=7)
            end_date = today

    overall_total_inspections = 0
    overall_inspections_with_defects = 0
    overall_total_defects = 0
    overall_bodies_without_offline_defects = 0

    posts_data = {}

    for key, info in posts_info.items():
        zone = info["zone"]
        post = info["post"]

        vins_with_entries = set()
        vins_with_defects = set()
        vins_without_offline = set()
        defect_counter = Counter()
        controller_counter = Counter()
        controller_vin_pairs = set()
        total_defects = 0

        for history in VINHistory.objects.all():
            vin = history.vin
            zone_data = history.history.get(zone, {})
            post_entries = zone_data.get(post, [])

            for entry in post_entries:
                raw_date = entry.get("date_added")
                if not raw_date:
                    continue

                entry_date = parse_date(raw_date[:10])
                if start_date and entry_date < start_date:
                    continue
                if end_date and entry_date > end_date:
                    continue

                vins_with_entries.add(vin)

                defects = entry.get("defects", [])
                controller = entry.get("controller", "Неизвестно")
                has_offline = False
                filtered = []

                for defect in defects:
                    if defect.get("repair_type") == "offline":
                        has_offline = True

                    detail = defect.get("detail", "Неизвестно")
                    name = defect.get("defect", "Неизвестно")
                    quantity = defect.get("quantity") or 1
                    try:
                        quantity = max(int(quantity), 1)
                    except (TypeError, ValueError):
                        quantity = 1

                    key = f"{detail} — {name}"
                    defect_counter[key] += quantity
                    total_defects += quantity
                    filtered.append(defect)

                if defects:
                    vins_with_defects.add(vin)

                if not has_offline:
                    vins_without_offline.add(vin)

                if filtered and (controller, vin) not in controller_vin_pairs:
                    controller_counter[controller] += 1
                    controller_vin_pairs.add((controller, vin))

        total_inspections = len(vins_with_entries)
        inspections_with_defects = len(vins_with_defects)
        inspections_without_defects = total_inspections - inspections_with_defects
        total_without_offline = len(vins_without_offline)

        dpu = round(total_defects / total_inspections, 2) if total_inspections else 0
        str_percentage = round((total_without_offline / total_inspections) * 100, 2) if total_inspections else 0

        posts_data[key] = {
            "title": info["title"],
            "total_inspections": total_inspections,
            "inspections_with_defects": inspections_with_defects,
            "inspections_without_defects": inspections_without_defects,
            "without_defects": total_inspections - inspections_with_defects,
            "total_defects": total_defects,
            "dpu": dpu,
            "str_percentage": str_percentage,
            "top_defects": defect_counter.most_common(5),
            "top_controllers": controller_counter.most_common(5),
        }

        # Общая статистика
        overall_total_inspections += total_inspections
        overall_inspections_with_defects += inspections_with_defects
        overall_total_defects += total_defects
        overall_bodies_without_offline_defects += total_without_offline

    inspections_without_defects = overall_total_inspections - overall_inspections_with_defects

    if overall_total_inspections == 0:
        overall_dpu = 0
        overall_str_percentage = 0
    else:
        overall_dpu = round(overall_total_defects / overall_total_inspections, 2)
        overall_str_percentage = round((overall_bodies_without_offline_defects / overall_total_inspections) * 100, 2)

    context = {
        "period": period,
        "start_date": start_date_str or '',
        "end_date": end_date_str or '',
        "overall_total_inspections": overall_total_inspections,
        "overall_inspections_with_defects": overall_inspections_with_defects,
        "overall_without_defects": inspections_without_defects,
        "overall_total_defects": overall_total_defects,
        "overall_dpu": overall_dpu,
        "overall_str_percentage": overall_str_percentage,
        "posts_data": posts_data,
    }

    return render(request, "users/incoming/incoming_general_report.html", context)


@login_required
@permission_required('users.access_to_pqa_data', raise_exception=True)
def body_inspection_report(request):
    zone = "Цех поставки"
    post = "Зона первичного осмотра кузовов, DKD"

    period = request.GET.get("period", "week")
    today = now().date()

    start_date_str = request.GET.get("start_date")
    end_date_str = request.GET.get("end_date")

    selected_brands = request.GET.getlist("brand")
    all_brands = TraceData.objects.values_list("brand", flat=True).distinct()

    try:
        custom_start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date() if start_date_str else None
        custom_end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date() if end_date_str else None
    except ValueError:
        custom_start_date = custom_end_date = None

    if custom_start_date and custom_end_date:
        start_date = custom_start_date
        end_date = custom_end_date
    elif period == "day":
        start_date = today
        end_date = today
    elif period == "month":
        start_date = today - timedelta(days=30)
        end_date = today
    elif period == "all":
        start_date = None
        end_date = None
    else:
        start_date = today - timedelta(days=7)
        end_date = today

    selected_grades = request.GET.getlist("grade")
    all_grades = ["V1+", "V1", "V2", "V3"]

    # ✅ Приводим бренды к нижнему регистру и без пробелов для надёжного сравнения
    normalized_selected_brands = [b.strip().lower() for b in selected_brands]

    total_defects = 0
    defect_detail_counter = Counter()
    controller_counter = Counter()
    controller_vin_pairs = set()

    vins_with_entries = set()
    defective_vins = set()
    vins_with_offline_defects = set()

    for history in VINHistory.objects.all():
        vin = history.vin

        # ✅ фильтрация по маркам через TraceData
        if normalized_selected_brands:
            trace = TraceData.objects.filter(vin_rk__iexact=vin).first()
            if not trace or trace.brand.strip().lower() not in normalized_selected_brands:
                continue

        zone_data = history.history.get(zone, {})
        post_entries = zone_data.get(post, [])

        for entry in post_entries:
            raw_date = entry.get("date_added")
            if not raw_date:
                continue

            entry_date = parse_date(raw_date[:10])
            if start_date and entry_date < start_date:
                continue
            if end_date and entry_date > end_date:
                continue

            vins_with_entries.add(vin)

            defects = entry.get("defects", [])
            controller = entry.get("controller", "Неизвестно")
            has_offline_defect = False

            filtered_defects = []

            for defect in defects:
                grade = defect.get("grade", "")
                if not selected_grades or grade in selected_grades:
                    filtered_defects.append(defect)
                    if defect.get("repair_type") == "offline":
                        has_offline_defect = True

            if defects:
                defective_vins.add(vin)

            if has_offline_defect:
                vins_with_offline_defects.add(vin)

            if filtered_defects:
                for defect in filtered_defects:
                    defect_name = defect.get("defect", "Неизвестно")
                    detail_name = defect.get("detail", "Неизвестно")
                    try:
                        quantity = int(defect.get("quantity", 1))
                        quantity = max(quantity, 1)
                    except (TypeError, ValueError):
                        quantity = 1

                    key = f"{detail_name} — {defect_name}"
                    defect_detail_counter[key] += quantity
                    total_defects += quantity

                if controller and vin and (controller, vin) not in controller_vin_pairs:
                    controller_counter[controller] += 1
                    controller_vin_pairs.add((controller, vin))

    total_inspections = len(vins_with_entries)
    inspections_with_defects = len(defective_vins)
    total_with_offline_defects = len(vins_with_offline_defects)

    if total_inspections == 0:
        dpu = 0
        str_percentage = 0
    else:
        dpu = round(total_defects / total_inspections, 2)
        str_percentage = round((total_inspections - total_with_offline_defects) / total_inspections * 100, 2)

    context = {
        "all_grades": all_grades,
        "selected_grades": selected_grades,
        "all_brands": all_brands,
        "selected_brands": selected_brands,  # ✅ передаём оригинальный список для шаблона
        "total_inspections": total_inspections,
        "total_defects": total_defects,
        "dpu": dpu,
        "str_percentage": str_percentage,
        "top_defects": defect_detail_counter.most_common(20),
        "top_controllers": controller_counter.most_common(20),
        "start_date": start_date_str or '',
        "end_date": end_date_str or '',
        "bodies_without_any_defects": total_inspections - inspections_with_defects,
        "inspections_with_defects": inspections_with_defects,
    }

    return render(request, "users/incoming/body_inspection_report.html", context)


@login_required
@permission_required('users.access_to_pqa_data', raise_exception=True)
def body_inspection_table(request):
    zone = "Цех поставки"
    post = "Зона первичного осмотра кузовов, DKD"

    records = []
    for history in VINHistory.objects.all():
        vin = history.vin
        zone_data = history.history.get(zone, {})
        post_entries = zone_data.get(post, [])

        for entry in post_entries:
            defects = entry.get("defects", [])
            raw_date = entry.get("date_added")
            date = parse_datetime(raw_date) if raw_date else None
            controller = entry.get("controller") or entry.get("extra_data", {}).get("controller", "")
            body_photos = entry.get("body_photos", []) or entry.get("extra_data", {}).get("body_photos", [])

            # 🔹 Новое: берём длительность и номер контейнера (с запасным источником из extra_data)
            inspection_time = (
                entry.get("inspection_duration_seconds")
                or entry.get("extra_data", {}).get("inspection_duration_seconds")
            )
            container_number = (
                entry.get("container_number")
                or entry.get("extra_data", {}).get("container_number")
            )

            group_key = f"{vin}_{raw_date}"

            if defects:
                for defect in defects:
                    records.append({
                        "vin": vin,
                        "date": date,
                        "controller": controller,
                        "grade": defect.get("grade"),
                        "defect": defect.get("defect"),
                        "detail": defect.get("detail"),
                        "quantity": defect.get("quantity"),
                        "repair_type": defect.get("repair_type"),
                        "responsible": defect.get("responsible"),
                        "comment": defect.get("custom_detail_explanation"),
                        "extra_comment": defect.get("custom_defect_explanation"),
                        "photos": defect.get("defect_photos", []),
                        "photos_json": json.dumps(defect.get("defect_photos", [])),
                        "body_photos": body_photos,
                        "body_photos_json": json.dumps(body_photos),
                        "has_defect": True,
                        "group_key": group_key,

                        # 🔹 Новые поля в записи
                        "inspection_time": inspection_time,
                        "container_number": container_number,
                    })
            else:
                records.append({
                    "vin": vin,
                    "date": date,
                    "controller": controller,
                    "grade": None,
                    "defect": "Без дефектов",
                    "detail": None,
                    "quantity": None,
                    "repair_type": None,
                    "responsible": None,
                    "comment": None,
                    "extra_comment": None,
                    "photos": [],
                    "photos_json": [],
                    "body_photos": body_photos,
                    "body_photos_json": json.dumps(body_photos),
                    "has_defect": False,
                    "group_key": group_key,

                    # 🔹 Тоже прокидываем сюда
                    "inspection_time": inspection_time,
                    "container_number": container_number,
                })

    records.sort(key=lambda r: r["date"] or timezone.datetime.min, reverse=True)

    # --- Пользователи контроллеров
    all_controllers = {r["controller"] for r in records if r.get("controller")}
    controller_users_qs = CustomUser.objects.filter(username__in=all_controllers)
    controller_users = {user.username: user for user in controller_users_qs}

    # --- TraceData по VIN
    vin_list = list({r["vin"] for r in records if r.get("vin")})
    trace_data = TraceData.objects.filter(vin_rk__in=vin_list)
    vin_trace_map = {
        td.vin_rk: {"brand": td.brand, "model": td.model, "config_code": td.config_code}
        for td in trace_data
    }

    return render(request, "users/incoming/body_table.html", {
        "records": records,
        "controller_users": controller_users,
        "vin_trace_map": vin_trace_map,
    })


@login_required
@permission_required('users.access_to_pqa_data', raise_exception=True)
def body_inspection_form(request):
    """Ручное внесение осмотра — Зона первичного осмотра кузовов (DKD)"""

    post = get_object_or_404(Post, name="Зона первичного осмотра кузовов, DKD")

    details = BodyDetail.objects.filter(posts=post)
    defects = Defect.objects.filter(posts=post)
    defect_grades = DefectGrade.objects.all()
    defect_responsibles = DefectResponsible.objects.filter(posts=post)

    if request.method == "POST":
        form = BodyUnloadingZoneDKDForm(request.POST, request.FILES, post_id=post.id)

        if form.is_valid():
            vin_number = form.cleaned_data.get("vin_number")
            container_number = form.cleaned_data.get("container_number", "").strip()
            manual_date_str = request.POST.get("manual_inspection_date")

            try:
                manual_date = datetime.strptime(manual_date_str, "%Y-%m-%d")
                manual_date_full = datetime.combine(manual_date.date(), time(0, 0))
                date_added = make_aware(manual_date_full).isoformat()
            except Exception:
                date_added = now_almaty_iso()

            history_entry, _ = VINHistory.objects.get_or_create(vin=vin_number)
            zone = post.location
            post_name = post.name

            body_photo_urls = []
            body_photos = form.cleaned_data["body_photos"]
            if body_photos:
                for file in body_photos:
                    compressed = compress_uploaded_image(file)
                    file_path = f"images/body_photos/{compressed.name}"
                    os.makedirs(os.path.join(settings.MEDIA_ROOT, "images/body_photos"), exist_ok=True)
                    with open(os.path.join(settings.MEDIA_ROOT, file_path), "wb+") as dest:
                        for chunk in compressed.chunks():
                            dest.write(chunk)
                    body_photo_urls.append(f"{settings.MEDIA_URL}{file_path}")

            has_defect = request.POST.get("has_defect", "")

            if has_defect == "no":
                inspection_data = {
                    "vin_number": vin_number,
                    "controller": request.user.username,
                    "defects": [],
                    "body_photos": body_photo_urls,
                    "date_added": date_added,
                }
                if container_number:
                    inspection_data["container_number"] = container_number

                try:
                    inspection_data["inspection_duration_seconds"] = int(request.POST.get("inspection_duration_seconds"))
                except (ValueError, TypeError):
                    inspection_data["inspection_duration_seconds"] = None

                history_entry.history.setdefault(zone, {}).setdefault(post_name, []).append(inspection_data)
                history_entry.save()

                messages.success(request, "✅ Осмотр без дефектов успешно сохранён!")
                return redirect("body_inspection_form")

            # Если есть дефекты
            defect_index = 1
            while f"responsible_{defect_index}" in request.POST:
                responsible_obj = DefectResponsible.objects.filter(id=request.POST.get(f"responsible_{defect_index}")).first()
                detail_obj = BodyDetail.objects.filter(id=request.POST.get(f"detail_{defect_index}")).first()
                defect_obj = Defect.objects.filter(id=request.POST.get(f"defect_{defect_index}")).first()
                grade_obj = DefectGrade.objects.filter(id=request.POST.get(f"grade_{defect_index}")).first()

                defect_data = {
                    "responsible": responsible_obj.name if responsible_obj else None,
                    "detail": detail_obj.name if detail_obj else None,
                    "custom_detail_explanation": request.POST.get(f"custom_detail_explanation_{defect_index}"),
                    "defect": defect_obj.name if defect_obj else None,
                    "custom_defect_explanation": request.POST.get(f"custom_defect_explanation_{defect_index}"),
                    "quantity": request.POST.get(f"quantity_{defect_index}", 1),
                    "grade": grade_obj.name if grade_obj else None,
                    "repair_type": request.POST.get(f"repair_type_{defect_index}"),
                    "defect_photos": [],
                }

                defect_photo_urls = []
                for file in request.FILES.getlist(f"defect_photos_{defect_index}"):
                    compressed = compress_uploaded_image(file)
                    file_path = f"images/defects/{compressed.name}"
                    os.makedirs(os.path.join(settings.MEDIA_ROOT, "images/defects"), exist_ok=True)
                    with open(os.path.join(settings.MEDIA_ROOT, file_path), "wb+") as dest:
                        for chunk in compressed.chunks():
                            dest.write(chunk)
                    defect_photo_urls.append(f"{settings.MEDIA_URL}{file_path}")

                defect_data["defect_photos"] = defect_photo_urls

                inspection_data = {
                    "vin_number": vin_number,
                    "controller": request.user.username,
                    "defects": [defect_data],
                    "body_photos": body_photo_urls,
                    "date_added": date_added,
                }

                if container_number:
                    inspection_data["container_number"] = container_number

                try:
                    inspection_data["inspection_duration_seconds"] = int(request.POST.get("inspection_duration_seconds"))
                except (ValueError, TypeError):
                    inspection_data["inspection_duration_seconds"] = None

                history_entry.history.setdefault(zone, {}).setdefault(post_name, []).append(inspection_data)
                defect_index += 1

            history_entry.save()
            messages.success(request, "✅ Инспекция с дефектами успешно сохранена!")
            return redirect("body_inspection_form")

        else:
            messages.error(request, "❌ Ошибка: Проверьте заполненные данные!")

    else:
        form = BodyUnloadingZoneDKDForm(post_id=post.id)

    return render(request, "users/incoming/body_inspection_form.html", {
        "form": form,
        "details": details,
        "defects": defects,
        "defect_grades": defect_grades,
        "defect_responsibles": defect_responsibles,
        "post": post,
    })


@login_required
@permission_required('users.access_to_pqa_data', raise_exception=True)
def body_inspection_export(request):
    return render(request, "users/incoming/body_inspection_export.html")


@login_required
@permission_required('users.access_to_pqa_data', raise_exception=True)
def parts_acceptance_report(request):
    zone = "Цех поставки"
    post = "Зона выгрузки комплектующих, DKD"

    # Фильтр по дате
    period = request.GET.get("period", "week")
    today = now().date()

    start_date_str = request.GET.get("start_date")
    end_date_str = request.GET.get("end_date")

    try:
        custom_start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date() if start_date_str else None
        custom_end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date() if end_date_str else None
    except ValueError:
        custom_start_date = custom_end_date = None

    if custom_start_date and custom_end_date:
        start_date = custom_start_date
        end_date = custom_end_date
    elif period == "day":
        start_date = today
        end_date = today
    elif period == "month":
        start_date = today - timedelta(days=30)
        end_date = today
    elif period == "all":
        start_date = None
        end_date = None
    else:
        start_date = today - timedelta(days=7)
        end_date = today

    selected_grades = request.GET.getlist("grade")
    all_grades = ["V1+", "V1", "V2", "V3"]

    total_defects = 0
    defect_detail_counter = Counter()
    controller_counter = Counter()
    controller_vin_pairs = set()

    vins_with_entries = set()
    defective_vins = set()
    vins_with_offline_defects = set()

    for history in VINHistory.objects.all():
        vin = history.vin
        zone_data = history.history.get(zone, {})
        post_entries = zone_data.get(post, [])

        for entry in post_entries:
            raw_date = entry.get("date_added")
            if not raw_date:
                continue

            entry_date = parse_date(raw_date[:10])
            if start_date and entry_date < start_date:
                continue
            if end_date and entry_date > end_date:
                continue

            vins_with_entries.add(vin)

            defects = entry.get("defects", [])
            controller = entry.get("controller", "Неизвестно")
            has_offline_defect = False

            filtered_defects = []

            for defect in defects:
                grade = defect.get("grade", "")
                if not selected_grades or grade in selected_grades:
                    filtered_defects.append(defect)
                    if defect.get("repair_type") == "offline":
                        has_offline_defect = True

            if defects:
                defective_vins.add(vin)

            if has_offline_defect:
                vins_with_offline_defects.add(vin)

            if filtered_defects:
                for defect in filtered_defects:
                    defect_name = defect.get("defect", "Неизвестно")
                    detail_name = defect.get("detail", "Неизвестно")
                    try:
                        quantity = int(defect.get("quantity", 1))
                        quantity = max(quantity, 1)
                    except (TypeError, ValueError):
                        quantity = 1

                    key = f"{detail_name} — {defect_name}"
                    defect_detail_counter[key] += quantity
                    total_defects += quantity

                if controller and vin and (controller, vin) not in controller_vin_pairs:
                    controller_counter[controller] += 1
                    controller_vin_pairs.add((controller, vin))

    total_inspections = len(vins_with_entries)
    inspections_with_defects = len(defective_vins)
    total_with_offline_defects = len(vins_with_offline_defects)

    if total_inspections == 0:
        dpu = 0
        str_percentage = 0
    else:
        dpu = round(total_defects / total_inspections, 2)
        str_percentage = round((total_inspections - total_with_offline_defects) / total_inspections * 100, 2)

    context = {
        "all_grades": all_grades,
        "selected_grades": selected_grades,
        "total_inspections": total_inspections,
        "total_defects": total_defects,
        "dpu": dpu,
        "str_percentage": str_percentage,
        "top_defects": defect_detail_counter.most_common(20),
        "top_controllers": controller_counter.most_common(20),
        "start_date": start_date_str or '',
        "end_date": end_date_str or '',
        "bodies_without_any_defects": total_inspections - inspections_with_defects,
        "inspections_with_defects": inspections_with_defects,
    }

    return render(request, "users/incoming/parts_acceptance_report.html", context)


@login_required
@permission_required('users.access_to_pqa_data', raise_exception=True)
def parts_acceptance_table(request):
    zone = "Цех поставки"
    post = "Зона выгрузки комплектующих, DKD"

    records = []

    for history in VINHistory.objects.all():
        vin = history.vin
        post_entries = history.history.get(zone, {}).get(post, [])

        for entry in post_entries:
            raw_date = entry.get("date_added")
            date = parse_datetime(raw_date) if raw_date else None
            controller = entry.get("controller", "") or entry.get("extra_data", {}).get("controller", "")
            engine_number = entry.get("engine_number", "")
            engine_photo = entry.get("engine_photo", "")
            component_photos = entry.get("component_photos", [])
            inspection_time = entry.get("inspection_duration_seconds", None)
            defects = entry.get("defects", [])
            group_key = f"{vin}_{raw_date}"

            if defects:
                for defect in defects:
                    records.append({
                        "vin": vin,
                        "date": date,
                        "controller": controller,
                        "engine_number": engine_number,
                        "engine_photo": engine_photo,
                        "component_photos": component_photos,
                        "grade": defect.get("grade"),
                        "defect": defect.get("defect"),
                        "detail": defect.get("detail"),
                        "quantity": defect.get("quantity"),
                        "repair_type": defect.get("repair_type"),
                        "responsible": defect.get("responsible"),
                        "comment": defect.get("custom_defect_explanation"),
                        "extra_comment": defect.get("custom_detail_explanation"),
                        "photos": defect.get("defect_photos", []),
                        "has_defect": True,
                        "group_key": group_key,
                        "inspection_time": inspection_time,
                    })
            else:
                records.append({
                    "vin": vin,
                    "date": date,
                    "controller": controller,
                    "engine_number": engine_number,
                    "engine_photo": engine_photo,
                    "component_photos": component_photos,
                    "grade": None,
                    "defect": "Без дефектов",
                    "detail": None,
                    "quantity": None,
                    "repair_type": None,
                    "responsible": None,
                    "comment": None,
                    "extra_comment": None,
                    "photos": [],
                    "has_defect": False,
                    "group_key": group_key,
                    "inspection_time": inspection_time,
                })

    records.sort(key=lambda r: r["date"] or timezone.datetime.min, reverse=True)

    return render(request, "users/incoming/parts_acceptance_table.html", {"records": records})


@login_required
@permission_required('users.access_to_pqa_data', raise_exception=True)
def parts_acceptance_export(request):
    return render(request, "users/incoming/part_acceptance_export.html")


@login_required
@permission_required('users.access_to_pqa_data', raise_exception=True)
def parts_acceptance_form(request):
    return render(request, "users/incoming/forms/parts_acceptance_form.html")


@login_required
@permission_required('users.access_to_pqa_data', raise_exception=True)
def final_acceptance_report(request):
    zone = "Цех поставки"
    post = "Зона основной приемки DKD"

    period = request.GET.get("period", "week")
    today = now().date()


    start_date_str = request.GET.get("start_date")
    end_date_str = request.GET.get("end_date")

    selected_brands = request.GET.getlist("brand")
    all_brands = TraceData.objects.values_list("brand", flat=True).distinct()

    try:
        custom_start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date() if start_date_str else None
        custom_end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date() if end_date_str else None
    except ValueError:
        custom_start_date = custom_end_date = None

    if custom_start_date and custom_end_date:
        start_date = custom_start_date
        end_date = custom_end_date
    elif period == "day":
        start_date = today
        end_date = today
    elif period == "month":
        start_date = today - timedelta(days=30)
        end_date = today
    elif period == "all":
        start_date = None
        end_date = None
    else:
        start_date = today - timedelta(days=7)
        end_date = today

    selected_grades = request.GET.getlist("grade")
    all_grades = ["V1+", "V1", "V2", "V3"]
    normalized_selected_brands = [b.strip().lower() for b in selected_brands]

    total_defects = 0
    defect_detail_counter = Counter()
    controller_counter = Counter()
    controller_vin_pairs = set()

    vins_with_entries = set()
    defective_vins = set()
    vins_with_offline_defects = set()

    for history in VINHistory.objects.all():
        zone_data = history.history.get(zone, {})
        post_entries = zone_data.get(post, [])

        for entry in post_entries:
            raw_date = entry.get("date_added")
            if not raw_date:
                continue

            entry_date = parse_date(raw_date[:10])
            if start_date and entry_date < start_date:
                continue
            if end_date and entry_date > end_date:
                continue

            vin_entry = entry.get("vin_number") or history.vin
            vins_with_entries.add(vin_entry)

            # ✅ фильтрация по бренду
            if normalized_selected_brands:
                trace = TraceData.objects.filter(vin_rk__iexact=vin_entry).first()
                if not trace or trace.brand.strip().lower() not in normalized_selected_brands:
                    continue

            defects = entry.get("defects", [])
            controller = entry.get("controller", "Неизвестно")
            has_offline_defect = False
            filtered_defects = []

            for defect in defects:
                grade = defect.get("grade", "")
                if not selected_grades or grade in selected_grades:
                    filtered_defects.append(defect)
                    if defect.get("repair_type") == "offline":
                        has_offline_defect = True

            if defects:
                defective_vins.add(vin_entry)

            if has_offline_defect:
                vins_with_offline_defects.add(vin_entry)

            if filtered_defects:
                for defect in filtered_defects:
                    defect_name = defect.get("defect", "Неизвестно")
                    detail_name = defect.get("detail", "Неизвестно")
                    try:
                        quantity = int(defect.get("quantity", 1))
                        quantity = max(quantity, 1)
                    except (TypeError, ValueError):
                        quantity = 1

                    key = f"{detail_name} — {defect_name}"
                    defect_detail_counter[key] += quantity
                    total_defects += quantity

                if controller and vin_entry and (controller, vin_entry) not in controller_vin_pairs:
                    controller_counter[controller] += 1
                    controller_vin_pairs.add((controller, vin_entry))

    total_inspections = len(vins_with_entries)
    inspections_with_defects = len(defective_vins)
    total_with_offline_defects = len(vins_with_offline_defects)

    if total_inspections == 0:
        dpu = 0
        str_percentage = 0
    else:
        dpu = round(total_defects / total_inspections, 2)
        str_percentage = round((total_inspections - total_with_offline_defects) / total_inspections * 100, 2)

    context = {
        "all_grades": all_grades,
        "selected_grades": selected_grades,
        "all_brands": all_brands,
        "selected_brands": selected_brands,
        "total_inspections": total_inspections,
        "total_defects": total_defects,
        "dpu": dpu,
        "str_percentage": str_percentage,
        "top_defects": defect_detail_counter.most_common(20),
        "top_controllers": controller_counter.most_common(20),
        "start_date": start_date_str or '',
        "end_date": end_date_str or '',
        "bodies_without_any_defects": total_inspections - inspections_with_defects,
        "inspections_with_defects": inspections_with_defects,
    }

    return render(request, "users/incoming/final_acceptance_report.html", context)


@login_required
@permission_required('users.access_to_pqa_data', raise_exception=True)
def final_acceptance_table(request):
    zone = "Цех поставки"
    post = "Зона основной приемки DKD"

    records = []

    for history in VINHistory.objects.all():
        vin = history.vin
        post_entries = history.history.get(zone, {}).get(post, [])

        for entry in post_entries:
            raw_date = entry.get("date_added")
            date = parse_datetime(raw_date) if raw_date else None
            controller = entry.get("controller", "") or entry.get("extra_data", {}).get("controller", "")
            body_photos = entry.get("body_photos", []) or entry.get("extra_data", {}).get("body_photos", [])
            inspection_time = entry.get("inspection_duration_seconds", None)
            defects = entry.get("defects", [])
            group_key = f"{vin}_{raw_date}"

            if defects:
                for defect in defects:
                    records.append({
                        "vin": vin,
                        "date": date,
                        "controller": controller,
                        "grade": defect.get("grade"),
                        "defect": defect.get("defect"),
                        "detail": defect.get("detail"),
                        "quantity": defect.get("quantity"),
                        "repair_type": defect.get("repair_type"),
                        "responsible": defect.get("responsible"),
                        "comment": defect.get("custom_detail_explanation"),
                        "extra_comment": defect.get("custom_defect_explanation"),
                        "photos": defect.get("defect_photos", []),
                        "body_photos": body_photos,
                        "inspection_time": inspection_time,
                        "has_defect": True,
                        "group_key": group_key,
                    })
            else:
                records.append({
                    "vin": vin,
                    "date": date,
                    "controller": controller,
                    "grade": None,
                    "defect": "Без дефектов",
                    "detail": None,
                    "quantity": None,
                    "repair_type": None,
                    "responsible": None,
                    "comment": None,
                    "extra_comment": None,
                    "photos": [],
                    "body_photos": body_photos,
                    "inspection_time": inspection_time,
                    "has_defect": False,
                    "group_key": group_key,
                })

    records.sort(key=lambda r: r["date"] or timezone.datetime.min, reverse=True)

    return render(request, "users/incoming/final_acceptance_table.html", {"records": records})


@login_required
@permission_required('users.access_to_pqa_data', raise_exception=True)
def final_acceptance_export(request):
    return render(request, "users/incoming/final_acceptance_export.html")


@login_required
@permission_required('users.access_to_pqa_data', raise_exception=True)
def final_acceptance_form(request):
    """Ручное внесение инспекции — Зона основной приёмки DKD"""

    post = get_object_or_404(Post, name="Зона основной приемки DKD")

    details = BodyDetail.objects.all()
    defects = Defect.objects.filter(posts=post)
    defect_grades = DefectGrade.objects.all()
    defect_responsibles = DefectResponsible.objects.filter(posts=post)

    vin_history = None

    if request.method == "POST":
        form = MainUnloadingZoneDKDForm(request.POST, request.FILES, post_id=post.id)

        if form.is_valid():
            vin_number = form.cleaned_data.get("vin_number")
            if not vin_number:
                messages.error(request, "❌ VIN номер обязателен.")
                return redirect("final_acceptance_form")

            history_entry, _ = VINHistory.objects.get_or_create(vin=vin_number)

            zone = post.location
            post_name = post.name

            # 📌 Дата, введённая вручную
            manual_date_str = request.POST.get("manual_inspection_date")
            try:
                manual_date = datetime.strptime(manual_date_str, "%Y-%m-%d")
                manual_date_aware = make_aware(manual_date)
                date_added = manual_date_aware.isoformat()
            except Exception:
                date_added = now_almaty_iso()

            # ✅ Фото кузова
            body_photo_urls = []
            for file in form.cleaned_data["body_photos"]:
                compressed = compress_uploaded_image(file)
                file_path = f"images/body_photos/{compressed.name}"
                os.makedirs(os.path.join(settings.MEDIA_ROOT, "images/body_photos"), exist_ok=True)
                with open(os.path.join(settings.MEDIA_ROOT, file_path), "wb+") as dest:
                    for chunk in compressed.chunks():
                        dest.write(chunk)
                body_photo_urls.append(f"{settings.MEDIA_URL}{file_path}")

            # 🔁 Дефекты
            inspection_data = {
                "vin_number": vin_number,
                "controller": request.user.username,
                "body_photos": body_photo_urls,
                "date_added": date_added,
                "defects": []
            }

            defect_index = 1
            while f"responsible_{defect_index}" in request.POST:
                responsible_obj = DefectResponsible.objects.filter(id=request.POST.get(f"responsible_{defect_index}")).first()
                detail_obj = BodyDetail.objects.filter(id=request.POST.get(f"detail_{defect_index}")).first()
                defect_obj = Defect.objects.filter(id=request.POST.get(f"defect_{defect_index}")).first()
                grade_obj = DefectGrade.objects.filter(id=request.POST.get(f"grade_{defect_index}")).first()

                defect_data = {
                    "responsible": responsible_obj.name if responsible_obj else None,
                    "detail": detail_obj.name if detail_obj else None,
                    "custom_detail_explanation": request.POST.get(f"custom_detail_explanation_{defect_index}"),
                    "defect": defect_obj.name if defect_obj else None,
                    "custom_defect_explanation": request.POST.get(f"custom_defect_explanation_{defect_index}"),
                    "quantity": request.POST.get(f"quantity_{defect_index}", 1),
                    "grade": grade_obj.name if grade_obj else None,
                    "repair_type": request.POST.get(f"repair_type_{defect_index}"),
                    "defect_photos": [],
                }

                for file in request.FILES.getlist(f"defect_photos_{defect_index}"):
                    compressed = compress_uploaded_image(file)
                    file_path = f"images/defects/{compressed.name}"
                    os.makedirs(os.path.join(settings.MEDIA_ROOT, "images/defects"), exist_ok=True)
                    with open(os.path.join(settings.MEDIA_ROOT, file_path), "wb+") as dest:
                        for chunk in compressed.chunks():
                            dest.write(chunk)
                    defect_data["defect_photos"].append(f"{settings.MEDIA_URL}{file_path}")

                inspection_data["defects"].append(defect_data)
                defect_index += 1

            # ⏺️ Сохраняем в историю
            history_entry.history.setdefault(zone, {}).setdefault(post_name, []).append(inspection_data)
            history_entry.save()

            messages.success(request, "✅ Инспекция сохранена задним числом!")
            return redirect("final_acceptance_form")
        else:
            messages.error(request, "❌ Ошибка в форме!")

    else:
        form = MainUnloadingZoneDKDForm(post_id=post.id)

    return render(request, "users/incoming/final_acceptance_form.html", {
        "form": form,
        "details": details,
        "defects": defects,
        "defect_grades": defect_grades,
        "defect_responsibles": defect_responsibles,
        "post": post,
        "vin_history": vin_history,
    })



@login_required
@permission_required('users.access_to_pqa_data', raise_exception=True)
def containers_acceptance_table(request):
    """
    Таблица для поста 'Приемка контейнеров' (Цех поставки).
    Поддерживаемый формат записи (пример):
    {
        "photos": [...],
        "controller": "BK-027",
        "date_added": "2025-09-09T08:48:26.557442+05:00",
        "has_defect": "no",
        "description": "",
        "container_number": "9367144"
    }
    """
    import json

    zone = "Цех поставки"
    post = "Приемка контейнеров"

    records = []

    for history in ContainerHistory.objects.all():
        # vin = history.vin  # 1) Remove this line
        post_entries = (history.history or {}).get(zone, {}).get(post, []) or []

        for entry in post_entries:
            raw_date = entry.get("date_added") or entry.get("date")
            date = parse_datetime(raw_date) if raw_date else None

            controller = entry.get("controller", "") or entry.get("extra_data", {}).get("controller", "")
            container_number = (
                entry.get("container_number", "")
                or entry.get("extra_data", {}).get("container_number", "")
            )

            photos = entry.get("photos", []) or []
            description = entry.get("description", "") or ""

            # Приводим has_defect к булеву виду
            has_defect_val = str(entry.get("has_defect", "")).lower()
            has_defect = has_defect_val in ("yes", "true", "1")

            # 2) Change group_key assignment to use container_number
            group_key = f"{container_number}_{raw_date}"

            # 3,4) Build record dict without "vin", and with "description": description
            records.append({
                "date": date,
                "controller": controller,
                "container_number": container_number,
                "description": description,
                "photos": photos,
                "photos_json": json.dumps(photos),
                "has_defect": has_defect,
                "group_key": group_key,
            })

    # Сортировка по дате убыванию
    records.sort(key=lambda r: r["date"] or timezone.datetime.min, reverse=True)

    # ---- Дополняем контекст: контроллеры и TraceData ----
    all_controllers = {r["controller"] for r in records if r.get("controller")}
    controller_users_qs = CustomUser.objects.filter(username__in=all_controllers)
    controller_users = {user.username: user for user in controller_users_qs}

    # 5) Remove vin_list, trace_data, vin_trace_map construction; instead set vin_trace_map = {}
    vin_trace_map = {}

    # 6) Keep vin_trace_map in context for template compatibility
    return render(request, "users/incoming/containers_acceptance_table.html", {
        "records": records,
        "controller_users": controller_users,
        "vin_trace_map": vin_trace_map,
    })


@login_required
@permission_required('users.access_to_pqa_data', raise_exception=True)
def containers_acceptance_export(request):
    return render(request, "users/incoming/containers_acceptance_export.html")







@login_required
@permission_required('users.access_to_vqa_data', raise_exception=True)
def assembly_workshop_dashboard(request):
    today_str = now().date().isoformat()
    zone = "Цех сборки"

    post_vin_set = defaultdict(set)
    post_defect_counts = defaultdict(int)

    for history in VINHistory.objects.all():
        vin = history.vin
        posts = history.history.get(zone, {})
        for post_name, entries in posts.items():
            for entry in entries:
                date_str = entry.get("date_added", "")[:10]
                if date_str != today_str:
                    continue
                post_vin_set[post_name].add(vin)

                defects = entry.get("defects")
                if isinstance(defects, list):
                    post_defect_counts[post_name] += len(defects)
                elif entry.get("has_defect") == "yes":
                    post_defect_counts[post_name] += 1

    # Маппинг: пост → участок
    POST_AREA_MAPPING = {
        "Пост момента затяжек": "Текущий контроль",
        "Chassis": "Текущий контроль",
        "Финал текущий контроль": "Текущий контроль",


        "Зазоры и перепады": "Первая инспекция",
        "Экстерьер": "Первая инспекция",
        "Интерьер": "Первая инспекция",
        "Багажник": "Первая инспекция",
        "Мотор": "Первая инспекция",
        "Функцонал": "Первая инспекция",

        "Геометрия колес": "Тестовая линия",
        "Регулировка света фар и калибровка руля": "Тестовая линия",
        "Тормозная система": "Тестовая линия",
        "Underbody": "Тестовая линия",
        "ADAS": "Тестовая линия",
        "AVM": "Тестовая линия",
        "Герметичность кузова": "Тестовая линия",

        "Диагностика": "Финал + Тест трек",
        "Тест трек": "Финал + Тест трек",
        "Документация": "Финал + Тест трек",
    }

    grouped_posts = defaultdict(list)

    for post_name, area in POST_AREA_MAPPING.items():
        encoded_post = urllib.parse.quote(post_name)

        # только для "Документация" ведём на отдельный view в приложении assembly
        if post_name == "Документация":
            table_url = reverse("assembly:documentation_table")
        else:
            table_url = reverse("assembly_post_table") + f"?post={encoded_post}"

        grouped_posts[area].append({
            "title": post_name,
            "icon": "",
            "table_url": table_url,
            "report_url": reverse("assembly_post_report") + f"?post={encoded_post}",
            "export_url": reverse("assembly_post_export") + f"?post={encoded_post}",
            "manual_entry_url": reverse("assembly_post_manual_entry") + f"?post={encoded_post}",
            "today_vins": len(post_vin_set.get(post_name, set())),
            "today_entries": post_defect_counts.get(post_name, 0),
        })

    return render(request, "users/assembly/assembly_workshop_dashboard.html", {
        "grouped_posts": dict(grouped_posts),
        "enable_particles": True,
        "enable_background": True,
        "enable_white_square": False,
    })


from datetime import datetime
import pytz

@login_required
@permission_required('users.access_to_vqa_data', raise_exception=True)
def assembly_post_manual_entry(request):
    zone = "Цех сборки"
    post_name = request.GET.get("post")

    if not post_name or post_name == "Текущий контроль":
        return HttpResponse("❌ Укажите корректный пост", status=400)

    if request.method == "POST":
        form = AssemblyTemplateForm(request.POST, request.FILES)

        if form.is_valid():
            vin_number = form.cleaned_data.get("vin_number")
            defect_description = form.cleaned_data.get("defect_description", "").strip()
            defect_photos = request.FILES.getlist("defect_photos")
            has_defect = form.cleaned_data.get("has_defect", "")
            custom_date = request.POST.get("custom_date")
            custom_time = request.POST.get("custom_time")

            if not vin_number or not custom_date or not custom_time:
                messages.error(request, "❌ VIN, дата и время обязательны.")
                return redirect(request.path + f"?post={post_name}")

            # 👇 Объединяем дату и время с часовым поясом
            try:
                combined_dt = datetime.strptime(f"{custom_date} {custom_time}", "%Y-%m-%d %H:%M:%S")
                localized_dt = pytz.timezone("Asia/Almaty").localize(combined_dt)
                date_added = localized_dt.isoformat()
            except Exception:
                messages.error(request, "❌ Неверный формат даты или времени.")
                return redirect(request.path + f"?post={post_name}")

            history_entry, _ = VINHistory.objects.get_or_create(vin=vin_number)

            defect_photo_urls = []
            if defect_photos:
                for file in defect_photos:
                    compressed = compress_uploaded_image(file)
                    path = f"images/defects/{compressed.name}"
                    full_path = os.path.join(settings.MEDIA_ROOT, path)
                    os.makedirs(os.path.dirname(full_path), exist_ok=True)
                    with open(full_path, "wb+") as destination:
                        for chunk in compressed.chunks():
                            destination.write(chunk)
                    defect_photo_urls.append(f"{settings.MEDIA_URL}{path}")

            inspection_data = {
                "vin_number": vin_number,
                "controller": request.user.username,
                "defect_description": defect_description,
                "defect_photos": defect_photo_urls,
                "has_defect": has_defect,
                "date_added": date_added,
            }

            if zone not in history_entry.history:
                history_entry.history[zone] = {}
            if post_name not in history_entry.history[zone]:
                history_entry.history[zone][post_name] = []

            history_entry.history[zone][post_name].append(inspection_data)
            history_entry.save()

            messages.success(request, "✅ Данные успешно сохранены!")
            return redirect(request.path + f"?post={post_name}")
        else:
            messages.error(request, "❌ Ошибка валидации данных.")
    else:
        form = AssemblyTemplateForm()

    return render(request, "users/assembly/manual_entry_form.html", {
        "form": form,
        "post_name": post_name,
    })





@login_required
@permission_required('users.access_to_vqa_data', raise_exception=True)
def assembly_post_report(request):
    zone = "Цех сборки"
    post_name = request.GET.get("post")

    if not post_name or post_name == "Текущий контроль":
        return HttpResponse("❌ Укажите корректный пост", status=400)

    today = now().date()

    start_date_str = request.GET.get("start_date")
    end_date_str = request.GET.get("end_date")

    selected_brands = request.GET.getlist("brand")
    selected_models = request.GET.getlist("model")

    try:
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date() if start_date_str else today - timedelta(days=7)
        end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date() if end_date_str else today
    except ValueError:
        start_date = today - timedelta(days=7)
        end_date = today

    all_brands = TraceData.objects.values_list("brand", flat=True).distinct()
    if selected_brands:
        all_models = TraceData.objects.filter(brand__in=selected_brands).values_list("model", flat=True).distinct()
    else:
        all_models = TraceData.objects.values_list("model", flat=True).distinct()

    filtered_vins = TraceData.objects.all()
    if selected_brands:
        filtered_vins = filtered_vins.filter(brand__in=selected_brands)
    if selected_models:
        filtered_vins = filtered_vins.filter(model__in=selected_models)

    vin_set = set(filtered_vins.values_list("vin_rk", flat=True))

    passed_vins = set()
    passed_with_defect = set()
    passed_without_defect = set()
    controller_counter = Counter()
    controller_vin_pairs = set()
    defect_counter = Counter()
    total_defects = 0

    histories = VINHistory.objects.all()
    if vin_set:
        histories = histories.filter(vin__in=vin_set)

    for history in histories:
        vin = history.vin
        post_entries = history.history.get(zone, {}).get(post_name, [])
        if not post_entries:
            continue

        entries_in_range = []
        for entry in post_entries:
            raw_date = entry.get("date_added")
            if not raw_date:
                continue
            entry_date = parse_date(raw_date[:10])
            if entry_date and start_date <= entry_date <= end_date:
                entries_in_range.append(entry)

        if not entries_in_range:
            continue

        passed_vins.add(vin)

        vin_has_defect = any(
            entry.get("has_defect") == "yes" or entry.get("defects")
            for entry in entries_in_range
        )

        if vin_has_defect:
            passed_with_defect.add(vin)
            for entry in entries_in_range:
                if entry.get("has_defect") == "yes" or entry.get("defects"):
                    controller = entry.get("controller", "Неизвестно")
                    if (controller, vin) not in controller_vin_pairs:
                        controller_counter[controller] += 1
                        controller_vin_pairs.add((controller, vin))

                    for defect in entry.get("defects", []):
                        name = defect.get("name", "?")
                        unit = defect.get("unit", "?")
                        combined = f"{name} ({unit})" if name and unit else name or unit or "Неизвестно"
                        defect_counter[combined] += 1
                        total_defects += 1
        else:
            passed_without_defect.add(vin)

    total_inspections = len(passed_vins)
    inspections_with_defects = len(passed_with_defect)
    inspections_without_defects = total_inspections - inspections_with_defects
    dpu = round(total_defects / total_inspections, 2) if total_inspections else 0
    str_percentage = round(inspections_without_defects / total_inspections * 100, 2) if total_inspections else 0

    context = {
        "post_name": post_name,
        "start_date": start_date_str or '',
        "end_date": end_date_str or '',
        "total_inspections": total_inspections,
        "inspections_with_defects": inspections_with_defects,
        "inspections_without_defects": inspections_without_defects,
        "total_defects": total_defects,
        "dpu": dpu,
        "str_percentage": str_percentage,
        "unique_vin_count": total_inspections,
        "top_controllers": controller_counter.most_common(20),
        "top_defects": defect_counter.most_common(20),
        "all_brands": all_brands,
        "selected_brands": selected_brands,
        "all_models": all_models,
        "selected_models": selected_models,
        "enable_particles": True,
        "enable_background": True,
        "enable_white_square": False,
    }

    return render(request, "users/assembly/assembly_post_report.html", context)






@login_required
@permission_required('users.access_to_vqa_data', raise_exception=True)
def assembly_current_report(request):
    # Заглушка для отчета
    return render(request, "assembly/assembly/torque_graph_dkd.html")


@login_required
@permission_required('users.access_to_vqa_data', raise_exception=True)
def assembly_current_table(request):
    zone = "Цех сборки"
    post = "Пост момента затяжки, DKD"

    records = []

    for history in VINHistory.objects.all():
        vin = history.vin
        post_entries = history.history.get(zone, {}).get(post, [])

        for entry in post_entries:
            raw_date = entry.get("date_added")
            date = parse_datetime(raw_date) if raw_date else None

            controller = entry.get("controller", "")
            mod = entry.get("modification", "")
            part = entry.get("assembly_part", "")
            torques = entry.get("torque_values", [])
            defects = entry.get("defects", [])
            group_key = f"{vin}_{raw_date}"

            if defects:
                for defect in defects:
                    records.append({
                        "vin": vin,
                        "date": date,
                        "controller": controller,
                        "modification": mod,
                        "assembly_part": part,
                        "torque_values": torques,
                        "defect_type": defect.get("type"),
                        "photos": defect.get("photos", []),
                        "quantity": defect.get("quantity", ""),
                        "has_defect": True,
                        "group_key": group_key,
                    })
            else:
                records.append({
                    "vin": vin,
                    "date": date,
                    "controller": controller,
                    "modification": mod,
                    "assembly_part": part,
                    "torque_values": torques,
                    "defect_type": "Без дефектов",
                    "photos": [],
                    "quantity": "",
                    "has_defect": False,
                    "group_key": group_key,
                })

    # Сортировка по дате убыванию
    records.sort(key=lambda r: r["date"] or timezone.datetime.min, reverse=True)

    return render(request, "users/assembly/assembly_current_table.html", {"records": records})


@login_required
@permission_required('users.access_to_vqa_data', raise_exception=True)
def assembly_current_export(request):
    # Заглушка для экспорта Excel
    return render(request, "users/assembly/assembly_current_export.html")












@login_required
@permission_required('users.access_to_vqa_data', raise_exception=True)
def assembly_post_table(request):
    zone = "Цех сборки"
    post = request.GET.get("post")

    valid_posts = [
        "Пост момента затяжек", "Chassis", "Финал текущий контроль", "Зазоры и перепады", "Экстерьер", "Интерьер", "Багажник",
        "Мотор", "Функцонал", "Геометрия колес", "Регулировка света фар и калибровка руля",
        "Тормозная система", "Underbody", "ADAS", "AVM", "Герметичность кузова", "Диагностика", "Тест трек", "Документация",
    ]

    if not post or post not in valid_posts:
        return HttpResponse("❌ Укажите корректный пост", status=400)

    records_grouped = defaultdict(list)

    for history in VINHistory.objects.all():
        vin = history.vin
        post_entries = history.history.get(zone, {}).get(post, [])

        for entry in post_entries:
            raw_date = entry.get("date_added")
            date = parse_datetime(raw_date) if raw_date else None
            controller = entry.get("controller", "") or entry.get("extra_data", {}).get("controller", "")
            group_key = f"{vin}_{raw_date}"
            line = entry.get("line", "")
            duration = entry.get("inspection_duration_seconds", "")

            # Новый формат: список дефектов
            if "defects" in entry and entry["defects"]:
                for defect in entry["defects"]:
                    records_grouped[vin].append({
                        "date": date,
                        "controller": controller,
                        "defect_description": defect.get("name", "Без описания"),
                        "photos": defect.get("photos", []),
                        "photos_json": json.dumps(defect.get("photos", [])),
                        "has_defect": True,
                        "group_key": group_key,
                        "unit": defect.get("unit", ""),
                        "grade": defect.get("grade", ""),
                        "comment": defect.get("comment", ""),
                        "responsible": defect.get("responsible", ""),
                        "line": line,
                        "duration": duration,
                    })

            # Старый формат: один дефект
            elif "defect_description" in entry or "defect_photos" in entry:
                records_grouped[vin].append({
                    "date": date,
                    "controller": controller,
                    "defect_description": entry.get("defect_description", "Без описания"),
                    "photos": entry.get("defect_photos", []),
                    "photos_json": json.dumps(entry.get("defect_photos", [])),
                    "has_defect": True,
                    "group_key": group_key,
                    "unit": "",
                    "grade": "",
                    "comment": "",
                    "responsible": "",
                    "line": line,
                    "duration": duration,
                })

            # Без дефектов
            else:
                records_grouped[vin].append({
                    "date": date,
                    "controller": controller,
                    "defect_description": "Без дефектов",
                    "photos": [],
                    "photos_json": "[]",
                    "has_defect": False,
                    "group_key": group_key,
                    "unit": "",
                    "grade": "",
                    "comment": "",
                    "responsible": "",
                    "line": line,
                    "duration": duration,
                })

    # Сортировка внутри VIN
    for vin in records_grouped:
        records_grouped[vin].sort(key=lambda r: r["date"] or timezone.datetime.min, reverse=True)

    # Сортировка VIN по дате первого дефекта
    records_grouped_sorted = sorted(
        records_grouped.items(),
        key=lambda item: item[1][0]["date"] or timezone.datetime.min,
        reverse=True
    )

    # Собираем всех уникальных логинов
    all_controllers = {
        entry["controller"]
        for entries in records_grouped.values()
        for entry in entries
        if entry["controller"]
    }

    # Получаем пользователей одним запросом
    controller_users_qs = CustomUser.objects.filter(username__in=all_controllers)
    controller_users = {user.username: user for user in controller_users_qs}

    # Получаем VIN'ы
    vin_list = list(records_grouped.keys())

    # Получаем данные из TraceData по VIN RK
    trace_data = TraceData.objects.filter(vin_rk__in=vin_list)
    vin_trace_map = {
        td.vin_rk: {
            "brand": td.brand,
            "model": td.model,
            "config_code": td.config_code,
        }
        for td in trace_data
    }

    return render(request, "users/assembly/assembly_post_table.html", {
        "records_grouped": records_grouped_sorted,
        "post": post,
        "controller_users": controller_users,
        "vin_trace_map": vin_trace_map,
    })




from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.http import HttpResponse, HttpResponseBadRequest
from django.conf import settings
from django.utils import timezone as dj_tz
from django.db.models import Q

from collections import defaultdict
from datetime import datetime
from PIL import Image
import xlsxwriter
import json, io, os


def _collect_dataset(post: str, rows_from_client: list | None):
    """
    Возвращает:
      records_grouped_sorted: list[(vin, [rows])]
      vin_meta: dict[vin] -> (brand, model, config)
      has_any_photos: bool  (есть ли хотя бы одно фото в данных)
    """
    zone = "Цех сборки"

    from collections import defaultdict
    from datetime import datetime
    from django.utils import timezone as dj_tz
    import json

    def _split_dt(dt):
        if not dt:
            return "", ""
        try:
            s = str(dt).replace("Z", "+00:00")
            d = datetime.fromisoformat(s)
            if d.tzinfo:
                d = d.astimezone(dj_tz.get_current_timezone())
            return d.strftime("%d.%m.%Y"), d.strftime("%H:%M")
        except Exception:
            try:
                if hasattr(dt, "strftime"):
                    return dt.strftime("%d.%m.%Y"), dt.strftime("%H:%M")
            except Exception:
                pass
            return "", ""

    def _norm(s: str) -> str:
        return (s or "").strip().lower()

    def _responsible_from_history(vin: str, date_s: str, time_s: str,
                                  unit_txt: str, defect_txt: str) -> str:
        """
        Пытаемся найти в истории VIN запись поста с тем же датой/временем
        и забрать оттуда extra.qrr_responsibles (по совпадению unit/defect).
        """
        vh = VINHistory.objects.filter(vin=vin).only("history").first()
        if not vh or not vh.history:
            return ""
        h = vh.history
        if isinstance(h, str):
            try:
                h = json.loads(h)
            except Exception:
                return ""

        entries = (h.get(zone, {}) or {}).get(post, []) or []
        if not isinstance(entries, list):
            return ""

        # кандидаты по точному совпадению даты/времени (как в таблице)
        candidates = []
        for e in entries:
            ds, ts = _split_dt(e.get("date_added"))
            if ds == (date_s or "") and ts == (time_s or ""):
                candidates.append(e)

        if not candidates:
            return ""

        unit_l = _norm(unit_txt)
        defect_l = _norm(defect_txt)

        bucket = []
        for e in candidates:
            defects = e.get("defects") or []
            if not isinstance(defects, list):
                continue

            # сначала пытаемся по точному совпадению unit/defect
            matched = False
            for d in defects:
                du = _norm(d.get("unit") or d.get("detail"))
                dn = _norm(d.get("name") or d.get("defect"))
                ok_u = (not unit_l) or (du == unit_l)
                ok_d = (not defect_l) or (dn == defect_l)
                if ok_u and ok_d:
                    extra = d.get("extra") or {}
                    resp = extra.get("qrr_responsibles") or []
                    if isinstance(resp, list):
                        bucket.extend([str(x) for x in resp if x is not None])
                    matched = True

            # если по текстам не нашли, но дефект один — берём его как fallback
            if not matched and len(defects) == 1:
                extra = (defects[0].get("extra") or {})
                resp = extra.get("qrr_responsibles") or []
                if isinstance(resp, list):
                    bucket.extend([str(x) for x in resp if x is not None])

        # уберём дубли, сохранив порядок
        bucket = list(dict.fromkeys(bucket))
        return ", ".join(bucket)

    records_grouped: dict[str, list[dict]] = defaultdict(list)
    vin_meta: dict[str, tuple[str, str, str]] = {}
    has_any_photos = False

    if isinstance(rows_from_client, list) and rows_from_client:
        # ✅ использовать отфильтрованные строки со страницы
        for r in rows_from_client:
            vin = (r.get("vin") or "").strip().upper()
            if not vin:
                continue
            brand  = (r.get("brand") or "").strip()
            model  = (r.get("model") or "").strip()
            config = (r.get("config_code") or "").strip()
            if vin not in vin_meta or not any(vin_meta[vin]):
                vin_meta[vin] = (brand, model, config)

            photos = r.get("photos") or []
            has_any_photos = has_any_photos or bool(photos)

            date_s = (r.get("date") or "").strip()
            time_s = (r.get("time") or "").strip()
            line   = (r.get("line") or "").strip()
            unit   = (r.get("unit") or "").strip()
            defect = (r.get("defect") or "").strip()
            comment = (r.get("comment") or "").strip()
            grade   = (r.get("grade") or "").strip()
            controller = (r.get("controller") or "").strip()

            # если ответственный не прислан — достаём из истории VIN
            responsible = (r.get("responsible") or "").strip()
            if not responsible:
                responsible = _responsible_from_history(vin, date_s, time_s, unit, defect)

            records_grouped[vin].append({
                "date": date_s, "time": time_s, "line": line,
                "unit": unit, "defect": defect, "comment": comment, "grade": grade,
                "controller": controller, "responsible": responsible,
                "photos": photos,
            })

        # дотянуть мету, если пустая
        need_fill = [v for v, (b, m, c) in vin_meta.items() if not (b and m and c)]
        if need_fill:
            for t in TraceData.objects.filter(vin_rk__in=need_fill).only("vin_rk", "brand", "model", "config_code"):
                v = (t.vin_rk or "").upper()
                old = vin_meta.get(v, ("", "", ""))
                vin_meta[v] = (old[0] or (t.brand or ""), old[1] or (t.model or ""), old[2] or (t.config_code or ""))

    else:
        # 🟨 запасной режим — собрать всё из БД (как было)
        for vh in VINHistory.objects.only("vin", "history"):
            vin = (vh.vin or "").strip().upper()
            if not vin:
                continue
            h = vh.history or {}
            if isinstance(h, str):
                try:
                    h = json.loads(h)
                except Exception:
                    continue

            post_entries = (h.get(zone, {}) or {}).get(post, []) or []
            for entry in post_entries:
                raw_dt = entry.get("date_added")
                date_s, time_s = _split_dt(raw_dt)
                controller = (entry.get("controller") or entry.get("extra_data", {}).get("controller") or "").strip()
                line = entry.get("line", "") or ""

                if isinstance(entry.get("defects"), list) and entry["defects"]:
                    for d in entry["defects"]:
                        name   = (d.get("name") or d.get("defect") or "").strip() or "Без описания"
                        unit   = (d.get("unit") or d.get("detail") or "").strip()
                        grade  = (d.get("grade") or "").strip()
                        comment = (d.get("comment") or "").strip()
                        photos = list(dict.fromkeys([str(p) for p in (d.get("photos") or [])]))
                        has_any_photos = has_any_photos or bool(photos)
                        extra = d.get("extra") or {}
                        responsibles = extra.get("qrr_responsibles") or []
                        responsible_str = ", ".join([str(x) for x in responsibles if x is not None]) \
                                          if isinstance(responsibles, list) else ""
                        records_grouped[vin].append({
                            "date": date_s, "time": time_s, "line": line,
                            "unit": unit, "defect": name, "comment": comment, "grade": grade,
                            "controller": controller, "responsible": responsible_str,
                            "photos": photos,
                        })
                else:
                    records_grouped[vin].append({
                        "date": date_s, "time": time_s, "line": line,
                        "unit": "", "defect": "Без дефектов", "comment": "", "grade": "",
                        "controller": controller, "responsible": "", "photos": [],
                    })

        # мета по VIN
        vin_list = list(records_grouped.keys())
        for t in TraceData.objects.filter(vin_rk__in=vin_list).only("vin_rk", "brand", "model", "config_code"):
            vin_meta[(t.vin_rk or "").upper()] = (t.brand or "", t.model or "", t.config_code or "")

    # сортировки
    def _dt_key(r: dict):
        try:
            return datetime.strptime(f'{r.get("date","")} {r.get("time","")}', "%d.%m.%Y %H:%M")
        except Exception:
            return datetime.min

    for v in records_grouped:
        records_grouped[v].sort(key=_dt_key, reverse=True)

    records_grouped_sorted = sorted(
        records_grouped.items(),
        key=lambda item: _dt_key(item[1][0]) if item[1] else datetime.min,
        reverse=True
    )

    return records_grouped_sorted, vin_meta, has_any_photos


def _build_excel(records_grouped_sorted, vin_meta, include_photos: bool, file_title: str) -> bytes:
    """Строит xlsx и возвращает bytes."""
    def _media_stream(abs_or_media_path: str) -> io.BytesIO | None:
        if not abs_or_media_path:
            return None
        rel = abs_or_media_path
        if rel.startswith("/media/"):
            rel = rel[len("/media/"):]
        rel = rel.lstrip("/")
        full = os.path.join(settings.MEDIA_ROOT, rel)
        if not os.path.exists(full) or not os.path.isfile(full):
            return None
        try:
            with open(full, "rb") as f:
                return io.BytesIO(f.read())
        except Exception:
            return None

    def _image_scale_for_draw(image_stream: io.BytesIO, max_w_px: int, max_h_px: int) -> tuple[float, float]:
        try:
            pos = image_stream.tell()
            image_stream.seek(0)
            im = Image.open(image_stream)
            w, h = im.size
            image_stream.seek(pos)
            if not w or not h:
                return 1.0, 1.0
            sx = min(max_w_px / float(w), 1.0)
            sy = min(max_h_px / float(h), 1.0)
            return sx, sy
        except Exception:
            return 1.0, 1.0

    # динамика по фото
    max_photos = 0
    if include_photos:
        for _, rows in records_grouped_sorted:
            for r in rows:
                max_photos = max(max_photos, len(r.get("photos") or []))
        max_photos = max(max_photos, 1)  # хотя бы одна колонка «Фото 1»

    bio = io.BytesIO()
    wb = xlsxwriter.Workbook(bio, {"in_memory": True, "strings_to_urls": False, "constant_memory": True})
    ws = wb.add_worksheet(file_title[:31] or "Отчёт")  # ограничение Excel в 31 символ

    fmt_head  = wb.add_format({"bold": True, "align": "center", "valign": "vcenter",
                               "bg_color": "#F2F2F2", "border": 1, "text_wrap": True})
    fmt_merge = wb.add_format({"align": "center", "valign": "vcenter", "border": 1})
    fmt_txt   = wb.add_format({"valign": "top", "border": 1, "text_wrap": True})
    fmt_ctr   = wb.add_format({"align": "center", "valign": "top", "border": 1})

    headers_fixed = [
        "VIN", "Бренд", "Модель", "Код комплектации",
        "Дата", "Время", "Линия",
        "Деталь", "Дефект", "Комментарий", "Грейд", "Контролер", "Ответственный",
    ]
    photo_headers = [f"Фото {i}" for i in range(1, max_photos + 1)] if include_photos else []
    headers = headers_fixed + photo_headers

    widths = [20, 14, 18, 18, 12, 10, 14, 22, 24, 28, 10, 16, 20] + ([18] * len(photo_headers))
    for i, (h, w) in enumerate(zip(headers, widths)):
        ws.write(0, i, h, fmt_head)
        ws.set_column(i, i, w)

    # если есть фото — повыше строки; если нет — стандарт
    if include_photos:
        ws.set_default_row(120)
    else:
        ws.set_default_row(18)

    PHOTO_MAX_W = 120
    PHOTO_MAX_H = 110

    # индексы
    COL_VIN, COL_BRAND, COL_MODEL, COL_CFG = 0, 1, 2, 3
    COL_DATE, COL_TIME, COL_LINE = 4, 5, 6
    COL_UNIT, COL_DEFECT, COL_COMMENT, COL_GRADE, COL_CONTROLLER, COL_RESP = 7, 8, 9, 10, 11, 12
    COL_PHOTOS_START = len(headers_fixed)

    row_xlsx = 1
    for vin, rows in records_grouped_sorted:
        brand, model, cfg = vin_meta.get(vin, ("", "", ""))
        start_row = row_xlsx

        for r in rows:
            ws.write(row_xlsx, COL_VIN,   vin,   fmt_txt)
            ws.write(row_xlsx, COL_BRAND, brand, fmt_txt)
            ws.write(row_xlsx, COL_MODEL, model, fmt_txt)
            ws.write(row_xlsx, COL_CFG,   cfg,   fmt_txt)

            ws.write(row_xlsx, COL_DATE, (r.get("date") or ""), fmt_ctr)
            ws.write(row_xlsx, COL_TIME, (r.get("time") or ""), fmt_ctr)
            ws.write(row_xlsx, COL_LINE, (r.get("line") or ""), fmt_txt)

            ws.write(row_xlsx, COL_UNIT,      (r.get("unit") or ""),      fmt_txt)
            ws.write(row_xlsx, COL_DEFECT,    (r.get("defect") or ""),    fmt_txt)
            ws.write(row_xlsx, COL_COMMENT,   (r.get("comment") or ""),   fmt_txt)
            ws.write(row_xlsx, COL_GRADE,     (r.get("grade") or ""),     fmt_ctr)
            ws.write(row_xlsx, COL_CONTROLLER,(r.get("controller") or ""),fmt_txt)
            ws.write(row_xlsx, COL_RESP,      (r.get("responsible") or ""), fmt_txt)

            if include_photos:
                photos = r.get("photos") or []
                for j in range(max_photos):
                    col = COL_PHOTOS_START + j
                    p = photos[j] if j < len(photos) else None
                    if not p:
                        continue
                    stream = _media_stream(p)
                    if stream is None:
                        ws.write(row_xlsx, col, os.path.basename(str(p)), fmt_txt)
                        continue
                    try:
                        sx, sy = _image_scale_for_draw(stream, PHOTO_MAX_W, PHOTO_MAX_H)
                        ws.insert_image(row_xlsx, col, "photo.bin", {
                            "image_data": stream, "x_offset": 4, "y_offset": 4,
                            "x_scale": sx, "y_scale": sy
                        })
                    except Exception:
                        ws.write(row_xlsx, col, os.path.basename(str(p)), fmt_txt)

            row_xlsx += 1

        # слить VIN/Бренд/Модель/Код на группу
        if row_xlsx - 1 > start_row:
            ws.merge_range(start_row, COL_VIN,   row_xlsx - 1, COL_VIN,   vin,   fmt_merge)
            ws.merge_range(start_row, COL_BRAND, row_xlsx - 1, COL_BRAND, brand, fmt_merge)
            ws.merge_range(start_row, COL_MODEL, row_xlsx - 1, COL_MODEL, model, fmt_merge)
            ws.merge_range(start_row, COL_CFG,   row_xlsx - 1, COL_CFG,   cfg,   fmt_merge)

    wb.close()
    bio.seek(0)
    return bio.getvalue()


def _export_common(request, include_photos: bool, name_suffix: str):
    post = (request.GET.get("post") or "").strip()
    if not post:
        return HttpResponseBadRequest("parameter 'post' is required")

    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except Exception:
        payload = {}
    rows_from_client = payload.get("rows")

    records_grouped_sorted, vin_meta, _ = _collect_dataset(post, rows_from_client)
    if not records_grouped_sorted:
        return HttpResponseBadRequest("Нет данных для экспорта по этому посту")

    data = _build_excel(
        records_grouped_sorted=records_grouped_sorted,
        vin_meta=vin_meta,
        include_photos=include_photos,
        file_title="Отчёт",
    )

    fname = f"post_export_{post}{name_suffix}_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"
    resp = HttpResponse(
        data,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{fname}"'
    return resp


# ===== публичные вьюшки =====
@login_required
@permission_required('users.access_to_vqa_data', raise_exception=True)
@require_POST
def assembly_post_export(request):
    """Скачать С ФОТО (как сейчас). Учитывает фильтры с фронта, если присланы rows."""
    return _export_common(request, include_photos=True, name_suffix="_with_photos")


@login_required
@permission_required('users.access_to_vqa_data', raise_exception=True)
@require_POST
def assembly_post_export_nophotos(request):
    """Скачать БЕЗ ФОТО — те же колонки, но без блока Фото."""
    return _export_common(request, include_photos=False, name_suffix="_no_photos")



def _build_excel_custom(records_grouped_sorted, vin_meta, columns: list[str],
                        photo_limit: int | None = None, sheet_title: str = "Отчёт") -> bytes:
    """
    columns — список ключей в нужном порядке. Допустимые ключи:
      'vin','brand','model','config_code','date','time','line',
      'unit','defect','comment','grade','controller','responsible','photos'

    Если 'photos' присутствует — добавляются N колонок «Фото i».
    N = max(1, min(photo_limit, max_фото_в_данных)) если photo_limit задан,
        иначе N = max(1, max_фото_в_данных).
    """
    allowed = {
        "vin","brand","model","config_code","date","time","line",
        "unit","defect","comment","grade","controller","responsible","photos"
    }
    bad = [c for c in columns if c not in allowed]
    if bad:
        raise ValueError(f"Недопустимые колонки: {', '.join(bad)}")

    include_photos = "photos" in columns

    # ----- утилиты для фото -----
    def _media_stream(abs_or_media_path: str) -> io.BytesIO | None:
        if not abs_or_media_path:
            return None
        rel = abs_or_media_path
        if rel.startswith("/media/"):
            rel = rel[len("/media/"):]
        rel = rel.lstrip("/")
        full = os.path.join(settings.MEDIA_ROOT, rel)
        if not os.path.exists(full) or not os.path.isfile(full):
            return None
        try:
            with open(full, "rb") as f:
                return io.BytesIO(f.read())
        except Exception:
            return None

    def _image_scale_for_draw(image_stream: io.BytesIO, max_w_px: int, max_h_px: int) -> tuple[float, float]:
        try:
            pos = image_stream.tell()
            image_stream.seek(0)
            im = Image.open(image_stream)
            w, h = im.size
            image_stream.seek(pos)
            if not w or not h:
                return 1.0, 1.0
            sx = min(max_w_px / float(w), 1.0)
            sy = min(max_h_px / float(h), 1.0)
            return sx, sy
        except Exception:
            return 1.0, 1.0

    # посчитаем максимальное число фото
    max_photos_in_data = 0
    if include_photos:
        for _, rows in records_grouped_sorted:
            for r in rows:
                max_photos_in_data = max(max_photos_in_data, len(r.get("photos") or []))
        max_photos_in_data = max(max_photos_in_data, 1)
    # итоговое число колонок Фото
    if include_photos:
        if isinstance(photo_limit, int) and photo_limit >= 1:
            photo_count = max(1, min(photo_limit, max_photos_in_data))
        else:
            photo_count = max_photos_in_data
    else:
        photo_count = 0

    # заголовки / ширины / выравнивание
    title_map = {
        "vin":"VIN", "brand":"Бренд", "model":"Модель", "config_code":"Код комплектации",
        "date":"Дата", "time":"Время", "line":"Линия",
        "unit":"Деталь", "defect":"Дефект", "comment":"Комментарий",
        "grade":"Грейд", "controller":"Контролер", "responsible":"Ответственный",
        "photos":"Фото"
    }
    width_map = {
        "vin":20, "brand":14, "model":18, "config_code":18,
        "date":12, "time":10, "line":14,
        "unit":22, "defect":24, "comment":28,
        "grade":10, "controller":16, "responsible":20,
        "photos":18
    }
    centered = {"date","time","grade"}

    # «плоский» список колонок без «photos» (он расширяется отдельно)
    base_cols = [c for c in columns if c != "photos"]

    # подготовка книги
    bio = io.BytesIO()
    wb = xlsxwriter.Workbook(bio, {"in_memory": True, "strings_to_urls": False, "constant_memory": True})
    ws = wb.add_worksheet((sheet_title or "Отчёт")[:31])

    fmt_head  = wb.add_format({"bold": True, "align": "center", "valign": "vcenter",
                               "bg_color": "#F2F2F2", "border": 1, "text_wrap": True})
    fmt_merge = wb.add_format({"align": "center", "valign": "vcenter", "border": 1})
    fmt_txt   = wb.add_format({"valign": "top", "border": 1, "text_wrap": True})
    fmt_ctr   = wb.add_format({"align": "center", "valign": "top", "border": 1})

    # сформировать заголовки и ширины
    headers: list[str] = []
    widths:  list[int] = []
    for k in base_cols:
        headers.append(title_map[k])
        widths.append(width_map.get(k, 16))
    if include_photos:
        for i in range(1, photo_count+1):
            headers.append(f"Фото {i}")
            widths.append(width_map["photos"])

    # выставить шапку и ширины
    for col, (h, w) in enumerate(zip(headers, widths)):
        ws.write(0, col, h, fmt_head)
        ws.set_column(col, col, w)

    ws.set_default_row(120 if include_photos else 18)
    PHOTO_MAX_W = 120
    PHOTO_MAX_H = 110

    # позиции специальных колонок (если они включены)
    idx_map = {k: i for i, k in enumerate(base_cols)}
    photos_col_start = len(base_cols)  # старт фото-блока (если он есть)

    # запись строк
    row_xlsx = 1
    for vin, rows in records_grouped_sorted:
        brand, model, cfg = vin_meta.get(vin, ("", "", ""))
        start_row = row_xlsx

        for r in rows:
            # пишем «базовые» колонки в заданном порядке
            for k, col in idx_map.items():
                if k == "vin":
                    val = vin
                elif k == "brand":
                    val = brand
                elif k == "model":
                    val = model
                elif k == "config_code":
                    val = cfg
                else:
                    val = (r.get(k) or "")
                ws.write(row_xlsx, col, val, fmt_ctr if k in centered else fmt_txt)

            # блок фото
            if include_photos:
                photos = r.get("photos") or []
                for j in range(photo_count):
                    col = photos_col_start + j
                    p = photos[j] if j < len(photos) else None
                    if not p:
                        continue
                    stream = _media_stream(p)
                    if stream is None:
                        ws.write(row_xlsx, col, os.path.basename(str(p)), fmt_txt)
                        continue
                    try:
                        sx, sy = _image_scale_for_draw(stream, PHOTO_MAX_W, PHOTO_MAX_H)
                        ws.insert_image(row_xlsx, col, "photo.bin", {
                            "image_data": stream, "x_offset": 4, "y_offset": 4,
                            "x_scale": sx, "y_scale": sy
                        })
                    except Exception:
                        ws.write(row_xlsx, col, os.path.basename(str(p)), fmt_txt)

            row_xlsx += 1

        # слияния по VIN/Бренд/Модель/Код (только если эти колонки включены)
        if row_xlsx - 1 > start_row:
            for k, v in (("vin", vin), ("brand", brand), ("model", model), ("config_code", cfg)):
                if k in idx_map:
                    c = idx_map[k]
                    ws.merge_range(start_row, c, row_xlsx - 1, c, v, fmt_merge)

    wb.close()
    bio.seek(0)
    return bio.getvalue()


@login_required
@permission_required('users.access_to_vqa_data', raise_exception=True)
@require_POST
def assembly_post_export_custom(request):
    """
    Кастом-экспорт по выбранным колонкам.
    Ожидает JSON в теле запроса:
      {
        "rows": [...],              # опционально, отфильтрованные строки со страницы
        "columns": ["vin","date",...,"photos"],   # ОБЯЗАТЕЛЬНО, порядок = порядок в файле
        "photo_limit": 3            # опционально, если есть "photos" (>=1)
      }
    Параметр GET ?post= обязателен (как и в других экспортерах).
    """
    post = (request.GET.get("post") or "").strip()
    if not post:
        return HttpResponseBadRequest("parameter 'post' is required")

    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except Exception:
        payload = {}

    columns = payload.get("columns")
    if not isinstance(columns, list) or not columns:
        return HttpResponseBadRequest("В теле запроса должен быть список 'columns'.")

    rows_from_client = payload.get("rows")
    photo_limit = payload.get("photo_limit")

    # собрать датасет (если rows присланы — учитываются все активные фильтры страницы)
    records_grouped_sorted, vin_meta, _ = _collect_dataset(post, rows_from_client)
    if not records_grouped_sorted:
        return HttpResponseBadRequest("Нет данных для экспорта по этому посту")

    try:
        data = _build_excel_custom(
            records_grouped_sorted=records_grouped_sorted,
            vin_meta=vin_meta,
            columns=columns,
            photo_limit=photo_limit if isinstance(photo_limit, int) else None,
            sheet_title="Отчёт",
        )
    except ValueError as e:
        return HttpResponseBadRequest(str(e))

    fname = f"post_export_{post}_custom_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"
    resp = HttpResponse(
        data,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename=\"{fname}\"'
    return resp










@login_required
@permission_required('users.access_to_the_shift_management', raise_exception=True)
def master_controller_panel(request):
    User = get_user_model()

    # Сохраняем редактируемого контроллера
    if request.method == 'POST' and 'user_id' in request.POST:
        user_id = request.POST.get('user_id')
        user = get_object_or_404(User, id=user_id, role='controller')

        form_edit = ControllerEditForm(request.POST, request.FILES, instance=user)
        if form_edit.is_valid():
            form_edit.save()
            messages.success(request, f"✅ Контроллер {user.username} успешно обновлён.")
            return redirect('master_controller_panel')
        else:
            messages.error(request, f"❌ Ошибка при обновлении контроллера {user.username}.")
            # Тут можно сохранить form_edit в context при необходимости
    # Создание нового контроллера
    elif request.method == 'POST':
        form = ControllerCreationForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ Контроллер успешно создан.")
            return redirect('master_controller_panel')
        else:
            for error in form.errors.get('__all__', []):
                if "Пароли не совпадают" in error:
                    messages.error(request, "❌ Пароли не совпадают. Повторите ввод.")
                elif "This password is too short" in error:
                    messages.error(request, "❌ Пароль слишком короткий. Минимум 8 символов.")
                elif "This password is too common" in error:
                    messages.error(request, "❌ Пароль слишком простой. Используйте более надёжный.")
                elif "This password is entirely numeric" in error:
                    messages.error(request, "❌ Пароль не должен состоять только из цифр.")
                else:
                    messages.error(request, f"❌ {error}")
        form_edit = None  # не нужен
    else:
        form = ControllerCreationForm()
        form_edit = None

    controllers = User.objects.filter(role='controller', is_superuser=False).order_by('username')

    context = {
        'form': form,
        'controllers': controllers,
        "enable_particles": True,
        "enable_background": True,
        "enable_white_square": False,
    }
    return render(request, 'users/controller_panel/master_controller_panel.html', context)


@login_required
@permission_required('users.access_to_the_shift_management', raise_exception=True)
def delete_controller(request, user_id):
    user = get_object_or_404(get_user_model(), id=user_id, role='controller', is_superuser=False)

    username = user.username
    user.delete()
    messages.success(request, f"🗑️ Контроллер '{username}' успешно удалён.")
    return redirect('master_controller_panel')


@login_required
@permission_required('users.access_to_the_shift_management', raise_exception=True)
def change_controller_password(request, user_id):
    user = get_object_or_404(get_user_model(), id=user_id, role='controller', is_superuser=False)

    if request.method == 'POST':
        form = ControllerPasswordChangeForm(request.POST)
        if form.is_valid():
            user.set_password(form.cleaned_data['new_password'])
            user.save()
            messages.success(request, f"✅ Пароль для {user.username} успешно изменён.")
            return redirect('master_controller_panel')
        else:
            for error in form.errors.get('__all__', []):
                if "Пароли не совпадают" in error:
                    messages.error(request, "❌ Пароли не совпадают.")
                elif "This password is too short" in error:
                    messages.error(request, "❌ Пароль слишком короткий. Минимум 8 символов.")
                elif "This password is too common" in error:
                    messages.error(request, "❌ Пароль слишком простой. Используйте более надёжный.")
                elif "This password is entirely numeric" in error:
                    messages.error(request, "❌ Пароль не должен состоять только из цифр.")
                else:
                    messages.error(request, f"❌ {error}")
    else:
        form = ControllerPasswordChangeForm()

    return render(request, 'users/controller_panel/change_controller_password.html', {
        'form': form,
        'controller': user,
    })


@login_required
@permission_required('users.access_to_vqa_data', raise_exception=True)
def assembly_post_edit(request, vin, post_name, timestamp):
    zone = "Цех сборки"

    try:
        timestamp_dt = datetime.fromisoformat(timestamp)
    except ValueError:
        return HttpResponse("❌ Неверный формат времени", status=400)

    history_entry = get_object_or_404(VINHistory, vin=vin)
    post_entries = history_entry.history.get(zone, {}).get(post_name, [])

    entry_to_edit = None
    for entry in post_entries:
        raw_date = entry.get("date_added")
        try:
            entry_dt = datetime.fromisoformat(raw_date)
        except Exception:
            continue

        if abs(entry_dt - timestamp_dt) <= timedelta(seconds=1):
            entry_to_edit = entry
            break

    if not entry_to_edit:
        return HttpResponse("❌ Запись не найдена", status=404)

    if request.method == "POST":
        VINHistoryBackup.objects.create(
            vin=vin,
            post=post_name,
            zone=zone,
            entry=entry_to_edit.copy(),
            action="edit"
        )

        new_vin = request.POST.get("vin_number", "").strip()
        has_defect = request.POST.get("has_defect", "")
        defect_description = request.POST.get("defect_description", "").strip()

        if not new_vin:
            messages.error(request, "❌ VIN обязателен.")
            return redirect(request.path)

        if new_vin == vin:
            # VIN не менялся
            entry_to_edit["has_defect"] = has_defect
            entry_to_edit["defect_description"] = defect_description
            history_entry.save()
            messages.success(request, f"✅ Изменения сохранены.")
        else:
            # VIN менялся
            post_entries.remove(entry_to_edit)

            if post_entries:
                history_entry.history[zone][post_name] = post_entries
            else:
                del history_entry.history[zone][post_name]
                if not history_entry.history[zone]:
                    del history_entry.history[zone]

            entry_to_edit["vin_number"] = new_vin
            entry_to_edit["has_defect"] = has_defect
            entry_to_edit["defect_description"] = defect_description

            new_history_entry, _ = VINHistory.objects.get_or_create(vin=new_vin)
            new_history_entry.history.setdefault(zone, {}).setdefault(post_name, []).append(entry_to_edit)

            history_entry.save()
            new_history_entry.save()
            messages.success(request, f"✅ Изменения сохранены. Запись перенесена в VIN {new_vin}.")

        return redirect(reverse("assembly_post_table") + f"?post={post_name}")

    # перед return render(...)
    try:
        entry_dt = datetime.fromisoformat(entry_to_edit["date_added"])
        entry_to_edit["date_display"] = entry_dt.strftime("%Y-%m-%d %H:%M")
    except Exception:
        entry_to_edit["date_display"] = entry_to_edit["date_added"]

    return render(request, "users/assembly/assembly_post_edit.html", {
        "post_name": post_name,
        "entry": entry_to_edit,
        "vin": vin,
        "timestamp": timestamp,
    })



@login_required
@permission_required('users.access_to_vqa_data', raise_exception=True)
def delete_entry_view(request, vin, post, timestamp):
    zone = "Цех сборки"

    try:
        timestamp_dt = datetime.fromisoformat(timestamp)
    except ValueError:
        return HttpResponse("❌ Неверный формат времени", status=400)

    history_entry = get_object_or_404(VINHistory, vin=vin)
    post_entries = history_entry.history.get(zone, {}).get(post, [])

    updated_entries = []
    deleted_entry = None  # ← сохраним удалённый элемент для бэкапа

    for entry in post_entries:
        try:
            entry_dt = datetime.fromisoformat(entry.get("date_added", ""))
            if abs(entry_dt - timestamp_dt) <= timedelta(seconds=1):
                deleted_entry = entry
                continue  # не добавляем в обновлённый список
        except Exception:
            pass
        updated_entries.append(entry)

    if deleted_entry:
        # ✅ сохраняем удалённую запись в VINHistoryBackup
        VINHistoryBackup.objects.create(
            vin=vin,
            post=post,
            zone=zone,
            entry=deleted_entry,
            action="delete"
        )

        # ✅ сохраняем оставшиеся записи в основной таблице
        history_entry.history[zone][post] = updated_entries
        history_entry.save()

        messages.success(request, "✅ Запись успешно удалена и сохранена в архив.")
    else:
        messages.error(request, "❌ Запись не найдена.")

    return redirect(reverse("assembly_post_table") + f"?post={post}")










@login_required
@permission_required('users.access_to_vqa_data', raise_exception=True)
def assembly_general_report(request):
    brands_selected = request.GET.getlist("brand")
    BRAND_MAP = {
        "gwm": ["haval", "tank"],
        "chery": ["chery"],
        "changan": ["changan"]
    }
    models_selected = request.GET.getlist("model")
    download_brand = request.GET.get("download_brand")  # 👈 добавлено

    mapped_brands = list(chain.from_iterable(BRAND_MAP.get(b, [b]) for b in brands_selected))

    trace_qs = TraceData.objects.all()

    if mapped_brands:
        trace_qs = trace_qs.filter(brand__in=mapped_brands)

    if models_selected:
        trace_qs = trace_qs.filter(model__in=models_selected)

    filtered_vins = set(trace_qs.values_list("vin_rk", flat=True))

    if mapped_brands:
        models_available = TraceData.objects.filter(brand__in=mapped_brands).values_list("model", flat=True).distinct()
    else:
        models_available = TraceData.objects.values_list("model", flat=True).distinct()

    brands_all = TraceData.objects.values_list("brand", flat=True).distinct()

    posts_info = {
        "torque_control": {"post": "Пост момента затяжек", "title": "Пост момента затяжек"},
        "chassis": {"post": "Chassis", "title": "Chassis"},
        "final_current": {"post": "Финал текущий контроль", "title": "Финал текущий контроль"},
        "gaps_and_drops": {"post": "Зазоры и перепады", "title": "Зазоры и перепады"},
        "exterior": {"post": "Экстерьер", "title": "Экстерьер"},
        "interior": {"post": "Интерьер", "title": "Интерьер"},
        "trunk": {"post": "Багажник", "title": "Багажник"},
        "motor": {"post": "Мотор", "title": "Мотор"},
        "functionality": {"post": "Функцонал", "title": "Функцонал"},
        "wheel_geometry": {"post": "Геометрия колес", "title": "Геометрия колес"},
        "light_and_steering": {"post": "Регулировка света фар и калибровка руля", "title": "Регулировка света фар и калибровка руля"},
        "brake_system": {"post": "Тормозная система", "title": "Тормозная система"},
        "underbody": {"post": "Underbody", "title": "Underbody"},
        "adas": {"post": "ADAS", "title": "ADAS"},
        "avm": {"post": "AVM", "title": "AVM"},
        "body_tightness": {"post": "Герметичность кузова", "title": "Герметичность кузова"},
        "diagnostics": {"post": "Диагностика", "title": "Диагностика"},
        "test_track": {"post": "Тест трек", "title": "Тест трек"},
        "documentation": {"post": "Документация", "title": "Документация"},
    }

    zone = "Цех сборки"
    today = now().date()

    start_date_str = request.GET.get("start_date")
    end_date_str = request.GET.get("end_date")

    try:
        start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date() if start_date_str else today - timedelta(days=7)
        end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date() if end_date_str else today
    except ValueError:
        start_date = today - timedelta(days=7)
        end_date = today

    vin_with_final = set()
    vin_with_defect_on_final = set()
    total_defects_on_final = 0
    overall_total_inspections = 0
    overall_inspections_with_defects = 0
    overall_total_defects = 0
    overall_without_defects = 0
    unique_vins = set()
    posts_data = {}

    for key, info in posts_info.items():
        post = info["post"]

        passed_vins = set()
        passed_with_defect = set()
        passed_without_defect = set()
        defect_counter = Counter()
        controller_counter = Counter()
        controller_vin_pairs = set()
        total_defects = 0

        histories = VINHistory.objects.all()
        if filtered_vins:
            histories = histories.filter(vin__in=filtered_vins)

        for history in histories:
            vin = history.vin
            post_entries = history.history.get(zone, {}).get(post, [])
            if not post_entries:
                continue

            entries_in_range = []
            for entry in post_entries:
                raw_date = entry.get("date_added")
                if not raw_date:
                    continue
                entry_date = parse_date(raw_date[:10])
                if entry_date and start_date <= entry_date <= end_date:
                    entries_in_range.append(entry)

            if not entries_in_range:
                continue

            passed_vins.add(vin)
            unique_vins.add(vin)

            vin_has_defect = any(
                entry.get("has_defect") == "yes" or entry.get("defects")
                for entry in entries_in_range
            )

            if vin_has_defect:
                passed_with_defect.add(vin)
                for entry in entries_in_range:
                    if entry.get("has_defect") == "yes" or entry.get("defects"):
                        controller = entry.get("controller", "Неизвестно")
                        if (controller, vin) not in controller_vin_pairs:
                            controller_counter[controller] += 1
                            controller_vin_pairs.add((controller, vin))

                        for defect in entry.get("defects", []):
                            name = defect.get("name")
                            unit = defect.get("unit")
                            combined = f"{name} ({unit})" if name and unit else name or unit or "Неизвестно"
                            defect_counter[combined] += 1
                            total_defects += 1
            else:
                passed_without_defect.add(vin)

        total_inspections = len(passed_vins)
        inspections_with_defects = len(passed_with_defect)
        inspections_without_defects = len(passed_without_defect)
        dpu = round(total_defects / total_inspections, 2) if total_inspections else 0
        str_percentage = round(inspections_without_defects / total_inspections * 100, 2) if total_inspections else 0

        encoded_post = urllib.parse.quote(info["post"])
        report_url = reverse("assembly_post_report") + f"?post={encoded_post}"
        table_url = reverse("assembly_post_table") + f"?post={encoded_post}"
        export_url = reverse("assembly_post_export") + f"?post={encoded_post}"

        posts_data[key] = {
            "title": info["title"],
            "total_inspections": total_inspections,
            "inspections_with_defects": inspections_with_defects,
            "without_defects": inspections_without_defects,
            "total_defects": total_defects,
            "dpu": dpu,
            "str_percentage": str_percentage,
            "top_defects": defect_counter.most_common(5),
            "top_controllers": controller_counter.most_common(5),
            "report_url": report_url,
            "table_url": table_url,
            "export_url": export_url,
        }

    # 📌 Анализ только по посту "Финал текущий контроль"
    # 📌 VINы, прошедшие финальный пост "Документация"
    # 📌 VINы, прошедшие финальный пост "Документация" в указанный период
    histories = VINHistory.objects.all()
    if filtered_vins:
        histories = histories.filter(vin__in=filtered_vins)

    vin_with_final = set()
    vin_with_defect_on_any_post = set()
    total_defects_for_final_vins = 0
    # --- begin: date map ---
    vin_final_date_map: dict[str, str] = {}
    final_post_name = "Документация"
    # --- end: date map ---

    for history in histories:
        vin = history.vin
        zone_data = history.history.get(zone, {})

        # Сначала ищем хотя бы одну запись на "Документация" в нужный период
        final_entries = zone_data.get(final_post_name, [])
        passed_final = False

        for entry in final_entries:
            raw_date = entry.get("date_added")
            if not raw_date:
                continue
            entry_date = parse_date(raw_date[:10])
            if entry_date and start_date <= entry_date <= end_date:
                # --- begin: set date in map if not already set ---
                if vin not in vin_final_date_map:
                    vin_final_date_map[vin] = (raw_date[:10] if isinstance(raw_date, str) else entry_date.strftime("%Y-%m-%d"))
                # --- end: set date in map ---
                passed_final = True
                break

        if not passed_final:
            continue  # VIN не прошёл "Документация" в выбранный период

        vin_with_final.add(vin)

        # ✅ Теперь считаем все дефекты на всех постах у этого VIN
        for post_entries in zone_data.values():
            for entry in post_entries:
                # 🔸 Не фильтруем по дате — считаем ВСЕ дефекты за всю историю VIN
                defects = entry.get("defects", [])
                if entry.get("has_defect") == "yes" or defects:
                    vin_with_defect_on_any_post.add(vin)
                    total_defects_for_final_vins += len(defects)

    overall_total_inspections = len(vin_with_final)
    overall_inspections_with_defects = len(vin_with_defect_on_any_post)
    overall_without_defects = overall_total_inspections - overall_inspections_with_defects
    overall_total_defects = total_defects_for_final_vins
    overall_dpu = round(overall_total_defects / overall_total_inspections, 2) if overall_total_inspections else 0

    # 📌 VIN-данные для всплывающего окна
    vins_with_defects = list(vin_with_defect_on_any_post)
    vins_without_defects = list(vin_with_final - vin_with_defect_on_any_post)
    vins_all = list(vin_with_final)

    request.session["vin_with_final"] = vins_all
    request.session["vin_with_defect_on_any_post"] = vins_with_defects
    request.session["vin_final_date_map"] = vin_final_date_map

    # Сопоставим VIN → бренд и модель
    trace_data_map = {
        obj.vin_rk: {
            "brand": obj.brand,
            "model": obj.model
        }
        for obj in TraceData.objects.filter(vin_rk__in=vin_with_final)
    }

    User = get_user_model()
    all_controllers = set()
    for pdata in posts_data.values():
        for username, _ in pdata["top_controllers"]:
            all_controllers.add(username)

    # Получаем данные пользователей
    controller_users = {
        user.username: user for user in User.objects.filter(username__in=all_controllers)
    }

    context = {
        "start_date": start_date_str or start_date.strftime("%Y-%m-%d"),
        "end_date": end_date_str or end_date.strftime("%Y-%m-%d"),
        "overall_total_inspections": overall_total_inspections,
        "overall_inspections_with_defects": overall_inspections_with_defects,
        "overall_without_defects": overall_without_defects,
        "overall_total_defects": overall_total_defects,
        "overall_dpu": overall_dpu,
        "vins_with_defects": vins_with_defects,
        "vins_without_defects": vins_without_defects,
        "vins_all": vins_all,
        "vin_brand_model_map": trace_data_map,
        # "overall_str_percentage": overall_str_percentage,
        "posts_data": posts_data,
        "controller_users": controller_users,
        "unique_vin_count": overall_total_inspections,
        "models_available": models_available,
        "brands_selected": brands_selected,
        "models_selected": models_selected,
        "brands_all": brands_all,
        "fixed_brands_excel": ['gwm', 'chery', 'changan'],
        "enable_particles": True,
        "enable_background": True,
        "enable_white_square": False,
    }

    return render(request, "users/assembly/assembly_general_report.html", context)


@login_required
@permission_required('users.access_to_vqa_data', raise_exception=True)
def vin_list_view(request, vin_type):
    # раньше было: pop(...)
    vins_all = request.session.get("vin_with_final")
    vins_with_defects = request.session.get("vin_with_defect_on_any_post")
    vin_final_date_map = request.session.get("vin_final_date_map", {}) or {}

    if not vins_all or vins_with_defects is None:
        messages.error(request, "❌ Сначала сформируйте отчёт.")
        return redirect("assembly_general_report")

    vins_without_defects = list(set(vins_all) - set(vins_with_defects))

    vins_map = {
        "with_defects": vins_with_defects,
        "without_defects": vins_without_defects,
        "all": vins_all,
    }

    titles = {
        "with_defects": "VIN с дефектами",
        "without_defects": "VIN без дефектов",
        "all": "Все VIN",
    }

    vins = vins_map.get(vin_type, [])
    title = titles.get(vin_type, "VIN номера")

    trace_data_map = {
        obj.vin_rk: {"brand": obj.brand, "model": obj.model}
        for obj in TraceData.objects.filter(vin_rk__in=vins)
    }

    return render(request, "users/vin_lists_modal.html", {
        "vins": vins,
        "title": title,
        "trace_data_map": trace_data_map,
        "vin_type": vin_type,   # 👈 добавили для кнопки экспорта
        "vin_final_date_map": vin_final_date_map,
    })


@login_required
@permission_required('users.access_to_vqa_data', raise_exception=True)
def export_vin_list_excel(request, vin_type):
    """
    Экспорт текущего списка VIN'ов (как в модалке) в Excel.
    Колонки: №, VIN, Бренд, Модель, Дата прохождения.
    vin_type: with_defects | without_defects | all
    """
    vins_all = request.session.get("vin_with_final")
    vins_with_defects = request.session.get("vin_with_defect_on_any_post")
    vin_final_date_map = request.session.get("vin_final_date_map", {}) or {}

    if not vins_all or vins_with_defects is None:
        return HttpResponse("❌ Сначала сформируйте отчёт.", status=400)

    vins_without_defects = list(set(vins_all) - set(vins_with_defects))

    vins_map = {
        "with_defects": list(vins_with_defects),
        "without_defects": vins_without_defects,
        "all": list(vins_all),
    }
    vins = vins_map.get(vin_type, [])
    if not isinstance(vins, list):
        vins = list(vins)

    # бренд/модель
    trace_map = {
        t.vin_rk: {"brand": t.brand or "", "model": t.model or ""}
        for t in TraceData.objects.filter(vin_rk__in=vins)
    }

    # Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "VIN-список"
    ws.append(["№", "VIN", "Бренд", "Модель", "Дата прохождения"])

    for idx, v in enumerate(vins, start=1):
        meta = trace_map.get(v, {"brand": "", "model": ""})
        ws.append([idx, v, meta.get("brand", ""), meta.get("model", ""), (vin_final_date_map.get(v, "") or "")])

    for col in ws.columns:
        for cell in col:
            cell.alignment = Alignment(vertical="center")

    filename = f"vin_list_{vin_type}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
@permission_required('users.access_to_vqa_data', raise_exception=True)
def download_summary_report(request):
    brand = request.GET.get("download_brand")
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
    end_date = datetime.strptime(end_date, "%Y-%m-%d").date()

    path = generate_summary_report(brand, start_date, end_date)

    with open(path, "rb") as f:
        response = HttpResponse(f.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f"attachment; filename={os.path.basename(path)}"
        return response





from django.utils.timezone import make_aware, is_naive
from django.utils.dateparse import parse_datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment

@login_required
@permission_required('users.access_to_vqa_data', raise_exception=True)
def export_all_defects_excel(request):
    start_date_str = request.GET.get("start_date")
    end_date_str   = request.GET.get("end_date")
    brand_filter   = request.GET.get("download_brand") or ""  # может прийти пустой

    # --- границы периода ---
    try:
        start_date = make_aware(datetime.combine(datetime.strptime(start_date_str, "%Y-%m-%d").date(), time.min)) if start_date_str else None
        end_date   = make_aware(datetime.combine(datetime.strptime(end_date_str,   "%Y-%m-%d").date(), time.max)) if end_date_str   else None
    except ValueError:
        return HttpResponse("❌ Неверный формат даты", status=400)

    def _iter_entries(history_data):
        """
        Универсальный проход по истории:
        - Новый формат: {zone: {post: [entries]}}
        - Полу-новый:   {zone: [entries]}
        - Старый:       [entries]
        Возвращает кортежи (zone, post, entry)
        """
        if isinstance(history_data, dict):
            for zone_name, posts in history_data.items():
                if isinstance(posts, dict):
                    for post_name, entries in posts.items():
                        if isinstance(entries, list):
                            for e in entries:
                                yield zone_name or "", post_name or "", e or {}
                elif isinstance(posts, list):
                    for e in posts:
                        yield zone_name or "", "", e or {}
        elif isinstance(history_data, list):
            for e in history_data:
                yield "", "", e or {}

    vin_histories = VINHistory.objects.all()
    data_rows = []

    for history in vin_histories:
        vin = getattr(history, 'vin', '')
        trace = TraceData.objects.filter(vin_rk=vin).first()
        if not trace:
            continue

        if brand_filter and (getattr(trace, 'brand', '') != brand_filter):
            continue

        brand       = getattr(trace, 'brand',       '') or ''
        model       = getattr(trace, 'model',       '') or ''
        config_code = getattr(trace, 'config_code', '') or ''

        history_data = getattr(history, 'history', {}) or {}

        for zone_name, post_name, entry in _iter_entries(history_data):
            try:
                # --- дата ---
                date_raw = (entry.get("date_added")
                            or entry.get("date")
                            or entry.get("extra_data", {}).get("date_added"))
                if not date_raw:
                    continue
                dt = parse_datetime(str(date_raw))
                if dt is None:
                    continue
                if is_naive(dt):
                    dt = make_aware(dt)

                if start_date and dt < start_date:
                    continue
                if end_date and dt > end_date:
                    continue

                # --- остальное ---
                controller = (entry.get("controller")
                              or entry.get("extra_data", {}).get("controller")
                              or "")
                line       = entry.get("line", "") or entry.get("extra_data", {}).get("line", "") or ""
                duration   = entry.get("inspection_duration_seconds", "") or entry.get("extra_data", {}).get("inspection_duration_seconds", "")

                defects = entry.get("defects")
                if isinstance(defects, list) and defects:
                    for d in defects:
                        data_rows.append({
                            "zone": zone_name, "post": post_name, "vin": vin,
                            "brand": brand, "model": model, "config_code": config_code,
                            "date": dt, "line": line,
                            "unit": d.get("unit", ""),
                            "defect": d.get("name", "") or d.get("defect", ""),
                            "comment": d.get("comment", ""),
                            "quantity": d.get("quantity", ""),
                            "grade": d.get("grade", ""),
                            "responsible": d.get("responsible", ""),
                            "controller": controller, "duration": duration,
                        })
                else:
                    # запись без списка defects (старый формат)
                    data_rows.append({
                        "zone": zone_name, "post": post_name, "vin": vin,
                        "brand": brand, "model": model, "config_code": config_code,
                        "date": dt, "line": line,
                        "unit": entry.get("unit", ""),
                        "defect": entry.get("name", "") or entry.get("defect_description", "") or entry.get("defect", ""),
                        "comment": entry.get("comment", ""),
                        "quantity": entry.get("quantity", ""),
                        "grade": entry.get("grade", ""),
                        "responsible": entry.get("responsible", ""),
                        "controller": controller, "duration": duration,
                    })
            except Exception:
                # Пропускаем битые записи, чтобы экспорт не падал
                continue

    # --- Excel ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Все дефекты"

    headers = [
        "Участок", "Пост", "VIN", "Бренд", "Модель", "Код комплектации", "Дата", "Линия",
        "Деталь", "Дефект", "Комментарий", "Количество", "Грейд", "Кто виноват",
        "Контроллер", "Время осмотра (сек)"
    ]
    ws.append(headers)

    for row in data_rows:
        try:
            ws.append([
                row.get("zone", ""),
                row.get("post", ""),
                row.get("vin", ""),
                row.get("brand", ""),
                row.get("model", ""),
                row.get("config_code", ""),
                row.get("date").strftime("%d.%m.%Y %H:%M") if row.get("date") else "",
                row.get("line", ""),
                row.get("unit", ""),
                row.get("defect", ""),
                row.get("comment", ""),
                row.get("quantity", ""),
                row.get("grade", ""),
                row.get("responsible", ""),
                row.get("controller", ""),
                row.get("duration", ""),
            ])
        except Exception:
            continue

    # выравнивание
    for col in ws.columns:
        for cell in col:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"defects_summary_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response






from django.db.models import Case, When, Value, IntegerField

@login_required
@permission_required('users.access_to_the_vqa_data_management', raise_exception=True)
def manage_post_visibility(request):
    zone_form = AssemblyZoneForm()
    unit_form = AssemblyUnitForm()
    defect_form = AssemblyDefectForm()

    if request.method == "POST":
        if request.headers.get("x-requested-with") == "XMLHttpRequest":
            action = request.POST.get("action")

            if action == "bulk_units_from_zone":
                post_id = request.POST.get("post_id")
                zone_id = request.POST.get("zone_id")

                post = PostAssembly.objects.filter(id=post_id).first()
                if not post:
                    return JsonResponse({"ok": False, "message": "Пост не найден"}, status=404)

                zone = AssemblyZone.objects.filter(id=zone_id).first()
                if not zone:
                    return JsonResponse({"ok": False, "message": "Подсистема не найдена"}, status=404)

                unit_ids_all = list(
                    AssemblyUnit.objects.filter(zone_id=zone_id).values_list("id", flat=True)
                )
                existing = set(post.assembly_units.values_list("id", flat=True))
                to_add_ids = [uid for uid in unit_ids_all if uid not in existing]

                if to_add_ids:
                    post.assembly_units.add(*AssemblyUnit.objects.filter(id__in=to_add_ids))

                return JsonResponse({
                    "ok": True,
                    "added_unit_ids": to_add_ids,
                    "total_in_zone": len(unit_ids_all),
                })

            if action == "bulk_remove_units_of_zone":
                post_id = request.POST.get("post_id")
                zone_id = request.POST.get("zone_id")

                post = PostAssembly.objects.filter(id=post_id).first()
                if not post:
                    return JsonResponse({"ok": False, "message": "Пост не найден"}, status=404)

                zone = AssemblyZone.objects.filter(id=zone_id).first()
                if not zone:
                    return JsonResponse({"ok": False, "message": "Подсистема не найдена"}, status=404)

                unit_ids_zone = list(
                    AssemblyUnit.objects.filter(zone_id=zone_id).values_list("id", flat=True)
                )
                to_remove_qs = post.assembly_units.filter(id__in=unit_ids_zone)
                removed_ids = list(to_remove_qs.values_list("id", flat=True))

                if removed_ids:
                    post.assembly_units.remove(*to_remove_qs)

                return JsonResponse({
                    "ok": True,
                    "removed_unit_ids": removed_ids,
                    "total_in_zone": len(unit_ids_zone),
                })

            # тумблер zone/unit/defect
            post_id = request.POST.get("post_id")
            item_id = request.POST.get("item_id")
            item_type = request.POST.get("item_type")

            post = PostAssembly.objects.filter(id=post_id).first()
            if not post:
                return JsonResponse({"status": "error", "message": "Пост не найден"}, status=404)

            model_map = {
                "zone": (post.assembly_zones, AssemblyZone),
                "unit": (post.assembly_units, AssemblyUnit),
                "defect": (post.assembly_defects, AssemblyDefect),
            }

            if item_type not in model_map:
                return JsonResponse({"status": "error", "message": "Неверный тип"}, status=400)

            m2m_field, model = model_map[item_type]
            item = model.objects.filter(id=item_id).first()
            if not item:
                return JsonResponse({"status": "error", "message": "Объект не найден"}, status=404)

            if m2m_field.filter(id=item_id).exists():
                m2m_field.remove(item)
                return JsonResponse({"status": "removed"})
            else:
                m2m_field.add(item)
                return JsonResponse({"status": "added"})

        # обычные формы
        form_type = request.POST.get("form_type")
        if form_type == "zone":
            zone_form = AssemblyZoneForm(request.POST)
            if zone_form.is_valid():
                zone_form.save()
                messages.success(request, "Подсистема успешно добавлена.")
            else:
                messages.error(request, f"Ошибка при сохранении зоны: {zone_form.errors}")
        elif form_type == "unit":
            unit_form = AssemblyUnitForm(request.POST)
            if unit_form.is_valid():
                name = unit_form.cleaned_data['name']
                zone = unit_form.cleaned_data['zone']
                if not AssemblyUnit.objects.filter(name=name, zone=zone).exists():
                    unit_form.save()
                    messages.success(request, "Деталь успешно добавлена.")
                else:
                    messages.warning(request, "Такая деталь уже существует в выбранной зоне.")
            else:
                messages.error(request, f"Ошибка при сохранении детали: {unit_form.errors}")
        elif form_type == "defect":
            defect_form = AssemblyDefectForm(request.POST)
            if defect_form.is_valid():
                defect_form.save()
                messages.success(request, "Дефект успешно добавлен.")
            else:
                messages.error(request, f"Ошибка при сохранении дефекта: {defect_form.errors}")

        return redirect(request.get_full_path())

    # отображение
    zones = AssemblyZone.objects.all().order_by("name")
    units = AssemblyUnit.objects.all().order_by("name")
    defects = AssemblyDefect.objects.all().order_by("name")

    order = [
        "Пост момента затяжек",
        "Chassis",
        "Финал текущий контроль",
        "Зазоры и перепады",
        "Экстерьер",
        "Интерьер",
        "Багажник",
        "Мотор",
        "Функцонал",  # ← проверь точное написание как в БД
        "Геометрия колес",
        "Регулировка света фар и калибровка руля",
        "Тормозная система",
        "Underbody",
        "ADAS",
        "AVM",
        "Герметичность кузова",
        "Диагностика",
        "Тест трек",
        "Документация",
    ]

    order_case = Case(
        *[When(name=n, then=Value(i)) for i, n in enumerate(order)],
        default=Value(999),
        output_field=IntegerField(),
    )

    posts_qs = (
        PostAssembly.objects
        .filter(name__in=order)
        .annotate(order_idx=order_case)
        .order_by("order_idx")
    )
    posts = list(posts_qs)  # список, чтобы переиспользовать

    post_visibility = {
        post.id: {
            "zones": list(post.assembly_zones.values_list("id", flat=True)),
            "units": list(post.assembly_units.values_list("id", flat=True)),
            "defects": list(post.assembly_defects.values_list("id", flat=True)),
        }
        for post in posts
    }

    return render(request, "users/assembly/manage_post_visibility.html", {
        "posts": posts,
        "zones": zones,
        "units": units,
        "defects": defects,
        "zone_form": zone_form,
        "unit_form": unit_form,
        "defect_form": defect_form,
        "post_visibility_json": json.dumps(post_visibility, cls=DjangoJSONEncoder),
    })



def _is_ajax(request):
    return request.headers.get("x-requested-with") == "XMLHttpRequest"


# ===== ZONES =====
@login_required
@require_POST
def assembly_zone_update(request):
    pk = request.POST.get("id")
    name = (request.POST.get("name") or "").strip()
    zone = get_object_or_404(AssemblyZone, pk=pk)

    if not name:
        if _is_ajax(request):
            return JsonResponse({"ok": False, "message": "Название подсистемы не может быть пустым."})
        messages.error(request, "Название подсистемы не может быть пустым.")
        return redirect("manage_post_visibility")

    if AssemblyZone.objects.filter(name__iexact=name).exclude(pk=pk).exists():
        if _is_ajax(request):
            return JsonResponse({"ok": False, "message": f"Подсистема «{name}» уже существует."})
        messages.warning(request, f"Подсистема «{name}» уже существует.")
        return redirect("manage_post_visibility")

    zone.name = name
    zone.save(update_fields=["name"])
    if _is_ajax(request):
        return JsonResponse({"ok": True, "message": f"Подсистема «{name}» сохранена."})
    messages.success(request, f"Подсистема «{name}» сохранена.")
    return redirect("manage_post_visibility")


@login_required
@require_POST
def assembly_zone_delete(request):
    pk = request.POST.get("id")
    zone = get_object_or_404(AssemblyZone, pk=pk)
    name = zone.name
    zone.delete()
    if _is_ajax(request):
        return JsonResponse({"ok": True, "removed": True, "message": f"Подсистема «{name}» удалена."})
    messages.success(request, f"Подсистема «{name}» удалена.")
    return redirect("manage_post_visibility")


# ===== UNITS =====
@login_required
@require_POST
def assembly_unit_update(request):
    pk = request.POST.get("id")
    name = (request.POST.get("name") or "").strip()
    zone_id = request.POST.get("zone")
    unit = get_object_or_404(AssemblyUnit, pk=pk)

    if not name:
        if _is_ajax(request):
            return JsonResponse({"ok": False, "message": "Название детали не может быть пустым."})
        messages.error(request, "Название детали не может быть пустым.")
        return redirect("manage_post_visibility")

    zone = get_object_or_404(AssemblyZone, pk=zone_id) if zone_id else None

    if AssemblyUnit.objects.filter(zone=zone, name__iexact=name).exclude(pk=pk).exists():
        if _is_ajax(request):
            return JsonResponse({"ok": False, "message": f"Деталь «{name}» уже существует в выбранной подсистеме."})
        messages.warning(request, f"Деталь «{name}» уже существует в выбранной подсистеме.")
        return redirect("manage_post_visibility")

    unit.name = name
    unit.zone = zone
    unit.save(update_fields=["name", "zone"])
    if _is_ajax(request):
        return JsonResponse({"ok": True, "message": f"Деталь «{name}» сохранена."})
    messages.success(request, f"Деталь «{name}» сохранена.")
    return redirect("manage_post_visibility")


@login_required
@require_POST
def assembly_unit_delete(request):
    pk = request.POST.get("id")
    unit = get_object_or_404(AssemblyUnit, pk=pk)
    name = unit.name
    unit.delete()
    if _is_ajax(request):
        return JsonResponse({"ok": True, "removed": True, "message": f"Деталь «{name}» удалена."})
    messages.success(request, f"Деталь «{name}» удалена.")
    return redirect("manage_post_visibility")


# ===== DEFECTS =====
@login_required
@require_POST
def assembly_defect_update(request):
    pk = request.POST.get("id")
    name = (request.POST.get("name") or "").strip()
    defect = get_object_or_404(AssemblyDefect, pk=pk)

    if not name:
        if _is_ajax(request):
            return JsonResponse({"ok": False, "message": "Название дефекта не может быть пустым."})
        messages.error(request, "Название дефекта не может быть пустым.")
        return redirect("manage_post_visibility")

    if AssemblyDefect.objects.filter(name__iexact=name).exclude(pk=pk).exists():
        if _is_ajax(request):
            return JsonResponse({"ok": False, "message": f"Дефект «{name}» уже существует."})
        messages.warning(request, f"Дефект «{name}» уже существует.")
        return redirect("manage_post_visibility")

    defect.name = name
    defect.save(update_fields=["name"])
    if _is_ajax(request):
        return JsonResponse({"ok": True, "message": f"Дефект «{name}» сохранён."})
    messages.success(request, f"Дефект «{name}» сохранён.")
    return redirect("manage_post_visibility")


@login_required
@require_POST
def assembly_defect_delete(request):
    pk = request.POST.get("id")
    defect = get_object_or_404(AssemblyDefect, pk=pk)
    name = defect.name
    defect.delete()
    if _is_ajax(request):
        return JsonResponse({"ok": True, "removed": True, "message": f"Дефект «{name}» удалён."})
    messages.success(request, f"Дефект «{name}» удалён.")
    return redirect("manage_post_visibility")





# QRQC
# QRQC отчет отдела VQA

# assembly/views_qrqc_demo.py
from datetime import date
from typing import Any, Dict, List

from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.utils.dateparse import parse_date
from django.db.models import Q

from assembly.services import utils_qrqc as uq

from qrr.models import QRRResponsible


# --- Counters (VIN stats) ---
from assembly.services.utils_counter import (
    counter_vins,
    counter_vins_from_vehicle_history,
    counter_vins_documentation,
    POST_AREA_MAPPING,
)

def _fmt_prev5_for_chart(rows: List[dict]) -> Dict[str, Any]:
    """
    Форматирует 5 предыдущих недель для графика.
    Поддерживает totals (unique_vins/dpu/str) и averages (avg_unique_vins/avg_dpu/avg_str).
    Также возвращает метаданные недель для кликов: iso_year, iso_week, start (Пн ISO) в ISO формате.
    """
    labels, dpu, str_, uniq = [], [], [], []
    weeks_meta: List[Dict[str, Any]] = []
    for r in rows:
        # метки
        labels.append(f"Нед. {r['iso_week']:02d}")
        # значения (兼容 totals/avg)
        dpu.append(r.get("dpu", r.get("avg_dpu", 0)))
        str_.append(r.get("str", r.get("avg_str", 0)))
        uniq.append(r.get("unique_vins", r.get("avg_unique_vins", 0)))
        # метаданные недели
        weeks_meta.append({
            "iso_year": r.get("iso_year"),
            "iso_week": r.get("iso_week"),
            "start": (r.get("start").isoformat() if hasattr(r.get("start"), "isoformat") else r.get("start")),
        })
    return {
        "labels": labels,
        "series": {"dpu": dpu, "str": str_, "unique_vins": uniq},
        "weeks": weeks_meta,
    }


def _fmt_current_week_for_chart(rows: List[dict]) -> Dict[str, Any]:
    """
    qrqc_current_week_daily -> {labels, series:{dpu,str,unique_vins}}
    """
    labels, dpu, str_, uniq = [], [], [], []
    for r in rows:
        d: date = r["date"]  # type: ignore[assignment]
        labels.append(d.strftime("%d.%m"))
        dpu.append(r["dpu"])
        str_.append(r["str"])
        uniq.append(r["unique_vins"])
    return {"labels": labels, "series": {"dpu": dpu, "str": str_, "unique_vins": uniq}}


@login_required
@permission_required('users.access_to_the_qrqc_report_vqa', raise_exception=True)
def qrqc_dashboard_view(request):
    """
    6 графиков: Chery/GWM/Changan × (5 прошлых недель totals / текущая неделя daily)
    GET:
      ?date=YYYY-MM-DD
      ?grade=V1|V2|V3
      ?post=Экстерьер|Интерьер|...
      ?models=Tiggo%202%20PRO,Tiggo%207  (опционально, список через запятую)
    """
    # базовые фильтры
    raw_date = request.GET.get("date")
    target = parse_date(raw_date) if raw_date else date.today()

    grade = request.GET.get("grade") or None
    post = request.GET.get("post") or None
    models_raw = request.GET.get("models") or ""
    models = [m.strip() for m in models_raw.split(",") if m.strip()] or None

    # weekly mode (sum/avg) from query param
    weekly_mode = request.GET.get("weekly")
    if weekly_mode not in {"sum", "avg"}:
        weekly_mode = "sum"

    brands = ["chery", "gwm", "changan"]
    charts: Dict[str, Dict[str, Any]] = {}

    for b in brands:
        prev5_totals = uq.qrqc_prev5_weeks_totals(target, brand=b, models=models, grade=grade, post=post)
        prev5_avg = uq.qrqc_prev5_weeks_avg(target, brand=b, models=models, grade=grade, post=post)
        current_week = uq.qrqc_current_week_daily(target, brand=b, models=models, grade=grade, post=post)

        charts[b] = {
            "prev5_sum": _fmt_prev5_for_chart(prev5_totals),  # суммы за 5 недель
            "prev5_avg": _fmt_prev5_for_chart(prev5_avg),  # средние по активным дням
            "current": _fmt_current_week_for_chart(current_week),
        }

    ctx = {
        "target_date": target.isoformat(),
        "active_filters": {"grade": grade, "post": post, "models": models or []},
        "charts": charts,
        "weekly_mode_default": weekly_mode,  # 'sum' | 'avg'
        "series_meta": {          # подсказка фронту по типам рядов и осям
            "dpu": {"type": "line", "yAxis": "left"},
            "str": {"type": "line", "yAxis": "right"},
            "unique_vins": {"type": "bar", "yAxis": "bar"},
        },
        "titles": {
            "chery":   {"prev5": "Chery — прошлые 5 недель",   "current": "Chery — текущая неделя"},
            "gwm":     {"prev5": "GWM — прошлые 5 недель",     "current": "GWM — текущая неделя"},
            "changan": {"prev5": "Changan — прошлые 5 недель", "current": "Changan — текущая неделя"},
        },
        "brands": brands,
        "enable_particles": True,
        "enable_background": True,
        "enable_white_square": False,
    }
    return render(request, "users/qrqc/qrqc_dashboard.html", ctx)

# ===== Brand-focused QRQC page (charts + tables) =====
from typing import Optional, Sequence

# --- Period helpers for QRQC brand page ---
from datetime import timedelta as _timedelta
from dateutil.relativedelta import relativedelta as _relativedelta

def _norm_list_param(raw: Optional[str]) -> Optional[list[str]]:
    """
    Accepts comma-separated string or None; returns list[str] or None.
    """
    if not raw:
        return None
    items = [x.strip() for x in raw.split(",") if x.strip()]
    return items or None


def _fmt_posts_for_chart(rows: List[dict]) -> Dict[str, Any]:
    """
    uq.qrqc_by_posts(...) -> chart payload
    rows item: {"post","inspected","defects","offline_defects","dpu","str",...}
    """
    labels = [r["post"] for r in rows]
    series = {
        "unique_vins": [r.get("inspected", 0) for r in rows],
        "dpu": [r.get("dpu", 0.0) for r in rows],
        "str": [r.get("str", 0.0) for r in rows],
    }
    return {"labels": labels, "series": series}


def _fmt_grades_for_chart(rows: List[dict]) -> Dict[str, Any]:
    """
    uq.qrqc_defects_by_grade(...) -> chart payload
    rows item: {"grade","defects", "vins":[...]}
    """
    labels = [r["grade"] for r in rows]
    series = {
        "defects": [r.get("defects", 0) for r in rows],
        "unique_vins": [len(r.get("vins", [])) for r in rows],
    }
    return {"labels": labels, "series": series}



def _fmt_models_for_chart(rows: List[dict]) -> Dict[str, Any]:
    """
    uq.qrqc_by_models(...) -> chart payload
    rows item: {"model","inspected","defects","offline_defects","dpu","str",...}
    """
    labels = [r["model"] or "—" for r in rows]
    series = {
        "unique_vins": [r.get("inspected", 0) for r in rows],
        "dpu": [r.get("dpu", 0.0) for r in rows],
        "str": [r.get("str", 0.0) for r in rows],
    }
    return {"labels": labels, "series": series}


# --- DPU по ответственным (for charts) ---
def _fmt_dpu_responsibles_for_chart(rows: List[dict], all_names: List[str] | None = None) -> Dict[str, Any]:
    """
    Формирует payload для графика DPU по ответственным.
    rows: результат uq.qrqc_dpu_by_responsible() -> [{"responsible","defects_weight","inspected","dpu"}, ...]
    all_names: если задан список всех ответственных, добавляет отсутствующих с dpu=0 в конец (по алфавиту).
    Возвращает {labels:[...], series:{dpu:[...]}, meta:{active_count:int, all_count:int}}
    """
    # активные (имеют данные)
    active_map = { (r.get("responsible") or "").strip(): float(r.get("dpu", 0.0) or 0.0) for r in (rows or []) }
    active_names = [name for name in active_map.keys() if name]
    # подготовим базовые значения
    labels: List[str] = []
    dpu_vals: List[float] = []
    # По умолчанию — отрисовываем только активных в алфавитном порядке
    for name in sorted(active_names, key=lambda x: x.lower()):
        labels.append(name)
        dpu_vals.append(round(active_map.get(name, 0.0), 6))
    active_count = len(labels)

    # Если требуется добить до «всех» — добавляем отсутствующих с нулями
    if all_names is not None:
        # Нормализуем входные имена
        all_norm = sorted({ (n or "").strip() for n in all_names if (n or "").strip() }, key=lambda x: x.lower())
        # Добавим специальные группы из расчёта (если встречаются)
        if "(В ожидании назначения)" in active_map and "(В ожидании назначения)" not in all_norm:
            all_norm.append("(В ожидании назначения)")
        if "ALL" in active_map and "ALL" not in all_norm:
            all_norm.append("ALL")
        # Пройдёмся по всему списку и сформируем итоговые ряды
        labels_all: List[str] = []
        dpu_all: List[float] = []
        for name in all_norm:
            labels_all.append(name)
            dpu_all.append(round(active_map.get(name, 0.0), 6))
        return {
            "labels": labels_all,
            "series": {"dpu": dpu_all},
            "meta": {"active_count": active_count, "all_count": len(labels_all)},
        }

    return {"labels": labels, "series": {"dpu": dpu_vals}, "meta": {"active_count": active_count, "all_count": active_count}}


# ==== Extra helpers for ranged daily series → chart payloads ====
from collections import defaultdict as _dd
from datetime import date as _date


def _fmt_daily_for_chart(rows: List[dict]) -> Dict[str, Any]:
    """
    rows items: {"date": date, "unique_vins": int, "defects": int, "dpu": float, "str": float}
    → {labels:["dd.mm",...], series:{unique_vins:[], dpu:[], str:[]}}
    """
    labels, uniq, dpu, str_ = [], [], [], []
    for r in rows or []:
        d: _date = r.get("date")  # type: ignore[assignment]
        if hasattr(d, "strftime"):
            labels.append(d.strftime("%d.%m"))
        else:
            labels.append(str(d))
        uniq.append(r.get("unique_vins", 0) or 0)
        dpu.append(r.get("dpu", 0.0) or 0.0)
        str_.append(r.get("str", 0.0) or 0.0)
    return {"labels": labels, "series": {"unique_vins": uniq, "dpu": dpu, "str": str_}}


def _aggregate_days_to_weeks(rows: List[dict]) -> Dict[str, Any]:
    """
    Группирует дневной ряд по ISO-неделям. Для каждой недели:
      unique_vins = Σ unique_vins
      defects = Σ defects
      dpu = defects / unique_vins (0 если unique_vins==0)
      str = 100 - ( (Σ offline_vins_count) / (Σ unique_vins) * 100 )
    offline_vins_count восстанавливаем из дневных STR: uniq * (100 - str) / 100.
    Возвращает chart payload {labels, series:{unique_vins,dpu,str}} с метками "Нед. WW".
    """
    buckets = _dd(lambda: {"uniq": 0, "defects": 0, "offline_vins": 0.0, "iso_year": None, "iso_week": None})
    for r in rows or []:
        d: _date = r.get("date")  # type: ignore[assignment]
        if not hasattr(d, "isocalendar"):
            continue
        y, w, _ = d.isocalendar()
        key = (y, w)
        uniq = int(r.get("unique_vins", 0) or 0)
        defects = int(r.get("defects", 0) or 0)
        str_val = float(r.get("str", 0.0) or 0.0)
        offline = (uniq * max(0.0, min(100.0, 100.0 - str_val)))/100.0
        b = buckets[key]
        b["iso_year"], b["iso_week"] = y, w
        b["uniq"] += uniq
        b["defects"] += defects
        b["offline_vins"] += offline

    # сортировка по неделям (возрастание)
    ordered = sorted(buckets.values(), key=lambda x: (x["iso_year"], x["iso_week"]))
    labels, uniq_s, dpu_s, str_s = [], [], [], []
    for b in ordered:
        labels.append(f"Нед. {int(b['iso_week']):02d}")
        uniq = int(b["uniq"]) or 0
        defects = int(b["defects"]) or 0
        dpu = round(defects/uniq, 2) if uniq else 0.0
        str_val = 0.0 if uniq == 0 else max(0.0, min(100.0, round(100.0 - ((b["offline_vins"]/uniq)*100.0), 2)))
        uniq_s.append(uniq)
        dpu_s.append(dpu)
        str_s.append(str_val)
    return {"labels": labels, "series": {"unique_vins": uniq_s, "dpu": dpu_s, "str": str_s}}


essential_month_names = ["01","02","03","04","05","06","07","08","09","10","11","12"]

def _aggregate_days_to_months(rows: List[dict]) -> Dict[str, Any]:
    """
    Группирует дневной ряд по месяцам (YYYY-MM). Математика как в неделях.
    Возвращает payload {labels:['YYYY-MM',...], series:{unique_vins,dpu,str}}.
    """
    buckets = _dd(lambda: {"uniq": 0, "defects": 0, "offline_vins": 0.0})
    for r in rows or []:
        d: _date = r.get("date")  # type: ignore[assignment]
        if not hasattr(d, "strftime"):
            continue
        ym = d.strftime("%Y-%m")
        uniq = int(r.get("unique_vins", 0) or 0)
        defects = int(r.get("defects", 0) or 0)
        str_val = float(r.get("str", 0.0) or 0.0)
        offline = (uniq * max(0.0, min(100.0, 100.0 - str_val)))/100.0
        b = buckets[ym]
        b["uniq"] += uniq
        b["defects"] += defects
        b["offline_vins"] += offline

    ordered_keys = sorted(buckets.keys())
    labels, uniq_s, dpu_s, str_s = [], [], [], []
    for k in ordered_keys:
        b = buckets[k]
        uniq = int(b["uniq"]) or 0
        defects = int(b["defects"]) or 0
        dpu = round(defects/uniq, 2) if uniq else 0.0
        str_val = 0.0 if uniq == 0 else max(0.0, min(100.0, round(100.0 - ((b["offline_vins"]/uniq)*100.0), 2)))
        labels.append(k)
        uniq_s.append(uniq)
        dpu_s.append(dpu)
        str_s.append(str_val)
    return {"labels": labels, "series": {"unique_vins": uniq_s, "dpu": dpu_s, "str": str_s}}



def _iso_week_bounds(_d: _date) -> tuple[_date, _date]:
    """
    Returns Monday..Sunday (inclusive) of the ISO week containing _d.
    """
    # Monday=0 .. Sunday=6
    wd = _d.weekday()
    start = _d - _timedelta(days=wd)
    end = start + _timedelta(days=6)
    return start, end

# --- New helpers for explicit ISO week selection ---
import re as _re

def _iso_week_bounds_yw(iso_year: int, iso_week: int) -> tuple[_date, _date]:
    """Return Monday..Sunday for a given ISO year/week."""
    # Find Monday of week 1
    jan4 = _date(iso_year, 1, 4)
    # Monday of the first ISO week
    week1_monday = jan4 - _timedelta(days=jan4.weekday())
    start = week1_monday + _timedelta(weeks=iso_week - 1)
    end = start + _timedelta(days=6)
    return start, end

def _parse_iso_week_param(request) -> tuple[int | None, int | None]:
    """Accepts ?iso=YYYY-Www or ?iso=YYYYWww or ?iso_year=&iso_week=; returns (year, week) or (None, None)."""
    raw = (request.GET.get("iso") or "").strip()
    if raw:
        m = _re.match(r"^(\d{4})-?[Ww](\d{1,2})$", raw)
        if m:
            try:
                y = int(m.group(1)); w = int(m.group(2))
                return y, w
            except Exception:
                pass
    y_raw = (request.GET.get("iso_year") or "").strip()
    w_raw = (request.GET.get("iso_week") or "").strip()
    if y_raw and w_raw and y_raw.isdigit() and w_raw.isdigit():
        try:
            return int(y_raw), int(w_raw)
        except Exception:
            return None, None
    return None, None


def _normalize_top_rows(rows: List[dict]) -> List[dict]:
    """
    Приводит строки TOP‑таблиц к единому формату:
      1) detail: str
      2) defect: str
      3) count: int
      4) grade: str
      5) responsibles: dict[str,float]
      6) defect_ids: list[str]
      7) vins: list[str]
    """
    normalized: List[dict] = []
    for r in (rows or []):
        try:
            normalized.append({
                "detail": r.get("detail") or "",
                "defect": r.get("defect") or "",
                "count": int(r.get("count", 0) or 0),
                "grade": r.get("grade") or "",
                "responsibles": r.get("responsibles") or {},
                "defect_ids": list(r.get("defect_ids") or []),
                "vins": list(r.get("vins") or []),
            })
        except Exception:
            # На всякий случай — пропускаем битую строку
            continue
    return normalized

@login_required
@permission_required('users.access_to_the_qrqc_report_vqa', raise_exception=True)
def qrqc_brand_view(request, brand=None):
    """
    Брендовая страница QRQC с наборами графиков и двумя таблицами:
      - график по неделям (суммарно) для бренда
      - график по дням (текущая ISO‑неделя target_date)
      - график по постам (DPU/STR/Проверено)
      - график по грейдам (кол-во дефектов, уникальные VIN)
      - график по моделям (DPU/STR/Проверено)
      - таблица «ТОП дефектов по грейдам»
      - таблица «ТОП дефектов по массовости»
    Поддерживаемые GET‑параметры (все опционально):
      ?brand=gwm|chery|changan
      ?date=YYYY-MM-DD    (если не задано start/end)
      ?start=YYYY-MM-DD&amp;end=YYYY-MM-DD
      ?models=A,B,C
      ?posts=P1,P2
      ?grades=V1,V2,V3
      ?weekly=sum|avg     (режим для «пред. 5 недель»: totals/averages)
    Примечание: недельные/дневные графики принимают только одиночные post/grade.
    Если выбрано несколько — в эти графики пост/грейд не применяются (берутся все).
    """
    # ---- Parse filters
    brand = ((request.GET.get("brand") or brand or "").strip().lower() or None)
    raw_date = request.GET.get("date")
    raw_start = request.GET.get("start")
    raw_end = request.GET.get("end")

    target = parse_date(raw_date) if raw_date else date.today()
    start_d = parse_date(raw_start) if raw_start else None
    end_d = parse_date(raw_end) if raw_end else None
    if raw_date:  # приоритет одиночного дня
        start_d = end_d = None

    # --- Parse ISO week explicit params ---
    iso_y, iso_w = _parse_iso_week_param(request)

    # Range & aggregation (affects ALL breakdown charts/tables when start/end are not provided)
    range_param = (request.GET.get("range") or "").strip().lower()
    agg_param   = (request.GET.get("agg") or "").strip().lower()  # reserved for UI, not used in backend math here

    # If an explicit ISO week is passed – use its Monday..Sunday as the working period
    if iso_y and iso_w:
        start_d, end_d = _iso_week_bounds_yw(iso_y, iso_w)
        # anchor target inside selected week to keep daily chart labels stable
        target = start_d
    elif start_d is None and end_d is None:
        # No explicit period – use range presets
        if range_param == "day":
            # current ISO week for the target date
            wn_start, wn_end = _iso_week_bounds(target)
            start_d, end_d = wn_start, wn_end
        elif range_param == "week":
            # last 5 ISO weeks including week of target
            wn_start, wn_end = _iso_week_bounds(target)
            start_d = wn_start - _timedelta(weeks=4)
            end_d = wn_end
        elif range_param == "month":
            start_d = target - _timedelta(days=29)
            end_d = target
        elif range_param == "halfyear":
            start_d = target - _relativedelta(months=6)
            end_d = target
        elif range_param == "year":
            start_d = target - _relativedelta(years=1)
            end_d = target
        else:
            # keep None – single-day mode (target_date only)
            pass

    models = _norm_list_param(request.GET.get("models"))
    posts = _norm_list_param(request.GET.get("posts"))
    grades = _norm_list_param(request.GET.get("grades"))

    weekly_mode = request.GET.get("weekly")
    if weekly_mode not in {"sum", "avg"}:
        weekly_mode = "sum"

    # Для недельных/дневных API — одиночные значения
    post_single = posts[0] if posts and len(posts) == 1 else None
    grade_single = grades[0] if grades and len(grades) == 1 else None

    # ---- Data assembly
    charts: Dict[str, Any] = {}

    # weekly totals / averages for brand
    prev5_totals = uq.qrqc_prev5_weeks_totals(target, brand=brand, models=models, grade=grade_single, post=post_single)
    prev5_avg = uq.qrqc_prev5_weeks_avg(target, brand=brand, models=models, grade=grade_single, post=post_single)
    # daily current-week chart: use selected ISO week if provided
    if iso_y and iso_w:
        try:
            current_week = uq.qrqc_week_daily(iso_y, iso_w, brand=brand, models=models, grade=grade_single, post=post_single)
        except Exception:
            current_week = uq.qrqc_current_week_daily(target, brand=brand, models=models, grade=grade_single, post=post_single)
    else:
        current_week = uq.qrqc_current_week_daily(target, brand=brand, models=models, grade=grade_single, post=post_single)

    charts["prev5_sum"] = _fmt_prev5_for_chart(prev5_totals)
    charts["prev5_avg"] = _fmt_prev5_for_chart(prev5_avg)
    charts["current"] = _fmt_current_week_for_chart(current_week)

    # posts breakdown (respects multi posts/grades)
    posts_rows = uq.qrqc_by_posts(
        target_date=target if (not start_d and not end_d) else None,
        start_d=start_d, end_d=end_d,
        brand=brand, models=models, grades=grades, posts=posts,
    )
    charts["by_posts"] = _fmt_posts_for_chart(posts_rows)

    # grades breakdown
    grades_rows = uq.qrqc_defects_by_grade(
        target_date=target if (not start_d and not end_d) else None,
        start_d=start_d, end_d=end_d,
        brand=brand, models=models, posts=posts, grades=grades,
    )
    charts["by_grades"] = _fmt_grades_for_chart(grades_rows)

    # models breakdown
    models_rows = uq.qrqc_by_models(
        target_date=target if (not start_d and not end_d) else None,
        start_d=start_d, end_d=end_d,
        brand=brand, models=models, grades=grades, posts=posts,
    )
    charts["by_models"] = _fmt_models_for_chart(models_rows)

    # ===== RANGED DAILY SERIES (month / halfyear / year) =====
    try:
        month_daily = uq.qrqc_last_month_daily(target, brand=brand, models=models, grade=grade_single, post=post_single)
    except Exception:
        month_daily = []
    try:
        halfyear_daily = uq.qrqc_last_halfyear_daily(target, brand=brand, models=models, grade=grade_single, post=post_single)
    except Exception:
        halfyear_daily = []
    try:
        year_daily = uq.qrqc_last_year_daily(target, brand=brand, models=models, grade=grade_single, post=post_single)
    except Exception:
        year_daily = []

    # For the UI: provide all aggregations explicitly so toggle (Днями/Неделями/Месяцами) works
    charts["month_by_days"]    = _fmt_daily_for_chart(month_daily)
    charts["month_by_weeks"]   = _aggregate_days_to_weeks(month_daily)
    charts["month_by_months"]  = _aggregate_days_to_months(month_daily)

    charts["halfyear_by_days"]   = _fmt_daily_for_chart(halfyear_daily)
    charts["halfyear_by_weeks"]  = _aggregate_days_to_weeks(halfyear_daily)
    charts["halfyear_by_months"] = _aggregate_days_to_months(halfyear_daily)

    charts["year_by_days"]   = _fmt_daily_for_chart(year_daily)
    charts["year_by_weeks"]  = _aggregate_days_to_weeks(year_daily)
    charts["year_by_months"] = _aggregate_days_to_months(year_daily)

    # tables
    top_by_grades = uq.qrqc_top_defects_by_grades(
        target_date=target if (not start_d and not end_d) else None,
        start_d=start_d, end_d=end_d,
        brand=brand, models=models, posts=posts, grades=grades,
    )
    top_by_mass = uq.qrqc_top_defects_by_mass(
        target_date=target if (not start_d and not end_d) else None,
        start_d=start_d, end_d=end_d,
        brand=brand, models=models, posts=posts, grades=grades,
    )
    # Нормализация в требуемый формат (7 полей)
    top_by_grades = _normalize_top_rows(top_by_grades)
    top_by_mass = _normalize_top_rows(top_by_mass)

    # --- DPU по ответственным ---
    dpu_resp_rows = uq.qrqc_dpu_by_responsible(
        target_date=target if (not start_d and not end_d) else None,
        start_d=start_d, end_d=end_d,
        brand=brand, models=models, posts=posts, grades=grades,
    )
    # все ответственные из справочника QRR
    try:
        all_resp_names = list(QRRResponsible.objects.order_by("name").values_list("name", flat=True))
    except Exception:
        all_resp_names = []
    # график только по активным (есть данные)
    chart_resp_active = _fmt_dpu_responsibles_for_chart(dpu_resp_rows)
    # график по всем из справочника (отсутствующие = 0)
    chart_resp_all = _fmt_dpu_responsibles_for_chart(dpu_resp_rows, all_names=all_resp_names)

    # агрегат ALL (все дефекты / все уникальные VIN за период)
    agg_all = next((r for r in dpu_resp_rows if (r.get("responsible") or "").upper() == "ALL"), None)
    dpu_all_value = float(agg_all.get("dpu", 0.0)) if agg_all else 0.0
    inspected_total = int(agg_all.get("inspected", 0) or 0) if agg_all else 0
    defects_weight_total = float(agg_all.get("defects_weight", 0.0) or 0.0) if agg_all else 0.0

    # ===== KTV (Отложенные дефекты) — отправляем в шаблон только видимые записи =====
    # Правило видимости: visible_from <= сегодня. Если выбраны грейды, отфильтруем по ним.
    today_ktv = timezone.localdate()
    ktv_qs = KTVDefect.objects.filter(visible_from__lte=today_ktv)

    # Если есть фильтр по грейдам — показываем только эти грейды (и пустые)
    if grades:
        ktv_qs = ktv_qs.filter(Q(grade__in=grades) | Q(grade__isnull=True) | Q(grade=""))

    ktv_by_grades_list = ktv_qs.filter(table_type="by_grades").order_by("-visible_from", "-created_at")
    ktv_by_mass_list   = ktv_qs.filter(table_type="by_mass").order_by("-visible_from", "-created_at")

    # Приводим к диктам для удобного рендера в шаблоне
    ktv_by_grades_ctx = [_ktv_to_dict(o) for o in ktv_by_grades_list]
    ktv_by_mass_ctx   = [_ktv_to_dict(o) for o in ktv_by_mass_list]

    ctx = {
        "brand": brand,
        "target_date": target.isoformat(),
        "start": (start_d.isoformat() if start_d else ""),
        "end": (end_d.isoformat() if end_d else ""),
        "iso_year": (iso_y or ""),
        "iso_week": (iso_w or ""),
        "range": range_param or "",
        "agg": agg_param or "",
        "filters": {
            "models": models or [],
            "posts": posts or [],
            "grades": grades or [],
            "weekly_mode": weekly_mode,
        },

        # графики
        "charts": charts,

        # таблицы
        "table_top_by_grades": top_by_grades,
        "table_top_by_mass": top_by_mass,

        # KTV: данные для брендовой страницы
        "ktv_by_grades": ktv_by_grades_ctx,
        "ktv_by_mass": ktv_by_mass_ctx,

        # график DPU по ответственным
        "dpu_responsibles_active": chart_resp_active,
        "dpu_responsibles_all": chart_resp_all,
        "responsibles_active_count": chart_resp_active.get("meta", {}).get("active_count", 0),
        "responsibles_all_count": chart_resp_all.get("meta", {}).get("all_count", 0),
        # агрегат ALL и знаменатель
        "dpu_all": dpu_all_value,
        "dpu_all_inspected": inspected_total,
        "dpu_all_defects_weight": defects_weight_total,

        # подсказка фронту по типам рядов и осям
        "series_meta": {
            # универсальный набор для комбо: колонка (unique_vins) + линии (dpu/str)
            "dpu": {"type": "line", "yAxis": "left"},
            "str": {"type": "line", "yAxis": "right"},
            "unique_vins": {"type": "bar", "yAxis": "bar"},
            "defects": {"type": "bar", "yAxis": "bar"},
        },

        # заголовки для шаблона (можете поменять на свои)
        "titles": {
            "prev5": "Прошлые 5 недель",
            "current": "Текущая неделя (по дням)",
            "by_posts": "По постам (DPU / STR / Проверено)",
            "by_grades": "По грейдам",
            "by_models": "По моделям (DPU / STR / Проверено)",
            "table_top_by_grades": "ТОП дефектов по грейдам",
            "table_top_by_mass": "ТОП дефектов по массовости",
        },

        # для переключателя totals/avg на шаблоне
        "weekly_mode_default": weekly_mode,
        "enable_particles": True,
        "enable_background": True,
        "enable_white_square": False,
    }

    return render(request, "users/qrqc/qrqc_brand_dashboard.html", ctx)


@login_required
@permission_required('users.access_to_the_qrqc_report_vqa', raise_exception=True)
def defect_details_view(request):
    """
    Детальная страница дефектов VIN.
    Поддерживает один или несколько дефектов.
    Query params:
      - vin=VIN или VIN1,VIN2 (опционально; если не задан, VIN извлекается из defect_id)
      - defect_id=ID  (можно перечислить через запятую, или повторить параметр несколько раз)
      - defect_ids=ID1,ID2,... (альтернативно)
    Шаблон ожидает контекст {"items": list[dict]} в формате карточек QRR.
    """
    from collections import defaultdict
    import re
    from django.http import HttpResponseBadRequest, HttpResponseNotFound
    from django.utils.dateparse import parse_datetime
    from django.utils import timezone

    # --- VIN param: допускаем CSV, пробелы
    raw_vin = (request.GET.get("vin") or "").strip()
    vin_list = [v.strip() for v in raw_vin.split(",") if v.strip()] if raw_vin else []

    # --- Собираем все варианты передачи id-шек
    ids: list[str] = []
    # 1) повторяющиеся defect_id
    ids += [x.strip() for x in request.GET.getlist("defect_id") if (x or "").strip()]
    # 2) одиночный defect_id, но в нём может быть CSV
    single = (request.GET.get("defect_id") or "").strip()
    if single and "," in single:
        ids += [x.strip() for x in single.split(",") if x.strip()]
    # 3) альтернативный параметр defect_ids=csv
    csv_ids = (request.GET.get("defect_ids") or "").strip()
    if csv_ids:
        ids += [x.strip() for x in csv_ids.split(",") if x.strip()]

    # Уникализуем, сохранив порядок
    seen = set()
    ids = [x for x in ids if not (x in seen or seen.add(x))]

    if not ids:
        return HttpResponseBadRequest("defect_id(ы) обязательны")

    # --- Разбивка id по VIN ---
    def _vin_from_defect_id(did: str) -> str | None:
        # ожидаем шаблон: VIN-цех-сборки-...
        m = re.match(r"^([A-Z0-9]+)-цех-сборки-", str(did), re.IGNORECASE)
        return m.group(1) if m else None

    vin_to_ids: dict[str, list[str]] = defaultdict(list)

    if len(vin_list) == 1:
        # Если задан один VIN — используем его по умолчанию, но если из id извлекли другой VIN, сгруппируем по извлечённому
        single_vin = vin_list[0]
        for did in ids:
            v = _vin_from_defect_id(did) or single_vin
            vin_to_ids[v].append(did)
    else:
        # VIN не задан или их несколько — пробуем извлекать из каждого id
        for did in ids:
            v = _vin_from_defect_id(did)
            if v:
                vin_to_ids[v].append(did)
        # если ни у одного id VIN не распознан, а vin_list задан (много VIN) — распределим первый VIN
        if not vin_to_ids and vin_list:
            for did in ids:
                vin_to_ids[vin_list[0]].append(did)

    if not vin_to_ids:
        return HttpResponseBadRequest("Не удалось определить VIN по переданным defect_id")

    # --- Получаем детали по группам VIN
    details_list: list[dict] = []
    try:
        if hasattr(uq, "get_defect_details_many"):
            for v, id_list in vin_to_ids.items():
                chunk = uq.get_defect_details_many(v, id_list) or []
                if chunk:
                    details_list.extend(chunk)
        else:
            # Фоллбэк: по одному id
            for v, id_list in vin_to_ids.items():
                for did in id_list:
                    item = uq.get_defect_details(v, did)
                    if item:
                        details_list.append(item)
    except Exception:
        details_list = []

    if not details_list:
        return HttpResponseNotFound("Дефекты не найдены")

    # Сортировка: по VIN, затем по дате (если есть)
    def _key(d: dict):
        vinv = d.get("vin") or ""
        dt = d.get("date")
        if not dt:
            raw = d.get("date_added") or d.get("date_str") or ""
            dt = parse_datetime(raw) if raw else None
        if dt is None:
            # минимально возможная aware-дата, чтобы такие элементы шли первыми
            try:
                return (vinv, timezone.make_aware(timezone.datetime.min))
            except Exception:
                return (vinv, timezone.now())
        return (vinv, timezone.localtime(dt) if timezone.is_aware(dt) else timezone.make_aware(dt, timezone.get_current_timezone()))

    details_list.sort(key=_key)

    # Если в запросе явно был один VIN — пробросим его, иначе оставим пустым (шаблону он не обязателен)
    vin_ctx = vin_list[0] if len(vin_list) == 1 else ""

    return render(request, "users/qrqc/defect_details.html", {"vin": vin_ctx, "items": details_list})


def _ktv_to_dict(o: KTVDefect):
    return {
        "id": o.id,
        "table_type": o.table_type,         # 'by_grades' | 'by_mass'
        "detail": o.detail,
        "defect": o.defect,
        "grade": o.grade,
        "count": o.count,
        "visible_from": o.visible_from.isoformat(),
        "created_at": o.created_at.isoformat(),
        "created_by": (o.created_by.get_full_name() or o.created_by.username) if o.created_by else None,
        "comment": o.comment or "",
    }


@login_required
def ktvdefect_list(request):
    """
    GET /api/ktvdefect/list
    Возвращает только видимые записи (visible_from <= сегодня), сгруппировано по типу.
    Параметры (опционально):
      ?from=YYYY-MM-DD  — нижняя граница видимости
      ?limit=10         — ограничение на количество в каждой группе
    """
    if request.method != "GET":
        return HttpResponseNotAllowed(["GET"])

    today = timezone.localdate()
    from_date = request.GET.get("from")
    try:
        from_date = parse_date(from_date) if from_date else None
    except Exception:
        from_date = None

    base_qs = KTVDefect.objects.filter(visible_from__lte=today)
    if from_date:
        base_qs = base_qs.filter(visible_from__gte=from_date)

    limit = request.GET.get("limit")
    try:
        limit = int(limit) if limit else None
    except Exception:
        limit = None

    by_grades = base_qs.filter(table_type="by_grades").order_by("-visible_from", "-created_at")
    by_mass   = base_qs.filter(table_type="by_mass").order_by("-visible_from", "-created_at")

    if limit:
        by_grades = by_grades[:limit]
        by_mass   = by_mass[:limit]

    data = {
        "by_grades": [_ktv_to_dict(o) for o in by_grades],
        "by_mass":   [_ktv_to_dict(o) for o in by_mass],
    }
    return JsonResponse(data, json_dumps_params={"ensure_ascii": False})


@login_required
@require_POST
def ktvdefect_upsert(request):
    """
    POST /api/ktvdefect/upsert  (JSON)
    Тело:
      {
        "table_type": "by_grades" | "by_mass",
        "detail": "string",
        "defect": "string",
        "grade": "string|null",
        "count": 123,                # новое значение (по умолчанию 0)
        "visible_from": "YYYY-MM-DD",
        "comment": "строка (опц.)"
      }
    Логика:
      - ключ уникальности: (table_type, detail, defect, grade)
      - если запись есть — обновляем count, visible_from и при необходимости дополняем comment.
      - если нет — создаём.
    """
    import json

    try:
        payload = json.loads(request.body.decode("utf-8"))
    except Exception:
        return HttpResponseBadRequest("Invalid JSON")

    table_type = payload.get("table_type")
    detail     = (payload.get("detail") or "").strip()
    defect     = (payload.get("defect") or "").strip()
    grade      = (payload.get("grade") or None)
    count      = payload.get("count", 0)
    visible_s  = payload.get("visible_from")
    comment    = (payload.get("comment") or "").strip()

    if table_type not in ("by_grades", "by_mass"):
        return HttpResponseBadRequest("table_type must be 'by_grades' or 'by_mass'")
    if not detail or not defect:
        return HttpResponseBadRequest("detail and defect are required")

    vis_date = parse_date(visible_s) if visible_s else None
    if not vis_date:
        vis_date = timezone.localdate()  # по умолчанию — сегодня

    obj, created = KTVDefect.objects.get_or_create(
        table_type=table_type,
        detail=detail,
        defect=defect,
        grade=grade if grade else None,
        defaults={
            "count": int(count) if isinstance(count, int) else 0,
            "visible_from": vis_date,
            "comment": comment,
            "created_by": request.user,
        }
    )

    # Если не новая — обновляем
    if not created:
        # политика обновления: перезаписываем count и видимость; комментарий дописываем с подписью автора
        obj.count = int(count) if isinstance(count, int) else obj.count
        obj.visible_from = vis_date
        if comment:
            stamp = timezone.now().strftime("%Y-%m-%d %H:%M")
            who = (request.user.get_full_name() or request.user.username)
            prev = (obj.comment or "").strip()
            addition = f"[{stamp} · {who}] {comment}"
            obj.comment = (prev + "\n" + addition).strip() if prev else addition
        if not obj.created_by:
            obj.created_by = request.user
        obj.save(update_fields=["count", "visible_from", "comment", "created_by"])

    return JsonResponse({"ok": True, "created": created, "item": _ktv_to_dict(obj)}, json_dumps_params={"ensure_ascii": False})


@login_required
@require_http_methods(["POST", "DELETE"])
def ktvdefect_delete(request, pk: int | None = None):
    """
    Удаляет запись KTV.
    Поддерживаемые варианты вызова:
      - POST /api/ktvdefect/delete/           с телом JSON {"id": 123} или формой id=123
      - DELETE /api/ktvdefect/delete/?id=123
      - DELETE /api/ktvdefect/<id>/delete/  (если настроен маршрут с pk)
    Требует права staff.
    """
    # Разрешаем удаление тем, у кого есть доступ:
    # - superuser / staff
    # - явное право users.delete_ktvdefect
    # - роли из нашего проекта: head_area, master
    user_role = getattr(request.user, "role", None)
    if not (
        request.user.is_superuser
        or request.user.is_staff
        or request.user.has_perm("users.delete_ktvdefect")
        or user_role in ("head_area", "master")
    ):
        return JsonResponse({"ok": False, "error": "forbidden"}, status=403)

    # Извлечь идентификатор из разных мест
    item_id = pk
    if item_id is None:
        # Пробуем JSON тело
        try:
            import json
            if request.body:
                data = json.loads(request.body.decode("utf-8"))
                item_id = data.get("id") or data.get("pk")
        except Exception:
            item_id = None

    if item_id is None:
        # Пробуем form-data
        item_id = request.POST.get("id") or request.POST.get("pk")

    if item_id is None:
        # Пробуем querystring (для DELETE)
        item_id = request.GET.get("id") or request.GET.get("pk")

    # Нормализуем к int
    try:
        item_id = int(item_id) if item_id is not None else None
    except (TypeError, ValueError):
        item_id = None

    if item_id is None:
        return JsonResponse({"ok": False, "error": "id required"}, status=400)

    # Удаление
    try:
        obj = KTVDefect.objects.get(pk=item_id)
    except KTVDefect.DoesNotExist:
        return JsonResponse({"ok": False, "error": "not found"}, status=404)

    obj.delete()
    return JsonResponse({"ok": True, "deleted_id": item_id})







# @login_required
# @role_required(["master", "head_area"])
# def export_all_defects_excel(request):
#     start_date_str = request.GET.get("start_date")
#     end_date_str = request.GET.get("end_date")
#     brand_filter = request.GET.get("download_brand")
#
#     try:
#         if start_date_str:
#             start_date = make_aware(datetime.combine(datetime.strptime(start_date_str, "%Y-%m-%d").date(), time.min))
#         else:
#             start_date = None
#
#         if end_date_str:
#             end_date = make_aware(datetime.combine(datetime.strptime(end_date_str, "%Y-%m-%d").date(), time.max))
#         else:
#             end_date = None
#     except ValueError:
#         return HttpResponse("❌ Неверный формат даты", status=400)
#
#     vin_histories = VINHistory.objects.all()
#     data_rows = []
#     max_photos = 0
#
#     for history in vin_histories:
#         vin = getattr(history, 'vin', '')
#         trace = TraceData.objects.filter(vin_rk=vin).first()
#         if not trace:
#             continue
#         if brand_filter and getattr(trace, 'brand', '') != brand_filter:
#             continue
#
#         brand = getattr(trace, 'brand', '')
#         model = getattr(trace, 'model', '')
#         config_code = getattr(trace, 'config_code', '')
#
#         history_data = getattr(history, 'history', {})
#
#         for zone_name, posts in history_data.items():
#             for post_name, entries in posts.items():
#                 for entry in entries:
#                     try:
#                         date_raw = entry.get("date_added")
#                         if not date_raw:
#                             continue
#                         date = parse_datetime(date_raw)
#                         if start_date and date < start_date:
#                             continue
#                         if end_date and date > end_date:
#                             continue
#
#                         controller = entry.get("controller", "") or entry.get("extra_data", {}).get("controller", "")
#                         line = entry.get("line", "")
#                         duration = entry.get("inspection_duration_seconds", "")
#
#                         if "defects" in entry and isinstance(entry["defects"], list) and entry["defects"]:
#                             for defect in entry["defects"]:
#                                 photos = defect.get("photos", [])
#                                 max_photos = max(max_photos, len(photos))
#                                 data_rows.append({
#                                     "zone": zone_name or "",
#                                     "post": post_name or "",
#                                     "vin": vin or "",
#                                     "brand": brand or "",
#                                     "model": model or "",
#                                     "config_code": config_code or "",
#                                     "date": date or "",
#                                     "line": line or "",
#                                     "unit": defect.get("unit", ""),
#                                     "defect": defect.get("name", ""),
#                                     "comment": defect.get("comment", ""),
#                                     "quantity": defect.get("quantity", ""),
#                                     "grade": defect.get("grade", ""),
#                                     "responsible": defect.get("responsible", ""),
#                                     "controller": controller or "",
#                                     "duration": duration or "",
#                                     "photos": photos,
#                                 })
#                         else:
#                             photos = entry.get("defect_photos", [])
#                             max_photos = max(max_photos, len(photos))
#                             data_rows.append({
#                                 "zone": zone_name or "",
#                                 "post": post_name or "",
#                                 "vin": vin or "",
#                                 "brand": brand or "",
#                                 "model": model or "",
#                                 "config_code": config_code or "",
#                                 "date": date or "",
#                                 "line": line or "",
#                                 "unit": "",
#                                 "defect": entry.get("defect_description", ""),
#                                 "comment": "",
#                                 "quantity": "",
#                                 "grade": "",
#                                 "responsible": "",
#                                 "controller": controller or "",
#                                 "duration": duration or "",
#                                 "photos": photos,
#                             })
#                     except Exception:
#                         continue
#
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Все дефекты"
#
#     headers = [
#         "Участок", "Пост", "VIN", "Бренд", "Модель", "Код комплектации", "Дата", "Линия",
#         "Деталь", "Дефект", "Комментарий", "Количество", "Грейд", "Кто виноват",
#         "Контроллер", "Время осмотра (сек)"
#     ] + [f"Фото {i+1}" for i in range(max_photos)]
#
#     ws.append(headers)
#
#     for row in data_rows:
#         try:
#             excel_row = [
#                 row.get("zone", ""),
#                 row.get("post", ""),
#                 row.get("vin", ""),
#                 row.get("brand", ""),
#                 row.get("model", ""),
#                 row.get("config_code", ""),
#                 row.get("date", "").strftime("%d.%m.%Y %H:%M") if row.get("date") else "",
#                 row.get("line", ""),
#                 row.get("unit", ""),
#                 row.get("defect", ""),
#                 row.get("comment", ""),
#                 row.get("quantity", ""),
#                 row.get("grade", ""),
#                 row.get("responsible", ""),
#                 row.get("controller", ""),
#                 row.get("duration", ""),
#             ]
#             excel_row += ["" for _ in range(max_photos)]
#             ws.append(excel_row)
#
#             row_idx = ws.max_row
#             for i, img_path in enumerate(row.get("photos", [])):
#                 insert_single_image(ws, row_idx, 17 + i, img_path)
#             ws.row_dimensions[row_idx].height = 75
#         except Exception:
#             continue
#
#     for col in ws.columns:
#         for cell in col:
#             cell.alignment = Alignment(wrap_text=True, vertical='top')
#
#     response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     filename = f"defects_summary_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
#     response['Content-Disposition'] = f'attachment; filename="{filename}"'
#     wb.save(response)
#     return response






# @login_required
# @role_required(["head_area"])
# def assembly_general_report_export(request):
#     start_date = request.GET.get("start_date")
#     end_date = request.GET.get("end_date")
#     brands = request.GET.getlist("brand")
#     models = request.GET.getlist("model")
#
#     # Используем тот же расчёт, что в assembly_general_report (можно вынести в отдельную функцию)
#     report_data = generate_report_data(start_date, end_date, brands, models)
#
#     # Генерация Excel
#     file_path = generate_excel_from_report(report_data, start_date, end_date, brands)
#
#     # Отдаем файл
#     with open(file_path, 'rb') as f:
#         response = HttpResponse(f.read(),
#                                 content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#         response['Content-Disposition'] = f'attachment; filename="report_{start_date}_{end_date}.xlsx"'
#         return response

@login_required
@permission_required('users.access_to_indicators', raise_exception=True)
def assembly_counter_dashboard(request):
    """
    Страница с тремя метриками и списками VIN:
      1) Пост-счётчик (AssemblyPassLog)
      2) Все посты от Первой инспекции до Финала (см. POST_AREA_MAPPING)
      3) Пост «Документация»
    Фильтры берём из TraceData: brand (мульти, через маппинг), model (мульти), start_date, end_date (YYYY-MM-DD)
    """
    # --- Параметры фильтров (как в assembly_general_report) ---
    brands_selected = request.GET.getlist("brand")  # ['gwm','chery',...]
    models_selected = request.GET.getlist("model")

    BRAND_MAP = {
        "gwm": ["haval", "tank"],
        "chery": ["chery"],
        "changan": ["changan"],
    }
    # нормализуем к нижнему регистру для сравнения в БД
    brands_selected = [b.strip().lower() for b in brands_selected if b.strip()]
    mapped_brands = list(chain.from_iterable(BRAND_MAP.get(b, [b]) for b in brands_selected))

    # Даты: если не заданы — «сегодня..сегодня»
    start_date = (request.GET.get("start_date") or "").strip() or None
    end_date = (request.GET.get("end_date") or "").strip() or None
    if not start_date and not end_date:
        today = timezone.localdate()
        start_date = today.isoformat()
        end_date = today.isoformat()

    # --- База фильтрации по трейсингу ---
    trace_qs = TraceData.objects.all()
    if mapped_brands:
        trace_qs = trace_qs.filter(brand__in=mapped_brands)
    if models_selected:
        trace_qs = trace_qs.filter(model__in=models_selected)
    # VIN'ы, разрешённые выбранными фильтрами
    allowed_vins = set(trace_qs.values_list("vin_rk", flat=True))

    # Опции для чекбоксов:
    #   бренды показываем верхнего уровня (ключи BRAND_MAP),
    #   модели формируем из уже отфильтрованного трейсинга
    brands_options = ["gwm", "chery", "changan"]
    models_options = list(
        trace_qs.exclude(model__isnull=True)
                .exclude(model="")
                .values_list("model", flat=True)
                .distinct()
                .order_by("model")
    )

    # --- Подсчёты по периодам (всегда считаем по дате), затем пересекаем с allowed_vins ---
    res_counter = counter_vins(
        brand=None, model=None,
        start_date=start_date, end_date=end_date,
        distinct=True,
    )
    res_vh_all = counter_vins_from_vehicle_history(
        posts=list(POST_AREA_MAPPING.keys()),
        brand=None, model=None,
        start_date=start_date, end_date=end_date,
    )
    res_docs = counter_vins_documentation(
        brand=None, model=None,
        start_date=start_date, end_date=end_date,
    )

    def _apply_trace_filter(payload: dict) -> dict:
        vins = payload.get("vins", []) or []
        if allowed_vins:
            vins = [v for v in vins if v in allowed_vins]
        return {"count": len(vins), "vins": vins}

    # Если установлен хотя бы один фильтр по трейсингу — применим пересечение.
    if brands_selected or models_selected:
        res_counter = _apply_trace_filter(res_counter)
        res_vh_all = _apply_trace_filter(res_vh_all)
        res_docs = _apply_trace_filter(res_docs)

    ctx = {
        "filters": {
            "brand": "", "model": "",
            "start_date": start_date or "",
            "end_date": end_date or "",
        },
        "counter": {"count": res_counter.get("count", 0), "vins": res_counter.get("vins", [])},
        "vehicle_history_all": {"count": res_vh_all.get("count", 0), "vins": res_vh_all.get("vins", [])},
        "documentation": {"count": res_docs.get("count", 0), "vins": res_docs.get("vins", [])},
        "brands_options": brands_options,
        "models_options": models_options,
        "enable_particles": True,
        "enable_background": True,
        "enable_white_square": False,
    }
    return render(request, "users/counter/counter_dashboard.html", ctx)


from django.http import JsonResponse

from django.http import JsonResponse
from django.utils import timezone
from django.views.decorators.cache import never_cache
from django.contrib.auth.decorators import login_required
from itertools import chain

# импортируй свои функции и мэппинги:
# from supplies.utils_qrqc import counter_vins, counter_vins_from_vehicle_history, counter_vins_documentation
# from supplies.constants import POST_AREA_MAPPING
# from users.decorators import role_required
from datetime import date

def _parse_date_or_none(s: str):
    if not s:
        return None
    try:
        return date.fromisoformat(s)
    except Exception:
        return None

def _ensure_shape(payload):
    """
    Нормализует любые ответы хелперов к формату:
    {"count": int, "vins": [..]}.
    """
    if payload is None:
        return {"count": 0, "vins": []}

    if isinstance(payload, dict):
        vins = list(payload.get("vins", []) or [])
        cnt = payload.get("count")
        if cnt is None:
            cnt = len(vins)
        return {"count": int(cnt), "vins": vins}

    # Если вдруг пришло число/итерируемое
    try:
        # iterable VINs
        vins = list(payload)
        return {"count": len(vins), "vins": vins}
    except TypeError:
        # scalar count
        try:
            return {"count": int(payload), "vins": []}
        except Exception:
            return {"count": 0, "vins": []}

@never_cache  # <- важно: отключаем кеш браузера/прокси/NGINX
@login_required
@permission_required('users.access_to_indicators', raise_exception=True)
def assembly_counter_api(request):
    """
    JSON API для трёх наборов данных:
      - "counter" (в процессе сборки)
      - "vehicle_history_all" (на линии тестов)
      - "documentation" (передано в СГП)

    Фильтры (GET/POST):
      - brand: gwm / chery / changan (мульти)
      - model: реальные модели из TraceData (мульти)
      - start_date, end_date: YYYY-MM-DD; если отсутствуют — используем сегодня.
    """
    data = request.POST if request.method == "POST" else request.GET

    BRAND_MAP = {
        "gwm": ["haval", "tank"],
        "chery": ["chery"],
        "changan": ["changan"],
    }

    brands_selected = [b.strip().lower() for b in data.getlist("brand") if b.strip()]
    models_selected = [m.strip() for m in data.getlist("model") if m.strip()]
    mapped_brands = list(chain.from_iterable(BRAND_MAP.get(b, [b]) for b in brands_selected))

    # Даты
    start_date_raw = (data.get("start_date") or "").strip()
    end_date_raw = (data.get("end_date") or "").strip()

    if not start_date_raw and not end_date_raw:
        today = timezone.localdate()
        start_date_dt = end_date_dt = today
    else:
        start_date_dt = _parse_date_or_none(start_date_raw) or timezone.localdate()
        end_date_dt = _parse_date_or_none(end_date_raw) or start_date_dt

    # Приводим к строкам ISO для ваших хелперов, если им так удобнее
    start_date = start_date_dt.isoformat()
    end_date = end_date_dt.isoformat()

    # ----- TraceData -> allowed VINs -----
    trace_qs = TraceData.objects.all()
    if mapped_brands:
        trace_qs = trace_qs.filter(brand__in=mapped_brands)
    if models_selected:
        trace_qs = trace_qs.filter(model__in=models_selected)
    allowed_vins = set(trace_qs.values_list("vin_rk", flat=True))

    # ----- Базовые расчёты без учёта Trace-фильтра -----
    res_counter = _ensure_shape(
        counter_vins(
            brand=None, model=None,
            start_date=start_date, end_date=end_date,
            distinct=True
        )
    )

    res_vh_all = _ensure_shape(
        counter_vins_from_vehicle_history(
            posts=list(POST_AREA_MAPPING.keys()),
            brand=None, model=None,
            start_date=start_date, end_date=end_date
        )
    )

    res_docs = _ensure_shape(
        counter_vins_documentation(
            brand=None, model=None,
            start_date=start_date, end_date=end_date
        )
    )

    # ----- Применяем Trace-фильтр, если он задан -----
    def _apply_trace_filter(payload):
        if not (brands_selected or models_selected):
            return payload  # фильтров нет — возвращаем как есть
        if not allowed_vins:
            return {"count": 0, "vins": []}
        vins = [v for v in payload.get("vins", []) if v in allowed_vins]
        return {"count": len(vins), "vins": vins}

    res_counter = _apply_trace_filter(res_counter)
    res_vh_all = _apply_trace_filter(res_vh_all)
    res_docs = _apply_trace_filter(res_docs)

    # ----- Ответ -----
    resp = JsonResponse(
        {
            "filters": {
                "brands": brands_selected,
                "models": models_selected,
                "start_date": start_date,
                "end_date": end_date,
            },
            "counter": res_counter,
            "vehicle_history_all": res_vh_all,
            "documentation": res_docs,
        },
        json_dumps_params={"ensure_ascii": False}
    )

    # Доп. защита от кеша (декоратор уже делает no-cache, но усилим заголовками)
    resp["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    resp["Pragma"] = "no-cache"
    resp["Expires"] = "0"
    return resp







from itertools import chain
from django.views.decorators.http import require_GET
from django.http import JsonResponse, HttpResponse
from django.utils.dateparse import parse_date, parse_datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment

# MES dashboard template views and helpers
from django.urls import reverse

# наши MES-калькуляторы
from assembly.services.utils_line_indicators import (
    get_trim_mes,
    get_bqa_mes,
    get_qa_mes,
    get_ves_mes,
    get_uud_mes,
)

# ==== MES: shared filters/parser for line counters ====
from django.utils import timezone as _tz
from datetime import datetime as _dt, time as _t


def _parse_mes_filters(request, *, allow_line: bool = True) -> dict:
    """
    Универсальный парсер фильтров для MES-апишек.
    - Даты: start_date/end_date принимают как YYYY-MM-DD, так и ISO datetime (YYYY-MM-DDTHH:MM).
      Если обе пустые — берём «сегодня».
    - Фильтры: brand (multi), model (multi), vin (точный), line (gwm|chery|changan).
    - Пагинация: page, per_page (ограничиваем до 2000).
    Возвращает dict с:
      start/end (aware datetimes), start_date/end_date (iso-строки), brands, models, vin, line, page, per_page.
    """
    data = request.GET if request.method == "GET" else request.POST

    # --- строки из запроса ---
    start_s = (data.get("start_date") or "").strip()
    end_s   = (data.get("end_date")   or "").strip()

    # --- сначала пробуем как datetime ---
    start_dt = parse_datetime(start_s) if start_s else None
    end_dt   = parse_datetime(end_s)   if end_s   else None

    # --- если не datetime, пробуем как date ---
    start_d = None
    end_d   = None
    if not start_dt and start_s:
        try:
            start_d = parse_date(start_s.split("T")[0])
        except Exception:
            pass
    if not end_dt and end_s:
        try:
            end_d = parse_date(end_s.split("T")[0])
        except Exception:
            pass

    # --- если ничего не передано, берём сегодня ---
    if not start_dt and not end_dt and not start_d and not end_d:
        today = _tz.localdate()
        start = _tz.make_aware(_dt.combine(today, _t.min))
        end   = _tz.make_aware(_dt.combine(today, _t.max))
        start_out = today.isoformat()
        end_out   = today.isoformat()
    else:
        # строим start
        if start_dt:
            start = _tz.make_aware(start_dt) if not _tz.is_aware(start_dt) else start_dt
            start_out = start.isoformat()
        elif start_d:
            start = _tz.make_aware(_dt.combine(start_d, _t.min))
            start_out = start_d.isoformat()
        else:
            start = None
            start_out = ""

        # строим end
        if end_dt:
            end = _tz.make_aware(end_dt) if not _tz.is_aware(end_dt) else end_dt
            end_out = end.isoformat()
        elif end_d:
            end = _tz.make_aware(_dt.combine(end_d, _t.max))
            end_out = end_d.isoformat()
        else:
            end = None
            end_out = ""

    # --- фильтры ---
    brands = [b.strip().lower() for b in data.getlist("brand") if b.strip()]
    models = [m.strip() for m in data.getlist("model") if m.strip()]
    vin = (data.get("vin") or "").strip().upper() or None
    line = (data.get("line") or "").strip().lower() if allow_line else None
    if line and line not in ("gwm", "chery", "changan"):
        line = None

    # --- пагинация ---
    try:
        page = int(data.get("page", 1))
    except Exception:
        page = 1
    try:
        per_page = int(data.get("per_page", 100))
    except Exception:
        per_page = 100
    per_page = max(1, min(per_page, 2000))

    return {
        "start": start,
        "end": end,
        "start_date": start_out,
        "end_date": end_out,
        "brands": brands,
        "models": models,
        "vin": vin,
        "line": line,
        "page": page,
        "per_page": per_page,
    }


# ==== helpers: time-based filtering for items ====
from typing import Optional

def _make_aware(dt):
    try:
        return _tz.make_aware(dt) if not _tz.is_aware(dt) else dt
    except Exception:
        return dt

ITEM_DT_PAIRS = [
    ("date", "time"),
    ("trim_out_date", "trim_out_time"),
    ("in_date", "in_time"),
    ("out_date", "out_time"),
    ("step1_date", "step1_time"),
    ("step3_date", "step3_time"),
]

def _extract_item_dt(it: dict) -> Optional[object]:
    """Try to build a timezone-aware datetime for a row item using common keys.
    Returns aware datetime or None.
    """
    # direct keys first
    for k in ("datetime", "dt", "timestamp"):
        v = it.get(k)
        if not v:
            continue
        try:
            d = parse_datetime(str(v))
            if d: return _make_aware(d)
        except Exception:
            pass
    # try pairs date+time
    for dkey, tkey in ITEM_DT_PAIRS:
        d = it.get(dkey)
        t = it.get(tkey)
        if not d or not t:
            continue
        try:
            iso = f"{d}T{t}"
            dt = parse_datetime(iso)
            if dt: return _make_aware(dt)
        except Exception:
            continue
    # last resort: just date
    for dkey, _ in ITEM_DT_PAIRS:
        d = it.get(dkey)
        if not d:
            continue
        try:
            dd = parse_date(str(d))
            if dd:
                return _make_aware(_dt.combine(dd, _t.min))
        except Exception:
            continue
    return None

def _filter_items_by_time(items: list[dict], start, end) -> list[dict]:
    if not (start and end):
        return items
    out = []
    for it in (items or []):
        dt = _extract_item_dt(it)
        if dt is None:
            # keep if we cannot determine — do not drop silently
            out.append(it)
        elif start <= dt <= end:
            out.append(it)
    return out


# ==== MES: reference lists & suggestions (brands/models/VIN) ====

from django.views.decorators.http import require_GET
from django.http import JsonResponse

@login_required
@require_GET
def mes_tracing_brands_api(request):
    """
    GET /api/mes/tracing/brands
    Optional: ?q=substr (case-insensitive)
    Returns distinct brand names from TraceData (as stored), sorted.
    """
    q = (request.GET.get("q") or "").strip().lower()
    try:
        qs = (TraceData.objects
                .exclude(brand__isnull=True)
                .exclude(brand="")
                .values_list("brand", flat=True)
                .distinct())
        brands = sorted({(b or "").strip() for b in qs if (b or "").strip()})
        if q:
            brands = [b for b in brands if q in b.lower()]
    except Exception:
        brands = []
    return JsonResponse({"brands": brands}, json_dumps_params={"ensure_ascii": False})


@login_required
@require_GET
def mes_tracing_models_api(request):
    """
    GET /api/mes/tracing/models
    Optional: ?brand=... (multi, accepts top-level gwm/chery/changan or raw brands like haval/tank)
              ?q=substr (case-insensitive)
    Returns distinct models from TraceData, optionally filtered by brand(s).
    """
    q = (request.GET.get("q") or "").strip()
    brands_selected = [b.strip().lower() for b in request.GET.getlist("brand") if b.strip()]
    BRAND_MAP = {"gwm": ["haval", "tank"], "chery": ["chery"], "changan": ["changan"]}
    mapped: set[str] = set()
    for b in brands_selected:
        mapped.update(BRAND_MAP.get(b, [b]))

    try:
        qs = TraceData.objects.all()
        if mapped:
            qs = qs.filter(brand__in=list(mapped))
        if q:
            qs = qs.filter(model__icontains=q)
        qs = (qs.exclude(model__isnull=True)
                .exclude(model="")
                .values_list("model", flat=True)
                .distinct()
                .order_by("model"))
        models = [m for m in qs if (m or "").strip()]
    except Exception:
        models = []
    return JsonResponse({"models": models}, json_dumps_params={"ensure_ascii": False})


@login_required
@require_GET
def mes_vin_suggest_api(request):
    """
    GET /api/mes/history/vins?q=MXM&limit=20
    VIN suggestions from history; falls back to TraceData if VINHistory is unavailable.
    """
    q = (request.GET.get("q") or "").strip().upper()
    try:
        limit = int(request.GET.get("limit", 20))
    except Exception:
        limit = 20
    if len(q) < 2:  # небольшая защита от лишних запросов
        return JsonResponse({"vins": []}, json_dumps_params={"ensure_ascii": False})

    vins: list[str] = []
    try:
        # prefer VINHistory if present
        from vehicle_history.models import VINHistory  # adjust import path if different
        qs = (VINHistory.objects.filter(vin__icontains=q)
                               .values_list("vin", flat=True)
                               .order_by("vin")[: max(1, limit)])
        vins = [v for v in qs if v]
    except Exception:
        try:
            qs = (TraceData.objects.filter(vin_rk__icontains=q)
                                   .values_list("vin_rk", flat=True)
                                   .order_by("vin_rk")[: max(1, limit)])
            vins = [v for v in qs if v]
        except Exception:
            vins = []

    # уникализуем и сортируем, чтобы не было дубликатов от разных источников
    vins = sorted(dict.fromkeys(vins))[: max(1, limit)]
    return JsonResponse({"vins": vins}, json_dumps_params={"ensure_ascii": False})


# ==== MES: BQA (Buffer QA) APIs ====

@login_required
@require_GET
def bqa_overview_api(request):
    """
    GET /api/mes/bqa/overview
    Параметры: start_date, end_date, brand (multi), model (multi), line, vin.
    Возвращает summary по BQA: bqa_in, bqa_wip_today, bqa_wip_all.
    """
    f = _parse_mes_filters(request, allow_line=True)
    res = get_bqa_mes(
        start=f["start"],
        end=f["end"],
        brands=f["brands"] or None,
        models=f["models"] or None,
        line=f["line"],
        vin=f["vin"],
    ) or {}

    payload = {
        "filters": {
            "start_date": f["start_date"],
            "end_date": f["end_date"],
            "brands": f["brands"],
            "models": f["models"],
            "vin": f["vin"] or "",
            "line": f["line"] or "",
        },
        "overall": res.get("overall", {}),
    }
    # если расчёт вернёт разбивку по линиям — пробрасываем
    if "by_line" in res:
        payload["by_line"] = res["by_line"]

    return JsonResponse(payload, json_dumps_params={"ensure_ascii": False})


@login_required
@require_GET
def bqa_list_api(request):
    """
    GET /api/mes/bqa/list?metric=bqa_in|bqa_wip_today|bqa_wip_all
    Возвращает пагинированный список VINов выбранной метрики.
    """
    f = _parse_mes_filters(request, allow_line=True)
    metric = (request.GET.get("metric") or "bqa_in").strip().lower()

    res = get_bqa_mes(
        start=f["start"],
        end=f["end"],
        brands=f["brands"] or None,
        models=f["models"] or None,
        line=f["line"],
        vin=f["vin"],
    ) or {}
    overall = res.get("overall", {})
    bucket = overall.get(metric, {"items": []})
    items = list(bucket.get("items", []) or [])
    items = _filter_items_by_time(items, f["start"], f["end"])

    total = len(items)
    lo = (f["page"] - 1) * f["per_page"]
    hi = lo + f["per_page"]

    return JsonResponse({
        "filters": {
            "start_date": f["start_date"],
            "end_date": f["end_date"],
            "brands": f["brands"],
            "models": f["models"],
            "vin": f["vin"] or "",
            "line": f["line"] or "",
            "metric": metric,
        },
        "pagination": {"page": f["page"], "per_page": f["per_page"], "total": total},
        "items": items[lo:hi],
    }, json_dumps_params={"ensure_ascii": False})


# ==== MES: QA (First Inspection → Sign Off) APIs ====

@login_required
@require_GET
def qa_overview_api(request):
    """
    GET /api/mes/qa/overview
    Параметры: start_date, end_date, brand (multi), model (multi), vin, include_items=0|1
    В QA одна общая линия, поэтому параметр line игнорируется.
    Возвращает summary: qa_in, wip_qa, sign_off (+ str, dpu).
    - qa_in/wip_qa/sign_off: всегда содержат поле count; при include_items=1 также возвращают items.
    - str: {percent, numerator, denominator}
    - dpu: {value, defects, unique_cars_today}
    """
    f = _parse_mes_filters(request, allow_line=False)
    res = get_qa_mes(
        start=f["start"],
        end=f["end"],
        brands=f["brands"] or None,
        models=f["models"] or None,
        vin=f["vin"],
    ) or {}

    # include_items flag like in TRIM overview
    include_items = (request.GET.get("include_items") or "").strip().lower() in {"1","true","yes","on"}

    overall_raw = res.get("overall", {}) or {}

    # Normalize the three core buckets (always return counts; items optionally)
    metrics = ("qa_in", "wip_qa", "sign_off")
    overall_norm: dict = {}
    for m in metrics:
        node = overall_raw.get(m, {}) or {}
        if isinstance(node, dict):
            cnt = int(node.get("count", 0) or 0)
            overall_norm[m] = {"count": cnt}
            if include_items:
                overall_norm[m]["items"] = list(node.get("items", []) or [])
        else:
            cnt = int(node or 0)
            overall_norm[m] = {"count": cnt}
            if include_items:
                overall_norm[m]["items"] = []

    # Pass-through STR and DPU aggregates if present
    if "str" in overall_raw:
        overall_norm["str"] = overall_raw["str"]
    if "dpu" in overall_raw:
        overall_norm["dpu"] = overall_raw["dpu"]

    payload = {
        "filters": {
            "start_date": f["start_date"],
            "end_date": f["end_date"],
            "brands": f["brands"],
            "models": f["models"],
            "vin": f["vin"] or "",
            "include_items": include_items,
        },
        "overall": overall_norm,
    }
    return JsonResponse(payload, json_dumps_params={"ensure_ascii": False})



@login_required
@require_GET
def qa_list_api(request):
    """
    GET /api/mes/qa/list?metric=qa_in|wip_qa|sign_off
    Возвращает пагинированный список VINов для выбранной метрики QA.
    """
    f = _parse_mes_filters(request, allow_line=False)
    metric = (request.GET.get("metric") or "qa_in").strip().lower()

    res = get_qa_mes(
        start=f["start"],
        end=f["end"],
        brands=f["brands"] or None,
        models=f["models"] or None,
        vin=f["vin"],
    ) or {}
    overall = res.get("overall", {})
    bucket = overall.get(metric, {"items": []})
    items = list(bucket.get("items", []) or [])
    items = _filter_items_by_time(items, f["start"], f["end"])

    total = len(items)
    lo = (f["page"] - 1) * f["per_page"]
    hi = lo + f["per_page"]

    return JsonResponse({
        "filters": {
            "start_date": f["start_date"],
            "end_date": f["end_date"],
            "brands": f["brands"],
            "models": f["models"],
            "vin": f["vin"] or "",
            "metric": metric,
        },
        "pagination": {"page": f["page"], "per_page": f["per_page"], "total": total},
        "items": items[lo:hi],
    }, json_dumps_params={"ensure_ascii": False})


# ==== MES: UUD (Участок устранения дефектов) APIs ====

@login_required
@require_GET
def uud_overview_api(request):
    """
    GET /api/mes/uud/overview
    Параметры: start_date, end_date, brand (multi), model (multi), vin.
    В УУД нет разбиения на линии (единая линия), поэтому параметр line игнорируется.
    Возвращает summary:
      - uud_in                — поступившие на УУД за период (уникальные VIN)
      - uud_in_wip_today      — из поступивших сегодня стоят в буфере (ещё не приняты УУД)
      - uud_in_wip_all        — все, кто в буфере УУД из любых дней (на шаге 1)
      - wip_uud               — сейчас в работе на УУД (шаг 2)
      - uud_out               — отремонтированы за период (переведены на шаг 3)
      - uud_out_wip_today     — сегодня отремонтированы, но ещё не приняты QA (шаг 3, за сегодня)
      - uud_out_wip_all       — все ожидающие приёма QA после ремонта (шаг 3, все дни)
    Замечание: точные ключи берутся из расчёта get_uud_mes(); если они отличаются, фронт получит то, что вернул расчёт.
    """
    f = _parse_mes_filters(request, allow_line=False)
    res = get_uud_mes(
        start=f["start"],
        end=f["end"],
        brands=f["brands"] or None,
        models=f["models"] or None,
        vin=f["vin"],
    ) or {}

    payload = {
        "filters": {
            "start_date": f["start_date"],
            "end_date": f["end_date"],
            "brands": f["brands"],
            "models": f["models"],
            "vin": f["vin"] or "",
        },
        # пробрасываем как есть; фронту доступны все поля, что вернул расчёт
        "overall": res.get("overall", {}),
    }
    return JsonResponse(payload, json_dumps_params={"ensure_ascii": False})


@login_required
@require_GET
def uud_list_api(request):
    """
    GET /api/mes/uud/list?metric=
      uud_in | uud_in_wip_today | uud_in_wip_all |
      wip_uud |
      uud_out | uud_out_wip_today | uud_out_wip_all

    Возвращает пагинированный список VIN-элементов для выбранной метрики УУД.
    Если ключ метрики отличается (например, в расчёте используется иной алиас),
    применяется карта псевдонимов ниже.
    """
    f = _parse_mes_filters(request, allow_line=False)
    metric = (request.GET.get("metric") or "uud_in").strip().lower()

    # Возможные алиасы на случай смены нейминга внутри расчёта
    aliases = {
        "uud_wip": "wip_uud",
        "wip": "wip_uud",
        "in_wip_today": "uud_in_wip_today",
        "in_wip_all": "uud_in_wip_all",
        "out_wip_today": "uud_out_wip_today",
        "out_wip_all": "uud_out_wip_all",
    }

    res = get_uud_mes(
        start=f["start"],
        end=f["end"],
        brands=f["brands"] or None,
        models=f["models"] or None,
        vin=f["vin"],
    ) or {}
    overall = res.get("overall", {}) or {}

    key = metric if metric in overall else aliases.get(metric, metric)
    if key not in overall:
        # Сформируем список доступных ключей, чтобы фронту было проще диагностировать
        return JsonResponse({
            "ok": False,
            "error": "unknown metric",
            "metric": metric,
            "available": sorted(list(overall.keys())),
        }, status=400)

    bucket = overall.get(key, {}) or {}
    items = list(bucket.get("items", []) or [])
    items = _filter_items_by_time(items, f["start"], f["end"])

    total = len(items)
    lo = (f["page"] - 1) * f["per_page"]
    hi = lo + f["per_page"]

    return JsonResponse({
        "filters": {
            "start_date": f["start_date"],
            "end_date": f["end_date"],
            "brands": f["brands"],
            "models": f["models"],
            "vin": f["vin"] or "",
            "metric": key,
        },
        "pagination": {"page": f["page"], "per_page": f["per_page"], "total": total},
        "items": items[lo:hi],
    }, json_dumps_params={"ensure_ascii": False})



# ==== MES: VES (Vehicle Evaluation Station) APIs ====

@login_required
@require_GET
def ves_overview_api(request):
    """
    GET /api/mes/ves/overview
    Параметры: start_date, end_date, brand (multi), model (multi), vin.
    В VES нет разбиения на линии, поэтому параметр line игнорируется.
    Возвращает summary по VES (как его отдаёт расчёт get_ves_mes):
      - ves_in    — переданы на VES за период (уникальные VIN)
      - wip_ves   — сейчас находятся на VES (отданы, но ещё не приняты обратно)
      - ves_out   — приняты обратно с VES за период
    Элементы метрик содержат поля, возвращаемые расчётом (например: vin, given_date/time, received_date/time,
    given_by, received_by, brand, model), пробрасываются без переименования.
    """
    f = _parse_mes_filters(request, allow_line=False)
    res = get_ves_mes(
        start=f["start"],
        end=f["end"],
        brands=f["brands"] or None,
        models=f["models"] or None,
        vin=f["vin"],
    ) or {}

    payload = {
        "filters": {
            "start_date": f["start_date"],
            "end_date": f["end_date"],
            "brands": f["brands"],
            "models": f["models"],
            "vin": f["vin"] or "",
        },
        "overall": res.get("overall", {}),
    }
    return JsonResponse(payload, json_dumps_params={"ensure_ascii": False})


@login_required
@require_GET
def ves_list_api(request):
    """
    GET /api/mes/ves/list?metric=ves_in|wip_ves|ves_out
    Возвращает пагинированный список VIN-элементов выбранной метрики.
    Поддерживаются алиасы: "in"->"ves_in", "out"->"ves_out", "wip"|"ves_wip"->"wip_ves".
    """
    f = _parse_mes_filters(request, allow_line=False)
    metric = (request.GET.get("metric") or "ves_in").strip().lower()

    # Алиасы на случай иного нейминга в запросах
    aliases = {
        "in": "ves_in",
        "out": "ves_out",
        "wip": "wip_ves",
        "ves_wip": "wip_ves",
    }

    res = get_ves_mes(
        start=f["start"],
        end=f["end"],
        brands=f["brands"] or None,
        models=f["models"] or None,
        vin=f["vin"],
    ) or {}
    overall = res.get("overall", {}) or {}

    key = metric if metric in overall else aliases.get(metric, metric)
    if key not in overall:
        return JsonResponse({
            "ok": False,
            "error": "unknown metric",
            "metric": metric,
            "available": sorted(list(overall.keys())),
        }, status=400)

    bucket = overall.get(key, {}) or {}
    items = list(bucket.get("items", []) or [])
    items = _filter_items_by_time(items, f["start"], f["end"])

    total = len(items)
    lo = (f["page"] - 1) * f["per_page"]
    hi = lo + f["per_page"]

    return JsonResponse({
        "filters": {
            "start_date": f["start_date"],
            "end_date": f["end_date"],
            "brands": f["brands"],
            "models": f["models"],
            "vin": f["vin"] or "",
            "metric": key,
        },
        "pagination": {"page": f["page"], "per_page": f["per_page"], "total": total},
        "items": items[lo:hi],
    }, json_dumps_params={"ensure_ascii": False})


# ==== MES: DASHBOARD SUMMARY (combined) ====

@login_required
@require_GET
def mes_summary_api(request):
    """
    GET /api/mes/summary

    Объединённый JSON по всем MES-секциям: TRIM, BQA, QA, UUD, VES.
    Фильтры периода (start_date/end_date) применяются ко всем.
    Фильтры brand/model/vin применяются только к нижнему блоку (bottom),
    как мы договаривались: верхний блок (top) — totals за все бренды/модели/линии.

    Query params:
      - start_date, end_date: YYYY-MM-DD (если оба пустые — сегодня)
      - brand: multi (?brand=gwm&brand=chery) — влияет ТОЛЬКО на bottom
      - model: multi — влияет ТОЛЬКО на bottom
      - vin: точный VIN — влияет ТОЛЬКО на bottom
      - line: gwm|chery|changan — применяется только к TRIM в bottom (в top игнорируем)
      - sections: csv из {trim,bqa,qa,uud,ves}; по умолчанию — все
      - include_items: 0|1 — если 1, отдаём списки items; иначе только агрегации без массивов

    Формат ответа:
      {
        "filters": {...},
        "summary": {
          "top": { "trim": {...}, "bqa": {...}, ... },
          "bottom": { "trim": {...}, "bqa": {...}, ... }
        }
      }
    """
    f = _parse_mes_filters(request, allow_line=True)

    include_items = (request.GET.get("include_items") or "").strip().lower() in {"1","true","yes","on"}
    sections_raw = (request.GET.get("sections") or "").strip()
    wanted = {s.strip().lower() for s in sections_raw.split(",") if s.strip()} or {"trim","bqa","qa","uud","ves"}
    allowed = {"trim","bqa","qa","uud","ves"}
    sections = wanted & allowed

    def _safe(calc, **kwargs):
        try:
            return calc(**kwargs) or {}
        except Exception:
            return {}

    # --- TOP: totals за период без brand/model/vin/line ---
    top: dict = {}
    if "trim" in sections:
        top["trim"] = _safe(
            get_trim_mes,
            start=f["start"], end=f["end"],
            line=None, brands=None, models=None, vin=None,
        )
    if "bqa" in sections:
        top["bqa"] = _safe(
            get_bqa_mes,
            start=f["start"], end=f["end"],
            brands=None, models=None, line=None, vin=None,
        )
    if "qa" in sections:
        top["qa"] = _safe(
            get_qa_mes,
            start=f["start"], end=f["end"],
            brands=None, models=None, vin=None,
        )
    if "uud" in sections:
        top["uud"] = _safe(
            get_uud_mes,
            start=f["start"], end=f["end"],
            brands=None, models=None, vin=None,
        )
    if "ves" in sections:
        top["ves"] = _safe(
            get_ves_mes,
            start=f["start"], end=f["end"],
            brands=None, models=None, vin=None,
        )

    # --- BOTTOM: применяем brand/model/vin; line только к TRIM ---
    bottom: dict = {}
    if "trim" in sections:
        bottom["trim"] = _safe(
            get_trim_mes,
            start=f["start"], end=f["end"],
            line=f["line"], brands=f["brands"] or None, models=f["models"] or None, vin=f["vin"],
        )
    if "bqa" in sections:
        bottom["bqa"] = _safe(
            get_bqa_mes,
            start=f["start"], end=f["end"],
            brands=f["brands"] or None, models=f["models"] or None, line=f["line"], vin=f["vin"],
        )
    if "qa" in sections:
        bottom["qa"] = _safe(
            get_qa_mes,
            start=f["start"], end=f["end"],
            brands=f["brands"] or None, models=f["models"] or None, vin=f["vin"],
        )
    if "uud" in sections:
        bottom["uud"] = _safe(
            get_uud_mes,
            start=f["start"], end=f["end"],
            brands=f["brands"] or None, models=f["models"] or None, vin=f["vin"],
        )
    if "ves" in sections:
        bottom["ves"] = _safe(
            get_ves_mes,
            start=f["start"], end=f["end"],
            brands=f["brands"] or None, models=f["models"] or None, vin=f["vin"],
        )

    # --- optionally strip items arrays to keep payload small ---
    def _strip_items(obj: dict) -> dict:
        if include_items:
            return obj
        def prune(x):
            if isinstance(x, dict):
                return {k: prune(v) for k, v in x.items() if k != "items"}
            if isinstance(x, list):
                return [prune(i) for i in x]
            return x
        return prune(obj)

    response = {
        "filters": {
            "start_date": f["start_date"],
            "end_date": f["end_date"],
            "brands": f["brands"],
            "models": f["models"],
            "vin": f["vin"] or "",
            "line": f["line"] or "",
            "sections": sorted(list(sections)),
            "include_items": bool(include_items),
        },
        "summary": {
            "top": _strip_items(top),
            "bottom": _strip_items(bottom),
        }
    }

    return JsonResponse(response, json_dumps_params={"ensure_ascii": False})


# ==== MES: DASHBOARD PAGE VIEWS (template-rendered) ====

@login_required
@permission_required('users.access_to_indicators', raise_exception=True)
def mes_dashboard_view(request):
    """
    Страница с общим графическим дашбордом MES.
    - Верхний блок (top): тоталы по всем брендам/моделям/линиям за период.
    - Нижний блок (bottom): те же метрики, но с фильтрами brand/model/vin (+ line для TRIM).
    Данные на страницу подгружаются фронтом через /api/mes/summary.
    Поддерживаемые GET-параметры (для предзаполнения фильтров и автoобновления):
      ?start_date=YYYY-MM-DD&amp;end_date=YYYY-MM-DD  (по умолчанию: сегодня..сегодня)
      ?brand=gwm&amp;brand=chery                      (мульти, верхнего уровня)
      ?model=Tiggo%204                              (мульти)
      ?line=gwm|chery|changan                       (только для TRIM в нижнем блоке)
      ?vin=VIN
      ?sections=trim,bqa,qa,uud,ves
      ?refresh_ms=15000
      ?include_items=0|1  (проброс для /api/mes/summary)
    """
    data = request.GET
    # Период
    start_date = (data.get("start_date") or "").strip()
    end_date   = (data.get("end_date") or "").strip()
    if not start_date and not end_date:
        today = _tz.localdate()
        start_date = end_date = today.isoformat()

    # Фильтры
    brands_selected = [b.strip().lower() for b in data.getlist("brand") if b.strip()]
    models_selected = [m.strip() for m in data.getlist("model") if m.strip()]
    vin  = (data.get("vin") or "").strip().upper()
    line = (data.get("line") or "").strip().lower()
    sections_req = [s.strip().lower() for s in (data.get("sections") or "").split(",") if s.strip()]
    sections = sections_req or ["trim", "bqa", "qa", "uud", "ves"]
    # Автообновление
    try:
        refresh_ms = int(data.get("refresh_ms", "15000"))
    except Exception:
        refresh_ms = 15000
    include_items = (data.get("include_items") or "").strip().lower() in {"1","true","yes","on"}

    # Опции селектов
    # Для моделей берём distinct из TraceData; если на странице уже выбраны бренды — учтём их маппинг.
    BRAND_MAP = {"gwm": ["haval", "tank"], "chery": ["chery"], "changan": ["changan"]}
    trace_qs = TraceData.objects.all()
    if brands_selected:
        mapped = list(chain.from_iterable(BRAND_MAP.get(b, [b]) for b in brands_selected))
        trace_qs = trace_qs.filter(brand__in=mapped)
    models_options = list(
        trace_qs.exclude(model__isnull=True)
                .exclude(model="")
                .values_list("model", flat=True)
                .distinct()
                .order_by("model")
    )

    def _rev(name: str, default: str) -> str:
        try:
            return reverse(name)
        except Exception:
            return default

    api_map = {
        "summary": _rev("mes_summary_api", "/api/mes/summary"),
        "trim": {
            "overview": _rev("trim_overview_api", "/api/mes/trim/overview"),
            "list":     _rev("trim_list_api", "/api/mes/trim/list"),
        },
        "bqa": {
            "overview": _rev("bqa_overview_api", "/api/mes/bqa/overview"),
            "list":     _rev("bqa_list_api", "/api/mes/bqa/list"),
        },
        "qa": {
            "overview": _rev("qa_overview_api", "/api/mes/qa/overview"),
            "list":     _rev("qa_list_api", "/api/mes/qa/list"),
        },
        "uud": {
            "overview": _rev("uud_overview_api", "/api/mes/uud/overview"),
            "list":     _rev("uud_list_api", "/api/mes/uud/list"),
        },
        "ves": {
            "overview": _rev("ves_overview_api", "/api/mes/ves/overview"),
            "list":     _rev("ves_list_api", "/api/mes/ves/list"),
        },
    }

    ctx = {
        "filters": {
            "start_date": start_date,
            "end_date": end_date,
            "brands": brands_selected,
            "models": models_selected,
            "vin": vin,
            "line": line,
            "sections": sections,
            "include_items": include_items,
        },
        "options": {
            "brands": ["gwm", "chery", "changan"],
            "models": models_options,
            "lines":  ["gwm", "chery", "changan"],
        },
        "api": api_map,
        "refresh_ms": refresh_ms,
        "enable_particles": True,
        "enable_background": True,
        "enable_white_square": False,
    }
    return render(request, "users/mes/mes_dashboard.html", ctx)


@login_required
@permission_required('users.access_to_indicators', raise_exception=True)
def mes_metric_list_view(request, section: str | None = None):
    """
    Страница «VIN‑лист по метрике» (по клику на карточку).
    Рендерит таблицу с пагинацией; данные тянутся фронтом из соответствующего list‑API.
    GET:
      ?section=trim|bqa|qa|uud|ves
      ?metric=...            (например: trim_in|wip_trim|trim_out, bqa_in|bqa_wip_all, qa_in|sign_off, ...)
      ?start_date, ?end_date
      ?brand, ?model (multi)
      ?vin
      ?line (используется только для TRIM)
    """
    section = (section or request.GET.get("section") or "trim").strip().lower()
    if section not in {"trim", "bqa", "qa", "uud", "ves"}:
        section = "trim"

    # вытащим фильтры и default метрику
    f = _parse_mes_filters(request, allow_line=(section == "trim"))
    default_metric = {
        "trim": "trim_in",
        "bqa":  "bqa_in",
        "qa":   "qa_in",
        "uud":  "uud_in",
        "ves":  "ves_in",
    }[section]
    metric = (request.GET.get("metric") or default_metric).strip().lower()

    def _rev(name: str, default: str) -> str:
        try:
            return reverse(name)
        except Exception:
            return default

    api_list_url = {
        "trim": _rev("trim_list_api", "/api/mes/trim/list"),
        "bqa":  _rev("bqa_list_api",  "/api/mes/bqa/list"),
        "qa":   _rev("qa_list_api",   "/api/mes/qa/list"),
        "uud":  _rev("uud_list_api",  "/api/mes/uud/list"),
        "ves":  _rev("ves_list_api",  "/api/mes/ves/list"),
    }[section]

    api_table_url = _rev("mes_metric_table_api", "/api/mes/table")

    ctx = {
        "section": section,
        "metric": metric,
        "api_list_url": api_list_url,
        "api_table_url": api_table_url,
        "filters": {
            "start_date": f["start_date"],
            "end_date":   f["end_date"],
            "brands":     f["brands"],
            "models":     f["models"],
            "vin":        f["vin"] or "",
            "line":       f["line"] or "",
            "page":       f["page"],
            "per_page":   f["per_page"],
        },
        # декоративные флаги, чтобы шаблон соответствовал стилю других страниц
        # "enable_particles": True,
        # "enable_background": True,
        # "enable_white_square": False,
    }
    return render(request, "users/mes/metric_list.html", ctx)


# ==== MES: Dynamic table API (columns + items) ====

KEY_TITLES = {
    "vin": "VIN",
    "brand": "Бренд",
    "model": "Модель",
    "line": "Линия",
    "date": "Дата",
    "time": "Время",
    "controller": "Контролёр",
    "controller_out": "Контролёр OUT",
    "controller_in": "Контролёр IN",
    "controller_give": "Кто передал",
    "controller_receive": "Кто принял",
    "controller_finish": "Завершил",
    "trim_out_date": "Дата TRIM OUT",
    "trim_out_time": "Время TRIM OUT",
    "in_date": "Дата передачи",
    "in_time": "Время передачи",
    "out_date": "Дата приёмки",
    "out_time": "Время приёмки",
    "step1_date": "Дата поступления",
    "step1_time": "Время поступления",
    "step3_date": "Дата ремонта",
    "step3_time": "Время ремонта",
    # резерв на будущее:
    "saved_by": "Сохранил",
    "controller_login": "Логин контролёра",
}

def _infer_columns_from_items(items: list[dict]) -> list[dict]:
    """
    Формирует список столбцов для таблицы: [{key, title}, ...]
    Берём только известные ключи (KEY_TITLES); порядок — предпочтительный, затем остальные по появлению.
    """
    if not items:
        # дефолтный набор — только VIN
        return [{"key": "vin", "title": KEY_TITLES["vin"]}]

    # ключи, встречающиеся в данных
    seen: list[str] = []
    for it in items:
        if not isinstance(it, dict):
            continue
        for k in it.keys():
            if k in KEY_TITLES and k not in seen:
                seen.append(k)

    preferred_order = [
        "vin", "brand", "model", "line",
        "date", "time",
        "trim_out_date", "trim_out_time",
        "in_date", "in_time",
        "out_date", "out_time",
        "step1_date", "step1_time",
        "step3_date", "step3_time",
        "controller", "controller_out", "controller_in",
        "controller_give", "controller_receive", "controller_finish",
        "saved_by", "controller_login",
    ]

    ordered = [k for k in preferred_order if k in seen] + [k for k in seen if k not in preferred_order]
    return [{"key": k, "title": KEY_TITLES.get(k, k)} for k in ordered]

def _resolve_metric_key(section: str, metric: str, overall: dict) -> str | None:
    """
    Приводим метрику к ключу, присутствующему в overall, с учётом алиасов.
    """
    section = (section or "").lower().strip()
    metric  = (metric or "").lower().strip()

    # дефолты по секциям
    defaults = {"trim": "trim_in", "bqa": "bqa_in", "qa": "qa_in", "uud": "uud_in", "ves": "ves_in"}
    if not metric:
        metric = defaults.get(section, "trim_in")

    aliases = {}
    if section == "ves":
        aliases = {"in": "ves_in", "out": "ves_out", "wip": "wip_ves", "ves_wip": "wip_ves"}
    elif section == "uud":
        aliases = {
            "uud_wip": "wip_uud",
            "wip": "wip_uud",
            "in_wip_today": "uud_in_wip_today",
            "in_wip_all": "uud_in_wip_all",
            "out_wip_today": "uud_out_wip_today",
            "out_wip_all": "uud_out_wip_all",
        }
    # для остальных секций алиасы не требуются

    key = metric if metric in (overall or {}) else aliases.get(metric, metric)
    return key if key in (overall or {}) else None

@login_required
@require_GET
def mes_metric_table_api(request):
    """
    GET /api/mes/table
      ?section=trim|bqa|qa|uud|ves
      ?metric=...
      ?start_date=YYYY-MM-DD&end_date=YYYY-MM-DD
      ?brand=...&model=... (multi)
      ?line=... (только для TRIM)
      ?vin=VIN
      ?page=1&per_page=100

    Возвращает:
    {
      "filters": {...},
      "pagination": {...},
      "columns": [{"key":"vin","title":"VIN"}, ...],
      "items": [ {...}, ... ]
    }
    """
    section = (request.GET.get("section") or "trim").strip().lower()
    allow_line = (section == "trim")
    f = _parse_mes_filters(request, allow_line=allow_line)

    # Получаем расчёт по секции
    calc_map = {
        "trim": lambda: get_trim_mes(start=f["start"], end=f["end"], line=f["line"], brands=f["brands"] or None, models=f["models"] or None, vin=f["vin"]),
        "bqa":  lambda: get_bqa_mes(start=f["start"], end=f["end"], line=f["line"], brands=f["brands"] or None, models=f["models"] or None, vin=f["vin"]),
        "qa":   lambda: get_qa_mes(start=f["start"], end=f["end"], brands=f["brands"] or None, models=f["models"] or None, vin=f["vin"]),
        "uud":  lambda: get_uud_mes(start=f["start"], end=f["end"], brands=f["brands"] or None, models=f["models"] or None, vin=f["vin"]),
        "ves":  lambda: get_ves_mes(start=f["start"], end=f["end"], brands=f["brands"] or None, models=f["models"] or None, vin=f["vin"]),
    }
    if section not in calc_map:
        section = "trim"

    try:
        res = calc_map[section]() or {}
    except Exception:
        res = {}

    overall = res.get("overall", {}) or {}
    metric = (request.GET.get("metric") or "").strip().lower()
    key = _resolve_metric_key(section, metric, overall)
    if not key:
        return JsonResponse({
            "ok": False,
            "error": "unknown metric",
            "available": sorted(list(overall.keys())),
        }, status=400)

    bucket = overall.get(key, {}) or {}
    items = list(bucket.get("items", []) or [])
    items = _filter_items_by_time(items, f["start"], f["end"])

    # Явная фильтрация по линии для TRIM (на случай, если аггрегатор вернул все линии)
    if allow_line and f["line"]:
        items = [it for it in items if (it.get("line") or "").strip().lower() == f["line"]]

    total = len(items)
    lo = (f["page"] - 1) * f["per_page"]
    hi = lo + f["per_page"]
    items_page = items[lo:hi]

    columns = _infer_columns_from_items(items_page if items_page else items)

    return JsonResponse({
        "filters": {
            "section": section,
            "metric": key,
            "start_date": f["start_date"],
            "end_date": f["end_date"],
            "brands": f["brands"],
            "models": f["models"],
            "vin": f["vin"] or "",
            "line": f["line"] or "",
        },
        "pagination": {"page": f["page"], "per_page": f["per_page"], "total": total},
        "columns": columns,
        "items": items_page,
    }, json_dumps_params={"ensure_ascii": False})


# ===== TRIM MES JSON APIs =====

from datetime import datetime, time
from django.utils import timezone

ALLOWED_LINES = {"gwm", "chery", "changan"}

def _parse_trim_filters(request):
    """
    Common filter parser for TRIM APIs.
    Accepts:
      - start_date, end_date: YYYY-MM-DD **или** ISO datetime (e.g. 2025-09-22T00:00)
      - brand: multi (?brand=gwm&brand=chery) — top-level brands
      - model: multi (?model=Tiggo%204)
      - line: one of gwm|chery|changan (applies mainly to TRIM IN/WIP; TRIM OUT uses brand/model mapping)
      - vin: exact VIN (optional)
      - page, per_page: for list endpoint
    Returns dict with timezone-aware datetimes for 'start'/'end'.
    """
    from django.utils.dateparse import parse_date, parse_datetime

    data = request.GET if request.method == "GET" else request.POST

    # --- Period ---
    start_s = (data.get("start_date") or "").strip()
    end_s   = (data.get("end_date") or "").strip()

    # 1) сначала пытаемся как ISO datetime; 2) иначе как date (берём границы суток)
    start_dt = parse_datetime(start_s) if start_s else None
    end_dt   = parse_datetime(end_s)   if end_s   else None

    if start_dt or end_dt:
        start = timezone.make_aware(start_dt) if (start_dt and not timezone.is_aware(start_dt)) else start_dt
        end   = timezone.make_aware(end_dt)   if (end_dt   and not timezone.is_aware(end_dt))   else end_dt
    else:
        # допускаем вход вида 2025-09-22T00:00 — возьмём часть до 'T'
        if "T" in start_s:
            start_s = start_s.split("T", 1)[0]
        if "T" in end_s:
            end_s = end_s.split("T", 1)[0]
        start_d = parse_date(start_s) if start_s else None
        end_d   = parse_date(end_s)   if end_s   else None
        if not start_d and not end_d:
            # default to today
            today = timezone.localdate()
            start = timezone.make_aware(datetime.combine(today, time.min))
            end   = timezone.make_aware(datetime.combine(today, time.max))
        else:
            start = timezone.make_aware(datetime.combine(start_d, time.min)) if start_d else None
            end   = timezone.make_aware(datetime.combine(end_d,   time.max)) if end_d   else None

    # --- Brand/model filters (top-level brands for our MES calculators) ---
    brands = [b.strip().lower() for b in data.getlist("brand") if (b or "").strip()] or None
    models = [m.strip() for m in data.getlist("model") if (m or "").strip()] or None

    # --- Line, VIN ---
    line = (data.get("line") or "").strip().lower() or None
    if line and line not in ALLOWED_LINES:
        line = None
    vin = (data.get("vin") or "").strip().upper() or None

    # --- Pagination for list endpoint ---
    try:
        per_page = int(data.get("per_page", 100))
    except Exception:
        per_page = 100
    try:
        page = max(1, int(data.get("page", 1)))
    except Exception:
        page = 1

    return {
        "start": start,
        "end": end,
        "brands": brands,
        "models": models,
        "line": line,
        "vin": vin,
        "page": page,
        "per_page": per_page,
        "raw": {"start_date": (data.get("start_date") or "").strip(), "end_date": (data.get("end_date") or "").strip()},
    }


@login_required
@require_GET
def trim_overview_api(request):
    """
    GET /api/mes/trim/overview
    Query:
      start_date=YYYY-MM-DD&amp;end_date=YYYY-MM-DD
      &brand=gwm&amp;brand=chery
      &model=Tiggo%204
      &line=gwm|chery|changan
      &vin=VIN
    Returns aggregated counts for TRIM: overall and by lines.
    """
    f = _parse_trim_filters(request)

    payload = get_trim_mes(
        start=f["start"], end=f["end"],
        line=f["line"], brands=f["brands"], models=f["models"], vin=f["vin"]
    ) or {}

    # Controls whether to include items (VIN lists) in response
    include_items = (request.GET.get("include_items") or "").strip().lower() in {"1", "true", "yes", "on"}

    overall_raw = payload.get("overall", {}) or {}
    by_line_raw = payload.get("by_line", {}) or {}

    # Normalize overall: always return counts; optionally include items
    metrics = ("trim_in", "wip_trim", "trim_out")
    overall_norm: dict = {}
    for m in metrics:
        node = overall_raw.get(m, {}) or {}
        # node may be a dict with "count"/"items" or a bare number
        if isinstance(node, dict):
            cnt = int(node.get("count", 0) or 0)
            if include_items:
                overall_norm[m] = {"count": cnt, "items": list(node.get("items", []) or [])}
            else:
                overall_norm[m] = {"count": cnt}
        else:
            # bare number fallback
            cnt = int(node or 0)
            overall_norm[m] = {"count": cnt} if not include_items else {"count": cnt, "items": []}

    # Compact per-line counts for UI (backward compatible: numbers only)
    lines = {ln: {"trim_in": 0, "wip_trim": 0, "trim_out": 0} for ln in ALLOWED_LINES}
    try:
        for ln in ALLOWED_LINES:
            lines[ln]["trim_in"]  = int(((by_line_raw.get("trim_in", {})  or {}).get(ln, {}) or {}).get("count", 0) or 0)
            lines[ln]["wip_trim"] = int(((by_line_raw.get("wip_trim", {}) or {}).get(ln, {}) or {}).get("count", 0) or 0)
            lines[ln]["trim_out"] = int(((by_line_raw.get("trim_out", {}) or {}).get(ln, {}) or {}).get("count", 0) or 0)
    except Exception:
        # fail-safe: keep zeros if structure differs
        pass

    # Optional detailed per-line VIN lists (non-breaking extra field)
    if include_items:
        lines_items = {ln: {"trim_in": [], "wip_trim": [], "trim_out": []} for ln in ALLOWED_LINES}
        try:
            for ln in ALLOWED_LINES:
                lines_items[ln]["trim_in"]  = list((((by_line_raw.get("trim_in", {})  or {}).get(ln, {}) or {}).get("items", []) or []))
                lines_items[ln]["wip_trim"] = list((((by_line_raw.get("wip_trim", {}) or {}).get(ln, {}) or {}).get("items", []) or []))
                lines_items[ln]["trim_out"] = list((((by_line_raw.get("trim_out", {}) or {}).get(ln, {}) or {}).get("items", []) or []))
        except Exception:
            pass

    resp = {
        "filters": {
            "brands": f["brands"] or [],
            "models": f["models"] or [],
            "line": f["line"] or "",
            "vin": f["vin"] or "",
            "start_date": f["raw"]["start_date"],
            "end_date": f["raw"]["end_date"],
        },
        "overall": overall_norm,
        "by_line": lines,
    }
    if include_items:
        resp["by_line_items"] = lines_items

    return JsonResponse(resp, json_dumps_params={"ensure_ascii": False})


@login_required
@require_GET
def trim_list_api(request):
    """
    GET /api/mes/trim/list
    Query:
      metric=trim_in|wip_trim|trim_out   (required)
      start_date, end_date, brand, model, line, vin
      page (default 1), per_page (default 100)
    Returns a paginated list of VIN items for the chosen metric.
    """
    metric = (request.GET.get("metric") or "").strip().lower()
    if metric not in {"trim_in", "wip_trim", "trim_out"}:
        return JsonResponse({"ok": False, "error": "metric must be trim_in|wip_trim|trim_out"}, status=400)

    f = _parse_trim_filters(request)
    payload = get_trim_mes(
        start=f["start"], end=f["end"],
        line=f["line"], brands=f["brands"], models=f["models"], vin=f["vin"]
    ) or {}

    items = list(((payload.get("overall", {}).get(metric, {}) or {}).get("items", []) or []))

    # Optional explicit line filter for UI (safe even if aggregator already handled this)
    if f["line"]:
        items = [it for it in items if (it.get("line") or "").strip().lower() == f["line"]]

    total = len(items)
    start_idx = (f["page"] - 1) * f["per_page"]
    end_idx = start_idx + f["per_page"]
    page_items = items[start_idx:end_idx]

    return JsonResponse({
        "filters": {
            "metric": metric,
            "brands": f["brands"] or [],
            "models": f["models"] or [],
            "line": f["line"] or "",
            "vin": f["vin"] or "",
            "start_date": f["raw"]["start_date"],
            "end_date": f["raw"]["end_date"],
        },
        "pagination": {"page": f["page"], "per_page": f["per_page"], "total": total},
        "items": page_items,
    }, json_dumps_params={"ensure_ascii": False})


@login_required
@permission_required('users.access_to_indicators', raise_exception=True)
def mes_metric_table_view(request):
    """
    HTML-страница с динамической таблицей по MES-метрике.
    Тянет данные из /api/mes/table с теми же GET-параметрами.
    """
    section = (request.GET.get("section") or "trim").strip().lower()
    allow_line = (section == "trim")

    f = _parse_mes_filters(request, allow_line=allow_line)

    default_metric = {
        "trim": "trim_in",
        "bqa":  "bqa_in",
        "qa":   "qa_in",
        "uud":  "uud_in",
        "ves":  "ves_in",
    }.get(section, "trim_in")

    metric = (request.GET.get("metric") or default_metric).strip().lower()

    try:
        api_table_url = reverse("mes_metric_table_api")
    except Exception:
        api_table_url = "/api/mes/table"

    try:
        api_table_export_url = reverse("mes_metric_table_export")
    except Exception:
        api_table_export_url = "/api/mes/table/export"

    ctx = {
        "section": section,
        "metric": metric,
        "api_table_url": api_table_url,
        "api_table_export_url": api_table_export_url,
        "filters": {
            "start_date": f["start_date"],
            "end_date":   f["end_date"],
            "brands":     f["brands"],
            "models":     f["models"],
            "vin":        f["vin"] or "",
            "line":       f["line"] or "",
            "page":       f["page"],
            "per_page":   f["per_page"],
        },
    }
    return render(request, "users/mes/metric_table.html", ctx)









@login_required
@require_GET
def mes_metric_table_export(request):
    """
    GET /api/mes/table/export
      Пробрасывает те же параметры, что и /api/mes/table,
      и возвращает Excel (.xlsx) с теми же колонками и данными (без пагинации).
    """
    section = (request.GET.get("section") or "trim").strip().lower()
    allow_line = (section == "trim")
    f = _parse_mes_filters(request, allow_line=allow_line)

    # Получаем расчёт по секции
    calc_map = {
        "trim": lambda: get_trim_mes(start=f["start"], end=f["end"], line=f["line"], brands=f["brands"] or None, models=f["models"] or None, vin=f["vin"]),
        "bqa":  lambda: get_bqa_mes(start=f["start"], end=f["end"], line=f["line"], brands=f["brands"] or None, models=f["models"] or None, vin=f["vin"]),
        "qa":   lambda: get_qa_mes(start=f["start"], end=f["end"], brands=f["brands"] or None, models=f["models"] or None, vin=f["vin"]),
        "uud":  lambda: get_uud_mes(start=f["start"], end=f["end"], brands=f["brands"] or None, models=f["models"] or None, vin=f["vin"]),
        "ves":  lambda: get_ves_mes(start=f["start"], end=f["end"], brands=f["brands"] or None, models=f["models"] or None, vin=f["vin"]),
    }
    if section not in calc_map:
        section = "trim"

    try:
        res = calc_map[section]() or {}
    except Exception:
        res = {}

    overall = res.get("overall", {}) or {}
    metric = (request.GET.get("metric") or "").strip().lower()
    key = _resolve_metric_key(section, metric, overall)
    if not key:
        return JsonResponse({"ok": False, "error": "unknown metric", "available": sorted(list(overall.keys()))}, status=400)

    bucket = overall.get(key, {}) or {}
    items = list(bucket.get("items", []) or [])

    # Явная фильтрация по линии для TRIM
    if allow_line and f["line"]:
        items = [it for it in items if (it.get("line") or "").strip().lower() == f["line"]]

    # Колонки по полному набору items
    columns = _infer_columns_from_items(items)

    # === Генерация Excel ===
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, Alignment as XLAlignment

    wb = Workbook()
    ws = wb.active
    ws.title = "Показатели"

    # Заголовок + метаданные фильтров
    title = f"Показатели {section.upper()} — {key}"
    ws["A1"] = title
    ws["A1"].font = Font(bold=True)
    ws.append([])  # пустая строка

    # строка с периодом/фильтрами (в одну строку для простоты)
    meta = f"Период: {f['start_date']} … {f['end_date']}"
    if f["brands"]:
        meta += f" | Бренды: {', '.join(f['brands'])}"
    if f["models"]:
        meta += f" | Модели: {', '.join(f['models'])}"
    if f["vin"]:
        meta += f" | VIN: {f['vin']}"
    if allow_line and f["line"]:
        meta += f" | Линия: {f['line']}"
    ws.append([meta])
    ws.append([])

    # Заголовки таблицы
    header_row = [c.get("title") or c.get("key") for c in columns]
    ws.append(header_row)
    for i in range(1, len(header_row) + 1):
        cell = ws.cell(row=ws.max_row, column=i)
        cell.font = Font(bold=True)
        cell.alignment = XLAlignment(horizontal="center")

    # Данные
    keys = [c["key"] for c in columns]
    for it in items:
        row = []
        for k in keys:
            v = it.get(k, "")
            # Преобразуем списки/сложные типы в строку
            if isinstance(v, (list, tuple, set)):
                v = ", ".join(str(x) for x in v)
            elif isinstance(v, dict):
                v = str(v)
            row.append(v if v is not None else "")
        ws.append(row)

    # Автоширина по максимальной длине в колонке (ограничим 60 символами)
    for idx, _ in enumerate(header_row, start=1):
        max_len = max((len(str(ws.cell(row=r, column=idx).value or "")) for r in range(1, ws.max_row + 1)), default=10)
        ws.column_dimensions[get_column_letter(idx)].width = min(max(10, max_len + 2), 60)

    # Ответ
    from django.http import HttpResponse
    filename = f"mes_{section}_{key}_{f['start_date']}_{f['end_date']}.xlsx".replace("..", ".").replace("/", "-")
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response








# UUD SHUCKRAT

from supplies.models import TraceData
from vehicle_history.models import VINHistory
from django.db.models import Q
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from django.utils import timezone as dj_tz
from openpyxl.styles import Alignment, Font, PatternFill




@login_required
@permission_required('users.access_to_the_uud_table', raise_exception=True)
def uud_report(request):
    """
    Вкладки формируются по steps ПОСЛЕДНЕЙ сессии:
      - step1 (Отдали на УУД) — без учёта периода (все застрявшие на step1)
      - step2/step3/done — по своим *_at в выбранном диапазоне.
    defects_by_vin содержит все дефекты без фильтра по датам и теперь включает 'comment'
    (приоритет: comment → custom_defect_explanation → custom_detail_explanation).
    """
    UUD_ZONE = "УУД"
    UUD_POST = "УУД"

    def _parse_dt(dt_str):
        if not dt_str:
            return None
        try:
            s = str(dt_str).replace("Z", "+00:00")
            dt = datetime.fromisoformat(s)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            return dt
        except Exception:
            return None

    def _date_or_none(dt):
        try:
            return dt.date()
        except Exception:
            return None

    def _brand_model(vin: str):
        if not vin:
            return "", ""
        t = (TraceData.objects
             .filter(Q(vin_rk__iexact=vin) | Q(vin_c__iexact=vin))
             .order_by("-date_added")
             .only("brand", "model")
             .first())
        return ((t.brand or "") if t else "", (t.model or "") if t else "")

    def _safe_entry_index(s: dict) -> int:
        idx = s.get("entry_index")
        if isinstance(idx, int):
            return idx
        sid = s.get("id") or ""
        try:
            tail = str(sid).rsplit("-", 1)[-1]
            return int(tail)
        except Exception:
            return 0

    def _last_session(sessions: list) -> dict | None:
        if not sessions:
            return None

        def keyer(s):
            return (_safe_entry_index(s),
                    _parse_dt(s.get("created_at")) or datetime.min.replace(tzinfo=timezone.utc))
        return max((s for s in sessions if isinstance(s, dict)), key=keyer, default=None)

    def _comment_from_defect(d: dict) -> str:
        """comment → custom_defect_explanation → custom_detail_explanation (если обе кастомные есть — берём custom_defect_explanation)"""
        if not isinstance(d, dict):
            return ""
        comment = (d.get("comment") or "").strip()
        if comment:
            return comment
        c_def = (d.get("custom_defect_explanation") or "").strip()
        c_det = (d.get("custom_detail_explanation") or "").strip()
        if c_def and c_det:
            return c_def
        return c_def or c_det or ""

    def _iter_defects_from_history(history: dict):
        """
        yield dict(zone, post, status, detail, defect, grade, date, time, by, comment)
        - Сборка:  detail <- unit,  defect <- name
        - Поставки: detail <- detail, defect <- defect
        - status из extra.UUD.status (если есть)
        - comment приоритет: comment → custom_defect_explanation → custom_detail_explanation
        - date/time/by из entry.date_added / entry.controller
        """
        if not isinstance(history, dict):
            return

        # --- Цех сборки
        assembly = history.get("Цех сборки") or {}
        if isinstance(assembly, str):
            try:
                assembly = json.loads(assembly)
            except Exception:
                assembly = {}
        if isinstance(assembly, dict):
            for post_name, entries in assembly.items():
                if not isinstance(entries, list):
                    continue
                for entry in entries:
                    controller = (entry.get("controller") or "").strip()
                    dt = _parse_dt(entry.get("date_added"))
                    for d in (entry.get("defects") or []):
                        extra = d.get("extra") or {}
                        uud = extra.get("UUD") or {}
                        yield {
                            "zone":   "Цех сборки",
                            "post":   post_name,
                            "status": (uud.get("status") or "").strip(),
                            "detail": (d.get("unit") or "").strip(),
                            "defect": (d.get("name") or "").strip(),
                            "grade":  (d.get("grade") or "").strip(),
                            "date":   (dt.date().isoformat() if dt else ""),
                            "time":   (dt.strftime("%H:%M") if dt else ""),
                            "by":     controller,
                            "comment": _comment_from_defect(d),
                        }

        # --- Цех поставки
        supplies = history.get("Цех поставки") or {}
        if isinstance(supplies, str):
            try:
                supplies = json.loads(supplies)
            except Exception:
                supplies = {}
        if isinstance(supplies, dict):
            for post_name, entries in supplies.items():
                if not isinstance(entries, list):
                    continue
                for entry in entries:
                    controller = (entry.get("controller") or "").strip()
                    dt = _parse_dt(entry.get("date_added"))
                    for d in (entry.get("defects") or []):
                        extra = d.get("extra") or {}
                        uud = extra.get("UUD") or {}
                        yield {
                            "zone":   "Цех поставки",
                            "post":   post_name,
                            "status": (uud.get("status") or "").strip(),
                            "detail": (d.get("detail") or "").strip(),
                            "defect": (d.get("defect") or "").strip(),
                            "grade":  (d.get("grade") or "").strip(),
                            "date":   (dt.date().isoformat() if dt else ""),
                            "time":   (dt.strftime("%H:%M") if dt else ""),
                            "by":     controller,
                            "comment": _comment_from_defect(d),
                        }

    # ---- период (для step2/3/4 и заголовка; step1 его не учитывает) ----
    q_date  = (request.GET.get("date")  or "").strip()
    q_start = (request.GET.get("start") or "").strip()
    q_end   = (request.GET.get("end")   or "").strip()

    if q_start and q_end:
        try:
            start_date = datetime.strptime(q_start, "%Y-%m-%d").date()
            end_date   = datetime.strptime(q_end,   "%Y-%m-%d").date()
            if end_date < start_date:
                start_date, end_date = end_date, start_date
        except ValueError:
            start_date = end_date = timezone.now().date()
        date_label = f"{start_date.isoformat()} … {end_date.isoformat()}"
    else:
        if q_date:
            try:
                start_date = end_date = datetime.strptime(q_date, "%Y-%m-%d").date()
            except ValueError:
                start_date = end_date = timezone.now().date()
        else:
            start_date = end_date = timezone.now().date()
        date_label = start_date.isoformat()

    def _in_range(dt):
        d = _date_or_none(dt)
        return (dt and d and start_date <= d <= end_date)

    # ---- сбор данных ----
    rows1, rows2, rows3, rows4 = [], [], [], []
    active_now = 0
    defects_by_vin = {}

    for vh in VINHistory.objects.only("vin", "history"):
        vin = (vh.vin or "").strip().upper()
        if not vin:
            continue

        history = vh.history or {}
        if isinstance(history, str):
            try:
                history = json.loads(history)
            except Exception:
                continue

        # все дефекты (без фильтра дат), теперь с comment
        vin_defects = list(_iter_defects_from_history(history))
        vin_defects.sort(key=lambda r: (r["date"], r["time"]), reverse=True)
        defects_by_vin[vin] = vin_defects

        # последняя сессия УУД
        zone = history.get(UUD_ZONE) or {}
        if isinstance(zone, str):
            try:
                zone = json.loads(zone)
            except Exception:
                zone = {}
        sessions = zone.get(UUD_POST) or []
        if not isinstance(sessions, list) or not sessions:
            continue

        if any((s.get("steps") or "").strip().lower() != "done" for s in sessions if isinstance(s, dict)):
            active_now += 1

        last = _last_session(sessions)
        if not last:
            continue

        extra = last.get("extra_data") or {}
        steps_raw = (last.get("steps") or "").strip().lower()
        brand, model = _brand_model(vin)

        t1 = _parse_dt(extra.get("step1_at"))
        t2 = _parse_dt(extra.get("step2_at"))
        t3 = _parse_dt(extra.get("step3_at"))
        t4 = _parse_dt(extra.get("step4_at"))

        if steps_raw == "step1":
            rows1.append({
                "vin": vin, "brand": brand, "model": model,
                "date": t1.date().isoformat() if t1 else "",
                "time": t1.strftime("%H:%M") if t1 else "",
                "by":   (extra.get("step1_by") or "").strip(),
            })
        elif steps_raw == "step2" and _in_range(t2):
            rows2.append({
                "vin": vin, "brand": brand, "model": model,
                "date": t2.date().isoformat() if t2 else "",
                "time": t2.strftime("%H:%M") if t2 else "",
                "by":   (extra.get("step2_by") or "").strip(),
            })
        elif steps_raw == "step3" and _in_range(t3):
            rows3.append({
                "vin": vin, "brand": brand, "model": model,
                "date": t3.date().isoformat() if t3 else "",
                "time": t3.strftime("%H:%M") if t3 else "",
                "by":   (extra.get("step3_by") or "").strip(),
            })
        elif steps_raw in ("done", "step4") and _in_range(t4):
            rows4.append({
                "vin": vin, "brand": brand, "model": model,
                "date": t4.date().isoformat() if t4 else "",
                "time": t4.strftime("%H:%M") if t4 else "",
                "by":   (extra.get("step4_by") or "").strip(),
            })

    # сортировки
    for rows in (rows1, rows2, rows3, rows4):
        rows.sort(key=lambda r: (r["date"], r["time"]), reverse=True)

    payload = {
        "date_local": date_label,
        "start_date": start_date.isoformat(),
        "end_date": end_date.isoformat(),
        "step1_rows": rows1, "step1_count": len(rows1),
        "step2_rows": rows2, "step2_count": len(rows2),
        "step3_rows": rows3, "step3_count": len(rows3),
        "step4_rows": rows4, "step4_count": len(rows4),
        "active_now": active_now,
        "defects_by_vin": defects_by_vin,  # включает comment
    }

    wants_json = (request.headers.get("X-Requested-With") == "XMLHttpRequest"
                  or request.GET.get("format") == "json")
    if wants_json:
        return JsonResponse(payload, status=200, safe=False)
    return render(request, "users/uud/uud_report.html", payload)


@login_required
@permission_required('users.access_to_the_uud_table', raise_exception=True)
def uud_report_export(request):
    """
    Экспорт Excel:
      - «Ожидают ремонта» (step1) — без периода (все застрявшие на step1 по последней сессии)
      - «УУД начала ремонт» (step2), «Ждёт приёма на линию» (step3), «Линия приняла» (done) — по своему *_at в диапазоне.
      - Колонка «Комментарий» по приоритету: comment → custom_defect_explanation → custom_detail_explanation.
      - Учитывает фильтры: VIN (q), Бренд (brand[]), Модель (model[]).
    """
    UUD_ZONE = "УУД"
    UUD_POST = "УУД"

    # ---------- helpers ----------
    def _parse_dt(dt_str):
        if not dt_str:
            return None
        try:
            s = str(dt_str).replace("Z", "+00:00")
            dt = datetime.fromisoformat(s)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=dj_tz.utc)
            return dt
        except Exception:
            return None

    def _date_or_none(dt):
        try:
            return dt.date()
        except Exception:
            return None

    def _brand_model(vin):
        if not vin:
            return "", ""
        t = (TraceData.objects
             .filter(Q(vin_rk__iexact=vin) | Q(vin_c__iexact=vin))
             .order_by("-date_added")
             .only("brand", "model")
             .first())
        return ((t.brand or "") if t else "", (t.model or "") if t else "")

    def _safe_entry_index(s: dict) -> int:
        idx = s.get("entry_index")
        if isinstance(idx, int):
            return idx
        sid = s.get("id") or ""
        try:
            tail = str(sid).rsplit("-", 1)[-1]
            return int(tail)
        except Exception:
            return 0

    def _last_session(sessions: list) -> dict | None:
        def keyer(s):
            return (_safe_entry_index(s),
                    _parse_dt(s.get("created_at")) or datetime.min.replace(tzinfo=dj_tz.utc))
        return max((s for s in sessions if isinstance(s, dict)), key=keyer, default=None)

    def _comment_from_defect(d: dict) -> str:
        comment = (d.get("comment") or "").strip()
        if comment:
            return comment
        c_def = (d.get("custom_defect_explanation") or "").strip()
        c_det = (d.get("custom_detail_explanation") or "").strip()
        if c_def and c_det:
            return c_def
        return c_def or c_det or ""

    # нормализованные дефекты (включая comment)
    def _iter_defects_from_history(history: dict):
        if not isinstance(history, dict):
            return

        # Цех сборки
        assembly = history.get("Цех сборки") or {}
        if isinstance(assembly, str):
            try:
                assembly = json.loads(assembly)
            except Exception:
                assembly = {}
        if isinstance(assembly, dict):
            for post_name, entries in assembly.items():
                if not isinstance(entries, list):
                    continue
                for entry in entries:
                    controller = (entry.get("controller") or "").strip()
                    dt = _parse_dt(entry.get("date_added"))
                    for d in (entry.get("defects") or []):
                        extra = d.get("extra") or {}
                        uud = extra.get("UUD") or {}
                        yield {
                            "zone": "Цех сборки",
                            "post": post_name,
                            "status": (uud.get("status") or "").strip(),
                            "detail": (d.get("unit") or "").strip(),
                            "defect": (d.get("name") or "").strip(),
                            "grade":  (d.get("grade") or "").strip(),
                            "date":   (dt.date().isoformat() if dt else ""),
                            "time":   (dt.strftime("%H:%M") if dt else ""),
                            "by":     controller,
                            "comment": _comment_from_defect(d),
                        }

        # Цех поставки
        supplies = history.get("Цех поставки") or {}
        if isinstance(supplies, str):
            try:
                supplies = json.loads(supplies)
            except Exception:
                supplies = {}
        if isinstance(supplies, dict):
            for post_name, entries in supplies.items():
                if not isinstance(entries, list):
                    continue
                for entry in entries:
                    controller = (entry.get("controller") or "").strip()
                    dt = _parse_dt(entry.get("date_added"))
                    for d in (entry.get("defects") or []):
                        extra = d.get("extra") or {}
                        uud = extra.get("UUD") or {}
                        yield {
                            "zone": "Цех поставки",
                            "post": post_name,
                            "status": (uud.get("status") or "").strip(),
                            "detail": (d.get("detail") or "").strip(),
                            "defect": (d.get("defect") or "").strip(),
                            "grade":  (d.get("grade") or "").strip(),
                            "date":   (dt.date().isoformat() if dt else ""),
                            "time":   (dt.strftime("%H:%M") if dt else ""),
                            "by":     controller,
                            "comment": _comment_from_defect(d),
                        }

    # ---------- фильтры из запроса ----------
    q = (request.GET.get("q") or "").strip().upper()

    # поддержим и brand=... и brand[]=...
    def _get_multi(name: str):
        vals = request.GET.getlist(name)
        if not vals:
            vals = request.GET.getlist(f"{name}[]")
        # нормализация (сохраняем пустые строки как валидные значения)
        return [ (v or "").strip() for v in vals if v is not None ]

    brands_raw = _get_multi("brand")
    models_raw = _get_multi("model")

    brands_set = set(brands_raw) if brands_raw else None
    models_set = set(models_raw) if models_raw else None

    # ---------- период ----------
    q_date  = (request.GET.get("date")  or "").strip()
    q_start = (request.GET.get("start") or "").strip()
    q_end   = (request.GET.get("end")   or "").strip()

    if q_start and q_end:
        try:
            start_date = datetime.strptime(q_start, "%Y-%m-%d").date()
            end_date   = datetime.strptime(q_end,   "%Y-%m-%d").date()
            if end_date < start_date:
                start_date, end_date = end_date, start_date
        except ValueError:
            start_date = end_date = dj_tz.now().date()
        date_label = f"{start_date.isoformat()} … {end_date.isoformat()}"
        filename_suffix = f"{start_date.isoformat()}_{end_date.isoformat()}"
    else:
        if q_date:
            try:
                start_date = end_date = datetime.strptime(q_date, "%Y-%m-%d").date()
            except ValueError:
                start_date = end_date = dj_tz.now().date()
        else:
            start_date = end_date = dj_tz.now().date()
        date_label = start_date.isoformat()
        filename_suffix = start_date.isoformat()

    # строки для листов: (vin, brand, model, date, time, by, defects_list)
    sheets_data = {"step1": [], "step2": [], "step3": [], "done": []}

    def _in_range(dt):
        d = _date_or_none(dt)
        return (dt and d and start_date <= d <= end_date)

    # ---------- сбор данных ----------
    for vh in VINHistory.objects.only("vin", "history"):
        vin = (vh.vin or "").strip().upper()
        if not vin:
            continue

        history = vh.history or {}
        if isinstance(history, str):
            try:
                history = json.loads(history)
            except Exception:
                continue

        defects = list(_iter_defects_from_history(history))
        defects.sort(key=lambda r: (r["date"], r["time"]), reverse=True)

        zone = history.get(UUD_ZONE) or {}
        if isinstance(zone, str):
            try:
                zone = json.loads(zone)
            except Exception:
                zone = {}
        sessions = zone.get(UUD_POST) or []
        if not isinstance(sessions, list) or not sessions:
            continue

        last = _last_session(sessions)
        if not last:
            continue

        extra = last.get("extra_data") or {}
        steps_raw = (last.get("steps") or "").strip().lower()
        brand, model = _brand_model(vin)

        t1 = _parse_dt(extra.get("step1_at"))
        t2 = _parse_dt(extra.get("step2_at"))
        t3 = _parse_dt(extra.get("step3_at"))
        t4 = _parse_dt(extra.get("step4_at"))

        row1 = (vin, brand, model, t1.date().isoformat() if t1 else "", t1.strftime("%H:%M") if t1 else "",
                (extra.get("step1_by") or "").strip(), defects)
        row2 = (vin, brand, model, t2.date().isoformat() if t2 else "", t2.strftime("%H:%M") if t2 else "",
                (extra.get("step2_by") or "").strip(), defects)
        row3 = (vin, brand, model, t3.date().isoformat() if t3 else "", t3.strftime("%H:%M") if t3 else "",
                (extra.get("step3_by") or "").strip(), defects)
        row4 = (vin, brand, model, t4.date().isoformat() if t4 else "", t4.strftime("%H:%M") if t4 else "",
                (extra.get("step4_by") or "").strip(), defects)

        if steps_raw == "step1":
            sheets_data["step1"].append(row1)
        elif steps_raw == "step2" and _in_range(t2):
            sheets_data["step2"].append(row2)
        elif steps_raw == "step3" and _in_range(t3):
            sheets_data["step3"].append(row3)
        elif steps_raw in ("done", "step4") and _in_range(t4):
            sheets_data["done"].append(row4)

    for k in sheets_data:
        sheets_data[k].sort(key=lambda r: (r[3], r[4]), reverse=True)

    # ---------- общий предикат фильтрации (VIN/Бренд/Модель) ----------
    def _row_passes_filters(row):
        # row = (vin, brand, model, date, time, by, defects)
        vin_val   = (row[0] or "").upper()
        brand_val = row[1] or ""
        model_val = row[2] or ""
        if q and q not in vin_val:
            return False
        if brands_set is not None and brand_val not in brands_set:
            return False
        if models_set is not None and model_val not in models_set:
            return False
        return True

    # применяем фильтры ДО записи на листы
    for key in ("step1", "step2", "step3", "done"):
        sheets_data[key] = [r for r in sheets_data[key] if _row_passes_filters(r)]

    # ---------- workbook ----------
    wb = Workbook()
    ws_step1 = wb.active
    ws_step1.title = "Ожидают ремонта"
    ws_step2 = wb.create_sheet("УУД начала ремонт")
    ws_step3 = wb.create_sheet("Ждёт приёма на линию")
    ws_done  = wb.create_sheet("Линия приняла")

    STATUS_MAP = {
        "impossible":   ("Невозможно",   "000000", "FFFFFF"),
        "resolved":     ("Устранено",    "198754", "FFFFFF"),
        "not_resolved": ("Не устранено", "DC3545", "FFFFFF"),
        "checking":     ("Проверяется",  "FFC107", "000000"),
        "":             ("—",            "E9ECEF", "6C757D"),
        None:           ("—",            "E9ECEF", "6C757D"),
    }

    def _write_sheet(ws, data_rows):
        # Период на всю ширину (1..10), счётчик в кол. 11
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
        c = ws.cell(row=1, column=1, value=f"Период: {date_label}")
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")
        ws.cell(row=1, column=11, value=f"Количество: {len(data_rows)}").font = Font(bold=True)

        headers = [
            "VIN", "Бренд", "Модель", "Дата (УУД)", "Время (УУД)", "Контролер (УУД)",
            "Статус дефекта", "Деталь", "Дефект", "Комментарий"
        ]
        ws.append(headers)

        cur = 3
        for vin, brand, model, date_s, time_s, by, defects in data_rows:
            dlist = defects if defects else [None]
            group_len = max(1, len(dlist))
            start_row = cur

            for i, d in enumerate(dlist):
                row = [
                    vin   if i == 0 else "",
                    brand if i == 0 else "",
                    model if i == 0 else "",
                    date_s if i == 0 else "",
                    time_s if i == 0 else "",
                    by if i == 0 else "",
                ]

                if d:
                    status_raw = (d.get("status") or "").strip() or ""
                    title, bg, fg = STATUS_MAP.get(status_raw, STATUS_MAP[""])
                    detail = d.get("detail") or "—"
                    defect = d.get("defect") or "—"
                    comment = d.get("comment") or "—"
                else:
                    title, bg, fg = STATUS_MAP[""]
                    detail = "—"; defect = "—"; comment = "—"

                row.extend([title, detail, defect, comment])
                ws.append(row)

                # оформление статуса
                status_cell = ws.cell(row=cur, column=7)
                status_cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
                status_cell.font = Font(color=fg, bold=True)
                status_cell.alignment = Alignment(horizontal="center")

                # перенос в комментарии
                ws.cell(row=cur, column=10).alignment = Alignment(wrap_text=True, vertical="top")

                cur += 1

            # merge первых 6 колонок (VIN..Контролер) на длину группы
            if group_len > 1:
                for col in range(1, 7):
                    ws.merge_cells(start_row=start_row, start_column=col,
                                   end_row=start_row + group_len - 1, end_column=col)
                    top_cell = ws.cell(row=start_row, column=col)
                    top_cell.alignment = Alignment(horizontal="center", vertical="center")

        # автоширина
        for col_cells in ws.columns:
            col_idx = col_cells[0].column
            letter = get_column_letter(col_idx)
            max_len = 0
            for c in col_cells:
                if c.value is not None:
                    max_len = max(max_len, len(str(c.value)))
            ws.column_dimensions[letter].width = min(max_len + 2, 60)

        # примечание (1..10)
        note_row = ws.max_row + 2
        ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=10)
        nc = ws.cell(row=note_row, column=1, value=(
            "Данная информация актуальна только на момент скачивания — "
            f"{dj_tz.localtime(dj_tz.now()).strftime('%Y-%m-%d %H:%M')}"
        ))
        nc.font = Font(italic=True, color="6C757D")
        nc.alignment = Alignment(horizontal="left", vertical="center")

    _write_sheet(ws_step1, sheets_data["step1"])
    _write_sheet(ws_step2, sheets_data["step2"])
    _write_sheet(ws_step3, sheets_data["step3"])
    _write_sheet(ws_done,  sheets_data["done"])

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    resp = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="uud_report_{filename_suffix}.xlsx"'
    return resp





