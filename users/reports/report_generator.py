# reports/report_generator.py
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from collections import defaultdict
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from django.utils.dateparse import parse_date
from datetime import datetime
from io import BytesIO
import os

from supplies.models import TraceData
from vehicle_history.models import VINHistory


def write_cell(ws, cell, value):
    """Безопасная запись даже в объединённые ячейки"""
    if ws[cell].__class__.__name__ == "MergedCell":
        for merged_range in ws.merged_cells.ranges:
            if cell in merged_range:
                top_left = merged_range.bounds[:2]  # (min_col, min_row)
                fixed_cell = f"{get_column_letter(top_left[0])}{top_left[1]}"
                ws[fixed_cell] = value
                return
    else:
        ws[cell] = value


def generate_summary_report(brand, start_date, end_date, template_dir="users/reports/templates", save_path=None):
    zone = "Цех сборки"
    post = "Финал"

    brand = brand or ""
    brands_to_query = ["Haval", "Tank"] if brand.lower() == "gwm" else [brand.capitalize()]

    BRAND_MODEL_COLUMN_MAP = {
        "gwm": {
            "E": ["Jolion"],
            "G": ["Tank 300"],
            "I": ["H6"],
            "K": ["Haval M6"],
            "M": ["H5"],
            "O": ["H9"],
            "Q": ["Haval Dargo"],
            "S": ["Haval H6"],
            "T": ["Tank 500"],
        },
        "changan": {
            "E": ["Alsvin"],
            "G": ["UNI-V"],
            "I": ["UNI-K"],
            "K": ["cs55plus"],
            "M": ["CS75 Plus"],
            "O": ["CS95"],
        },
        "chery": {
            "E": ["Tiggo 2 PRO"],
            "G": ["Tiggo 4"],
            "I": ["Tiggo 7"],
            "K": ["Tiggo 8 Pro"],
            "M": ["Tiggo 8"],
            "O": ["Tiggo 9"],
            "Q": ["Arrizo 6"],
            "S": ["Arrizo 8"],
        }
    }

    print("📋 Поисковый бренд:", brands_to_query)
    trace_qs = TraceData.objects.filter(
        brand__in=brands_to_query,
        date_added__date__range=(start_date, end_date)
    )

    vin_to_model = {}
    vins_by_model = defaultdict(set)
    print("🔍 vins_by_model keys:", list(vins_by_model.keys()))
    for item in trace_qs:
        norm_model = item.model.strip().lower().replace(" ", "")
        norm_vin = item.vin_rk.strip().lower().replace(" ", "")
        vins_by_model[norm_model].add(norm_vin)
        vin_to_model[norm_vin] = norm_model
        print("📄 model from TraceData:", repr(item.model))

    histories = VINHistory.objects.filter(vin__in=trace_qs.values_list("vin_rk", flat=True))

    template_path = os.path.join(template_dir, f"{brand.lower()}_summary_report.xlsx")
    wb = load_workbook(template_path)
    ws = wb.active

    total_col_letter = "W"
    overall_total_inspections = 0
    overall_total_defects = 0

    column_model_map = BRAND_MODEL_COLUMN_MAP.get(brand.lower(), {})

    for col, model_variants in column_model_map.items():
        model_key = model_variants[0].strip().lower().replace(" ", "")
        vin_set = vins_by_model.get(model_key, set())

        print(f"📦 Модель: {model_key} → VIN-ов: {len(vin_set)}")

        if not vin_set:
            continue

        total_inspections = 0
        total_defects = 0
        vins_with_defects = set()

        for history in histories:
            try:
                norm_vin = history.vin.strip().lower().replace(" ", "")
                if norm_vin not in vin_set:
                    continue

                post_entries = history.history.get(zone, {}).get(post, [])
                valid_entries = [
                    entry for entry in post_entries
                    if "date_added" in entry and start_date <= parse_date(entry["date_added"][:10]) <= end_date
                ]

                if not valid_entries:
                    continue

                total_inspections += 1
                has_defect = any(entry.get("has_defect") == "yes" for entry in valid_entries)
                if has_defect:
                    vins_with_defects.add(norm_vin)
                    for entry in valid_entries:
                        total_defects += len(entry.get("defects", []))
            except Exception as e:
                print(f"⚠️ Ошибка при обработке VIN {history.vin}: {e}")
                continue

        dpu = round(total_defects / total_inspections, 2) if total_inspections else 0
        str_percent = 0  # 🔧 Пока просто 0, как ты просил

        write_cell(ws, f"{col}7", total_inspections)
        write_cell(ws, f"{col}8", total_defects)
        write_cell(ws, f"{col}10", dpu)
        write_cell(ws, f"{col}11", str_percent)

        overall_total_inspections += total_inspections
        overall_total_defects += total_defects

    overall_dpu = round(overall_total_defects / overall_total_inspections, 2) if overall_total_inspections else 0

    write_cell(ws, f"{total_col_letter}7", overall_total_inspections)
    write_cell(ws, f"{total_col_letter}8", overall_total_defects)
    write_cell(ws, f"{total_col_letter}9", overall_dpu)
    write_cell(ws, f"{total_col_letter}10", 0)  # 🔧 STR = 0
    write_cell(ws, "P1", start_date.strftime("%Y-%m-%d"))
    write_cell(ws, "S1", end_date.strftime("%Y-%m-%d"))

    if not save_path:
        save_path = f"summary_{brand}_{start_date}_{end_date}.xlsx"

    wb.save(save_path)
    return save_path